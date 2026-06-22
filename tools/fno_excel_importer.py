#!/usr/bin/env python3
"""
fno_excel_importer.py  -  Trendlyne F&O build-up Excel/CSV -> fno_signals
========================================================================
Reads a Trendlyne F&O data export (the daily futures sheet) and writes the
STRONG build-up rows into paper/events.db -> fno_signals, which the already-
running fno_paper_study.py consumes and paper-trades. The F&O bot is NOT
modified; this only puts signals in front of it.

Build-up -> direction (strong build-ups only, per user):
    "Long Build Up"   -> BULLISH  (bot buys ATM CE)
    "Short Build Up"  -> BEARISH  (bot buys ATM PE)
    Short Covering / Long Unwinding / blank  -> ignored

It uses the NEAR-MONTH future row per underlying for the build-up read. Stocks
AND indices are included. No daily cap. The bot picks its own option expiry/strike.

SAFETY: dry-run by default - prints the signals it WOULD write and writes NOTHING.
Pass --write to actually insert into fno_signals. Never places a trade itself.

Usage:
  python3 fno_excel_importer.py --file contracts.csv            # preview only
  python3 fno_excel_importer.py --inbox /home/globalbot/data/fno_inbox   # newest file
  python3 fno_excel_importer.py --inbox <dir> --write           # insert for the bot
"""

import os
import sys
import csv
import glob
import time
import sqlite3
import argparse
from datetime import datetime, date

HOME = "/home/globalbot"
EVENTS_DB = os.environ.get("EQ_EVENTS_DB", f"{HOME}/paper/events.db")
DEFAULT_INBOX = os.environ.get("FNO_INBOX", f"{HOME}/data/fno_inbox")
# remembers the last file we loaded so the daily cron imports each distinct file
# ONCE - and never re-trades a file you already used (e.g. Fri's file on Mon+Tue).
STATE_FILE = os.environ.get("FNO_IMPORT_STATE", f"{HOME}/data/.fno_last_import")

# strong build-ups only
BUILDUP_MAP = {
    "long build up": ("BULLISH", "EXCEL_LONG_BUILDUP"),
    "short build up": ("BEARISH", "EXCEL_SHORT_BUILDUP"),
}

# optional sanity filters (defaults keep ~everything; the tag already implies
# price+OI are aligned). Tighten via env if you want fewer, stronger signals.
MIN_OI_PCT = float(os.environ.get("FNO_MIN_OI_PCT", "0"))      # |%OI change| floor
MIN_DAY_PCT = float(os.environ.get("FNO_MIN_DAY_PCT", "0"))    # |%day change| floor


def _norm(s):
    return " ".join((s or "").strip().lower().split())


def _f(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return 0.0


def _exp_date(s):
    """Parse Trendlyne expiry -> date (CSV gives 'DD-MM-YY' string; XLSX may give
    a real datetime). Used to pick the nearest-expiry future."""
    if hasattr(s, "year"):                      # datetime/date from xlsx
        try:
            return s.date() if hasattr(s, "date") else s
        except Exception:
            return date.max
    for fmt in ("%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except Exception:
            continue
    return date.max


def pick_newest(inbox):
    pats = ["*.csv", "*.CSV", "*.xlsx", "*.XLSX", "*.xlsm"]
    files = []
    for p in pats:
        files += glob.glob(os.path.join(inbox, p))
    files = sorted(set(files), key=os.path.getmtime, reverse=True)
    return files[0] if files else None


def read_rows(path):
    """Read a CSV or XLSX export into a list of dict rows (headers from row 1).
    Trendlyne sometimes gives .csv, sometimes .xlsx - handle both."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("ERROR: this is an .xlsx file but 'openpyxl' isn't installed.\n"
                  "  Fix: pip3 install openpyxl   (or download the data as CSV instead)")
            sys.exit(2)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            wb.close(); return []
        out = []
        for r in it:
            out.append({header[i]: (r[i] if i < len(r) else None)
                        for i in range(len(header))})
        wb.close()
        return out
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def score(day_pct, oi_pct):
    """Bot trades score >= 65; keep build-ups in 65-95 by strength."""
    return round(min(95.0, 65.0 + min(15.0, abs(day_pct) * 5.0)
                     + min(15.0, abs(oi_pct) * 0.5)), 1)


def build_signals(path):
    """Return list of signal dicts from the near-month future of each underlying."""
    rows = read_rows(path)
    futs = [r for r in rows if _norm(r.get("OPTION TYPE")) == "future"]
    # near-month future per symbol
    nearest = {}
    for r in futs:
        sym = (r.get("SYMBOL") or "").strip().upper()
        if not sym:
            continue
        exp = _exp_date(r.get("EXPIRY"))
        if sym not in nearest or exp < nearest[sym][0]:
            nearest[sym] = (exp, r)

    signals = []
    for sym, (exp, r) in nearest.items():
        mapped = BUILDUP_MAP.get(_norm(r.get("BUILD UP")))
        if not mapped:
            continue
        direction, pattern = mapped
        day_pct = _f(r.get("%DAY CHANGE"))
        oi_pct = _f(r.get("%OI CHANGE"))
        if abs(oi_pct) < MIN_OI_PCT or abs(day_pct) < MIN_DAY_PCT:
            continue
        signals.append({
            "symbol": sym, "direction": direction, "pattern": pattern,
            "score": score(day_pct, oi_pct), "day_pct": day_pct,
            "oi_pct": oi_pct, "oi_chg": _f(r.get("OI CHANGE")),
            "expiry": exp.isoformat(),
        })
    signals.sort(key=lambda s: -abs(s["oi_pct"]))
    return signals


def already_imported_today(conn):
    """Symbols+directions already loaded from Excel today (avoid re-insert if the
    importer runs twice in a day)."""
    try:
        today = datetime.utcnow().strftime("%Y-%m-%d")
        rows = conn.execute(
            "SELECT symbol, direction FROM fno_signals "
            "WHERE mode='OI_EXCEL' AND substr(ts_utc,1,10)=?", (today,)).fetchall()
        return {(a, b) for a, b in rows}
    except Exception:
        return set()


def write_signals(signals):
    conn = sqlite3.connect(EVENTS_DB, timeout=10)
    conn.execute("CREATE TABLE IF NOT EXISTS fno_signals ("
                 "id INTEGER PRIMARY KEY AUTOINCREMENT, ts_utc TEXT, symbol TEXT, "
                 "direction TEXT, pattern TEXT, score REAL, spot_change_pct REAL, "
                 "ce_oi_change_pct REAL, pe_oi_change_pct REAL, mode TEXT, "
                 "consumed INTEGER DEFAULT 0)")
    dup = already_imported_today(conn)
    ins = skipped = 0
    ts = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    for s in signals:
        if (s["symbol"], s["direction"]) in dup:
            skipped += 1
            continue
        conn.execute("INSERT INTO fno_signals "
                     "(ts_utc,symbol,direction,pattern,score,spot_change_pct,"
                     "ce_oi_change_pct,pe_oi_change_pct,mode,consumed) "
                     "VALUES (?,?,?,?,?,?,?,?,?,0)",
                     (ts, s["symbol"], s["direction"], s["pattern"], s["score"],
                      s["day_pct"], s["oi_pct"], 0.0, "OI_EXCEL"))
        ins += 1
    conn.commit(); conn.close()
    return ins, skipped


def main():
    ap = argparse.ArgumentParser(description="Trendlyne F&O build-up importer")
    ap.add_argument("--file", help="specific CSV file")
    ap.add_argument("--inbox", default=DEFAULT_INBOX, help="folder; use newest CSV")
    ap.add_argument("--write", action="store_true", help="insert into fno_signals (default: preview only)")
    ap.add_argument("--force", action="store_true",
                    help="with --write, import even if this exact file was already loaded")
    args = ap.parse_args()

    path = args.file or pick_newest(args.inbox)
    if not path or not os.path.exists(path):
        print(f"ERROR: no CSV found (file={args.file}, inbox={args.inbox})")
        sys.exit(1)
    age_h = (time.time() - os.path.getmtime(path)) / 3600.0
    print(f"reading: {path}  (age {age_h:.1f}h)")

    # import-each-file-ONCE guard: signature = name + last-modified time. If the
    # newest file is the same one we already loaded (you forgot to upload a new
    # one), do nothing - never re-trade an already-used file.
    sig = f"{os.path.basename(path)}|{int(os.path.getmtime(path))}"
    if args.write and not args.force:
        try:
            if open(STATE_FILE).read().strip() == sig:
                print(f"\nALREADY IMPORTED this file - nothing written. "
                      f"Upload a new Trendlyne file to load fresh signals (or use --force).")
                sys.exit(0)
        except Exception:
            pass

    signals = build_signals(path)
    bull = sum(1 for s in signals if s["direction"] == "BULLISH")
    bear = sum(1 for s in signals if s["direction"] == "BEARISH")
    print(f"\nstrong build-up signals: {len(signals)}  (BULLISH/CE: {bull}, BEARISH/PE: {bear})\n")
    print(f"  {'SYMBOL':<14}{'DIR':<9}{'PATTERN':<22}{'day%':>8}{'OI%':>9}{'score':>7}")
    for s in signals:
        print(f"  {s['symbol']:<14}{s['direction']:<9}{s['pattern']:<22}"
              f"{s['day_pct']:>8.2f}{s['oi_pct']:>9.2f}{s['score']:>7.1f}")

    if args.write:
        ins, skipped = write_signals(signals)
        try:
            open(STATE_FILE, "w").write(sig)   # remember we've now loaded this file
        except Exception as e:
            print(f"(warning: could not write state file: {e})")
        print(f"\nWROTE {ins} signals into fno_signals (skipped {skipped} already loaded today).")
        print("The running fno_paper_study.py will pick these up (score >= 65).")
    else:
        print(f"\n[PREVIEW ONLY] nothing written. Re-run with --write to load into fno_signals.")


if __name__ == "__main__":
    main()
