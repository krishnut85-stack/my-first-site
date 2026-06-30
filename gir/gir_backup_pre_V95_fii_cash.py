#!/usr/bin/env python3
"""
GIR v26.0 — Autonomous NSE Trader
Owner: <redacted> | Kite: <redacted> | Server: <server-ip> (BLR1)

Architecture:
  EquityModule — CNC delivery, news-driven, RSI+Volume+ATR confirmation
  FnoModule    — NFO NRML, news-driven CE/PE, 218 F&O stocks, DTE cascade

Rules:
  - ZERO paper trading. Every order is real Kite or nothing.
  - Equity and F&O are completely independent. No shared candidates/filters.
  - Only shared: Kite access token + Telegram bot token.
  - GTT: ALWAYS check get_gtts() before placing. Modify if exists.
  - SL only moves UP, never down. Must be BELOW current LTP always.
  - F&O has NO equity-style regime filter.

Deploy:
  scp gir.py root@<server-ip>:/home/globalbot/gir.py
  systemctl restart globaleye
"""

VERSION = "GIR v28 + patches through V77"  # V77_FIX2

# ═══════════════════════════════════════════════════════════════════════════════
#  V28 CHANGES (Apr 18, 2026) — applied after comprehensive senior review
# ═══════════════════════════════════════════════════════════════════════════════
#  A1  Equity _can_enter checks Kite holdings (not stale local JSON)
#  A2  AMO scan runs ONCE per evening (20:00 IST) + checks Kite holdings
#  A4  KITE_API_KEY fallback removed (must come from .env)
#  A5  GTT modify failure returns None (not fake success gtt_id)
#  A7  Duplicate order check uses exact match (not .startswith)
#  A8  Bare except: replaced with except Exception: (4 locations)
#  A10 EQUITY REGIME FILTER: block bullish buys when market is falling AND stock
#      not outperforming, block chasing when stock already ran >3% today
#  C3  Freak trade protection (reject LTP >20% from prev close)
#  C10 Edge-weighted position sizing (score/60 multiplier, cap 3% risk)
#  D1  EQ_MAX_POS 10 -> 5 (concentrate into quality positions)
#  D3  Catalyst-based dead money exit (BUYBACK=60d, ORDER_WIN=30d, etc.)
#  D4  Smart FNO_CAP sizing: respect lot-cost constraint (was fixed 0.95)
#  D5  Two-stage F&O trail: 10-25% profit = 30% trail, 25%+ = 20% trail
#  D7  Capital utilization via pressing winners (add to existing winners if idle > 15%)
#
#  NOT applied (require backtest validation first):
#  A9  Kelly fraction (kept at 0.80 pending backtest)
#  D2  ATR-based trail (kept at fixed 2.5% pending backtest)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

import os
import sys
import json
import time
import math
import hashlib
import logging
import threading
import traceback
from datetime import datetime, timedelta, date, timezone
from collections import defaultdict
from pathlib import Path

import requests
import pyotp

try:
    from kiteconnect import KiteConnect
except ImportError:
    print("FATAL: pip install kiteconnect")
    sys.exit(1)

# V27 Stage 2: import NewsBrain (LLM news analyzer)
try:
    from news_brain import NewsBrain
    _NEWS_BRAIN_AVAILABLE = True
except ImportError as _nbe:
    NewsBrain = None
    _NEWS_BRAIN_AVAILABLE = False
    print(f"WARNING: news_brain.py not importable: {_nbe} - LLM mode disabled")

# V78: Trade recorder (silent observation, never blocks orders)
try:
    from trade_recorder import TradeRecorder
    _TRADE_RECORDER_AVAILABLE = True
except ImportError as _tre:
    TradeRecorder = None
    _TRADE_RECORDER_AVAILABLE = False
    print(f"WARNING: trade_recorder.py not importable: {_tre} - recording disabled")

try:
    import feedparser
except ImportError:
    print("FATAL: pip install feedparser")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("GIR")

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

DATA_DIR = Path("/home/globalbot/data")
DATA_DIR.mkdir(parents=True, exist_ok=True)

# ── Kite credentials ──
# V28 A4: KITE_API_KEY must come from .env (no hardcoded fallback)
KITE_API_KEY = os.getenv("KITE_API_KEY", "")
KITE_API_SECRET = os.getenv("KITE_API_SECRET", "")
KITE_TOTP_SECRET = os.getenv("KITE_TOTP_SECRET", "")
ZERODHA_USER_ID = os.getenv("ZERODHA_USER_ID", "")
ZERODHA_PASSWORD = os.getenv("ZERODHA_PASSWORD", "")
KITE_TOKEN_FILE = DATA_DIR / "kite_token.json"

# ── Telegram ──
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

# ── Capital wall ──
EQUITY_PCT = 0.75
FNO_PCT = 0.25

# ── Equity config ──
EQ_MAX_POS = 10                 # FIX_V43_MAXPOS: raised 5->10 (11 legacy holdings blocking scan, ₹80K idle)
EQ_MAX_PER_SECTOR = 3           # V79_FIX3: 2->3 (allow more sector concentration when signals cluster)
EQ_RISK_PER_TRADE = 0.02
EQ_INITIAL_SL_PCT = 0.025       # V79_FIX1: 0.04 -> 0.025 reverted (V77 widening was wrong fix; tight SL + time-stagnation is the right combo)
EQ_TRAIL_SL_PCT = 0.025         # V79_FIX2: 0.035 -> 0.025 reverted with V79_FIX1
EQ_DEAD_MONEY_DAYS = 15
EQ_KELLY_FRACTION = 0.80
EQ_MIN_CONFIDENCE = 42
EQ_T2_ATR_MULT = 14             # Target 2 = 14 x ATR (from v9.1 backtest)
EQ_PROFIT_LOCK_MIN = 15         # Minutes after entry before profit lock
EQ_SCAN_INTERVAL_MIN = 30       # Scan every 30 minutes

# ── F&O config ──
FNO_MAX_POS = 5
FNO_MAX_PER_STOCK = 1
FNO_MAX_PER_SECTOR = 2          # PATCH_V3_SECTOR_CAP: max 2 F&O positions per sector
FNO_VIX_MAX = 22.0              # PATCH_V4_VIX_FILTER: block F&O entries when India VIX > 22

# ── OI tracking (PATCH_V5_OI_LOGGER) ──
OI_SNAPSHOT_INTERVAL_MIN = 60   # Snapshot every 60 minutes during market hours
OI_RETENTION_DAYS = 14          # Keep 14 days of snapshots, auto-delete older
OI_STRIKES_RANGE = 5            # ATM +/- 5 strikes per option type
OI_BUILDUP_MIN_PCT = 15.0       # Minimum % OI change to trigger buildup signal
OI_SIGNAL_MIN_SCORE = 65        # Minimum score for OI signal to enter pipeline
OI_INDICES = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"]

# PATCH_V73_INDEX_SPOT_KEY: map internal index symbols to Kite's NSE quote symbols
_INDEX_SPOT_KEY_MAP = {
    "NIFTY":      "NSE:NIFTY 50",
    "BANKNIFTY":  "NSE:NIFTY BANK",
    "FINNIFTY":   "NSE:NIFTY FIN SERVICE",
    "MIDCPNIFTY": "NSE:NIFTY MID SELECT",
}
def _index_spot_key(sym):
    """Return the correct Kite quote key for index symbols, NSE:{sym} for stocks."""
    return _INDEX_SPOT_KEY_MAP.get(sym, f"NSE:{sym}")

# ── Gap detection (PATCH_V6_GAP) ──
GAP_INDEX_THRESHOLD = 0.5       # Index gap threshold: 0.5%
GAP_STOCK_THRESHOLD = 1.0       # Stock gap threshold: 1.0%
GAP_SCORE_BOOST = 10            # Score boost for aligned signals

# ── Sectoral rotation (PATCH_V8_SECTOR_ROTATION) ──
SECTOR_ROTATION_REFRESH_MIN = 15    # Refresh sector ranking every 15 minutes
SECTOR_ROTATION_TOP_BOOST = 10      # Score boost for stocks in top-3 sectors
SECTOR_ROTATION_BOTTOM_BLOCK = True # Block entries for stocks in bottom-2 sectors
SECTOR_INDICES_MAP = {
    "BANKING": "NSE:NIFTY BANK",
    "IT": "NSE:NIFTY IT",
    "PHARMA": "NSE:NIFTY PHARMA",
    "FMCG": "NSE:NIFTY FMCG",
    "AUTO": "NSE:NIFTY AUTO",
    "METALS": "NSE:NIFTY METAL",
    "ENERGY": "NSE:NIFTY ENERGY",
    "REALTY": "NSE:NIFTY REALTY",
    "SERVICES": "NSE:NIFTY FIN SERVICE",
    "HEALTHCARE": "NSE:NIFTY HEALTHCARE",
    "CONSUMPTION": "NSE:NIFTY CONSUMPTION",
    "PSUBANK": "NSE:NIFTY PSU BANK",
    "PVTBANK": "NSE:NIFTY PVT BANK",
    "MEDIA": "NSE:NIFTY MEDIA",
    "INFRA": "NSE:NIFTY INFRA",
    "OILGAS": "NSE:NIFTY OIL & GAS",
}
def FNO_CAP_PER_TRADE_FN(capital):
    # FIX_V29_ADAPTIVE: scale per-trade allocation with capital
    if capital < 10000:   return 0.30
    if capital < 50000:   return 0.50
    if capital < 200000:  return 0.70
    return 0.95
FNO_CAP_PER_TRADE = 0.30  # default for small capital; overridden at runtime        # 30% of F&O capital per trade
FNO_SL_PCT = 0.30               # 30% SL on premium
FNO_GTT_LIMIT_OFFSET = 0.005  # FIX_V60: was 5%
EQ_GTT_LIMIT_OFFSET = 0.05  # FIX_V60
FNO_TRAIL_ACTIVATE = 0.10     # V27 PATCH (Bug #10): trail only after 10% profit captured
FNO_TRAIL_PCT = 0.20          # 20% below LTP
FNO_TRAIL_DRY_RUN = False

# V27 Stage 2: LLM news brain integration (dry-run = log only, no trades)
USE_LLM_NEWS = True
LLM_DRY_RUN = False  # True = LLM logs decisions only (no trades). False = LLM candidates added to live trade list
FNO_MIN_DTE = 20  # V77_FIX8: 10->20 (data: 6/9 F&O exits were DTE_CUT/EMERGENCY at <14 DTE)
FNO_MAX_DTE = 45  # PATCH_V28_OPTION_PICKER: was 28. Opens May expiry (DTE~40) for longer-hold catalyst plays
FNO_EMERGENCY_DTE = 12          # Emergency exit
FNO_DTE_CUT_LOSERS = 15         # Cut losers at DTE <= 15
FNO_DTE_WARN = 20               # Warning at DTE <= 20
FNO_MIN_OI = 50000
FNO_MAX_SPREAD = 0.08  # V75: 5% -> 8%           # 5% max bid-ask spread
FNO_MIN_CATALYST = 60           # PATCH_V28_SECTOR_LOG: sector over-cap logs demoted to DEBUG           # Minimum catalyst score
FNO_SCAN_INTERVAL_MIN = 15      # Scan every 15 minutes

# ── Risk ──
CRASH_NIFTY_DROP = 0.05         # Block if Nifty drops > 5% intraday
MONTHLY_LOSS_LIMIT = 0.09       # 9% monthly
DAILY_LOSS_AVG_LIMIT = 0.15  # FIX_V29: was 0.05, relaxed for small F&O capital     # 5% average loss (not cumulative)

# ── Blacklist ──
BLACKLIST = {"AQYLON"}
BLOCKED_PREFIXES = ["GS20", "GS19", "GSEC", "AFIL", "SSCL", "SDL", "TBILL", "SGBOND"]
BLOCKED_SUFFIXES = ["-SM", "-ST", "-BE", "-BZ", "-IL", "-BL"]

# ── NSE holidays 2026 ──
NSE_HOLIDAYS = {
    date(2026, 1, 26), date(2026, 3, 10), date(2026, 3, 30), date(2026, 3, 31),
    date(2026, 4, 14), date(2026, 5, 1), date(2026, 7, 17), date(2026, 8, 15),
    date(2026, 8, 17), date(2026, 10, 2), date(2026, 10, 20), date(2026, 10, 21),
    date(2026, 11, 5), date(2026, 11, 18), date(2026, 12, 25),
}

# ── RSS feeds (Indian financial news + global macro) ──
# PATCH_V62_GLOBAL_FEEDS: added 5 global sources for Fed/crude/US/geopolitical catalysts
NEWS_FEEDS = {
    "ET Markets": "https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms",
    "ET Corporate": "https://economictimes.indiatimes.com/industry/rssfeeds/13352306.cms",
    "MC Markets": "https://www.moneycontrol.com/rss/marketreports.xml",
    "MC Business": "https://www.moneycontrol.com/rss/business.xml",
    "BS Markets": "https://www.business-standard.com/rss/markets-106.rss",
    "Mint Markets": "https://www.livemint.com/rss/markets",
    "Mint Companies": "https://www.livemint.com/rss/companies",
    "FE Market": "https://www.financialexpress.com/market/feed/",
    "NDTV Profit": "https://www.ndtv.com/business/feeds",
    "CNBC TV18 Market": "https://www.cnbctv18.com/rss/market.xml",
    "CNBC TV18 Companies": "https://www.cnbctv18.com/rss/companies.xml",
    "ZeeBiz Markets": "https://www.zeebiz.com/rss/markets.xml",
    "ZeeBiz Stocks": "https://www.zeebiz.com/rss/stocks.xml",
    "BT Markets": "https://www.businesstoday.in/rss/markets",
    "RBI": "https://www.rbi.org.in/scripts/rss.aspx?id=2",
    # PATCH_V62_GLOBAL_FEEDS — global catalysts that move NIFTY/BANKNIFTY overnight
    "Reuters Business": "https://feeds.reuters.com/reuters/businessNews",
    "Reuters World": "https://feeds.reuters.com/Reuters/worldNews",
    "CNBC Top News": "https://search.cnbc.com/rs/search/combinedcms/view.xml?partnerId=wrss01&id=100003114",
    "Investing.com News": "https://www.investing.com/rss/news.rss",
    "FT Companies": "https://www.ft.com/rss/companies",
}

# ── News keywords for direction detection ──
BULLISH_KEYWORDS = [
    "buy rating", "upgrade", "outperform", "overweight", "top pick",
    "accumulate", "raises target", "target raised", "order win", "wins order",
    "record profit", "beats estimate", "above estimate", "strong results",
    "revenue growth", "profit growth", "expansion", "new plant", "capex",
    "buyback", "bonus issue", "stock split", "promoter buying", "promoter buy",
    "insider buying", "bulk deal buy", "block deal buy", "dividend declared",
    "fdi inflow", "rate cut", "rbi cut", "stimulus", "reform", "fii buying",
    "dii buying", "contract win", "deal win", "partnership", "collaboration",
    "acquisition", "merger positive", "stake increase", "credit upgrade",
    "new order", "capacity expansion", "margin improvement", "debt reduction",
    "breakthrough", "patent", "approval received", "clearance",
]

BEARISH_KEYWORDS = [
    "sell rating", "downgrade", "underperform", "underweight", "reduce",
    "cuts target", "target cut", "fraud", "scam", "penalty", "fine imposed",
    "loss widens", "misses estimate", "below estimate", "weak results",
    "revenue decline", "profit decline", "shutdown", "closure", "layoff",
    "promoter selling", "promoter sell", "insider selling", "bulk deal sell",
    "fii selling", "dii selling", "credit downgrade", "rating downgrade",
    "debt default", "npa", "bad loan", "sebi action", "ban", "suspension",
    "recall", "investigation", "raid", "probe", "stake decrease",
    "margin pressure", "cost overrun", "delay", "cancellation",
    "downgrade outlook", "negative outlook", "ofs", "qip dilution",
]

# ── Filing types and their signals ──
FILING_SIGNALS = {
    "BUYBACK":          {"direction": "BULLISH",  "score": 95},
    "PROMOTER_BUY":     {"direction": "BULLISH",  "score": 90},
    "ORDER_WIN":        {"direction": "BULLISH",  "score": 80},
    "DIVIDEND":         {"direction": "BULLISH",  "score": 70},
    "BONUS":            {"direction": "BULLISH",  "score": 75},
    "STOCK_SPLIT":      {"direction": "BULLISH",  "score": 70},
    "ACQUISITION":      {"direction": "BULLISH",  "score": 75},
    "EXPANSION":        {"direction": "BULLISH",  "score": 70},
    "RESULTS_BEAT":     {"direction": "BULLISH",  "score": 80},
    "PROMOTER_SELL":    {"direction": "BEARISH",  "score": 75},
    "QIP":              {"direction": "BEARISH",  "score": 60},
    "OFS":              {"direction": "BEARISH",  "score": 70},
    "CREDIT_DOWNGRADE": {"direction": "BEARISH",  "score": 85},
    "FRAUD":            {"direction": "BEARISH",  "score": 95},
    "RESULTS_MISS":     {"direction": "BEARISH",  "score": 80},
    "PENALTY":          {"direction": "BEARISH",  "score": 70},
}

# ── Sector pod config (for equity sector limits) ──
# ======================================================================
#  TRENDLYNE SECTOR LOADER — auto-populate SECTOR_MAP from xlsx
# ======================================================================

# Trendlyne sector_name -> our internal sector label
_TRENDLYNE_SECTOR_TO_LABEL = {
    "Banking and Finance":              "BANKING",  # will be split by override lists below
    "Software & Services":              "IT",
    "Hardware Technology & Equipment":  "IT",
    "Telecommunications Equipment":     "SERVICES",
    "Pharmaceuticals & Biotechnology":  "PHARMA",
    "Healthcare":                       "HEALTHCARE",
    "FMCG":                             "FMCG",
    "Food, Beverages & Tobacco":        "FMCG",
    "Consumer Durables":                "CONSUMPTION",
    "Retailing":                        "CONSUMPTION",
    "Automobiles & Auto Components":    "AUTO",
    "General Industrials":              "INFRA",
    "Oil & Gas":                        "OILGAS",
    "Utilities":                        "ENERGY",
    "Metals & Mining":                  "METALS",
    "Cement and Construction":          "INFRA",
    "Realty":                           "REALTY",
    "Chemicals & Petrochemicals":       "CHEMICALS",
    "Fertilizers":                      "CHEMICALS",
    "Transportation":                   "SERVICES",
    "Telecom Services":                 "SERVICES",
    "Hotels Restaurants & Tourism":     "SERVICES",
    "Media":                            "MEDIA",
    "Commercial Services & Supplies":   "SERVICES",
    "Diversified Consumer Services":    "SERVICES",
    "Textiles Apparels & Accessories":  "CONSUMPTION",
    "Forest Materials":                 "INFRA",
    "Diversified":                      "SERVICES",
    "Others":                           "SERVICES",
}

# Override: PSU banks -> PSUBANK (Trendlyne lumps them under "Banking and Finance")
_PSU_BANKS = {
    "SBIN", "PNB", "BANKBARODA", "CANBK", "IOB", "UNIONBANK", "INDIANB",
    "BANKINDIA", "CENTRALBK", "MAHABANK", "UCOBANK", "PSB", "IDBI"
}

# Override: Non-bank financials -> SERVICES (Trendlyne also puts these in Banking and Finance)
_NBFC_AND_FINANCIALS = {
    "BAJFINANCE", "BAJAJFINSV", "BAJAJHLDNG", "CHOLAFIN", "M&MFIN", "MUTHOOTFIN",
    "SHRIRAMFIN", "PFC", "RECLTD", "LICHSGFIN", "CANFINHOME", "HUDCO", "PEL",
    "L&TFH", "ABCAPITAL", "IIFL", "MOTILALOFS", "ANGELONE", "360ONE", "POLICYBZR",
    "PAYTM", "SBICARD", "MFSL", "SBILIFE", "HDFCLIFE", "ICICIPRULI", "ICICIGI",
    "LICI", "MAX FINANCIAL", "HDFCAMC", "NIPPON LIFE", "ABSLAMC", "UTIAMC",
    "BSE", "MCX", "CDSL", "CAMS", "KFINTECH", "IEX", "IREDA"
}


def _load_trendlyne_sector_map():
    """Build {NSE_SYMBOL: SECTOR_LABEL} from trendlyne_data.xlsx.

    Returns dict. On any failure, returns empty dict (caller falls back).
    Does NOT raise exceptions — safe to call at import time.
    """
    result = {}
    try:
        import pandas as pd
        import os
        path = "/home/globalbot/data/trendlyne_data.xlsx"
        if not os.path.exists(path):
            return result
        df = pd.read_excel(path)
        if "NSE Code" not in df.columns or "sector_name" not in df.columns:
            return result
        mapped = 0
        for _, row in df.iterrows():
            sym = row.get("NSE Code")
            sec = row.get("sector_name")
            if not isinstance(sym, str) or not isinstance(sec, str):
                continue
            sym = sym.strip().upper()
            sec = sec.strip()
            if not sym or not sec:
                continue
            label = _TRENDLYNE_SECTOR_TO_LABEL.get(sec, "SERVICES")
            # Banking override: split into PSUBANK / PVTBANK / SERVICES
            if sec == "Banking and Finance":
                if sym in _PSU_BANKS:
                    label = "PSUBANK"
                elif sym in _NBFC_AND_FINANCIALS:
                    label = "SERVICES"
                else:
                    label = "PVTBANK"
            result[sym] = label
            mapped += 1
        return result
    except Exception:
        return {}


# Load Trendlyne-derived sector map at import time (silent on failure)
_TRENDLYNE_SECTOR_MAP = _load_trendlyne_sector_map()


# Hardcoded overrides — these win over Trendlyne
SECTOR_MAP = {
    "HDFCBANK": "BANKING", "ICICIBANK": "BANKING", "SBIN": "BANKING",
    "KOTAKBANK": "BANKING", "AXISBANK": "BANKING", "INDUSINDBK": "BANKING",
    "BANDHANBNK": "BANKING", "FEDERALBNK": "BANKING", "IDFCFIRSTB": "BANKING",
    "PNB": "BANKING", "BANKBARODA": "BANKING", "CANBK": "BANKING",
    "TCS": "IT", "INFY": "IT", "WIPRO": "IT", "HCLTECH": "IT",
    "TECHM": "IT", "LTIM": "IT", "MPHASIS": "IT", "COFORGE": "IT",
    "SUNPHARMA": "PHARMA", "DRREDDY": "PHARMA", "CIPLA": "PHARMA",
    "DIVISLAB": "PHARMA", "BIOCON": "PHARMA", "AUROPHARMA": "PHARMA",
    "HINDUNILVR": "FMCG", "ITC": "FMCG", "NESTLEIND": "FMCG",
    "BRITANNIA": "FMCG", "DABUR": "FMCG", "MARICO": "FMCG",
    "MARUTI": "AUTO", "TATAMOTORS": "AUTO", "M&M": "AUTO",
    "BAJAJ-AUTO": "AUTO", "HEROMOTOCO": "AUTO", "ASHOKLEY": "AUTO",
    "RELIANCE": "ENERGY", "ONGC": "ENERGY", "BPCL": "ENERGY",
    "IOC": "ENERGY", "GAIL": "ENERGY", "NTPC": "ENERGY",
    "TATASTEEL": "METALS", "JSWSTEEL": "METALS", "HINDALCO": "METALS",
    "COALINDIA": "METALS", "VEDL": "METALS", "NMDC": "METALS",
    "PIDILITIND": "CHEMICALS", "UPL": "CHEMICALS", "SRF": "CHEMICALS",
    "LT": "INFRA", "ADANIENT": "INFRA", "ADANIPORTS": "INFRA",
    "BHARTIARTL": "SERVICES", "BAJFINANCE": "SERVICES", "BAJAJFINSV": "SERVICES",
}

# Merge Trendlyne-derived map into SECTOR_MAP.
# Trendlyne WINS over hardcoded (authoritative + current industry labels + granular sectors).
# Hardcoded entries serve as fallback for stocks Trendlyne does not cover.
_hardcoded_only = {_s: _l for _s, _l in SECTOR_MAP.items() if _s not in _TRENDLYNE_SECTOR_MAP}
SECTOR_MAP = dict(_TRENDLYNE_SECTOR_MAP)
SECTOR_MAP.update(_hardcoded_only)

try:
    import logging as _logging
    _logging.getLogger().info(
        f'[SECTOR_MAP] total={len(SECTOR_MAP)} '
        f'trendlyne={len(_TRENDLYNE_SECTOR_MAP)} '
        f'hardcoded_overrides={sum(1 for s in SECTOR_MAP if s not in _TRENDLYNE_SECTOR_MAP or SECTOR_MAP[s] != _TRENDLYNE_SECTOR_MAP.get(s))}'
    )
except Exception:
    pass



# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def now_ist():
    """Current time in IST."""
    return datetime.now(timezone(timedelta(hours=5, minutes=30)))


def today_ist():
    """Current date in IST."""
    return now_ist().date()


def is_trading_day(d=None):
    """Check if given date is an NSE trading day."""
    d = d or today_ist()
    if d.weekday() >= 5:  # Saturday/Sunday
        return False
    if d in NSE_HOLIDAYS:
        return False
    return True


def is_market_hours():
    """Check if NSE market is currently open (9:15 to 15:30 IST)."""
    if not is_trading_day():
        return False
    n = now_ist()
    market_open = n.replace(hour=9, minute=15, second=0, microsecond=0)
    market_close = n.replace(hour=15, minute=30, second=0, microsecond=0)
    return market_open <= n <= market_close


def is_amo_window():
    """Check if AMO orders can be placed (15:45-01:00 or 05:30-09:15)."""
    n = now_ist()
    h, m = n.hour, n.minute
    t = h * 60 + m
    # 15:45 to 23:59
    if t >= 945:
        return True
    # 00:00 to 01:00
    if t <= 60:
        return True
    # 05:30 to 09:14
    if 330 <= t <= 554:
        return True
    return False


def load_json(path, default=None):
    """Load JSON file, return default if missing/corrupt."""
    try:
        with open(path, "r") as f:
            return json.load(f)
    except Exception:
        return default if default is not None else {}


def save_json(path, data):
    """Save data to JSON file atomically."""
    tmp = str(path) + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2, default=str)
    os.replace(tmp, str(path))


def is_valid_symbol(sym):
    """Check if symbol is a valid equity stock (not bond, ETF, SME, etc)."""
    if not sym or not isinstance(sym, str):
        return False
    sym_upper = sym.upper()
    if sym_upper in BLACKLIST:
        return False
    # Numeric prefix = government security
    if sym_upper[0].isdigit():
        return False
    # Blocked prefixes
    for bp in BLOCKED_PREFIXES:
        if bp in sym_upper:
            return False
    # Blocked suffixes
    for bs in BLOCKED_SUFFIXES:
        if sym_upper.endswith(bs):
            return False
    # Special characters (bonds, series)
    if "%" in sym_upper:
        return False
    return True


def get_sector(symbol):
    """Get sector for a symbol, default SERVICES."""
    return SECTOR_MAP.get(symbol, "SERVICES")


# ═══════════════════════════════════════════════════════════════════════════════
#  KITE SESSION — One login, both modules use the token independently
# ═══════════════════════════════════════════════════════════════════════════════

# FIX_V29_RATE_LIMIT: exponential backoff retry wrapper
def _kite_retry(fn, *args, max_retries=3, backoff=1.0, **kwargs):
    """FIX_V53_RETRY: Exponential backoff wrapper for Kite API calls.

    Retries on transient errors (rate limit, network). Does NOT retry on
    permanent errors (InputException, OrderException, etc.) as those are bugs,
    not transient failures. Re-raises after max retries so caller's try/except
    still fires.

    Usage: _kite_retry(kite.place_order, variety="regular", ...)
    """
    import time as _t
    last_exc = None
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            last_exc = e
            name = type(e).__name__
            msg = str(e)
            # Don't retry permanent errors
            if name in ("InputException", "OrderException", "TokenException",
                        "HoldingException", "PermissionException"):
                raise
            # Retry transient ones
            is_transient = (
                "Too many requests" in msg or
                "timeout" in msg.lower() or
                "connection" in msg.lower() or
                name in ("NetworkException", "GeneralException")
            )
            if is_transient and attempt < max_retries - 1:
                sleep_s = backoff * (2 ** attempt)
                try:
                    log.warning(f"_kite_retry: attempt {attempt+1}/{max_retries} failed "
                                f"({name}: {msg[:80]}), sleeping {sleep_s}s")
                except Exception:
                    pass
                _t.sleep(sleep_s)
                continue
            raise
    if last_exc:
        raise last_exc
    return None


# ═══════════════════════════════════════════════════════════════
# FIX_V29_ZERODHA_RULES: central exchange-rule compliance helpers
# Every price/qty touching Kite API must pass through these
# ═══════════════════════════════════════════════════════════════
import math as _math_zr

# FIX_V56_TICK_SIZE: NSE tick rules revised April 15 2025
# <250=0.01  251-1000=0.05  1001-5000=0.10  5001-10000=0.50  10001-20000=1.00  >20001=5.00
def _fallback_tick_by_price(price, is_option=False):
    # FIX_V60_TICK_SL_UP: option-aware. NFO options ALWAYS 0.05.
    if is_option:
        return 0.05
    try:
        p = float(price or 0)
    except Exception:
        p = 0
    if p <= 0: return 0.05
    if p < 250: return 0.01
    if p < 1000: return 0.05
    if p < 5000: return 0.10
    if p < 10000: return 0.50
    if p < 20000: return 1.00
    return 5.00


class TickSizeCache:
    # FIX_V56_TICK_SIZE: authoritative per-symbol tick size cache from Kite
    _cache = {}
    _exch_cache = {}
    _loaded_at = 0.0
    _TTL = 86400

    @classmethod
    def load(cls, kite=None, force=False):
        import time as _t
        now = _t.time()
        if not force and cls._cache and (now - cls._loaded_at) < cls._TTL:
            return len(cls._cache)
        if kite is None:
            try:
                kite = KiteSession.kite()
            except Exception:
                return 0
        if kite is None:
            return 0
        loaded = 0
        for exch in ("NSE", "NFO", "BSE"):
            try:
                instr = kite.instruments(exch) or []
                for i in instr:
                    sym = i.get("tradingsymbol", "")
                    ts = i.get("tick_size", 0) or 0
                    if sym and ts > 0:
                        if sym not in cls._cache:
                            cls._cache[sym] = float(ts)
                        cls._exch_cache[(exch, sym)] = float(ts)
                        loaded += 1
            except Exception as _e:
                try:
                    log.warning(f"[FIX_V56_TICK_SIZE] instruments({exch}) failed: {_e}")
                except Exception:
                    pass
        cls._loaded_at = now
        try:
            log.info(f"[FIX_V56_TICK_SIZE] loaded {loaded} tick entries (unique: {len(cls._cache)})")
        except Exception:
            pass
        return loaded

    @classmethod
    def get(cls, symbol, exchange=None, price=None):
        if not cls._cache:
            try:
                cls.load()
            except Exception:
                pass
        if exchange:
            ts = cls._exch_cache.get((exchange, symbol))
            if ts and ts > 0:
                return ts
        ts = cls._cache.get(symbol)
        if ts and ts > 0:
            return ts
        return _fallback_tick_by_price(price)


def get_tick(symbol=None, exchange=None, price=None):
    # FIX_V56_TICK_SIZE: public helper
    if symbol:
        return TickSizeCache.get(symbol, exchange=exchange, price=price)
    return _fallback_tick_by_price(price)


def _snap_tick(price, tick=None, symbol=None, exchange=None):
    # FIX_V56_TICK_SIZE: symbol-aware snap-down. Backward-compatible.
    if price is None or price <= 0:
        return price
    if tick is None or tick <= 0:
        if symbol:
            tick = TickSizeCache.get(symbol, exchange=exchange, price=price)
        else:
            tick = _fallback_tick_by_price(price)
    if not tick or tick <= 0:
        tick = 0.05
    decimals = max(0, -int(_math_zr.floor(_math_zr.log10(tick)))) if tick < 1 else 0
    return round(_math_zr.floor(float(price) / tick) * tick, decimals if decimals > 0 else 2)


def _snap_tick_up(price, tick=None, symbol=None, exchange=None):
    # FIX_V56_TICK_SIZE: symbol-aware snap-up. Backward-compatible.
    if price is None or price <= 0:
        return price
    if tick is None or tick <= 0:
        if symbol:
            tick = TickSizeCache.get(symbol, exchange=exchange, price=price)
        else:
            tick = _fallback_tick_by_price(price)
    if not tick or tick <= 0:
        tick = 0.05
    decimals = max(0, -int(_math_zr.floor(_math_zr.log10(tick)))) if tick < 1 else 0
    return round(_math_zr.ceil(float(price) / tick) * tick, decimals if decimals > 0 else 2)

# FIX_V60_TICK_SL_UP
def _snap_sl_tick_up(price, symbol, exchange):
    if not symbol or not exchange:
        raise ValueError("[V60] _snap_sl_tick_up needs symbol+exchange")
    return _snap_tick_up(price, symbol=symbol, exchange=exchange)

def _validate_gtt_tick(symbol, exchange, trigger, limit_price):
    try:
        tick = TickSizeCache.get(symbol, exchange=exchange, price=trigger)
    except Exception:
        tick = None
    if not tick or tick <= 0:
        is_opt = (exchange == "NFO") and (symbol.endswith("CE") or symbol.endswith("PE"))
        tick = _fallback_tick_by_price(trigger, is_option=is_opt)
    try:
        t_ok = abs(round(float(trigger)/tick)*tick - float(trigger)) < 0.001
        p_ok = abs(round(float(limit_price)/tick)*tick - float(limit_price)) < 0.001
    except Exception:
        return False
    if not t_ok:
        log.error(f"[V60_TICK_GUARD] BLOCK {exchange}:{symbol} trig={trigger} tick={tick}")
    if not p_ok:
        log.error(f"[V60_TICK_GUARD] BLOCK {exchange}:{symbol} price={limit_price} tick={tick}")
    return t_ok and p_ok

# FIX_V60P2_TICK_GUARD: blocks tick-non-compliant GTT API calls
def _safe_place_gtt(kite, **kwargs):
    try:
        tsym = kwargs.get("tradingsymbol")
        exch = kwargs.get("exchange")
        trigs = kwargs.get("trigger_values") or []
        orders = kwargs.get("orders") or []
        for tv in trigs:
            for o in orders:
                if not _validate_gtt_tick(tsym, exch, tv, o.get("price", 0)):
                    log.error(f"[V60P2_GUARD] place_gtt BLOCKED {exch}:{tsym} trig={tv} price={o.get('price')}")
                    return None
        return kite.place_gtt(**kwargs)
    except Exception as e:
        log.error(f"[V60P2_GUARD] place_gtt raised: {e}")
        raise

def _safe_modify_gtt(kite, **kwargs):
    try:
        tsym = kwargs.get("tradingsymbol")
        exch = kwargs.get("exchange")
        trigs = kwargs.get("trigger_values") or []
        orders = kwargs.get("orders") or []
        for tv in trigs:
            for o in orders:
                if not _validate_gtt_tick(tsym, exch, tv, o.get("price", 0)):
                    log.error(f"[V60P2_GUARD] modify_gtt BLOCKED {exch}:{tsym} trig={tv} price={o.get('price')}")
                    return None
        return kite.modify_gtt(**kwargs)
    except Exception as e:
        log.error(f"[V60P2_GUARD] modify_gtt raised: {e}")
        raise

def _gtt_gap_ok(trigger, ltp, min_gap_pct=0.0025):
    """Kite requires GTT trigger >= 0.25% away from LTP."""
    if ltp <= 0 or trigger <= 0:
        return False
    return abs(trigger - ltp) / ltp >= min_gap_pct

def _gtt_safe_trigger(trigger, ltp, side="SELL", min_gap_pct=0.0030, symbol=None, exchange=None):
    # FIX_V53_GAP + FIX_V56_TICK_SIZE + PATCH_V66_TICK_FLOOR
    # PATCH_V66: enforce absolute 3-tick gap to avoid sub-rupee NSE rejection.
    # Apr 24 TATASTEEL bug: ltp=0.31 + min_gap_pct=0.30% -> snap to 0.30 (1 paise gap)
    # NSE rejected ("difference should be more than 0.09"). Now: max(3*tick, pct).
    if ltp <= 0 or trigger <= 0:
        return 0.0
    try:
        _tick = TickSizeCache.get(symbol, exchange=exchange, price=ltp) if symbol else None
    except Exception:
        _tick = None
    if not _tick or _tick <= 0:
        _tick = _fallback_tick_by_price(ltp) or 0.05
    abs_min_gap = 3.0 * _tick
    pct_min_gap = ltp * min_gap_pct
    eff_gap = max(abs_min_gap, pct_min_gap)
    if side == "SELL":
        max_allowed = ltp - eff_gap
        if max_allowed <= 0:
            return 0.0
        if trigger > max_allowed:
            return _snap_tick(max_allowed, symbol=symbol, exchange=exchange)
        return _snap_tick(trigger, symbol=symbol, exchange=exchange)
    else:
        min_allowed = ltp + eff_gap
        if trigger < min_allowed:
            return _snap_tick_up(min_allowed, symbol=symbol, exchange=exchange)
        return _snap_tick(trigger, symbol=symbol, exchange=exchange)


def _gtt_safe_limit(trigger, side="SELL", offset_pct=0.005, symbol=None, exchange=None):
    # FIX_V56_TICK_SIZE: correctly tick-aligned GTT limit price
    if trigger is None or trigger <= 0:
        return 0.0
    if side == "SELL":
        return _snap_tick(trigger * (1 - offset_pct), symbol=symbol, exchange=exchange)
    else:
        return _snap_tick_up(trigger * (1 + offset_pct), symbol=symbol, exchange=exchange)

def _fno_execution_range_ok(price, ref_price, band=0.40):
    """F&O execution range is typically +/-40% of reference price."""
    if ref_price <= 0:
        return True
    lo, hi = ref_price * (1 - band), ref_price * (1 + band)
    return lo <= price <= hi

def _fno_safe_limit_price(desired, ref_price, band=0.40):
    """FIX_V53_RANGE: Clamp F&O LIMIT order price to Kite's +/-40% execution range.

    Zerodha F&O rejects orders outside +/-40% of reference price.
    If desired price is outside range, clamp to safe boundary.

    Returns tick-aligned price. Returns desired unchanged if ref_price invalid.
    """
    if ref_price <= 0 or desired <= 0:
        return desired
    lo = ref_price * (1 - band) * 1.01  # 1% safety margin inside boundary
    hi = ref_price * (1 + band) * 0.99
    if desired < lo:
        return _snap_tick(lo)
    if desired > hi:
        return _snap_tick(hi)
    return _snap_tick(desired)


class KiteSession:
    """
    Manages Kite login and access token.
    Both EquityModule and FnoModule call kite() to get a KiteConnect instance.
    They share the access token but make independent API calls.
    """

    _kite = None
    _lock = threading.Lock()
    _last_login = None

    @classmethod
    def login(cls):
        """Full TOTP auto-login to Kite."""
        with cls._lock:
            try:
                kite = KiteConnect(api_key=KITE_API_KEY)

                # Step 1: Try saved token first
                token_data = load_json(KITE_TOKEN_FILE, {})
                saved_date = token_data.get("date", "")
                saved_token = token_data.get("access_token", "")

                if saved_date == str(today_ist()) and saved_token:
                    kite.set_access_token(saved_token)
                    try:
                        profile = kite.profile()
                        log.info(f"Kite: reused token for {profile.get('user_name', ZERODHA_USER_ID)}")
                        cls._kite = kite
                        cls._last_login = now_ist()
                        return True
                    except Exception:
                        log.info("Kite: saved token expired, doing fresh login")

                # Step 2: Fresh TOTP login
                if not ZERODHA_PASSWORD or not KITE_TOTP_SECRET:
                    log.error("Kite: missing password or TOTP secret in env")
                    return False

                session = requests.Session()
                session.headers.update({"User-Agent": "Mozilla/5.0"})

                # Login step 1
                r1 = session.post(
                    "https://kite.zerodha.com/api/login",
                    data={"user_id": ZERODHA_USER_ID, "password": ZERODHA_PASSWORD},
                    timeout=30,
                )
                if r1.status_code != 200:
                    log.error(f"Kite login step 1 failed: {r1.status_code}")
                    return False

                request_id = r1.json().get("data", {}).get("request_id", "")
                if not request_id:
                    log.error("Kite login: no request_id")
                    return False

                # Login step 2: TOTP
                time.sleep(1)
                totp = pyotp.TOTP(KITE_TOTP_SECRET).now()
                r2 = session.post(
                    "https://kite.zerodha.com/api/twofa",
                    data={
                        "user_id": ZERODHA_USER_ID,
                        "request_id": request_id,
                        "twofa_value": totp,
                        "twofa_type": "totp",
                    },
                    timeout=30,
                    allow_redirects=True,
                )

                # Extract request_token from redirect URL
                final_url = r2.url
                if "request_token=" not in final_url:
                    # Check response for token
                    if r2.status_code == 200 and "status" in r2.text:
                        # Try the session approach
                        login_url = kite.login_url()
                        r3 = session.get(login_url, allow_redirects=True, timeout=30)
                        final_url = r3.url

                if "request_token=" not in final_url:
                    log.error(f"Kite TOTP: no request_token in {final_url[:100]}")
                    return False

                request_token = final_url.split("request_token=")[1].split("&")[0]

                # Generate session
                data = kite.generate_session(request_token, api_secret=KITE_API_SECRET)
                access_token = data.get("access_token", "")
                if not access_token:
                    log.error("Kite: no access_token from generate_session")
                    return False

                kite.set_access_token(access_token)
                profile = kite.profile()
                log.info(f"Kite: logged in as {profile.get('user_name', ZERODHA_USER_ID)}")

                # Save token
                save_json(KITE_TOKEN_FILE, {
                    "access_token": access_token,
                    "date": str(today_ist()),
                })

                cls._kite = kite
                cls._last_login = now_ist()
                return True

            except Exception as e:
                log.error(f"Kite login failed: {e}")
                cls._kite = None
                return False

    @classmethod
    def kite(cls):
        """Get KiteConnect instance. Returns None if not logged in."""
        if cls._kite is None:
            cls.login()
        # V78: install recorder on the live kite object (idempotent, never raises)
        if cls._kite is not None and _TRADE_RECORDER_AVAILABLE:
            try:
                TradeRecorder.install(cls._kite)
            except Exception:
                pass  # Never let recorder failure block kite access
        return cls._kite

    @classmethod
    def ensure_connected(cls):
        """Make sure we have a valid Kite connection."""
        k = cls.kite()
        if k is None:
            return False
        try:
            k.profile()
            return True
        except Exception:
            log.info("Kite: token expired, re-logging in")
            # V78: uninstall recorder so it cleanly re-installs on the new kite object
            if _TRADE_RECORDER_AVAILABLE:
                try:
                    TradeRecorder.uninstall()
                except Exception:
                    pass
            cls._kite = None
            return cls.login()


# ═══════════════════════════════════════════════════════════════════════════════
#  TELEGRAM — Simple, direct. No shared gatekeeper.
# ═══════════════════════════════════════════════════════════════════════════════

def send_telegram(msg, silent=False):
    """Send a Telegram message. Returns True on success."""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return False
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        resp = requests.post(url, json={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": msg[:4000],
            "parse_mode": "Markdown",
            "disable_notification": silent,
        }, timeout=15)
        return resp.status_code == 200
    except Exception:
        return False


class TelegramThrottle:
    """
    Per-module message throttle. Prevents spam.
    Each module (equity/fno) has its own instance.
    """

    def __init__(self, name, max_per_hour=10):
        self.name = name
        self.max_per_hour = max_per_hour
        self._sent = []

    def can_send(self):
        now = time.time()
        self._sent = [t for t in self._sent if now - t < 3600]
        return len(self._sent) < self.max_per_hour

    def send(self, msg, force=False, silent=False):
        if not force and not self.can_send():
            return False
        prefix = f"*[{self.name}]*\n"
        result = send_telegram(prefix + msg, silent=silent)
        if result:
            self._sent.append(time.time())
        return result


# ═══════════════════════════════════════════════════════════════════════════════
#  GTT MANAGER — The bulletproof version. ALWAYS checks before placing.
# ═══════════════════════════════════════════════════════════════════════════════

# ======================================================================
#  M11 HEALTH MONITOR — critical alerts, no throttle
# ======================================================================

class HealthAlert:
    """Critical system alerts. No throttle. Fires on safety-critical events only."""
    def __init__(self):
        self._reconnect_alerted = False
        self.counters = {
            "orders_placed": 0,
            "orders_rejected": 0,
            "orders_cancelled": 0,
            "kite_login_fails": 0,
            "trailing_sl_errors": 0,
            "last_alert_ts": 0,
        }

    def fire(self, msg):
        """Send critical alert immediately (no throttle)."""
        try:
            send_telegram(f"⚠️ *[M11 HEALTH]*\n{msg}", silent=False)
            self.counters["last_alert_ts"] = time.time()
        except Exception as _e:
            log.error(f"M11 HealthAlert send failed: {_e}")

    def on_reconnect_ok(self):
        self._reconnect_alerted = False

    def on_reconnect_fail(self, count):
        if count >= 3 and not self._reconnect_alerted:
            self.fire(f"Kite reconnect failing ({count} attempts). Bot may sleep if it hits 5.")
            self._reconnect_alerted = True

    def reset_daily_counters(self):
        for k in ["orders_placed", "orders_rejected", "orders_cancelled",
                  "kite_login_fails", "trailing_sl_errors"]:
            self.counters[k] = 0


# Module-level singleton
health = HealthAlert()


# ======================================================================
#  M1 IV TRACKER — DTE-normalized ATM straddle proxy for IV Rank
# ======================================================================

class IVTracker:
    """Tracks daily IV proxy for all F&O underlyings.

    IV proxy = (ATM_CE_LTP + ATM_PE_LTP) / spot / sqrt(DTE/252)

    This is DTE-normalized so theta decay doesn't create fake IV trend.
    Stored in iv_history.json as: {symbol: [[date, proxy], ...]}
    Rolling 252 entries per symbol.
    """

    HISTORY_FILE = "/home/globalbot/data/iv_history.json"
    MAX_HISTORY = 252
    MIN_HISTORY_FOR_RANK = 30

    def __init__(self):
        self.history = {}
        self._load()

    def _load(self):
        try:
            import json, os
            if os.path.exists(self.HISTORY_FILE):
                with open(self.HISTORY_FILE, 'r') as f:
                    self.history = json.load(f)
                log.info(f"[M1] IV history loaded: {len(self.history)} symbols")
            else:
                log.info("[M1] IV history: starting fresh")
        except Exception as e:
            log.error(f"[M1] IV history load failed: {e}")
            self.history = {}

    def _save(self):
        try:
            import json
            with open(self.HISTORY_FILE, 'w') as f:
                json.dump(self.history, f)
        except Exception as e:
            log.error(f"[M1] IV history save failed: {e}")

    def record(self, symbol, iv_proxy, date_str=None):
        """Record today's IV proxy. date_str ISO format, default today."""
        if date_str is None:
            date_str = now_ist().date().isoformat()
        if symbol not in self.history:
            self.history[symbol] = []
        # Dedupe same date
        self.history[symbol] = [e for e in self.history[symbol] if e[0] != date_str]
        self.history[symbol].append([date_str, round(float(iv_proxy), 6)])
        # Trim to MAX_HISTORY
        self.history[symbol] = self.history[symbol][-self.MAX_HISTORY:]

    def iv_rank(self, symbol):
        """0-100. None if insufficient history."""
        h = self.history.get(symbol, [])
        if len(h) < self.MIN_HISTORY_FOR_RANK:
            return None
        values = [e[1] for e in h]
        current = values[-1]
        lo, hi = min(values), max(values)
        if hi == lo:
            return 50.0
        return round((current - lo) / (hi - lo) * 100, 2)

    def iv_percentile(self, symbol):
        """0-100. % of days current IV was higher than. None if insufficient."""
        h = self.history.get(symbol, [])
        if len(h) < self.MIN_HISTORY_FOR_RANK:
            return None
        values = [e[1] for e in h]
        current = values[-1]
        below = sum(1 for v in values if v < current)
        return round(below / len(values) * 100, 2)

    def stats(self):
        """Summary for logging."""
        ready = sum(1 for s, h in self.history.items() if len(h) >= self.MIN_HISTORY_FOR_RANK)
        total = len(self.history)
        avg_len = sum(len(h) for h in self.history.values()) / total if total else 0
        return {"symbols": total, "rank_ready": ready, "avg_history_days": round(avg_len, 1)}


# Module-level singleton
iv_tracker = IVTracker()


def compute_atm_iv_proxy(symbol, kite, fno_module):
    """Fetch ATM CE+PE, compute DTE-normalized straddle/spot proxy.
    Returns float or None if fetch fails.
    """
    try:
        import math
        # Get spot price
        # PATCH_V74: use global _index_spot_key() helper (handles FINNIFTY/MIDCPNIFTY)
        spot_key = _index_spot_key(symbol)
        spot_q = kite.quote([spot_key]).get(spot_key, {})
        spot = spot_q.get("last_price") or 0
        if not spot or spot <= 0:
            log.warning(f"[M1_DIAG] {symbol}: spot fetch failed (spot={spot})")
            return None

        # PATCH_V53_IV_PROXY_STRIKE_FIX: reuse OISnapshotEngine._get_atm_strikes
        # (single source of truth) instead of hardcoded step-table that did not
        # match real NSE strike intervals. Old code returned empty for every
        # symbol because guessed atm rarely matched a real strike.
        oi = getattr(fno_module, "_oi_engine", None)
        if oi is None:
            log.warning(f"[M1_DIAG] {symbol}: oi_engine is None on fno_module (type={type(fno_module).__name__})")
            return None

        strikes = oi._get_atm_strikes(symbol, spot)
        if not strikes:
            log.warning(f"[M1_DIAG] {symbol}: _get_atm_strikes returned empty (spot={spot})")
            return None

        instruments = oi._get_current_expiry_instruments(symbol, strikes)
        if not instruments or len(instruments) < 2:
            log.warning(f"[M1_DIAG] {symbol}: instruments empty/short (n={len(instruments) if instruments else 0}, strikes={strikes})")
            return None

        # Find CE and PE
        ce_sym = pe_sym = None
        expiry = None
        for inst in instruments:
            its = inst.get("instrument_type")
            if its == "CE":
                ce_sym = f"NFO:{inst['tradingsymbol']}"
                expiry = inst.get("expiry")
            elif its == "PE":
                pe_sym = f"NFO:{inst['tradingsymbol']}"

        if not ce_sym or not pe_sym or not expiry:
            log.warning(f"[M1_DIAG] {symbol}: ce/pe/expiry missing (ce={ce_sym}, pe={pe_sym}, exp={expiry})")
            return None

        # DTE
        from datetime import datetime  # PATCH_V70_AUDIT_CLEANUP: removed unused 'date' (only .date() methods used)
        if isinstance(expiry, str):
            exp_date = datetime.strptime(expiry, "%Y-%m-%d").date()
        elif hasattr(expiry, "date"):
            exp_date = expiry.date()
        else:
            exp_date = expiry
        dte = max((exp_date - now_ist().date()).days, 1)

        # Fetch CE + PE LTP in one batch
        opt_q = kite.quote([ce_sym, pe_sym])
        ce_ltp = opt_q.get(ce_sym, {}).get("last_price") or 0
        pe_ltp = opt_q.get(pe_sym, {}).get("last_price") or 0
        if ce_ltp <= 0 or pe_ltp <= 0:
            log.warning(f"[M1_DIAG] {symbol}: ltp fail (ce_ltp={ce_ltp}, pe_ltp={pe_ltp})")
            return None

        straddle = ce_ltp + pe_ltp
        # DTE-normalized: proxy = (straddle / spot) / sqrt(DTE/252)
        proxy = (straddle / spot) / math.sqrt(dte / 252.0)
        return proxy
    except Exception as e:
        log.warning(f"[M1] IV proxy compute failed for {symbol}: {e}")
        return None


def daily_iv_snapshot(kite, fno_module):
    """Fetch and record IV proxy for all tracked F&O symbols. Called once/day."""
    try:
        symbols = ["NIFTY", "BANKNIFTY"] + sorted(fno_module.fno_stocks)
        ok = 0
        fail = 0
        for sym in symbols:
            proxy = compute_atm_iv_proxy(sym, kite, fno_module)
            if proxy is not None:
                iv_tracker.record(sym, proxy)
                ok += 1
            else:
                fail += 1
            time.sleep(0.3)  # FIX_V50: Kite rate limit — 220 symbols x 0.3s = 66s total
        iv_tracker._save()
        s = iv_tracker.stats()
        log.info(f"[M1] Daily snapshot: ok={ok} fail={fail} | stats: {s}")
        return ok, fail
    except Exception as e:
        log.error(f"[M1] Daily snapshot crashed: {e}")
        return 0, 0




class GTTManager:
    """
    Manages GTT orders on Kite. NEVER places blind GTTs.

    Every operation:
      1. Calls kite.get_gtts() to get ALL active GTTs
      2. Builds symbol → gtt_id map
      3. If symbol has GTT → modify (not create new)
      4. If symbol has no GTT → place new
      5. Cleans orphan GTTs (stock sold but GTT still active)
      6. Removes duplicate GTTs (keeps newest, deletes older)

    Equity and F&O each have their own GTTManager instance.
    They independently call Kite API — no shared state.
    """

    def __init__(self, name):
        self.name = name  # "EQUITY" or "FNO"

    def _get_active_gtts(self):
        """Fetch all active GTTs from Kite, grouped by symbol."""
        kite = KiteSession.kite()
        if not kite:
            return {}
        try:
            all_gtts = kite.get_gtts() or []
            by_symbol = defaultdict(list)
            for g in all_gtts:
                if g.get("status") == "active":
                    sym = g.get("condition", {}).get("tradingsymbol", "")
                    if sym:
                        by_symbol[sym].append(g)
            return dict(by_symbol)
        except Exception as e:
            log.error(f"GTT {self.name}: fetch failed: {e}")
            return {}

    def ensure_gtt(self, symbol, exchange, qty, sl_price, target_price=0):
        """
        Ensure exactly ONE GTT exists for this symbol.
        If GTT exists → modify SL/target (only if SL moved UP).
        If no GTT → place new.
        If duplicates → delete extras.

        Returns gtt_id or None.
        """
        kite = KiteSession.kite()
        if not kite or qty <= 0 or sl_price <= 0:
            return None

        # PATCH_V1_INLOSS_FLOOR: prevent placing SL above LTP (broken stop)
        try:
            _ltp_data = kite.ltp([f"{exchange}:{symbol}"])
            _ltp = _ltp_data.get(f"{exchange}:{symbol}", {}).get("last_price", 0)
            if _ltp > 0 and sl_price >= _ltp:
                # FIX_V57_EQ_TIGHT_CAP: equity cap 15% -> 2.5%. F&O retains 15%.
                # Uses self.name ("EQUITY" or "FNO") for reliable module detection.
                _cap_pct = 0.975 if self.name == "EQUITY" else 0.85
                _new_sl = _snap_tick(_ltp * _cap_pct, symbol=symbol, exchange=exchange)
                log.warning(f"[PATCH_V1_INLOSS][FIX_V57_EQ_TIGHT_CAP] {self.name} {symbol}: SL={sl_price} >= LTP={_ltp}, cap_pct={_cap_pct} -> {_new_sl}")
                sl_price = _new_sl
        except Exception as _e:
            log.error(f"[PATCH_V1_INLOSS] {self.name} {symbol}: LTP check failed: {_e}")
        # END_PATCH_V1_INLOSS_FLOOR

        active = self._get_active_gtts()
        existing = active.get(symbol, [])

        # ── Delete duplicates (keep newest) ──
        if len(existing) > 1:
            existing.sort(key=lambda g: g.get("id", 0), reverse=True)
            for dup in existing[1:]:
                try:
                    kite.delete_gtt(dup["id"])
                    log.info(f"GTT {self.name}: deleted duplicate {dup['id']} for {symbol}")
                except Exception as _e71:
                    log.warning(f"GTT {self.name}: failed to delete duplicate {dup['id']} for {symbol}: {_e71} [PATCH_V71_RESILIENCE]")
            existing = [existing[0]]

        # ── Modify existing GTT ──
        if existing:
            gtt = existing[0]
            gtt_id = gtt["id"]
            old_triggers = gtt.get("trigger_values", [])
            old_sl = old_triggers[0] if old_triggers else 0

            # Only modify if new SL is HIGHER than old SL (SL only moves UP)
            if sl_price > old_sl:
                try:
                    # FIX_V53_GAP: fetch real LTP and ensure trigger gap
                    try:
                        _lt = kite.ltp([f"{exchange}:{symbol}"])
                        _ltp = _lt.get(f"{exchange}:{symbol}", {}).get("last_price", 0) or (sl_price * 1.05)
                    except Exception:
                        _ltp = sl_price * 1.05
                    safe_sl = _gtt_safe_trigger(sl_price, _ltp, side="SELL")
                    if safe_sl <= 0:
                        log.error(f"GTT {self.name}: {symbol} modify safe_sl invalid (sl={sl_price} ltp={_ltp}) [FIX_V53_GAP]")
                        return None
                    if abs(safe_sl - sl_price) > 0.01:
                        log.warning(f"GTT {self.name}: {symbol} modify SL adjusted {sl_price}→{safe_sl} for gap rule [FIX_V53_GAP]")

                    triggers = [safe_sl]
                    orders = [
                        {
                            "transaction_type": "SELL",
                            "quantity": qty,
                            "price": round(safe_sl * 0.98, 1),
                            "order_type": "LIMIT",
                            "product": "CNC" if exchange == "NSE" else "NRML",
                        }
                    ]
                    # FIX_V29_SINGLE_ONLY: always SINGLE, no OCO. Software handles target exit.
                    gtt_type = kite.GTT_TYPE_SINGLE
                    _safe_modify_gtt(kite, 
                        trigger_id=gtt_id,
                        trigger_type=gtt_type,
                        tradingsymbol=symbol,
                        exchange=exchange,
                        trigger_values=triggers,
                        last_price=_ltp,
                        orders=orders,
                    )
                    log.info(f"GTT {self.name}: modified {symbol} SL {old_sl}→{safe_sl}")
                    return gtt_id
                except Exception as e:
                    # V27 PATCH (Bug #3): NEVER delete-and-replace with looser SL on modify failure.
                    # If modify fails (rate limit, concurrent edit, network), leave existing GTT
                    # intact and retry on next tick. One stale tick beats losing protection.
                    if "trigger type" in str(e).lower():
                        try:
                            kite.delete_gtt(gtt_id)
                            log.info(f"GTT {self.name}: deleted old GTT {gtt_id} for {symbol} (trigger type change)")
                            return self.ensure_gtt(symbol, exchange, qty, sl_price, target_price)
                        except Exception as e2:
                            log.error(f"GTT {self.name}: delete+recreate failed for {symbol} ({e2}) - NO GTT PROTECTION")
                            return None  # V28 A5: signal failure, caller must retry
                    else:
                        # V28 A5: return None instead of fake-success gtt_id
                        # Caller's trailing SL logic will retry on next tick.
                        log.error(f"GTT {self.name}: modify {symbol} failed ({e}) - leaving existing GTT intact (will retry next tick)")
                        return None
            else:
                # PATCH_V70_AUDIT_CLEANUP: was unreachable after both try and except returned
                log.debug(f"GTT {self.name}: {symbol} SL {sl_price} <= existing {old_sl}, no change")
            return gtt_id

        # ── Place new GTT ──
        try:
            # FIX_V53_GAP: ensure trigger >= 0.25% away from LTP (Kite requirement)
            ltp_data = kite.ltp([f"{exchange}:{symbol}"])
            last_price = ltp_data.get(f"{exchange}:{symbol}", {}).get("last_price", sl_price * 1.05)
            safe_sl = _gtt_safe_trigger(sl_price, last_price, side="SELL")
            if safe_sl <= 0:
                log.error(f"GTT {self.name}: {symbol} safe_sl invalid (sl={sl_price} ltp={last_price}) [FIX_V53_GAP]")
                return None
            if abs(safe_sl - sl_price) > 0.01:
                log.warning(f"GTT {self.name}: {symbol} SL adjusted {sl_price}→{safe_sl} for Kite gap rule [FIX_V53_GAP]")

            triggers = [safe_sl]
            orders = [
                {
                    "transaction_type": "SELL",
                    "quantity": qty,
                    "price": round(safe_sl * 0.98, 1),
                    "order_type": "LIMIT",
                    "product": "CNC" if exchange == "NSE" else "NRML",
                }
            ]
            # FIX_V29_SINGLE_ONLY: always SINGLE, no OCO. Software handles target exit.
            gtt_type = kite.GTT_TYPE_SINGLE

            gtt_id = _safe_place_gtt(kite, 
                trigger_type=gtt_type,
                tradingsymbol=symbol,
                exchange=exchange,
                trigger_values=triggers,
                last_price=last_price,
                orders=orders,
            )
            log.info(f"GTT {self.name}: placed {symbol} SL={safe_sl} Target={target_price} ID={gtt_id}")
            return gtt_id
        except Exception as e:
            log.error(f"GTT {self.name}: place {symbol} failed: {e}")
            return None




# ═══════════════════════════════════════════════════════════════════════════════
#  TECHNICAL INDICATORS — Just 3. Each does ONE job.
# ═══════════════════════════════════════════════════════════════════════════════

def calc_rsi(closes, period=14):
    """
    Calculate RSI from a list of closing prices.
    Returns RSI value (0-100) or 50 if insufficient data.
    """
    if len(closes) < period + 1:
        return 50.0

    gains = []
    losses = []
    for i in range(1, len(closes)):
        delta = closes[i] - closes[i - 1]
        if delta > 0:
            gains.append(delta)
            losses.append(0)
        else:
            gains.append(0)
            losses.append(abs(delta))

    if len(gains) < period:
        return 50.0

    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period

    for i in range(period, len(gains)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period

    if avg_loss == 0:
        return 100.0

    rs = avg_gain / avg_loss
    return 100.0 - (100.0 / (1.0 + rs))


def calc_volume_ratio(volumes, period=20):
    """
    Calculate today's volume vs 20-day average.
    Returns ratio (1.0 = average, 2.0 = double average).
    """
    if len(volumes) < period + 1:
        return 1.0
    avg_vol = sum(volumes[-period - 1:-1]) / period
    if avg_vol <= 0:
        return 1.0
    return volumes[-1] / avg_vol


def calc_atr(highs, lows, closes, period=14):
    """
    Calculate Average True Range.
    Returns ATR value (in price units).
    """
    if len(closes) < period + 1:
        return closes[-1] * 0.02 if closes else 0  # Default 2% of price

    true_ranges = []
    for i in range(1, len(closes)):
        tr = max(
            highs[i] - lows[i],
            abs(highs[i] - closes[i - 1]),
            abs(lows[i] - closes[i - 1]),
        )
        true_ranges.append(tr)

    if len(true_ranges) < period:
        return sum(true_ranges) / len(true_ranges) if true_ranges else 0

    atr = sum(true_ranges[:period]) / period
    for i in range(period, len(true_ranges)):
        atr = (atr * (period - 1) + true_ranges[i]) / period

    return atr


# ═══════════════════════════════════════════════════════════════════════════════
#  NEWS SCANNER — Each module has its own instance, own interpretation
# ═══════════════════════════════════════════════════════════════════════════════

class NewsScanner:
    """
    Scans RSS feeds for stock-moving news.
    Returns candidates with symbol, direction, score, catalyst text.

    Each module (equity/fno) creates its own NewsScanner.
    Same feeds, but each module filters and interprets independently.
    """

    def __init__(self, name, valid_symbols=None, name_map=None, trendlyne_name_map=None):
        self.name = name
        self.valid_symbols = valid_symbols or set()
        self.name_map = name_map or {}  # PATCH_V28_SYMBOL_REGEX: company name → symbol
        self.trendlyne_name_map = trendlyne_name_map or {}  # FIX_V48: Stock Name to NSE Code
        self._seen = set()  # Dedup hashes
        self._seen_file = DATA_DIR / f"news_seen_{name.lower()}.json"
        self._load_seen()

    def _load_seen(self):
        data = load_json(self._seen_file, {"hashes": []})
        self._seen = set(data.get("hashes", [])[-5000:])  # Keep last 5000

    def _save_seen(self):
        save_json(self._seen_file, {"hashes": list(self._seen)[-5000:]})

    def _hash(self, title, link):
        raw = f"{title.strip().lower()}|{link.strip()}"
        return hashlib.md5(raw.encode()).hexdigest()[:16]

    def _extract_symbols(self, text):
        """Try to find NSE symbols mentioned in text.
        PATCH_V28_SYMBOL_REGEX: word-boundary matching to prevent
        'BEL' matching 'BELOW', 'HAL' matching 'SHALL', etc.
        Also maps common company names to NSE symbols.
        PATCH_V62_INDENT_FIX: _NAME_MAP now assigned once, not per-iteration.
        """
        import re as _re
        found = []
        text_upper = text.upper()
        for sym in self.valid_symbols:
            if len(sym) < 2:
                continue
            # Word-boundary match: sym must be a standalone word
            if _re.search(r'\b' + _re.escape(sym) + r'\b', text_upper):
                found.append(sym)
        # Company name -> symbol from Kite instruments (all 3000+ stocks)
        # PATCH_V62_INDENT_FIX: dedented out of the for-sym loop
        _NAME_MAP = self.name_map if self.name_map else {
            "TATA CONSULTANCY": "TCS", "HDFC BANK": "HDFCBANK",
            "INFOSYS": "INFY", "RELIANCE INDUSTRIES": "RELIANCE",
            "STATE BANK": "SBIN", "ICICI BANK": "ICICIBANK",
            "AXIS BANK": "AXISBANK", "KOTAK MAHINDRA": "KOTAKBANK",
            "LARSEN": "LT", "BAJAJ FINANCE": "BAJFINANCE",
            "HINDUSTAN UNILEVER": "HINDUNILVR", "TITAN COMPANY": "TITAN",
            "POWER FINANCE": "PFC", "BHARTI AIRTEL": "BHARTIARTL",
            "TATA MOTORS": "TATAMOTORS", "TATA STEEL": "TATASTEEL",
            "MARUTI SUZUKI": "MARUTI", "SUN PHARMA": "SUNPHARMA",
            "WIPRO": "WIPRO", "TECH MAHINDRA": "TECHM",
            "JSW STEEL": "JSWSTEEL", "ADANI PORTS": "ADANIPORTS",
            "ULTRA TECH": "ULTRACEMCO", "POWER GRID": "POWERGRID",
            "COAL INDIA": "COALINDIA", "INDIAN OIL": "IOC",
            "SHRIRAM FINANCE": "SHRIRAMFIN", "CROMPTON GREAVES": "CROMPTON",
        }
        for name, sym in _NAME_MAP.items():
            if name in text_upper and sym not in found and sym in self.valid_symbols:
                found.append(sym)
        # FIX_V48 Patch B: Trendlyne name fallback -- longest-match-wins
        if self.trendlyne_name_map and not found:
            _best_name = ""
            _best_sym = ""
            for tname, tsym in self.trendlyne_name_map.items():
                if tname in text_upper and len(tname) > len(_best_name):
                    if tsym in self.valid_symbols:
                        _best_name = tname
                        _best_sym = tsym
            if _best_sym and _best_sym not in found:
                found.append(_best_sym)
        return found

    def _detect_direction(self, text):
        """
        PATCH_V7_SIGNAL_QUALITY
        Detect BULLISH/BEARISH from full text (title + summary).
        Requires minimum 2 keyword matches OR clear margin (diff >= 2).
        Weak single-keyword matches return NEUTRAL to avoid false signals.
        """
        text_lower = (text or "").lower()
        bull_score = sum(1 for kw in BULLISH_KEYWORDS if kw in text_lower)
        bear_score = sum(1 for kw in BEARISH_KEYWORDS if kw in text_lower)

        # Require conviction: either 2+ matches on winning side, or clear margin
        if bull_score >= 2 and bull_score > bear_score:
            return "BULLISH", min(55 + bull_score * 10, 95)
        elif bear_score >= 2 and bear_score > bull_score:
            return "BEARISH", min(55 + bear_score * 10, 95)
        elif bull_score - bear_score >= 2:
            return "BULLISH", min(55 + bull_score * 10, 95)
        elif bear_score - bull_score >= 2:
            return "BEARISH", min(55 + bear_score * 10, 95)
        return "NEUTRAL", 30

    def scan(self):
        """
        Scan all RSS feeds. Return list of candidates:
        [{"symbol": "ONGC", "direction": "BULLISH", "score": 80, "catalyst": "..."}]
        """
        candidates = []
        total_checked = 0

        for feed_name, feed_url in NEWS_FEEDS.items():
            try:
                parsed = feedparser.parse(feed_url)
                for entry in parsed.entries[:20]:
                    total_checked += 1
                    title = entry.get("title", "")
                    link = entry.get("link", "")

                    h = self._hash(title, link)
                    if h in self._seen:
                        continue
                    self._seen.add(h)

                    # PATCH_V7_SIGNAL_QUALITY: read full article text, not just title
                    summary = entry.get("summary", "") or entry.get("description", "") or ""
                    # Strip HTML tags crudely
                    import re as _re
                    summary_clean = _re.sub(r"<[^>]+>", " ", summary)
                    full_text = f"{title} {summary_clean}"

                    # Find symbols (check both title and summary)
                    symbols = self._extract_symbols(full_text)
                    if not symbols:
                        continue

                    # Detect direction from full text
                    direction, score = self._detect_direction(full_text)
                    if direction == "NEUTRAL":
                        continue

                    _now_iso = now_ist().isoformat()
                    for sym in symbols:
                        candidates.append({
                            "symbol": sym,
                            "direction": direction,
                            "score": score,
                            "catalyst": f"{feed_name}: {title[:80]}",
                            "source": "NEWS",
                            "timestamp": _now_iso,
                            # FIX_V39_SCOUT_VALIDATION
                            "title": title,
                            "summary": summary_clean,
                            "link": link,
                            "source_feed": feed_name,
                        })

            except Exception as e:
                log.debug(f"News {self.name}: {feed_name} error: {e}")

        self._save_seen()
        if candidates:
            log.info(f"News {self.name}: {len(candidates)} candidates from {total_checked} headlines")
        return candidates


# ═══════════════════════════════════════════════════════════════════════════════
#  FILING MONITOR — Scrapes NSE corporate filings for catalysts
# ═══════════════════════════════════════════════════════════════════════════════

class FilingMonitor:
    """
    Monitors NSE corporate filings for trading catalysts.
    Catches promoter buying, buybacks, order wins, earnings BEFORE news RSS.
    Each module has its own instance — interprets filings independently.
    """

    NSE_FILING_URL = "https://www.nseindia.com/api/corporate-announcements?index=equities&from_date={from_date}&to_date={to_date}"

    # Filing keywords mapped to signal type
    FILING_KW_MAP = {
        # BULLISH signals
        "buyback": "BUYBACK", "buy back": "BUYBACK", "buy-back": "BUYBACK",
        "promoter buy": "PROMOTER_BUY", "promoter acqui": "PROMOTER_BUY",
        "increase in shareholding": "PROMOTER_BUY", "acquired shares": "PROMOTER_BUY",
        "order win": "ORDER_WIN", "order received": "ORDER_WIN", "bagged order": "ORDER_WIN",
        "wins order": "ORDER_WIN", "contract awarded": "ORDER_WIN", "new order": "ORDER_WIN",
        "dividend": "DIVIDEND", "interim dividend": "DIVIDEND", "final dividend": "DIVIDEND",
        "bonus": "BONUS", "bonus issue": "BONUS",
        "stock split": "STOCK_SPLIT", "sub-division": "STOCK_SPLIT",
        "acquisition": "ACQUISITION", "acquired": "ACQUISITION",
        "expansion": "EXPANSION", "new plant": "EXPANSION", "capacity addition": "EXPANSION",
        "revenue growth": "RESULTS_BEAT", "profit growth": "RESULTS_BEAT",
        "beats estimate": "RESULTS_BEAT", "above estimate": "RESULTS_BEAT",
        # BEARISH signals
        "promoter sell": "PROMOTER_SELL", "promoter disp": "PROMOTER_SELL",
        "decrease in shareholding": "PROMOTER_SELL", "sold shares": "PROMOTER_SELL",
        "qip": "QIP", "qualified institution": "QIP",
        "ofs": "OFS", "offer for sale": "OFS",
        "downgrade": "CREDIT_DOWNGRADE", "rating downgrade": "CREDIT_DOWNGRADE",
        "fraud": "FRAUD", "scam": "FRAUD", "penalty": "PENALTY", "fine imposed": "PENALTY",
        "loss widen": "RESULTS_MISS", "below estimate": "RESULTS_MISS",
        "misses estimate": "RESULTS_MISS", "revenue decline": "RESULTS_MISS",
    }

    def __init__(self, name, valid_symbols=None):
        self.name = name
        self.valid_symbols = valid_symbols or set()
        self._seen = set()
        self._seen_file = DATA_DIR / f"filing_seen_{name.lower()}.json"
        self._load_seen()
        self._session = requests.Session()
        self._session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/json",
            "Accept-Language": "en-US,en;q=0.9",
        })

    def _load_seen(self):
        data = load_json(self._seen_file, {"seen": []})
        self._seen = set(data.get("seen", [])[-3000:])

    def _save_seen(self):
        save_json(self._seen_file, {"seen": list(self._seen)[-3000:]})

    def scan(self):
        """
        Fetch recent NSE filings and extract trading signals.
        Returns list of candidates with symbol, direction, score, catalyst.
        PATCH_V62_FILING_RETRY: 3 retries with cookie refresh. NSE API 403s frequently.
        """
        candidates = []
        today = today_ist()
        from_date = (today - timedelta(days=1)).strftime("%d-%m-%Y")
        to_date = today.strftime("%d-%m-%Y")

        url = self.NSE_FILING_URL.format(from_date=from_date, to_date=to_date)
        filings = None
        for attempt in range(3):
            try:
                # Refresh NSE cookies each attempt
                self._session.get("https://www.nseindia.com", timeout=10)
                time.sleep(0.5 + attempt * 0.5)  # backoff: 0.5s, 1.0s, 1.5s
                resp = self._session.get(url, timeout=15)

                if resp.status_code == 200:
                    try:
                        filings = resp.json()
                        if isinstance(filings, list):
                            break  # success
                        filings = None
                    except Exception:
                        filings = None
                else:
                    log.debug(f"Filing {self.name} attempt {attempt+1}: NSE returned {resp.status_code}")
            except Exception as e:
                log.debug(f"Filing {self.name} attempt {attempt+1}: {e}")

        if not filings:
            log.debug(f"Filing {self.name}: all 3 attempts failed")
            return candidates

        try:

            for filing in filings[:100]:  # Process latest 100
                symbol = filing.get("symbol", "")
                # PATCH_V11_FILING_TEXT: read attchmntText (real content) in addition to desc (category label)
                # NSE puts the real filing summary in attchmntText; desc is just a category like "General Updates"
                _desc = filing.get("desc", "") or ""
                _body = filing.get("attchmntText", "") or ""
                subject = (_desc + " " + _body).lower()
                filing_id = filing.get("an_dt", "") + symbol

                if not symbol or not subject:
                    continue

                # Dedup
                h = hashlib.md5(filing_id.encode()).hexdigest()[:12]
                if h in self._seen:
                    continue
                self._seen.add(h)

                # Skip if not in our valid symbols
                if self.valid_symbols and symbol not in self.valid_symbols:
                    continue

                # Match filing keywords
                matched_type = None
                for keyword, filing_type in self.FILING_KW_MAP.items():
                    if keyword in subject:
                        matched_type = filing_type
                        break

                if not matched_type:
                    continue

                signal = FILING_SIGNALS.get(matched_type, {})
                direction = signal.get("direction", "NEUTRAL")
                score = signal.get("score", 50)

                if direction == "NEUTRAL":
                    continue

                candidates.append({
                    "symbol": symbol,
                    "direction": direction,
                    "score": score,
                    "catalyst": f"FILING {matched_type}: {subject[:80]}",
                    "source": "FILING",
                    "filing_type": matched_type,
                    # FIX_V39_SCOUT_VALIDATION
                    "title": _desc[:200] if _desc else f"FILING {matched_type} {symbol}",
                    "summary": _body[:4000] if _body else _desc[:2000],
                    "link": "",
                    "source_feed": "NSE_FILINGS",
                })

        except Exception as e:
            log.debug(f"Filing {self.name}: scan error: {e}")

        self._save_seen()
        if candidates:
            log.info(f"Filing {self.name}: {len(candidates)} signals from NSE filings")
        return candidates


# ═══════════════════════════════════════════════════════════════════════════════
#  BROKERAGE MONITOR — Catches analyst upgrades/downgrades from RSS
# ═══════════════════════════════════════════════════════════════════════════════

class BrokerageMonitor:
    """
    Scans RSS feeds specifically for brokerage calls.
    "HDFC Securities upgrades TCS to BUY with target Rs.4500"
    Each module has its own instance.
    """

    BROKERAGE_KEYWORDS = [
        "buy rating", "sell rating", "target price", "price target",
        "upgrade", "downgrade", "outperform", "underperform",
        "top pick", "accumulate", "overweight", "underweight",
        "raises target", "cuts target", "initiates coverage",
        "maintains buy", "reiterates buy", "conviction buy",
        "stocks to buy", "brokerage picks", "analyst pick",
    ]

    TIER1_BROKERAGES = [
        "hdfc securities", "motilal oswal", "clsa", "jm financial",
        "nuvama", "icici direct", "kotak securities", "jefferies",
        "goldman sachs", "morgan stanley", "nomura", "macquarie",
        "ubs", "bernstein", "citi", "bofa", "jp morgan",
    ]

    def __init__(self, name, valid_symbols=None):
        self.name = name
        self.valid_symbols = valid_symbols or set()
        self._seen = set()

    def scan(self, news_candidates=None):
        """
        Independent RSS scan for brokerage-specific headlines.
        PATCH_V62_BROKERAGE_DOC: news_candidates param is kept for backward compat
        but is not consumed. All call sites work — RSS scan is the primary path.
        If in future we want to enhance already-fetched news items with brokerage
        tier info, implement that here. For now, RSS-only is fine.
        Returns list of candidates with brokerage tier info.
        """
        candidates = []

        # Scan RSS feeds for brokerage-specific headlines
        for feed_name, feed_url in NEWS_FEEDS.items():
            try:
                parsed = feedparser.parse(feed_url)
                for entry in parsed.entries[:15]:
                    title = entry.get("title", "")
                    title_lower = title.lower()

                    # Must contain brokerage keyword
                    if not any(kw in title_lower for kw in self.BROKERAGE_KEYWORDS):
                        continue

                    link = entry.get("link", "")
                    h = hashlib.md5(f"{title}{link}".encode()).hexdigest()[:12]
                    if h in self._seen:
                        continue
                    self._seen.add(h)

                    # Extract symbols
                    text_upper = title.upper()
                    symbols = [s for s in self.valid_symbols if len(s) >= 2 and __import__("re").search(r"\b" + __import__("re").escape(s) + r"\b", text_upper)]  # PATCH_V28_SYMBOL_REGEX
                    if not symbols:
                        continue

                    # Detect direction
                    is_upgrade = any(kw in title_lower for kw in
                                     ["buy", "upgrade", "outperform", "overweight",
                                      "raises target", "top pick", "accumulate"])
                    is_downgrade = any(kw in title_lower for kw in
                                       ["sell", "downgrade", "underperform", "underweight",
                                        "cuts target", "reduce"])

                    if not is_upgrade and not is_downgrade:
                        continue

                    # Tier-1 brokerage bonus
                    is_tier1 = any(b in title_lower for b in self.TIER1_BROKERAGES)
                    score = 80 if is_tier1 else 65

                    direction = "BULLISH" if is_upgrade else "BEARISH"

                    for sym in symbols:
                        candidates.append({
                            "symbol": sym,
                            "direction": direction,
                            "score": score,
                            "catalyst": f"BROKERAGE {'T1 ' if is_tier1 else ''}{title[:80]}",
                            "source": "BROKERAGE",
                            # FIX_V39_SCOUT_VALIDATION
                            "title": title,
                            "summary": "",
                            "link": link,
                            "source_feed": feed_name,
                        })

            except Exception:
                pass

        if candidates:
            log.info(f"Brokerage {self.name}: {len(candidates)} analyst calls detected")
        return candidates


# ═══════════════════════════════════════════════════════════════════════════════
#  MACRO EVENT DETECTOR — RBI, war, oil, geopolitical events
# ═══════════════════════════════════════════════════════════════════════════════

class MacroDetector:
    """
    Detects macro events from news and maps them to sector/stock plays.
    "RBI rate cut" → BULLISH BANKING → CE on SBIN/HDFCBANK
    "Iran tensions" → BULLISH ENERGY → CE on ONGC/BPCL
    "IT spending cut" → BEARISH IT → PE on TCS/INFY

    Each module has its own instance.
    """

    # Macro theme → affected stocks and direction
    MACRO_MAP = {
        # FIX_PATH3_COMPREHENSIVE: 198 themes with avg 14 stocks each (Apr 24 2026)
        # Full NSE sector mapping via Trendlyne industry data, min Rs.500Cr mcap filter
        "500 gw target": {"direction": "BULLISH", "score": 70, "stocks": ["POWERGRID", "NTPCGREEN", "TORNTPOWER", "INOXGREEN", "JSWENERGY", "JPPOWER", "SUZLON", "NAVA", "WAAREERTL", "ADANIPOWER", "KPIGREEN", "NTPC", "ORIANA", "SOLARWORLD", "WAAREEENER", "CESC", "ACMESOLAR", "ADANIGREEN", "TATAPOWER", "INOXWIND"]},
        "5g rollout": {"direction": "BULLISH", "score": 70, "stocks": ["BHARTIARTL", "IDEA", "TATACOMM", "BHARTIHEXA", "HFCL", "STLTECH", "ITI", "RELIANCE"]},
        "5g spectrum": {"direction": "BULLISH", "score": 70, "stocks": ["BHARTIARTL", "IDEA", "TATACOMM", "BHARTIHEXA", "HFCL", "STLTECH", "ITI", "RELIANCE"]},
        "acc pli": {"direction": "BULLISH", "score": 75, "stocks": ["EXIDEIND", "ARE&M", "RELIANCE", "OLAELEC"]},
        "accounting fraud": {"direction": "BEARISH", "score": 90, "stocks": []},
        "aerospace": {"direction": "BULLISH", "score": 70, "stocks": ["HAL", "BHARATFORG", "MIDHANI", "BEL", "MTARTECH", "PARAS", "ASTRAMICRO", "DATAPATTNS"]},
        "affordable housing": {"direction": "BULLISH", "score": 70, "stocks": ["LICHSGFIN", "CANFINHOME", "APTUS", "AADHARHFC", "HOMEFIRST", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC", "RAMCOCEM", "JSWCEMENT", "INDIACEM"]},
        "agri growth": {"direction": "BULLISH", "score": 60, "stocks": ["ITC", "MARICO", "DABUR", "COROMANDEL", "FACT", "UPL", "PIIND", "SUMICHEM", "BAYERCROP", "CHAMBLFERT", "PARADEEP", "ESCORTS", "M&M"]},
        "agrochem": {"direction": "BULLISH", "score": 65, "stocks": ["UPL", "PIIND", "SUMICHEM", "BAYERCROP", "SHARDACROP", "RALLIS", "DHANUKA", "NACLIND", "BHAGCHEM", "BHARATRAS", "INSECTICID", "GSPCROP"]},
        "ai capex": {"direction": "BULLISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI"]},
        "ai investment": {"direction": "BULLISH", "score": 65, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI"]},
        "airline": {"direction": "BULLISH", "score": 55, "stocks": ["INDIGO"]},
        "airport expansion": {"direction": "BULLISH", "score": 70, "stocks": ["GMRAIRPORT", "ADANIPORTS"]},
        "aluminium price": {"direction": "BULLISH", "score": 70, "stocks": ["VEDL", "HINDZINC", "HINDALCO", "NATIONALUM", "JAINREC", "GRAVITA", "POCL", "ARFIN", "MAANALU", "MMP", "AMCL", "BAHETI"]},
        "atmanirbhar": {"direction": "BULLISH", "score": 70, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "DIXON", "BHARATFORG"]},
        "atmanirbhar bharat": {"direction": "BULLISH", "score": 75, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "DIXON", "BHARATFORG"]},
        "auto sales": {"direction": "BULLISH", "score": 65, "stocks": ["MARUTI", "M&M", "BAJAJ-AUTO", "EICHERMOT", "TVSMOTOR", "TMCV", "HYUNDAI", "MOTHERSON", "TMPV", "BOSCHLTD", "HEROMOTOCO", "ASHOKLEY", "UNOMINDA", "TIINDIA", "MRF", "BALKRISIND", "ESCORTS", "SONACOMS", "ATHERENERG", "ENDURANCE", "EXIDEIND", "TVSHLTD", "ZFCVINDIA", "APOLLOTYRE", "FORCEMOT"]},
        "ayushman bharat": {"direction": "BULLISH", "score": 70, "stocks": ["APOLLOHOSP", "MAXHEALTH", "FORTIS", "NH", "ASTERDM", "MEDANTA", "KIMS", "IKS", "LALPATHLAB", "AGARWALEYE", "RAINBOW", "INDGN", "VIJAYA", "PARKHOSPS", "METROPOLIS", "SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO", "HDFCLIFE", "SBILIFE", "ICICIGI", "ICICIPRULI", "LICI"]},
        "battery storage": {"direction": "BULLISH", "score": 75, "stocks": ["EXIDEIND", "ARE&M", "RELIANCE", "TATACHEM", "LIKHITHA"]},
        "bess": {"direction": "BULLISH", "score": 70, "stocks": ["EXIDEIND", "ARE&M", "RELIANCE", "TATAPOWER", "WAAREEENER"]},
        "bio-pharma shakti": {"direction": "BULLISH", "score": 75, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO"]},
        "biopharma": {"direction": "BULLISH", "score": 70, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO"]},
        "block deal": {"direction": "BULLISH", "score": 60, "stocks": []},
        "brahmos": {"direction": "BULLISH", "score": 80, "stocks": ["BDL", "BEL", "HAL", "BHARATFORG", "MIDHANI", "DATAPATTNS"]},
        "bta": {"direction": "BULLISH", "score": 65, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "TMCV", "TMPV", "SUNPHARMA"]},
        "budget 2026": {"direction": "BULLISH", "score": 65, "stocks": ["LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "ABB", "CUMMINSIND"]},
        "budget capex": {"direction": "BULLISH", "score": 75, "stocks": ["ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER", "BHEL", "GVT&D", "WAAREEENER", "HAVELLS", "SUZLON", "LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT"]},
        "bullet train": {"direction": "BULLISH", "score": 65, "stocks": ["LT", "RVNL", "IRCON", "TITAGARH", "BEML", "BHEL", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM"]},
        "cancer drug": {"direction": "BULLISH", "score": 70, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO"]},
        "capex push": {"direction": "BULLISH", "score": 75, "stocks": ["ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER", "BHEL", "GVT&D", "WAAREEENER", "HAVELLS", "SUZLON", "THERMAX", "APARINDS", "PREMIERENE", "3MINDIA", "KAYNES", "LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES"]},
        "capex revival": {"direction": "BULLISH", "score": 75, "stocks": ["ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER", "BHEL", "GVT&D", "WAAREEENER", "HAVELLS", "SUZLON", "THERMAX", "APARINDS", "PREMIERENE", "3MINDIA", "KAYNES", "LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES"]},
        "cement demand": {"direction": "BULLISH", "score": 70, "stocks": ["ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC", "RAMCOCEM", "JSWCEMENT", "INDIACEM", "NUVOCO", "STARCEMENT", "JKLAKSHMI", "BIRLACORPN", "PRSMJOHNSN", "MIDWESTLTD", "HEIDELBERG", "ORIENTCEM", "POKARNA", "BUILDPRO"]},
        "cement price hike": {"direction": "BULLISH", "score": 70, "stocks": ["ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC", "RAMCOCEM", "JSWCEMENT", "INDIACEM", "NUVOCO", "STARCEMENT", "JKLAKSHMI", "BIRLACORPN", "PRSMJOHNSN"]},
        "china dump": {"direction": "BEARISH", "score": 70, "stocks": ["JSWSTEEL", "TATASTEEL", "JINDALSTEL", "SAIL", "JSL", "APLAPOLLO", "WELCORP", "KIOCL", "SHYAMMETL", "SARDAEN", "VEDL", "HINDZINC", "HINDALCO", "NATIONALUM", "JAINREC"]},
        "china dumping": {"direction": "BEARISH", "score": 70, "stocks": ["JSWSTEEL", "TATASTEEL", "JINDALSTEL", "SAIL", "JSL", "APLAPOLLO", "WELCORP", "KIOCL", "SHYAMMETL", "SARDAEN", "VEDL", "HINDZINC", "HINDALCO", "NATIONALUM", "JAINREC"]},
        "china plus one": {"direction": "BULLISH", "score": 75, "stocks": ["DIXON", "SYRMA", "KAYNES", "BHARATFORG", "MOTHERSON", "SUNPHARMA", "AUROPHARMA", "DRREDDY", "CIPLA", "LAURUSLABS", "DIVISLAB", "PIDILITIND", "SRF", "FLUOROCHEM", "NAVINFLUOR", "HSCL", "DEEPAKNTR", "ATUL", "TATACHEM"]},
        "china+1": {"direction": "BULLISH", "score": 75, "stocks": ["DIXON", "SYRMA", "KAYNES", "BHARATFORG", "MOTHERSON", "SUNPHARMA", "AUROPHARMA", "DRREDDY", "CIPLA", "LAURUSLABS", "DIVISLAB", "PIDILITIND", "SRF", "FLUOROCHEM", "NAVINFLUOR", "HSCL", "DEEPAKNTR", "ATUL", "TATACHEM"]},
        "chip manufacturing": {"direction": "BULLISH", "score": 80, "stocks": ["VINYAS", "TATATECH", "SYMPHONY", "HIRECT", "MARINE", "NETWEB", "SYRMA", "LTTS", "MOSCHIP", "LGEINDIA", "HPL", "AIMTRON", "PGEL", "DIXON", "KPITTECH"]},
        "chip plant": {"direction": "BULLISH", "score": 80, "stocks": ["VINYAS", "TATATECH", "SYMPHONY", "HIRECT", "MARINE", "NETWEB", "SYRMA", "LTTS", "MOSCHIP", "LGEINDIA"]},
        "chip shortage": {"direction": "BEARISH", "score": 70, "stocks": ["MARUTI", "TMCV", "TMPV", "M&M", "BAJAJ-AUTO", "HEROMOTOCO", "DIXON", "AMBER"]},
        "cloud computing": {"direction": "BULLISH", "score": 65, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE"]},
        "coal output": {"direction": "BULLISH", "score": 60, "stocks": ["COALINDIA", "LLOYDSME", "NMDC", "GMDCLTD", "BHARATCOAL", "TEGA", "SANDUMA", "MOIL", "ASHAPURMIN", "ORISSAMINE"]},
        "consumption revival": {"direction": "BULLISH", "score": 70, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "RADICO", "UBL", "PGHH", "DMART", "ETERNAL", "TRENT", "MEESHO", "SWIGGY", "NYKAA", "NAUKRI", "VMM", "TITAN", "ASIANPAINT"]},
        "copper price": {"direction": "BULLISH", "score": 65, "stocks": ["HINDCOPPER", "BHAGYANGR", "SUNLITE", "VEDL"]},
        "cotton": {"direction": "BULLISH", "score": 60, "stocks": ["KPRMILL", "VTL", "TRIDENT", "WELSPUNLIV", "ARVIND", "ALOKINDS", "GARFIBRES", "GOKEX", "ICIL", "RAYMONDLSL"]},
        "credit downgrade": {"direction": "BEARISH", "score": 80, "stocks": []},
        "credit growth": {"direction": "BULLISH", "score": 70, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "critical minerals": {"direction": "BULLISH", "score": 70, "stocks": ["HINDZINC", "VEDL", "NATIONALUM", "NMDC", "HINDCOPPER", "COALINDIA", "LLOYDSME", "GMDCLTD", "BHARATCOAL"]},
        "crop msp": {"direction": "BULLISH", "score": 70, "stocks": ["ITC", "DABUR", "MARICO", "COROMANDEL", "FACT", "UPL", "PIIND", "SUMICHEM", "BAYERCROP", "CHAMBLFERT", "PARADEEP", "SHARDACROP", "RCF"]},
        "crr cut": {"direction": "BULLISH", "score": 75, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "crude oil": {"direction": "BULLISH", "score": 65, "stocks": ["ONGC", "OIL", "RELIANCE", "PRABHA", "ANTELOPUS", "HINDOILEXP"]},
        "crude spike": {"direction": "BEARISH", "score": 70, "stocks": ["MARUTI", "INDIGO", "ASIANPAINT", "HINDUNILVR", "HEROMOTOCO", "TITAN", "BRITANNIA", "BERGEPAINT", "KANSAINER"]},
        "data center": {"direction": "BULLISH", "score": 70, "stocks": ["RELIANCE", "POWERGRID", "NTPC", "CUMMINSIND", "BLUESTARCO", "HFCL", "NETWEB", "TATACOMM", "ADANIENT"]},
        "defence budget": {"direction": "BULLISH", "score": 80, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL", "NIBE", "ROSSTECH", "ZENTEC", "MIDHANI", "BHARATFORG"]},
        "defence deal": {"direction": "BULLISH", "score": 75, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL"]},
        "defence export": {"direction": "BULLISH", "score": 80, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL"]},
        "defence indigenization": {"direction": "BULLISH", "score": 85, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL", "NIBE", "ROSSTECH", "ZENTEC", "MIDHANI", "BHARATFORG"]},
        "defence order": {"direction": "BULLISH", "score": 80, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL", "NIBE", "ROSSTECH", "ZENTEC", "MIDHANI", "BHARATFORG"]},
        "digital payment": {"direction": "BULLISH", "score": 60, "stocks": ["PAYTM", "POLICYBZR", "IIFL", "MOTILALOFS", "SBICARD", "CDSL", "BSE"]},
        "dii support": {"direction": "BULLISH", "score": 65, "stocks": ["RELIANCE", "HDFCBANK", "BHARTIARTL", "TCS", "ICICIBANK", "SBIN", "INFY", "LT", "HINDUNILVR", "ITC", "AXISBANK", "KOTAKBANK", "BAJFINANCE", "M&M", "SUNPHARMA"]},
        "disinvestment": {"direction": "BULLISH", "score": 65, "stocks": ["SBIN", "PNB", "BANKBARODA", "CANBK", "UNIONBANK", "INDIANB", "IDBI", "BANKINDIA", "IOB", "MAHABANK", "NTPC", "NHPC", "SJVN", "POWERGRID", "ONGC", "OIL", "IOC", "BPCL", "HINDPETRO", "GAIL"]},
        "dollar strength": {"direction": "BULLISH", "score": 65, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "KPITTECH", "CYIENT", "SUNPHARMA"]},
        "e-commerce": {"direction": "BULLISH", "score": 60, "stocks": ["DMART", "NYKAA", "ETERNAL", "DELHIVERY", "SWIGGY", "TRENT", "MEESHO", "NAUKRI"]},
        "edible oil": {"direction": "BEARISH", "score": 60, "stocks": ["HINDUNILVR", "MARICO", "DABUR", "BRITANNIA", "PATANJALI", "AWL", "GOKULAGRO", "SUNDROP", "BCLIND"]},
        "electronics manufacturing": {"direction": "BULLISH", "score": 75, "stocks": ["LGEINDIA", "DIXON", "VOLTAS", "BLUESTARCO", "AMBER", "PGEL", "WHIRLPOOL", "EUREKAFORB", "KRN", "SYMPHONY", "CENTUM", "ORIENTELEC", "BOSCH-HCIL", "MARINE", "HIRECT", "KAYNES", "SYRMA"]},
        "ev adoption": {"direction": "BULLISH", "score": 70, "stocks": ["TMCV", "TMPV", "M&M", "HEROMOTOCO", "TVSMOTOR", "EICHERMOT", "BAJAJ-AUTO", "ATHERENERG", "EXIDEIND", "ARE&M"]},
        "ev push": {"direction": "BULLISH", "score": 70, "stocks": ["TMCV", "TMPV", "M&M", "HEROMOTOCO", "TVSMOTOR", "EICHERMOT", "BAJAJ-AUTO", "ATHERENERG", "EXIDEIND", "ARE&M", "BOSCHLTD"]},
        "ev sales": {"direction": "BULLISH", "score": 70, "stocks": ["TMCV", "TMPV", "M&M", "HEROMOTOCO", "TVSMOTOR", "ATHERENERG", "EXIDEIND"]},
        "ev subsidy": {"direction": "BULLISH", "score": 75, "stocks": ["TMCV", "TMPV", "M&M", "HEROMOTOCO", "TVSMOTOR", "EICHERMOT", "ATHERENERG", "EXIDEIND", "ARE&M"]},
        "fame scheme": {"direction": "BULLISH", "score": 75, "stocks": ["TMCV", "TMPV", "M&M", "HEROMOTOCO", "TVSMOTOR", "EICHERMOT", "ATHERENERG"]},
        "fda approval": {"direction": "BULLISH", "score": 85, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO", "ANTHEM", "IPCALAB", "AJANTPHARM", "JBCHEPHARM", "EMCURE"]},
        "fda warning": {"direction": "BEARISH", "score": 85, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO", "ANTHEM", "IPCALAB", "AJANTPHARM", "JBCHEPHARM", "EMCURE"]},
        "fertilizer subsidy": {"direction": "BULLISH", "score": 70, "stocks": ["COROMANDEL", "FACT", "UPL", "PIIND", "SUMICHEM", "BAYERCROP", "CHAMBLFERT", "PARADEEP", "SHARDACROP", "RCF", "GNFC", "GSFC", "RALLIS", "DHANUKA", "MBAPL"]},
        "festive demand": {"direction": "BULLISH", "score": 65, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "TITAN", "DMART", "HEROMOTOCO", "MARUTI", "TMPV", "BAJAJ-AUTO", "HAVELLS", "VOLTAS"]},
        "fii buying": {"direction": "BULLISH", "score": 70, "stocks": ["RELIANCE", "HDFCBANK", "BHARTIARTL", "TCS", "ICICIBANK", "SBIN", "INFY", "LT", "HINDUNILVR", "ITC", "AXISBANK", "KOTAKBANK", "BAJFINANCE", "M&M", "SUNPHARMA", "ASIANPAINT", "MARUTI", "NESTLEIND", "HCLTECH", "TITAN"]},
        "fii inflow": {"direction": "BULLISH", "score": 70, "stocks": ["RELIANCE", "HDFCBANK", "BHARTIARTL", "TCS", "ICICIBANK", "SBIN", "INFY", "LT", "HINDUNILVR", "ITC", "AXISBANK", "KOTAKBANK", "BAJFINANCE", "M&M", "SUNPHARMA", "ASIANPAINT", "MARUTI", "NESTLEIND", "HCLTECH", "TITAN"]},
        "fii outflow": {"direction": "BEARISH", "score": 70, "stocks": ["RELIANCE", "HDFCBANK", "BHARTIARTL", "TCS", "ICICIBANK", "SBIN", "INFY", "LT", "HINDUNILVR", "ITC", "AXISBANK", "KOTAKBANK", "BAJFINANCE", "M&M", "SUNPHARMA", "ASIANPAINT", "MARUTI", "NESTLEIND", "HCLTECH", "TITAN"]},
        "fii selling": {"direction": "BEARISH", "score": 70, "stocks": ["RELIANCE", "HDFCBANK", "BHARTIARTL", "TCS", "ICICIBANK", "SBIN", "INFY", "LT", "HINDUNILVR", "ITC", "AXISBANK", "KOTAKBANK", "BAJFINANCE", "M&M", "SUNPHARMA", "ASIANPAINT", "MARUTI", "NESTLEIND", "HCLTECH", "TITAN"]},
        "fiscal deficit": {"direction": "BULLISH", "score": 60, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN"]},
        "fmcg revival": {"direction": "BULLISH", "score": 70, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "RADICO", "UBL", "PGHH", "GODFRYPHLP", "GILLETTE", "AWL", "HATSUN", "EMAMILTD", "AVANTIFEED", "BIKAJI", "ZYDUSWELL", "ABDL", "EIDPARRY"]},
        "fraud allegation": {"direction": "BEARISH", "score": 85, "stocks": []},
        "freight corridor": {"direction": "BULLISH", "score": 70, "stocks": ["RVNL", "IRCON", "IRFC", "IRCTC", "RAILTEL", "RITES", "TITAGARH", "TEXRAIL", "BEML", "JWL", "CONCOR", "LT", "BHEL"]},
        "gcc expansion": {"direction": "BULLISH", "score": 65, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI"]},
        "generic drug": {"direction": "BULLISH", "score": 65, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO"]},
        "global slowdown": {"direction": "BEARISH", "score": 65, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "HEXT", "TATAELXSI", "TATATECH", "NETWEB", "APLAPOLLO", "WELCORP", "USHAMART", "TIIL", "SURYAROSNI", "BANSALWIRE", "GOODLUCK", "AEROFLEX", "SAMBHV", "JTLIND"]},
        "gold price": {"direction": "BULLISH", "score": 65, "stocks": ["TITAN", "MUTHOOTFIN", "MANAPPURAM", "RAJESHEXPO", "KALYANKJIL", "THANGAMAYL", "PCJEWELLER", "PNGJL", "BLUESTONE", "SKYGOLD", "SENCO"]},
        "green hydrogen": {"direction": "BULLISH", "score": 75, "stocks": ["RELIANCE", "NTPC", "ADANIGREEN", "LT", "GAIL", "GUJGASLTD", "IGL", "TATAPOWER"]},
        "grid infrastructure": {"direction": "BULLISH", "score": 70, "stocks": ["POWERGRID", "ADANIENSOL", "TATAPOWER", "POLYCAB", "KEI", "RRKABEL", "FINCABLES", "UNIVCABLES", "ADVAIT", "VMARCIND", "DYCL"]},
        "gst 2.0": {"direction": "BULLISH", "score": 75, "stocks": ["MARUTI", "TMPV", "HEROMOTOCO", "BAJAJ-AUTO", "TVSMOTOR", "EICHERMOT", "TITAN", "HAVELLS", "DIXON", "AMBER", "VOLTAS", "HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR"]},
        "gst cut": {"direction": "BULLISH", "score": 75, "stocks": ["MARUTI", "TMPV", "HEROMOTOCO", "BAJAJ-AUTO", "TVSMOTOR", "EICHERMOT", "TITAN", "HAVELLS", "DIXON", "AMBER", "VOLTAS", "HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR"]},
        "gst hike": {"direction": "BEARISH", "score": 70, "stocks": ["MARUTI", "TITAN", "HINDUNILVR", "ITC", "HAVELLS"]},
        "gst rationalization": {"direction": "BULLISH", "score": 75, "stocks": ["MARUTI", "TMPV", "HEROMOTOCO", "BAJAJ-AUTO", "TVSMOTOR", "EICHERMOT", "TITAN", "HAVELLS", "DIXON", "AMBER", "VOLTAS", "HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR"]},
        "gst reform": {"direction": "BULLISH", "score": 65, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "MARUTI", "TITAN", "HAVELLS", "DIXON"]},
        "h1b visa": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "KPITTECH", "CYIENT"]},
        "health insurance": {"direction": "BULLISH", "score": 70, "stocks": ["HDFCLIFE", "SBILIFE", "ICICIGI", "ICICIPRULI", "LICI", "STARHEALTH", "NIACL", "GICRE", "APOLLOHOSP", "MAXHEALTH", "FORTIS", "NH", "ASTERDM"]},
        "highway": {"direction": "BULLISH", "score": 70, "stocks": ["IRB", "GRINFRA", "DBL", "PNCINFRA", "ASHOKA", "LT"]},
        "hospital expansion": {"direction": "BULLISH", "score": 70, "stocks": ["APOLLOHOSP", "MAXHEALTH", "FORTIS", "NH", "ASTERDM", "MEDANTA", "KIMS", "IKS", "LALPATHLAB", "AGARWALEYE", "RAINBOW", "INDGN", "VIJAYA", "PARKHOSPS", "METROPOLIS", "HDFCLIFE", "SBILIFE", "ICICIGI", "ICICIPRULI", "LICI"]},
        "hotel demand": {"direction": "BULLISH", "score": 65, "stocks": ["INDHOTEL", "IRCTC", "ITCHOTELS", "JUBLFOOD", "EIHOTEL", "CHALET", "TRAVELFOOD", "THELEELA", "VENTIVE", "DEVYANI", "TBOTEK", "BLS", "LEMONTREE", "IXIGO", "WESTLIFE"]},
        "housing demand": {"direction": "BULLISH", "score": 70, "stocks": ["DLF", "LODHA", "PHOENIXLTD", "OBEROIRLTY", "PRESTIGE", "GODREJPROP", "BRIGADE", "ANANTRAJ", "SOBHA", "SIGNATURE", "NESCO", "LOTUSDEV", "MAHLIFE", "KALPATARU", "EMBDL", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC", "RAMCOCEM", "JSWCEMENT", "INDIACEM", "ASIANPAINT", "BERGEPAINT", "KANSAINER", "JSWDULUX", "INDIGOPNTS"]},
        "hydrogen mission": {"direction": "BULLISH", "score": 75, "stocks": ["RELIANCE", "NTPC", "ADANIGREEN", "LT", "GAIL", "IGL"]},
        "income tax cut": {"direction": "BULLISH", "score": 70, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "MARUTI", "TITAN", "TMCV", "TMPV", "HEROMOTOCO", "DMART"]},
        "indigenous weapons": {"direction": "BULLISH", "score": 80, "stocks": ["BEL", "HAL", "MAZDOCK", "BDL", "GRSE", "DATAPATTNS", "MTARTECH", "ZENTEC", "ASTRAMICRO", "APOLLO", "AXISCADES", "MIDHANI", "PARAS", "ROSSTECH", "IDEAFORGE"]},
        "inflation spike": {"direction": "BEARISH", "score": 65, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "MARUTI", "HEROMOTOCO"]},
        "infra spending": {"direction": "BULLISH", "score": 75, "stocks": ["LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES", "NCC", "GRINFRA", "POWERMECH", "DBL", "WELENT", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC", "RAMCOCEM", "JSWCEMENT", "INDIACEM", "ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER"]},
        "infrastructure spend": {"direction": "BULLISH", "score": 75, "stocks": ["LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES", "NCC", "GRINFRA", "POWERMECH", "DBL", "WELENT", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC", "RAMCOCEM", "JSWCEMENT", "INDIACEM", "ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER"]},
        "iran": {"direction": "BEARISH", "score": 65, "stocks": ["IOC", "BPCL", "HINDPETRO", "RELIANCE", "MRPL"]},
        "iron ore": {"direction": "BULLISH", "score": 65, "stocks": ["NMDC", "TATASTEEL", "JSWSTEEL", "SAIL", "JINDALSTEL"]},
        "ism 2.0": {"direction": "BULLISH", "score": 85, "stocks": ["VINYAS", "TATATECH", "SYMPHONY", "HIRECT", "MARINE", "NETWEB", "SYRMA", "LTTS", "MOSCHIP", "LGEINDIA", "HPL", "AIMTRON", "PGEL", "DIXON", "KPITTECH"]},
        "israel gaza": {"direction": "BEARISH", "score": 55, "stocks": ["IOC", "BPCL", "HINDPETRO"]},
        "it layoff": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "HEXT", "TATAELXSI", "TATATECH", "NETWEB", "KPITTECH", "FSL", "ECLERX", "ZENSARTECH", "BSOFT"]},
        "it spending cut": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "HEXT", "TATAELXSI", "TATATECH", "NETWEB", "KPITTECH", "FSL", "ECLERX", "ZENSARTECH", "BSOFT"]},
        "lithium battery": {"direction": "BULLISH", "score": 70, "stocks": ["EXIDEIND", "ARE&M", "RELIANCE", "TATACHEM"]},
        "lithium ion": {"direction": "BULLISH", "score": 70, "stocks": ["EXIDEIND", "ARE&M", "RELIANCE", "TATACHEM"]},
        "logistics growth": {"direction": "BULLISH", "score": 65, "stocks": ["ADANIPORTS", "JSWINFRA", "COCHINSHIP", "CONCOR", "DELHIVERY", "AEGISVOPAK", "GESHIP", "SCI", "BLUEDART", "SWANDEF", "SHADOWFAX", "GPPL", "TCI", "SHREEJISPG", "VRLLOG"]},
        "make in india": {"direction": "BULLISH", "score": 65, "stocks": ["BEL", "HAL", "MAZDOCK", "BDL", "GRSE", "DATAPATTNS", "MTARTECH", "ZENTEC", "ASTRAMICRO", "APOLLO", "DIXON", "KAYNES", "SYRMA", "BHARATFORG", "TATATECH"]},
        "metro project": {"direction": "BULLISH", "score": 65, "stocks": ["LT", "RVNL", "IRCON", "TITAGARH", "BEML", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "DALBHARAT", "ACC"]},
        "military": {"direction": "BULLISH", "score": 70, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL"]},
        "monsoon bad": {"direction": "BEARISH", "score": 70, "stocks": ["ITC", "HEROMOTOCO", "M&M", "MARICO", "DABUR", "ESCORTS", "VSTIND", "TVSMOTOR", "COROMANDEL", "FACT", "UPL", "PIIND", "SUMICHEM", "BAYERCROP", "CHAMBLFERT", "PARADEEP"]},
        "monsoon good": {"direction": "BULLISH", "score": 70, "stocks": ["ITC", "HEROMOTOCO", "M&M", "MARICO", "DABUR", "ESCORTS", "VSTIND", "TVSMOTOR", "COROMANDEL", "FACT", "UPL", "PIIND", "SUMICHEM", "BAYERCROP", "CHAMBLFERT", "PARADEEP"]},
        "msp hike": {"direction": "BULLISH", "score": 70, "stocks": ["ITC", "DABUR", "MARICO", "COROMANDEL", "FACT", "UPL", "PIIND", "SUMICHEM", "BAYERCROP", "CHAMBLFERT", "PARADEEP", "SHARDACROP", "RCF"]},
        "mutual fund inflow": {"direction": "BULLISH", "score": 65, "stocks": ["RELIANCE", "HDFCBANK", "BHARTIARTL", "TCS", "ICICIBANK", "SBIN", "INFY", "LT", "HINDUNILVR", "ITC", "AXISBANK", "KOTAKBANK", "BAJFINANCE", "M&M", "SUNPHARMA"]},
        "naval order": {"direction": "BULLISH", "score": 80, "stocks": ["MAZDOCK", "COCHINSHIP", "GRSE", "HAL", "BEL", "BDL"]},
        "neobank": {"direction": "BULLISH", "score": 55, "stocks": ["PAYTM", "POLICYBZR"]},
        "net zero": {"direction": "BULLISH", "score": 65, "stocks": ["POWERGRID", "NTPCGREEN", "TORNTPOWER", "INOXGREEN", "JSWENERGY", "JPPOWER", "SUZLON", "NAVA", "WAAREERTL", "ADANIPOWER", "KPIGREEN", "NTPC", "ORIANA", "SOLARWORLD", "WAAREEENER"]},
        "non-fossil fuel": {"direction": "BULLISH", "score": 65, "stocks": ["POWERGRID", "NTPCGREEN", "TORNTPOWER", "INOXGREEN", "JSWENERGY", "JPPOWER", "SUZLON", "NAVA", "WAAREERTL", "ADANIPOWER", "KPIGREEN", "NTPC", "ORIANA", "SOLARWORLD", "WAAREEENER", "CESC", "ACMESOLAR", "ADANIGREEN", "TATAPOWER", "INOXWIND"]},
        "npa reduction": {"direction": "BULLISH", "score": 70, "stocks": ["SBIN", "PNB", "BANKBARODA", "CANBK", "UNIONBANK", "INDIANB", "IOB", "BANKINDIA", "MAHABANK", "UCOBANK", "PSB", "CENTRALBK", "IDBI", "ICICIBANK"]},
        "nuclear deal": {"direction": "BULLISH", "score": 75, "stocks": ["NTPC", "BHEL", "LT", "NHPC"]},
        "nuclear mission": {"direction": "BULLISH", "score": 80, "stocks": ["NTPC", "BHEL", "LT", "TATAPOWER", "KKCL"]},
        "nuclear power": {"direction": "BULLISH", "score": 75, "stocks": ["NTPC", "BHEL", "LT", "NHPC", "TATAPOWER"]},
        "oil price": {"direction": "BULLISH", "score": 65, "stocks": ["ONGC", "OIL", "RELIANCE", "PRABHA", "ANTELOPUS", "HINDOILEXP"]},
        "opec": {"direction": "BULLISH", "score": 60, "stocks": ["ONGC", "OIL", "RELIANCE"]},
        "opec cut": {"direction": "BULLISH", "score": 60, "stocks": ["ONGC", "OIL", "RELIANCE"]},
        "paint demand": {"direction": "BULLISH", "score": 65, "stocks": ["ASIANPAINT", "BERGEPAINT", "KANSAINER", "JSWDULUX", "INDIGOPNTS", "SIRCA"]},
        "paint price hike": {"direction": "BULLISH", "score": 65, "stocks": ["ASIANPAINT", "BERGEPAINT", "KANSAINER", "JSWDULUX", "INDIGOPNTS", "SIRCA"]},
        "peak power demand": {"direction": "BULLISH", "score": 65, "stocks": ["ADANIPOWER", "NTPC", "POWERGRID", "ADANIENSOL", "TATAPOWER", "JSWENERGY", "TORNTPOWER", "NLCINDIA", "CESC", "NAVA", "TECHNOE", "RPOWER", "WAAREERTL", "GMRP&UI", "SGMART", "COALINDIA"]},
        "pharma export": {"direction": "BULLISH", "score": 75, "stocks": ["SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA", "ZYDUSLIFE", "MANKIND", "AUROPHARMA", "ALKEM", "GLENMARK", "LAURUSLABS", "BIOCON", "ABBOTINDIA", "GLAXO", "ANTHEM", "IPCALAB", "AJANTPHARM", "JBCHEPHARM", "EMCURE"]},
        "pli scheme": {"direction": "BULLISH", "score": 75, "stocks": ["LGEINDIA", "DIXON", "VOLTAS", "BLUESTARCO", "AMBER", "PGEL", "WHIRLPOOL", "EUREKAFORB", "KRN", "SYMPHONY", "KAYNES", "SYRMA", "BHARATFORG", "MOTHERSON", "MSUMI"]},
        "pm e-drive": {"direction": "BULLISH", "score": 75, "stocks": ["TMCV", "TMPV", "M&M", "HEROMOTOCO", "TVSMOTOR", "EICHERMOT", "ATHERENERG"]},
        "port expansion": {"direction": "BULLISH", "score": 70, "stocks": ["ADANIPORTS", "JSWINFRA", "GMRP&UI", "GPPL"]},
        "power transmission": {"direction": "BULLISH", "score": 75, "stocks": ["POWERGRID", "ADANIENSOL", "TATAPOWER", "JSWENERGY", "POLYCAB", "KEI", "RRKABEL", "FINCABLES", "UNIVCABLES", "ADVAIT", "VMARCIND", "DYCL"]},
        "private bank": {"direction": "BULLISH", "score": 65, "stocks": ["HDFCBANK", "ICICIBANK", "KOTAKBANK", "AXISBANK", "INDUSINDBK", "IDFCFIRSTB", "FEDERALBNK", "CSBBANK", "DCBBANK", "RBLBANK", "YESBANK", "BANDHANBNK"]},
        "private capex": {"direction": "BULLISH", "score": 70, "stocks": ["ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER", "BHEL", "GVT&D", "WAAREEENER", "HAVELLS", "SUZLON", "THERMAX", "APARINDS", "PREMIERENE", "3MINDIA", "KAYNES", "LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES"]},
        "psu bank": {"direction": "BULLISH", "score": 70, "stocks": ["SBIN", "PNB", "BANKBARODA", "CANBK", "UNIONBANK", "INDIANB", "IOB", "BANKINDIA", "MAHABANK", "UCOBANK", "PSB", "CENTRALBK", "IDBI"]},
        "psu privatization": {"direction": "BULLISH", "score": 70, "stocks": ["GRSE", "RVNL", "IRCON", "IRFC", "IRCTC", "RAILTEL", "RITES", "BEML", "NBCC"]},
        "quick commerce": {"direction": "BULLISH", "score": 60, "stocks": ["ETERNAL", "NYKAA", "DMART", "SWIGGY"]},
        "railway": {"direction": "BULLISH", "score": 70, "stocks": ["RVNL", "IRCON", "IRFC", "IRCTC", "RAILTEL", "RITES", "TITAGARH", "TEXRAIL", "BEML", "JWL", "CONCOR", "LT", "BHEL"]},
        "rare earth": {"direction": "BULLISH", "score": 75, "stocks": ["HINDZINC", "NATIONALUM", "VEDL", "HINDCOPPER", "NMDC", "COALINDIA", "LLOYDSME", "GMDCLTD", "BHARATCOAL"]},
        "rate cut": {"direction": "BULLISH", "score": 75, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "rating downgrade": {"direction": "BEARISH", "score": 75, "stocks": []},
        "rbi cut": {"direction": "BULLISH", "score": 80, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "rbi hike": {"direction": "BEARISH", "score": 75, "stocks": ["BAJFINANCE", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "CHOLAFIN", "BAJAJHFL", "LTF", "SBICARD", "HDBFS", "SUNDARMFIN", "AIIL", "PIRAMALFIN", "M&MFIN", "POONAWALLA", "LICHSGFIN", "PNBHOUSING", "MANAPPURAM", "AADHARHFC", "CGCL", "SAMMAANCAP", "FIVESTAR", "APTUS", "HOMEFIRST", "CANFINHOME"]},
        "rbi policy": {"direction": "BULLISH", "score": 55, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK"]},
        "rbi rate cut": {"direction": "BULLISH", "score": 80, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "rbi rate hike": {"direction": "BEARISH", "score": 75, "stocks": ["BAJFINANCE", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "CHOLAFIN", "AUBANK", "BAJAJHFL", "LTF", "SBICARD", "HDBFS", "SUNDARMFIN", "AIIL", "PIRAMALFIN", "M&MFIN", "POONAWALLA", "LICHSGFIN", "PNBHOUSING", "MANAPPURAM", "AADHARHFC", "CREDITACC", "CGCL", "SAMMAANCAP", "FIVESTAR", "APTUS"]},
        "recession": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "HEXT", "TATAELXSI", "TATATECH", "NETWEB", "APLAPOLLO", "WELCORP", "USHAMART", "TIIL", "SURYAROSNI", "BANSALWIRE", "GOODLUCK", "AEROFLEX", "SAMBHV", "JTLIND"]},
        "red sea": {"direction": "BEARISH", "score": 65, "stocks": ["CONCOR", "DELHIVERY", "GESHIP", "SCI", "BLUEDART", "SWANDEF", "SHADOWFAX", "SHREEJISPG", "VRLLOG", "MAHLOG"]},
        "renewable energy": {"direction": "BULLISH", "score": 70, "stocks": ["POWERGRID", "NTPCGREEN", "TORNTPOWER", "INOXGREEN", "JSWENERGY", "JPPOWER", "SUZLON", "NAVA", "WAAREERTL", "ADANIPOWER", "KPIGREEN", "NTPC", "ORIANA", "SOLARWORLD", "WAAREEENER", "CESC", "ACMESOLAR", "ADANIGREEN", "TATAPOWER", "INOXWIND", "NHPC", "TECHNOE", "SGMART", "CLEANMAX", "GMRP&UI"]},
        "repo cut": {"direction": "BULLISH", "score": 80, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "repo rate cut": {"direction": "BULLISH", "score": 80, "stocks": ["HDFCBANK", "SBIN", "ICICIBANK", "BAJFINANCE", "AXISBANK", "KOTAKBANK", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "BANKBARODA", "UNIONBANK", "CHOLAFIN", "PNB", "CANBK", "INDIANB", "IDBI", "BAJAJHFL", "LTF", "FEDERALBNK", "BANKINDIA", "IOB", "INDUSINDBK", "SBICARD", "YESBANK", "MAHABANK", "IDFCFIRSTB", "HDBFS", "SUNDARMFIN", "AIIL"]},
        "repo rate hike": {"direction": "BEARISH", "score": 75, "stocks": ["BAJFINANCE", "SHRIRAMFIN", "JIOFIN", "TATACAP", "MUTHOOTFIN", "CHOLAFIN", "AUBANK", "BAJAJHFL", "LTF", "SBICARD", "HDBFS", "SUNDARMFIN", "AIIL", "PIRAMALFIN", "M&MFIN", "POONAWALLA", "LICHSGFIN", "PNBHOUSING", "MANAPPURAM", "AADHARHFC", "CREDITACC", "CGCL", "SAMMAANCAP", "FIVESTAR", "APTUS"]},
        "road project": {"direction": "BULLISH", "score": 70, "stocks": ["LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "CEMPRO", "RITES", "NCC", "GRINFRA", "POWERMECH", "DBL", "WELENT"]},
        "rupee appreciation": {"direction": "BEARISH", "score": 60, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "KPITTECH", "CYIENT", "SUNPHARMA"]},
        "rupee depreciation": {"direction": "BULLISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "KPITTECH", "CYIENT", "SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA"]},
        "rupee strong": {"direction": "BEARISH", "score": 60, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "KPITTECH", "CYIENT", "SUNPHARMA"]},
        "rupee weak": {"direction": "BULLISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "KPITTECH", "CYIENT", "SUNPHARMA", "DIVISLAB", "TORNTPHARM", "DRREDDY", "LUPIN", "CIPLA"]},
        "rural demand": {"direction": "BULLISH", "score": 75, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "RADICO", "UBL", "PGHH", "GODFRYPHLP", "GILLETTE", "AWL", "HATSUN", "EMAMILTD", "HEROMOTOCO", "M&M", "TVSMOTOR", "ESCORTS"]},
        "rural growth": {"direction": "BULLISH", "score": 75, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "RADICO", "UBL", "PGHH", "GODFRYPHLP", "GILLETTE", "AWL", "HATSUN", "EMAMILTD", "HEROMOTOCO", "M&M", "TVSMOTOR", "ESCORTS"]},
        "rural revival": {"direction": "BULLISH", "score": 75, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "RADICO", "UBL", "PGHH", "GODFRYPHLP", "GILLETTE", "AWL", "HATSUN", "EMAMILTD", "HEROMOTOCO", "M&M", "TVSMOTOR", "ESCORTS"]},
        "russia oil": {"direction": "BULLISH", "score": 65, "stocks": ["RELIANCE", "IOC", "BPCL", "HINDPETRO", "ONGC", "OIL"]},
        "sebi action": {"direction": "BEARISH", "score": 75, "stocks": []},
        "sebi ban": {"direction": "BEARISH", "score": 80, "stocks": []},
        "semiconductor mission": {"direction": "BULLISH", "score": 85, "stocks": ["VINYAS", "TATATECH", "SYMPHONY", "HIRECT", "MARINE", "NETWEB", "SYRMA", "LTTS", "MOSCHIP", "LGEINDIA", "HPL", "AIMTRON", "PGEL", "DIXON", "KPITTECH", "CENTUM", "CYIENT", "VOLTAS", "AMBER", "ORIENTELEC"]},
        "small modular reactor": {"direction": "BULLISH", "score": 80, "stocks": ["NTPC", "BHEL", "LT", "TATAPOWER"]},
        "solar capacity": {"direction": "BULLISH", "score": 75, "stocks": ["INOXGREEN", "JSWENERGY", "KPIGREEN", "NTPC", "SOLARWORLD", "WAAREEENER", "ACMESOLAR", "ADANIGREEN", "TATAPOWER", "NHPC", "SJVN"]},
        "solar energy": {"direction": "BULLISH", "score": 70, "stocks": ["INOXGREEN", "JSWENERGY", "KPIGREEN", "NTPC", "SOLARWORLD", "WAAREEENER", "ACMESOLAR", "ADANIGREEN", "TATAPOWER", "NHPC", "SJVN"]},
        "specialty chemical": {"direction": "BULLISH", "score": 65, "stocks": ["PIDILITIND", "SRF", "FLUOROCHEM", "NAVINFLUOR", "HSCL", "DEEPAKNTR", "ATUL", "TATACHEM", "AARTIIND", "BASF", "DEEPAKFERT", "AETHER", "ANURAS", "FINEORG", "VINATIORGA", "PRIVISCL", "JUBLINGREA", "CLEAN", "ALKYLAMINE", "ACI"]},
        "steel price": {"direction": "BULLISH", "score": 70, "stocks": ["JSWSTEEL", "TATASTEEL", "JINDALSTEL", "SAIL", "JSL", "APLAPOLLO", "WELCORP", "KIOCL", "SHYAMMETL", "SARDAEN", "GALLANTT", "GPIL", "RATNAMANI", "JINDALSAW", "USHAMART"]},
        "steel price hike": {"direction": "BULLISH", "score": 70, "stocks": ["JSWSTEEL", "TATASTEEL", "JINDALSTEL", "SAIL", "JSL", "APLAPOLLO", "WELCORP", "KIOCL", "SHYAMMETL", "SARDAEN", "GALLANTT", "GPIL", "RATNAMANI", "JINDALSAW", "USHAMART"]},
        "stimulus": {"direction": "BULLISH", "score": 70, "stocks": ["LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "ABB", "CUMMINSIND", "POWERINDIA", "SIEMENS", "CGPOWER", "HDFCBANK", "SBIN"]},
        "stt hike": {"direction": "BEARISH", "score": 65, "stocks": ["ANGELONE", "MOTILALOFS", "BSE", "CDSL", "GEOJITFSL", "GROWW", "MCX", "360ONE"]},
        "suez disruption": {"direction": "BEARISH", "score": 65, "stocks": ["CONCOR", "DELHIVERY", "GESHIP", "SCI", "BLUEDART", "SWANDEF", "SHADOWFAX", "SHREEJISPG", "VRLLOG", "MAHLOG"]},
        "supply chain shift": {"direction": "BULLISH", "score": 70, "stocks": ["DIXON", "SYRMA", "KAYNES", "BHARATFORG", "MOTHERSON", "SUNPHARMA", "AUROPHARMA", "DRREDDY", "CIPLA", "LAURUSLABS", "DIVISLAB", "PIDILITIND", "SRF", "FLUOROCHEM", "NAVINFLUOR"]},
        "tariff hike": {"direction": "BULLISH", "score": 80, "stocks": ["BHARTIARTL", "IDEA", "BHARTIHEXA"]},
        "tax relief": {"direction": "BULLISH", "score": 70, "stocks": ["HINDUNILVR", "ITC", "NESTLEIND", "VBL", "BRITANNIA", "TATACONSUM", "GODREJCP", "MARICO", "UNITDSPR", "DABUR", "COLPAL", "PATANJALI", "MARUTI", "TITAN", "TMCV", "TMPV", "HEROMOTOCO", "DMART"]},
        "telecom capex": {"direction": "BULLISH", "score": 70, "stocks": ["BHARTIARTL", "IDEA", "TATACOMM", "BHARTIHEXA", "HFCL", "STLTECH", "ITI", "RELIANCE"]},
        "tourism boost": {"direction": "BULLISH", "score": 60, "stocks": ["INDHOTEL", "IRCTC", "ITCHOTELS", "JUBLFOOD", "EIHOTEL", "CHALET", "TRAVELFOOD", "THELEELA", "VENTIVE", "DEVYANI", "TBOTEK", "BLS", "LEMONTREE", "IXIGO", "WESTLIFE", "INDIGO"]},
        "trade deal": {"direction": "BULLISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "TATASTEEL", "JSWSTEEL", "SUNPHARMA", "CIPLA"]},
        "trade war": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "TATASTEEL", "JSWSTEEL", "HINDALCO"]},
        "trump tariff": {"direction": "BEARISH", "score": 75, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "TATASTEEL", "JSWSTEEL", "HINDALCO", "SUNPHARMA"]},
        "ukraine war": {"direction": "BULLISH", "score": 65, "stocks": ["BEL", "HAL", "MAZDOCK", "BDL", "GRSE", "DATAPATTNS", "MTARTECH", "ZENTEC", "ASTRAMICRO", "APOLLO", "AXISCADES", "MIDHANI", "PARAS", "ROSSTECH", "IDEAFORGE", "ONGC", "OIL", "RELIANCE"]},
        "union budget": {"direction": "BULLISH", "score": 65, "stocks": ["LT", "GMRAIRPORT", "RVNL", "IRB", "NBCC", "KPIL", "IRCON", "AFCONS", "ULTRACEMCO", "GRASIM", "AMBUJACEM", "SHREECEM", "JKCEMENT", "ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "ABB", "CUMMINSIND"]},
        "upi growth": {"direction": "BULLISH", "score": 60, "stocks": ["PAYTM", "POLICYBZR", "IIFL", "MOTILALOFS", "SBICARD", "CDSL", "BSE"]},
        "us recession": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "HEXT", "TATAELXSI", "TATATECH", "NETWEB", "KPITTECH", "FSL", "ECLERX", "ZENSARTECH", "BSOFT"]},
        "us tariff": {"direction": "BEARISH", "score": 70, "stocks": ["TCS", "INFY", "HCLTECH", "WIPRO", "TECHM", "LTM", "PERSISTENT", "OFSS", "MPHASIS", "COFORGE", "LTTS", "TATAELXSI", "TATASTEEL", "JSWSTEEL", "HINDALCO"]},
        "war": {"direction": "BULLISH", "score": 75, "stocks": ["ASTRAMICRO", "IDEAFORGE", "GRSE", "LTTS", "BHEL", "PARAS", "DATAPATTNS", "JAYKAY", "MTARTECH", "AXISCADES", "SOLARINDS", "MAZDOCK", "HAL", "C2C", "APOLLO", "BEL", "KRISHNADEF", "BDL", "DCXINDIA", "DCMSIL", "ONGC", "OIL"]},
        "wedding season": {"direction": "BULLISH", "score": 60, "stocks": ["TITAN", "TRENT", "DMART", "KALYANKJIL", "THANGAMAYL", "PCJEWELLER", "PNGJL"]},
        "wind capacity": {"direction": "BULLISH", "score": 70, "stocks": ["SUZLON", "INOXWIND", "ADANIGREEN", "TATAPOWER"]},
        "wind energy": {"direction": "BULLISH", "score": 70, "stocks": ["SUZLON", "INOXWIND", "ADANIGREEN", "TATAPOWER", "JSWENERGY"]},
    }

    def __init__(self, name, valid_symbols=None, trendlyne_scorer=None):
        self.name = name
        self.valid_symbols = valid_symbols or set()
        self._seen = set()
        # FIX_V4_DYNAMIC_MACRO: use Trendlyne for industry-based theme resolution
        self.trendlyne = trendlyne_scorer
        self._industry_cache = {}
        self._industry_cache_ttl = 3600

    def _resolve_stocks(self, theme, config):
        """FIX_V4_DYNAMIC_MACRO: resolve theme stocks via industry OR fallback_stocks OR legacy stocks."""
        import time as _time
        now = _time.time()
        if theme in self._industry_cache:
            ts, cached = self._industry_cache[theme]
            if now - ts < self._industry_cache_ttl:
                return cached
        industries = config.get("industries")
        resolved = []
        if industries and self.trendlyne:
            resolved = self.trendlyne.resolve_by_industry(
                industries,
                min_mcap_cr=config.get("min_mcap_cr", 500),
                max_stocks=config.get("max_stocks", 30),
                valid_symbols=self.valid_symbols if self.valid_symbols else None,
            )
        if not resolved:
            fallback = config.get("fallback_stocks") or config.get("stocks", [])
            resolved = [s for s in fallback if not self.valid_symbols or s in self.valid_symbols]
        self._industry_cache[theme] = (now, resolved)
        return resolved

    def scan(self):
        """FIX_V39_SCOUT_VALIDATION: scout only; real validation via NewsBrain v3.
        PATCH_V62_MACRO_BODY: match themes on title+summary, not title only.
        Previously 'rate cut' in article body but not headline was missed."""
        import re as _re
        candidates = []

        for feed_name, feed_url in NEWS_FEEDS.items():
            try:
                parsed = feedparser.parse(feed_url)
                for entry in parsed.entries[:15]:
                    title = entry.get("title", "")
                    # PATCH_V70_AUDIT_CLEANUP: removed unused title_lower
                    link = entry.get("link", "")

                    summary = entry.get("summary", "") or entry.get("description", "") or ""
                    summary_clean = _re.sub(r"<[^>]+>", " ", summary)
                    # PATCH_V62_MACRO_BODY: combined text for theme matching
                    full_lower = (title + " " + summary_clean).lower()

                    h = hashlib.md5(f"macro_{title}".encode()).hexdigest()[:12]
                    if h in self._seen:
                        continue

                    for theme, config in self.MACRO_MAP.items():
                        if theme in full_lower:
                            self._seen.add(h)
                            # FIX_V4_DYNAMIC_MACRO: industry-based resolution with fallback
                            # PATCH_V68_LLM_FIRST_DECISION: emit ARTICLE-level hint item.
                            # Stocks listed in MACRO_MAP for this theme become DISCOVERY HINTS.
                            # LLM reads full article and decides direction/score itself.
                            _hint_stocks = self._resolve_stocks(theme, config)
                            candidates.append({
                                "symbol": "_ARTICLE_",
                                "direction": "",
                                "score": 0,
                                "catalyst": f"MACRO {theme.upper()}: {title[:60]}",
                                "source": "MACRO",
                                "title": title,
                                "summary": summary_clean,
                                "link": link,
                                "source_feed": feed_name,
                                "theme_matched": theme,
                                "hints": _hint_stocks,
                                "v68_article_item": True,
                            })
                            break

            except Exception:
                pass

        if candidates:
            log.info(f"Macro {self.name}: {len(candidates)} scout candidates (will be validated)")
        return candidates


# ═══════════════════════════════════════════════════════════════════════════════
#  INDEX MACRO DETECTOR — PATCH_V63_INDEX_ROUTING
#  Scans news for themes that move indices directly (NIFTY, BANKNIFTY, FINNIFTY).
#  Emits candidates with symbol=index_name so FnoModule.select_option picks up
#  index CE/PE via is_index=True path.
#
#  Two detection layers:
#    1. Index-specific keywords ("nifty rally", "banknifty support broken")
#    2. Broad macro themes that move the whole index (inherited from MacroDetector
#       but routed to index symbols). Only F&O module uses this — equity has no
#       tradeable index instrument.
# ═══════════════════════════════════════════════════════════════════════════════

class IndexMacroDetector:
    """
    F&O only. Emits index candidates (NIFTY/BANKNIFTY/FINNIFTY) from:
      - Direct index-level news ("nifty hits record", "banknifty support broken")
      - Broad macro themes that move whole indices (Fed decision, FII flows, VIX spikes)

    Output format same as MacroDetector: {symbol, direction, score, catalyst, ...}.
    Goes through same NewsBrain validation + Expert Panel FNO scoring.
    is_index=True in select_option → uses 5-40 day DTE (vs 15-40 for stocks).
    """

    # Direct index-movers: keyword → (index, direction, score)
    # NIFTY: broad market moves. BANKNIFTY: banking/rate sensitive.
    # FINNIFTY: financial services. MIDCPNIFTY: mid-cap flows.
    INDEX_DIRECT_MAP = {
        # NIFTY direct
        "nifty hits record":        ("NIFTY", "BULLISH", 80),
        "nifty all-time high":      ("NIFTY", "BULLISH", 80),
        "nifty rally":              ("NIFTY", "BULLISH", 70),
        "nifty surges":             ("NIFTY", "BULLISH", 70),
        "nifty breakout":           ("NIFTY", "BULLISH", 75),
        "nifty correction":         ("NIFTY", "BEARISH", 70),
        "nifty crash":              ("NIFTY", "BEARISH", 85),
        "nifty tumbles":            ("NIFTY", "BEARISH", 75),
        "nifty plunges":            ("NIFTY", "BEARISH", 80),
        "nifty support broken":     ("NIFTY", "BEARISH", 75),
        "sensex hits record":       ("NIFTY", "BULLISH", 75),
        "sensex rallies":           ("NIFTY", "BULLISH", 70),
        "sensex crashes":           ("NIFTY", "BEARISH", 80),
        # BANKNIFTY direct
        "banknifty record":         ("BANKNIFTY", "BULLISH", 80),
        "banknifty all-time high":  ("BANKNIFTY", "BULLISH", 80),
        "banknifty rallies":        ("BANKNIFTY", "BULLISH", 75),
        "banknifty breakout":       ("BANKNIFTY", "BULLISH", 75),
        "banknifty crashes":        ("BANKNIFTY", "BEARISH", 80),
        "banknifty tumbles":        ("BANKNIFTY", "BEARISH", 75),
        "bank stocks rally":        ("BANKNIFTY", "BULLISH", 70),
        "bank stocks fall":         ("BANKNIFTY", "BEARISH", 70),
        "psu banks surge":          ("BANKNIFTY", "BULLISH", 70),
        "private banks rally":     ("BANKNIFTY", "BULLISH", 70),
        # FINNIFTY
        "financial services rally": ("FINNIFTY", "BULLISH", 70),
        "nbfc rally":               ("FINNIFTY", "BULLISH", 70),
        "nbfc crash":               ("FINNIFTY", "BEARISH", 75),
        # VIX signals (inverse — VIX spike = market fear = BEARISH NIFTY)
        "vix spikes":               ("NIFTY", "BEARISH", 70),
        "vix jumps":                ("NIFTY", "BEARISH", 65),
        "fear gauge":               ("NIFTY", "BEARISH", 60),
        "vix crashes":              ("NIFTY", "BULLISH", 65),  # calm = bullish
        "vix eases":                ("NIFTY", "BULLISH", 55),
    }

    # Broad macro themes that move whole indices. These layer on top of MacroDetector
    # which routes to individual stocks. IndexMacroDetector additionally routes to
    # the relevant index symbol for F&O options.
    INDEX_THEME_MAP = {
        # Rate-sensitive → BANKNIFTY dominant, NIFTY secondary
        "rbi rate cut":     [("BANKNIFTY", "BULLISH", 80), ("NIFTY", "BULLISH", 65)],
        "rbi cut":          [("BANKNIFTY", "BULLISH", 80), ("NIFTY", "BULLISH", 65)],
        "rate cut":         [("BANKNIFTY", "BULLISH", 70), ("NIFTY", "BULLISH", 55)],
        "rbi rate hike":    [("BANKNIFTY", "BEARISH", 75), ("NIFTY", "BEARISH", 60)],
        "rbi hike":         [("BANKNIFTY", "BEARISH", 75), ("NIFTY", "BEARISH", 60)],
        "repo rate cut":    [("BANKNIFTY", "BULLISH", 80), ("NIFTY", "BULLISH", 65)],
        "repo rate hike":   [("BANKNIFTY", "BEARISH", 75), ("NIFTY", "BEARISH", 60)],
        # FII flows → NIFTY dominant
        "fii buying":       [("NIFTY", "BULLISH", 70), ("BANKNIFTY", "BULLISH", 60)],
        "fii inflow":       [("NIFTY", "BULLISH", 70), ("BANKNIFTY", "BULLISH", 60)],
        "fii selling":      [("NIFTY", "BEARISH", 70), ("BANKNIFTY", "BEARISH", 60)],
        "fii outflow":      [("NIFTY", "BEARISH", 70), ("BANKNIFTY", "BEARISH", 60)],
        # Fed / US → NIFTY (IT-heavy)
        "fomc rate cut":    [("NIFTY", "BULLISH", 70), ("BANKNIFTY", "BULLISH", 55)],
        "fomc rate hike":   [("NIFTY", "BEARISH", 70), ("BANKNIFTY", "BEARISH", 55)],
        "fed cuts rates":   [("NIFTY", "BULLISH", 75), ("BANKNIFTY", "BULLISH", 60)],
        "fed hikes rates":  [("NIFTY", "BEARISH", 75), ("BANKNIFTY", "BEARISH", 60)],
        "fed dovish":       [("NIFTY", "BULLISH", 65)],
        "fed hawkish":      [("NIFTY", "BEARISH", 65)],
        "us recession":     [("NIFTY", "BEARISH", 75)],
        "us cpi cool":      [("NIFTY", "BULLISH", 65)],
        "us cpi hot":       [("NIFTY", "BEARISH", 65)],
        # Geopolitical / risk-off
        "war escalates":    [("NIFTY", "BEARISH", 70)],
        "risk-off":         [("NIFTY", "BEARISH", 65)],
        "risk off":         [("NIFTY", "BEARISH", 65)],
        "risk on":          [("NIFTY", "BULLISH", 60)],
        # India macro
        "union budget":     [("NIFTY", "BULLISH", 65), ("BANKNIFTY", "BULLISH", 60)],
        "india cpi cool":   [("NIFTY", "BULLISH", 65), ("BANKNIFTY", "BULLISH", 70)],
        "india cpi hot":    [("NIFTY", "BEARISH", 60), ("BANKNIFTY", "BEARISH", 70)],
        "gdp growth":       [("NIFTY", "BULLISH", 65)],
        "gdp slowdown":     [("NIFTY", "BEARISH", 65)],
        # GST / policy
        "gst cut":          [("NIFTY", "BULLISH", 60)],
        "gst hike":         [("NIFTY", "BEARISH", 55)],
    }

    # Only the 3 index symbols we actually trade F&O on (MIDCPNIFTY excluded due to low liquidity)
    TRADEABLE_INDICES = {"NIFTY", "BANKNIFTY", "FINNIFTY"}

    def __init__(self, name, fno_stocks=None):
        self.name = name
        self.fno_stocks = fno_stocks or set()
        self._seen = set()
        self._seen_file = DATA_DIR / f"index_macro_seen_{name.lower()}.json"
        self._load_seen()

    def _load_seen(self):
        data = load_json(self._seen_file, {"hashes": []})
        self._seen = set(data.get("hashes", [])[-3000:])

    def _save_seen(self):
        save_json(self._seen_file, {"hashes": list(self._seen)[-3000:]})

    def scan(self):
        """Returns index candidates. Only returns indices that are in fno_stocks
        (Kite-confirmed tradeable). Goes through NewsBrain validation + Expert Panel FNO."""
        import re as _re
        candidates = []
        hits_direct = 0
        hits_theme = 0

        for feed_name, feed_url in NEWS_FEEDS.items():
            try:
                parsed = feedparser.parse(feed_url)
                for entry in parsed.entries[:15]:
                    title = entry.get("title", "")
                    # PATCH_V70_AUDIT_CLEANUP: removed unused title_lower
                    link = entry.get("link", "")
                    summary = entry.get("summary", "") or entry.get("description", "") or ""
                    summary_clean = _re.sub(r"<[^>]+>", " ", summary)
                    full_lower = (title + " " + summary_clean).lower()

                    h = hashlib.md5(f"idxmacro_{title}|{link}".encode()).hexdigest()[:12]
                    if h in self._seen:
                        continue

                    matched = False

                    # Layer 1: direct index keywords
                    for kw, (idx, direction, score) in self.INDEX_DIRECT_MAP.items():
                        if kw in full_lower and idx in self.TRADEABLE_INDICES:
                            if self.fno_stocks and idx not in self.fno_stocks:
                                continue  # Kite doesn't have it
                            candidates.append({
                                "symbol": idx,
                                "direction": direction,
                                "score": score,
                                "catalyst": f"INDEX_DIRECT {kw.upper()}: {title[:60]}",
                                "source": "INDEX_MACRO",
                                "title": title,
                                "summary": summary_clean,
                                "link": link,
                                "source_feed": feed_name,
                                "theme_matched": kw,
                                "is_index": True,
                            })
                            hits_direct += 1
                            matched = True
                            break  # one match per headline enough

                    if matched:
                        self._seen.add(h)
                        continue

                    # Layer 2: broad macro themes → index routing
                    for theme, targets in self.INDEX_THEME_MAP.items():
                        if theme in full_lower:
                            for idx, direction, score in targets:
                                if idx not in self.TRADEABLE_INDICES:
                                    continue
                                if self.fno_stocks and idx not in self.fno_stocks:
                                    continue
                                candidates.append({
                                    "symbol": idx,
                                    "direction": direction,
                                    "score": score,
                                    "catalyst": f"INDEX_THEME {theme.upper()}: {title[:60]}",
                                    "source": "INDEX_MACRO",
                                    "title": title,
                                    "summary": summary_clean,
                                    "link": link,
                                    "source_feed": feed_name,
                                    "theme_matched": theme,
                                    "is_index": True,
                                })
                                hits_theme += 1
                            self._seen.add(h)
                            break  # one theme per headline enough

            except Exception as e:
                log.debug(f"IndexMacro {self.name}: {feed_name} error: {e}")

        self._save_seen()
        if candidates:
            log.info(f"IndexMacro {self.name}: {len(candidates)} index candidates (direct={hits_direct} theme={hits_theme})")
        return candidates


# ═══════════════════════════════════════════════════════════════════════════════
#  EARNINGS CALENDAR — Tracks upcoming quarterly results for F&O timing
# ═══════════════════════════════════════════════════════════════════════════════

class EarningsCalendar:
    """
    Tracks upcoming quarterly results dates.
    F&O module uses this to position BEFORE earnings (buy options 2-4 weeks ahead).
    Sources: NSE board meeting filings (purpose: quarterly results).
    """

    def __init__(self):
        self._calendar = {}  # symbol -> {"date": date_str, "quarter": "Q4"}
        self._cal_file = DATA_DIR / "earnings_calendar.json"
        self._load()

    def _load(self):
        self._calendar = load_json(self._cal_file, {})

    def _save(self):
        save_json(self._cal_file, self._calendar)

    def update_from_filings(self):  # PATCH_V70_AUDIT_CLEANUP: removed unused filings_data param
        """
        PATCH_V65_EARNINGS_FIX: corrected NSE endpoint and parser.
        """
        url = "https://www.nseindia.com/api/event-calendar?index=equities"
        last_err = None
        # PATCH_V65: 3-retry with cookie refresh (V62 FilingMonitor pattern)
        for attempt in range(3):
            try:
                session = requests.Session()
                session.headers.update({
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    "Accept": "application/json, text/plain, */*",
                    "Accept-Language": "en-US,en;q=0.9",
                    "Referer": "https://www.nseindia.com/",
                })
                session.get("https://www.nseindia.com", timeout=10)
                time.sleep(0.5)

                resp = session.get(url, timeout=15)
                if resp.status_code != 200:
                    last_err = f"HTTP {resp.status_code}"
                    log.warning(f"[PATCH_V65_EARNINGS] attempt {attempt+1}/3 NSE returned {resp.status_code}")
                    time.sleep(2)
                    continue

                events = resp.json()
                if not isinstance(events, list):
                    last_err = f"unexpected JSON type {type(events).__name__}"
                    log.warning(f"[PATCH_V65_EARNINGS] attempt {attempt+1}/3 {last_err}")
                    time.sleep(2)
                    continue

                # PATCH_V66_CALENDAR_PURGE: drop entries with date older than 2 days.
                # NSE rotates symbols off the feed post-earnings, leaving stale entries
                # in self._calendar forever. Purge defensively before re-populating.
                purged = 0
                _today_purge = today_ist()
                _stale_syms = []
                for _sym, _info in self._calendar.items():
                    _ds = (_info.get("date", "") or "").strip()
                    if not _ds:
                        _stale_syms.append(_sym)
                        continue
                    _ed = None
                    for _fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                        try:
                            _ed = datetime.strptime(_ds, _fmt).date()
                            break
                        except ValueError:
                            continue
                    if _ed is None:
                        _stale_syms.append(_sym)
                        continue
                    if (_today_purge - _ed).days > 2:
                        _stale_syms.append(_sym)
                for _sym in _stale_syms:
                    self._calendar.pop(_sym, None)
                    purged += 1

                added = 0
                for ev in events:
                    purpose = (ev.get("purpose", "") or "")
                    symbol = (ev.get("symbol", "") or "").strip()
                    earn_date = (ev.get("date", "") or "").strip()
                    if not symbol or not earn_date:
                        continue
                    if "Financial Result" not in purpose:
                        continue
                    self._calendar[symbol] = {
                        "date": earn_date,
                        "purpose": purpose[:120],
                        "company": (ev.get("company", "") or "")[:80],
                        "bm_desc": (ev.get("bm_desc", "") or "")[:200],
                    }
                    added += 1

                self._save()
                log.info(f"[PATCH_V65_EARNINGS] {added} earnings dates tracked from {len(events)} events (filter=Financial Result)")
                if purged:
                    log.info(f"[PATCH_V66_CALENDAR_PURGE] purged {purged} stale entries (>2 days past), calendar size now {len(self._calendar)}")
                return

            except Exception as e:
                last_err = str(e)
                log.warning(f"[PATCH_V65_EARNINGS] attempt {attempt+1}/3 exception: {e}")
                time.sleep(2)

        log.error(f"[PATCH_V65_EARNINGS] all 3 retries failed, last_err={last_err}")

    def get_upcoming(self, days_ahead=14):
        """Get stocks with earnings in next N days."""
        upcoming = []
        today = today_ist()

        for sym, info in self._calendar.items():
            try:
                date_str = info.get("date", "")
                for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        earn_date = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    continue

                days_to = (earn_date - today).days
                if 0 <= days_to <= days_ahead:
                    upcoming.append({
                        "symbol": sym,
                        "date": str(earn_date),
                        "days_to": days_to,
                    })
            except Exception:
                pass

        return sorted(upcoming, key=lambda x: x["days_to"])

    def get_earnings_plays(self, fno_stocks):
        """
        Find the BEST earnings plays — stocks where we should BUY options NOW
        to profit from IV expansion before results.
        
        Sweet spot: 14-21 days before results (IV is still low, enough time)
        Danger zone: 0-3 days before results (IV already peaked, IV crush risk)
        
        Returns list of play candidates with timing info.
        """
        plays = []
        today = today_ist()

        for sym, info in self._calendar.items():
            if sym not in fno_stocks:
                continue
            
            try:
                date_str = info.get("date", "")
                for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        earn_date = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    continue

                days_to = (earn_date - today).days

                # ── ENTRY ZONE: 14-21 days before results ──
                # IV is low, options are cheap, buy now
                if 14 <= days_to <= 21:
                    plays.append({
                        "symbol": sym,
                        "date": str(earn_date),
                        "days_to": days_to,
                        "action": "BUY_IV_PLAY",
                        "score": 85,  # High score — best timing
                        "direction": "BULLISH",  # Default, can be overridden by news
                        "catalyst": f"EARNINGS PLAY: {sym} results in {days_to}d ({earn_date}). IV expansion expected.",
                        "reason": "Buy options now while IV is low. Sell before results for IV profit.",
                    })

                # ── MODERATE ZONE: 7-13 days before ──
                # IV starting to rise, still OK to enter with strong catalyst
                elif 7 <= days_to <= 13:
                    plays.append({
                        "symbol": sym,
                        "date": str(earn_date),
                        "days_to": days_to,
                        "action": "BUY_IF_STRONG",
                        "score": 70,  # Moderate — IV already rising
                        "direction": "BULLISH",
                        "catalyst": f"EARNINGS MODERATE: {sym} results in {days_to}d. IV rising — need strong catalyst.",
                        "reason": "IV rising. Only enter with score 75+.",
                    })

                # ── DANGER ZONE: 0-3 days before ──
                # IV at peak. DON'T enter new positions. EXIT existing.
                elif 0 <= days_to <= 3:
                    plays.append({
                        "symbol": sym,
                        "date": str(earn_date),
                        "days_to": days_to,
                        "action": "EXIT_IV_CRUSH",
                        "score": 0,  # Don't buy
                        "direction": "NEUTRAL",
                        "catalyst": f"IV CRUSH WARNING: {sym} results in {days_to}d. EXIT options now.",
                        "reason": "IV will crush after results. Sell before.",
                    })

            except Exception:
                pass

        return plays


# ═══════════════════════════════════════════════════════════════════════════════
#  ECON CALENDAR — PATCH_V62_ECON_CALENDAR
#  Blocks new F&O entries near high-impact macro events (FOMC, RBI, CPI, NFP).
#  Event risk can blow fresh positions. Small additive module, no tick/GTT changes.
# ═══════════════════════════════════════════════════════════════════════════════

class EconCalendar:
    """
    Hard-block F&O entries within N hours of known macro event windows.
    No external API dependency — we maintain a rolling list of event times in IST.
    User updates this list weekly (or we fetch from forexfactory RSS in future).

    Usage:
        econ = EconCalendar()
        if econ.is_blocked():
            log.info(f"[ECON] Blocking F&O entry: {econ.reason()}")
            return  # skip this scan cycle's entries

    Default block window: 2 hours before and 1 hour after each event.
    """

    # Upcoming known events (IST). Update weekly via Telegram command or manual edit.
    # Format: (datetime_ist, "EVENT_NAME", "IMPACT")   impact in {HIGH, MEDIUM}
    # HIGH = block F&O entries. MEDIUM = log only.
    # PATCH_V62_ECON_CALENDAR_SEED: initial seed. Rama to maintain going forward.
    _DEFAULT_EVENTS = [
        # US Fed (FOMC statements released 23:30 IST 8x/year — update as scheduled)
        # RBI MPC (usually 10:00 IST — typically announced weeks in advance)
        # India CPI (17:30 IST monthly, ~12th of each month)
        # India WPI (12:00 IST monthly, ~14th of each month)
        # US CPI (18:00 IST monthly, ~10th–15th)
        # US NFP (18:00 IST 1st Friday of month)
        # ECB (17:45 IST 8x/year)
        # Placeholders only. Rama edits econ_events.json on droplet.
    ]

    def __init__(self, block_hours_before=2.0, block_hours_after=1.0):
        self.block_before = block_hours_before
        self.block_after = block_hours_after
        self._events_file = DATA_DIR / "econ_events.json"
        self._events = []
        self._reason = ""
        self._load()

    def _load(self):
        """Load events from JSON. Format:
        {"events": [{"iso": "2026-05-07T23:30:00+05:30", "name": "FOMC", "impact": "HIGH"}, ...]}
        """
        try:
            data = load_json(self._events_file, {"events": []})
            raw = data.get("events", [])
            parsed = []
            for e in raw:
                try:
                    iso = e.get("iso", "")
                    dt = datetime.fromisoformat(iso)
                    parsed.append((dt, e.get("name", "EVENT"), e.get("impact", "HIGH").upper()))
                except Exception:
                    continue
            self._events = parsed
            if parsed:
                log.info(f"[ECON] Loaded {len(parsed)} events from {self._events_file}")
        except Exception as e:
            log.debug(f"[ECON] load failed: {e}")
            self._events = []

    def add_event(self, dt_ist, name, impact="HIGH"):
        """Add an event programmatically (e.g., from Telegram command)."""
        self._events.append((dt_ist, name, impact.upper()))
        self._persist()

    def _persist(self):
        out = {"events": [{"iso": dt.isoformat(), "name": n, "impact": imp}
                          for dt, n, imp in self._events]}
        save_json(self._events_file, out)

    def is_blocked(self, now=None):
        """Return True if we're in a HIGH-impact event window."""
        now = now or now_ist()
        for dt, name, impact in self._events:
            if impact != "HIGH":
                continue
            delta = (dt - now).total_seconds() / 3600.0  # hours until event
            if -self.block_after <= delta <= self.block_before:
                self._reason = f"{name} at {dt.strftime('%Y-%m-%d %H:%M')} IST (Δ={delta:+.1f}h)"
                return True
        self._reason = ""
        return False

    def reason(self):
        return self._reason

    def next_event(self, now=None):
        """Return the next upcoming event tuple, or None."""
        now = now or now_ist()
        future = [(dt, n, imp) for dt, n, imp in self._events if dt > now]
        if not future:
            return None
        future.sort(key=lambda x: x[0])
        return future[0]


# ═══════════════════════════════════════════════════════════════════════════════
#  GLOBAL PRE-MARKET — PATCH_V62_PREMARKET
#  Reads SGX Nifty / GIFT Nifty and overnight Dow/Nasdaq futures for morning bias.
#  Consumed by AMO morning path to raise panel threshold on risk-off days.
# ═══════════════════════════════════════════════════════════════════════════════

class GlobalPreMarket:
    """
    Lightweight pre-market reader.
    Reads overnight moves of:
      - GIFT Nifty (NSEIX, successor to SGX Nifty): leading indicator for NIFTY open
      - Dow / Nasdaq futures: US close/after-hours direction
      - Crude (CL=F): energy/inflation read
      - DXY (DX-Y.NYB): dollar = inverse NIFTY signal
    Returns a simple bias: BULLISH / NEUTRAL / BEARISH + magnitude %.
    """

    # yfinance tickers
    TICKERS = {
        "GIFT_NIFTY": "^NSEI",         # fallback if GIFT symbol not available via yf
        "DOW_FUT": "YM=F",
        "NAS_FUT": "NQ=F",
        "SP_FUT": "ES=F",
        "CRUDE": "CL=F",
        "DXY": "DX-Y.NYB",
    }

    def __init__(self):
        self._last_reading = None
        self._last_ts = 0

    def read(self, force=False):
        """Fetch overnight changes. Cached for 30 min."""
        import time as _t
        if not force and self._last_reading and (_t.time() - self._last_ts) < 1800:
            return self._last_reading

        try:
            import yfinance as yf
        except ImportError:
            log.warning("[PREMARKET] yfinance not installed; skipping")
            return None

        reading = {"ts": now_ist().isoformat(), "items": {}, "bias": "NEUTRAL", "score": 0}

        for key, tkr in self.TICKERS.items():
            try:
                t = yf.Ticker(tkr)
                hist = t.history(period="5d", interval="1d")
                if hist is None or len(hist) < 2:
                    continue
                prev = float(hist["Close"].iloc[-2])
                last = float(hist["Close"].iloc[-1])
                if prev <= 0:
                    continue
                pct = (last - prev) / prev * 100.0
                reading["items"][key] = {"prev": prev, "last": last, "pct": round(pct, 2)}
            except Exception as e:
                log.debug(f"[PREMARKET] {key} fetch failed: {e}")

        # PATCH_V54: dropped broken GIFT_NIFTY weight (^NSEI was yesterday's NIFTY close).
        # Real GIFT comes from Kite NSEIX in _fno_amo_premarket_place. V62 = pure US/dollar.
        score = 0.0
        items = reading["items"]
        if "DOW_FUT" in items:
            score += items["DOW_FUT"]["pct"] * 0.25
        if "NAS_FUT" in items:
            score += items["NAS_FUT"]["pct"] * 0.25
        if "SP_FUT" in items:
            score += items["SP_FUT"]["pct"] * 0.25
        if "DXY" in items:
            score -= items["DXY"]["pct"] * 0.25  # inverse
        # CRUDE excluded — yfinance CL=F has contract-roll artifacts (Apr 30: +8.33% spurious)

        reading["score"] = round(score, 2)
        if score >= 0.4:
            reading["bias"] = "BULLISH"
        elif score <= -0.4:
            reading["bias"] = "BEARISH"
        else:
            reading["bias"] = "NEUTRAL"

        self._last_reading = reading
        self._last_ts = _t.time()

        log.info(f"[PREMARKET] bias={reading['bias']} score={reading['score']} items={list(items.keys())}")
        return reading


# ═══════════════════════════════════════════════════════════════════════════════
#  TRENDLYNE SCORER — Fundamental quality filter
# ═══════════════════════════════════════════════════════════════════════════════

class AvgVolumeLoader:
    """FIX_V52: Pre-market 20-day average volume cache"""
    def __init__(self):
        self._avg_vol = {}
        self._loaded_date = None
        self._cache_file = DATA_DIR / "avg_volume_20d.json"
        self._lock = threading.Lock()
        self._loading = False

    def load_from_disk(self):
        try:
            data = load_json(self._cache_file, {})
            self._avg_vol = {k: int(v) for k, v in data.get("map", {}).items()}
            self._loaded_date = data.get("date")
            log.info(f"[FIX_V52 AVGVOL] loaded {len(self._avg_vol)} symbols from disk (date={self._loaded_date})")
        except Exception as _e:
            log.warning(f"[FIX_V52 AVGVOL] disk load failed: {_e}")

    def get(self, sym):
        return self._avg_vol.get(sym, 0)

    def is_fresh(self):
        return self._loaded_date == str(today_ist())

    def refresh_async(self, kite, symbols):
        with self._lock:
            if self._loading or self.is_fresh():
                return
            self._loading = True
        t = threading.Thread(target=self._refresh_worker, args=(kite, list(symbols)), daemon=True)
        t.start()
        log.info(f"[FIX_V52 AVGVOL] background refresh started for {len(symbols)} symbols")

    def _refresh_worker(self, kite, symbols):
        try:
            try:
                instruments = kite.instruments("NSE")
                tok_map = {i["tradingsymbol"]: i["instrument_token"] for i in instruments if i.get("instrument_type") == "EQ"}
            except Exception as _e:
                log.error(f"[FIX_V52 AVGVOL] instrument lookup failed: {_e}")
                with self._lock: self._loading = False
                return
            to_dt = datetime.combine(today_ist(), datetime.min.time())
            from_dt = to_dt - timedelta(days=35)
            done = 0
            failed = 0
            new_map = {}
            for sym in symbols:
                tok = tok_map.get(sym)
                if not tok:
                    continue
                try:
                    h = kite.historical_data(tok, from_dt, to_dt, "day")
                    if h and len(h) >= 5:
                        vols = [x.get("volume", 0) for x in h[-20:] if x.get("volume", 0) > 0]
                        if vols:
                            new_map[sym] = int(sum(vols) / len(vols))
                    done += 1
                except Exception:
                    failed += 1
                time.sleep(0.35)
                if done % 500 == 0 and done > 0:
                    log.info(f"[FIX_V52 AVGVOL] progress: {done}/{len(symbols)}, cached={len(new_map)}")
            with self._lock:
                self._avg_vol = new_map
                self._loaded_date = str(today_ist())
                self._loading = False
            try:
                save_json(self._cache_file, {"date": self._loaded_date, "map": new_map})
            except Exception:
                pass
            log.info(f"[FIX_V52 AVGVOL] refresh complete: {len(new_map)} symbols, {failed} failures")
        except Exception as _e:
            log.error(f"[FIX_V52 AVGVOL] worker: {_e}")
            with self._lock: self._loading = False


class TrendlyneScorer:
    """
    Uses Trendlyne StratQ CSV data for fundamental quality scoring.
    Filters: PE, ROE, promoter holding, debt ratios.
    Prevents buying garbage stocks with good news.

    Each module has its own instance.
    """

    def __init__(self):
        self._data = {}  # symbol -> {pe, roe, promoter_pct, debt_equity}
        self._csv_file = DATA_DIR / "trendlyne_stratq.csv"
        self._load()

    def _load(self):
        """Load Trendlyne CSV if available."""
        try:
            # Try xlsx first, then csv
            xlsx_file = DATA_DIR / "trendlyne_data.xlsx"
            csv_file = self._csv_file
            
            if xlsx_file.exists():
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        log.info("Trendlyne: xlsx empty")
                        return
                    headers = [str(h).strip() if h else "" for h in rows[0]]
                    for row in rows[1:]:
                        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                        sym = str(row_dict.get("NSE Code", row_dict.get("Symbol", row_dict.get("Stock Name", "")))).strip()
                        if not sym:
                            continue
                        def safe_float(val):
                            try: return float(val) if val else 0
                            except Exception: return 0  # V28 A8
                        # FIX_V4_INDUSTRY: capture industry + sector for dynamic theme resolution
                        self._data[sym] = {
                            "durability": safe_float(row_dict.get("Trendlyne Durability Score", 0)),
                            "valuation": safe_float(row_dict.get("Trendlyne Valuation Score", 0)),
                            "momentum": safe_float(row_dict.get("Trendlyne Momentum Score", 0)),
                            "promoter_pct": safe_float(row_dict.get("Promoter holding pledge percentage % Qtr", 0)),
                            "market_cap": safe_float(row_dict.get("Market Capitalization", 0)),
                            "industry": str(row_dict.get("Industry Name", "") or "").strip(),
                            "sector": str(row_dict.get("sector_name", "") or "").strip(),
                        }
                    wb.close()
                    log.info(f"Trendlyne: {len(self._data)} stocks loaded from xlsx")
                    return
                except ImportError:
                    log.info("Trendlyne: openpyxl not installed, trying csv")
                except Exception as xe:
                    log.warning(f"Trendlyne: xlsx read error: {xe}")
            
            if not csv_file.exists():
                log.info("Trendlyne: no data file found, skipping fundamental filter")
                return

            import csv
            with open(csv_file, "r") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    sym = row.get("Symbol", row.get("symbol", "")).strip()
                    if not sym:
                        continue
                    self._data[sym] = {
                        "pe": float(row.get("PE", row.get("pe", 0)) or 0),
                        "roe": float(row.get("ROE", row.get("roe", 0)) or 0),
                        "promoter_pct": float(row.get("Promoter%", row.get("promoter_pct", 0)) or 0),
                        "debt_equity": float(row.get("Debt/Equity", row.get("debt_equity", 0)) or 0),
                    }
            log.info(f"Trendlyne: {len(self._data)} stocks loaded from csv")
        except Exception as e:
            log.debug(f"Trendlyne load error: {e}")

    def score(self, symbol):
        """
        Return fundamental quality score (0-100) for a symbol.
        100 = excellent fundamentals, 0 = unknown/poor.
        Returns 50 if no data (neutral — don't block, don't boost).
        """
        info = self._data.get(symbol)
        if not info:
            return 50  # No data = neutral

        # Use Trendlyne DVM scores directly (0-100 each)
        d = info.get("durability", 0)
        v = info.get("valuation", 0)
        m = info.get("momentum", 0)
        
        # If DVM scores available, use weighted average
        if d > 0 or v > 0 or m > 0:
            # Durability 40% + Valuation 30% + Momentum 30%
            score = d * 0.4 + v * 0.3 + m * 0.3
            return max(0, min(100, score))
        
        # Fallback: no DVM data
        return 50

    def is_quality(self, symbol, min_score=40):
        """Quick check: is this stock above minimum quality threshold?"""
        return self.score(symbol) >= min_score

    # FIX_V4_INDUSTRY_RESOLVE: dynamic theme resolution via industry lookup
    def resolve_by_industry(self, industries, min_mcap_cr=500, max_stocks=30, valid_symbols=None):
        """Return NSE symbols from specified industries, filtered by mcap, sorted desc by mcap."""
        if not self._data or not industries:
            return []
        industry_set = set(industries)
        matches = []
        for sym, data in self._data.items():
            if data.get("industry") in industry_set:
                mcap = data.get("market_cap", 0) or 0
                if mcap >= min_mcap_cr:
                    if not valid_symbols or sym in valid_symbols:
                        matches.append((sym, mcap))
        matches.sort(key=lambda x: x[1], reverse=True)
        return [s for s, _ in matches[:max_stocks]]

    def get_industry(self, symbol):
        return self._data.get(symbol, {}).get("industry", "")

    def get_sector(self, symbol):
        return self._data.get(symbol, {}).get("sector", "")


# ═══════════════════════════════════════════════════════════════════════════════
#  AMO ENGINE — Places after-market orders for next day
# ═══════════════════════════════════════════════════════════════════════════════

class AMOEngine:
    """
    Places After Market Orders (AMO) for next day execution.
    Runs after 15:45 when evening scan finds strong catalysts.
    AMO orders are LIMIT orders at previous close ± buffer.

    Rules:
    - Total AMO cost must NOT exceed available capital
    - Cancel unfilled AMOs at 09:20 next morning
    - GTT only placed AFTER AMO fills (not before)
    """

    def __init__(self, name):
        self.name = name
        self._amo_file = DATA_DIR / f"amo_{name.lower()}.json"
        self._placed = []  # Track placed AMOs for morning cleanup

    def place_amo(self, symbol, price, qty, exchange="NSE", product="CNC"):
        """Place an AMO LIMIT order."""
        if not is_amo_window():
            return None

        kite = KiteSession.kite()
        if not kite:
            # Try emergency login for after-hours
            if not KiteSession.login():
                return None
            kite = KiteSession.kite()
            if not kite:
                return None

        try:
            limit_price = round(price * 1.01, 1)  # 1% above last close

            order_id = kite.place_order(
                variety="amo",
                exchange=exchange,
                tradingsymbol=symbol,
                transaction_type="BUY",
                quantity=qty,
                order_type="LIMIT",
                price=limit_price,
                product=product,
                validity="DAY",
                market_protection=5,  # SEBI mandatory
            )
            log.info(f"AMO {self.name}: placed {symbol} x{qty} @ {limit_price} id={order_id}")

            self._placed.append({
                "symbol": symbol,
                "qty": qty,
                "price": limit_price,
                "order_id": str(order_id),
                "time": now_ist().isoformat(),
            })
            self._save()
            return order_id

        except Exception as e:
            log.error(f"AMO {self.name}: place {symbol} failed: {e}")
            return None

    def cancel_stale(self):
        """Cancel unfilled AMOs at 09:20. Free up capital."""
        kite = KiteSession.kite()
        if not kite:
            return 0

        cancelled = 0
        try:
            orders = kite.orders()
            for o in orders:
                if (o.get("variety") == "amo" and
                    o.get("status") in ("OPEN", "TRIGGER PENDING") and
                    o.get("transaction_type") == "BUY"):
                    try:
                        kite.cancel_order(variety="amo", order_id=o["order_id"])
                        cancelled += 1
                        log.info(f"AMO {self.name}: cancelled stale {o.get('tradingsymbol')} id={o['order_id']}")
                    except Exception:
                        pass
        except Exception as e:
            log.error(f"AMO {self.name}: cancel stale failed: {e}")

        if cancelled:
            log.info(f"AMO {self.name}: cancelled {cancelled} stale orders")
        self._placed = []
        self._save()
        return cancelled

    def get_total_pending_cost(self):
        """Total cost of pending AMOs — used to prevent over-ordering."""
        return sum(p.get("price", 0) * p.get("qty", 0) for p in self._placed)

    def _save(self):
        save_json(self._amo_file, {
            "placed": self._placed,
            "date": str(today_ist()),
        })

    def _load(self):
        data = load_json(self._amo_file, {})
        if data.get("date") == str(today_ist()):
            self._placed = data.get("placed", [])
        else:
            self._placed = []


# ═══════════════════════════════════════════════════════════════════════════════
#  EOD REPORTER — End of day summary to Telegram
# ═══════════════════════════════════════════════════════════════════════════════

class EODReporter:
    """
    Generates end-of-day summary for Telegram.
    Runs at 15:35 on trading days.
    """

    @staticmethod
    def generate_equity_report(equity_module):
        """Generate equity EOD summary."""
        positions = equity_module.positions
        if not positions:
            return "Equity: No open positions"

        kite = KiteSession.kite()
        if not kite:
            return "Equity: Kite not connected for EOD"

        try:
            sym_list = [f"NSE:{s}" for s in positions]
            ltp_data = kite.ltp(sym_list)

            lines = [f"*EQUITY EOD — {len(positions)} positions*\n"]
            total_pnl = 0

            for sym, pos in sorted(positions.items()):
                ltp = ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                entry = pos.get("entry_price", 0)
                qty = pos.get("qty", 0)
                pnl = (ltp - entry) * qty if ltp > 0 and entry > 0 else 0
                pnl_pct = ((ltp - entry) / entry * 100) if entry > 0 and ltp > 0 else 0
                total_pnl += pnl

                emoji = "+" if pnl >= 0 else ""
                lines.append(f"`{sym:12s}` {emoji}Rs.{pnl:,.0f} ({emoji}{pnl_pct:.1f}%)")

            # PATCH_V89_AUTHFIX: ensure fresh kite client from saved token file
            try:
                import json as _ajson, os as _aos
                from kiteconnect import KiteConnect as _KC
                _ak = _aos.getenv("KITE_API_KEY", "")
                with open("/home/globalbot/data/kite_token.json") as _atf:
                    _atok = _ajson.load(_atf)
                kite = _KC(api_key=_ak)
                kite.set_access_token(_atok["access_token"])
            except Exception as _ae:
                lines.append(f"_(auth fallback failed: {_ae})_")

            # PATCH_V89: Correct day P&L computation
            from datetime import datetime as _dt
            today_str = _dt.now().strftime("%Y-%m-%d")

            # 1. REALIZED P&L from today's completed orders (BUY+SELL pairs)
            realized_total = 0
            realized_lines = []
            try:
                from collections import defaultdict
                _by_sym = defaultdict(lambda: {'b': [], 's': []})
                for o in kite.orders():
                    if o.get('status') != 'COMPLETE': continue
                    ts = str(o.get('order_timestamp', ''))
                    if today_str not in ts: continue
                    q = o.get('filled_quantity', 0); p = o.get('average_price', 0)
                    if q == 0 or p == 0: continue
                    side = 'b' if o.get('transaction_type') == 'BUY' else 's'
                    _by_sym[o['tradingsymbol']][side].append((q, p))

                for _sym, _t in _by_sym.items():
                    _bq = sum(q for q,_ in _t['b']); _sq = sum(q for q,_ in _t['s'])
                    if _bq and _sq:
                        _bv = sum(q*p for q,p in _t['b']); _sv = sum(q*p for q,p in _t['s'])
                        _ab, _as_ = _bv/_bq, _sv/_sq
                        _matched = min(_bq, _sq)
                        _r = (_as_ - _ab) * _matched
                        realized_total += _r
                        realized_lines.append(f"  {_sym[:18]:18s} {_matched:>4} @ {_ab:.2f}->{_as_:.2f} = Rs.{_r:+,.0f}")
                    elif _sq and not _bq:
                        # PATCH_V90: SOLD-HOLDING (sold what we held from before).
                        # The day P&L for this is ALREADY in Kite's day_change×qty on the
                        # holding for the period before the sell. Showing lifetime gain
                        # here would double-count and inflate "today's profit".
                        # Just record the trade for visibility, don't add to realized total.
                        _sv = sum(q*p for q,p in _t['s'])
                        _as_ = _sv/_sq
                        realized_lines.append(f"  {_sym[:18]:18s} sold {_sq} @ {_as_:.2f} (lifetime gain not counted as day profit)")
            except Exception as _e:
                realized_lines.append(f"  (realized calc error: {str(_e)[:60]})")

            # 2. UNREALIZED TODAY from kite.holdings day_change
            day_unrealized = 0
            lifetime_unrealized = 0
            try:
                for _h in kite.holdings():
                    _q = _h['quantity'] + _h.get('t1_quantity', 0)
                    if _q == 0: continue
                    day_unrealized += _h.get('day_change', 0) * _q
                    lifetime_unrealized += _h.get('pnl', 0)
            except Exception as _e:
                pass

            # 3. F&O P&L from positions (intraday + carry)
            fno_pnl = 0
            try:
                for _p in kite.positions().get('net', []):
                    if _p['quantity'] != 0 or _p.get('m2m', 0):
                        fno_pnl += _p.get('pnl', 0)
            except Exception:
                pass

            day_total = realized_total + day_unrealized + fno_pnl

            # Build clean message
            lines.append("")
            lines.append("─" * 30)
            lines.append("*📊 DAY P&L — {}*".format(today_str))
            lines.append("─" * 30)

            if realized_lines:
                lines.append(f"*Realized ({len(realized_lines)} trades):*")
                lines.extend(realized_lines)
                lines.append(f"*Realized total: Rs.{realized_total:+,.0f}*")
            else:
                lines.append("Realized today: Rs.0 (no closed trades)")

            lines.append("")
            lines.append(f"Equity day move:  Rs.{day_unrealized:+,.0f}")
            lines.append(f"F&O P&L:          Rs.{fno_pnl:+,.0f}")
            lines.append(f"*DAY TOTAL:       Rs.{day_total:+,.0f}*")
            lines.append("")
            lines.append(f"_Lifetime unrealized: Rs.{lifetime_unrealized:+,.0f}_")
            lines.append(f"_Capital deployed: Rs.{equity_module.capital:,.0f}_")

            # 4. Write daily ledger for weekly review
            try:
                import json as _json
                _ledger = {
                    "date": today_str,
                    "realized_total": round(realized_total, 2),
                    "realized_trades": realized_lines,
                    "equity_day_unrealized": round(day_unrealized, 2),
                    "fno_pnl": round(fno_pnl, 2),
                    "day_total": round(day_total, 2),
                    "lifetime_unrealized": round(lifetime_unrealized, 2),
                    "capital": equity_module.capital,
                }
                _ledger_path = Path(f"/home/globalbot/data/daily_pnl_{today_str.replace('-','')}.json")
                _ledger_path.write_text(_json.dumps(_ledger, indent=2))
            except Exception:
                pass

            return "\n".join(lines)

        except Exception as e:
            return f"Equity EOD error: {e}"

    @staticmethod
    def generate_lifetime_report(equity_module, fno_module):
        """PATCH_V73: Since-inception performance summary."""
        try:
            from datetime import datetime

            def _load(module):
                try:
                    return load_json(module._trades_file, {"trades": []}).get("trades", [])
                except Exception:
                    return []

            def _parse_dt(s):
                try:
                    return datetime.fromisoformat(str(s).replace("Z","")).replace(tzinfo=None)
                except Exception:
                    return None

            def _section(label, trades):
                closed = [t for t in trades if t.get("pnl") is not None and t.get("exit_time")]
                if not closed:
                    return f"*{label}*: No closed trades\n"
                pnls = [t.get("pnl", 0) for t in closed]
                wins = [p for p in pnls if p > 0]
                losses = [p for p in pnls if p < 0]
                total = sum(pnls)
                wr = (len(wins) / len(closed) * 100) if closed else 0
                avg_w = (sum(wins) / len(wins)) if wins else 0
                avg_l = (sum(losses) / len(losses)) if losses else 0
                pf = (sum(wins) / abs(sum(losses))) if losses and sum(losses) != 0 else 0
                # max drawdown on cumulative curve
                sorted_t = sorted(closed, key=lambda t: t.get("exit_time",""))
                running = 0; peak = 0; max_dd = 0
                for t in sorted_t:
                    running += t.get("pnl", 0)
                    if running > peak: peak = running
                    dd = peak - running
                    if dd > max_dd: max_dd = dd
                dates = [d for d in (_parse_dt(t.get("exit_time")) for t in closed) if d]
                first = min(dates).strftime("%Y-%m-%d") if dates else "?"
                last  = max(dates).strftime("%Y-%m-%d") if dates else "?"
                days = ((max(dates) - min(dates)).days + 1) if len(dates) >= 2 else 1
                best = max(closed, key=lambda t: t.get("pnl", 0))
                worst = min(closed, key=lambda t: t.get("pnl", 0))
                lines = [
                    f"*{label}*",
                    f"Period: {first} to {last} ({days}d)",
                    f"Trades: {len(closed)} | {len(wins)}W / {len(losses)}L | WR {wr:.0f}%",
                    f"Total P&L: Rs.{total:+,.0f}",
                    f"Avg W: Rs.{avg_w:+,.0f} | Avg L: Rs.{avg_l:+,.0f} | PF: {pf:.2f}",
                    f"Max DD: Rs.{max_dd:,.0f}",
                    f"Best: {best.get('symbol', best.get('tradingsymbol','?'))} Rs.{best.get('pnl',0):+,.0f}",
                    f"Worst: {worst.get('symbol', worst.get('tradingsymbol','?'))} Rs.{worst.get('pnl',0):+,.0f}",
                ]
                # by source
                by_src = {}
                for t in closed:
                    src = (t.get("source") or "UNKNOWN").split("|")[0].split("(")[0].strip()[:20]
                    if src not in by_src:
                        by_src[src] = {"n":0, "w":0, "pnl":0.0}
                    by_src[src]["n"] += 1
                    by_src[src]["pnl"] += t.get("pnl", 0)
                    if t.get("pnl", 0) > 0: by_src[src]["w"] += 1
                if by_src:
                    lines.append("By source:")
                    for src, s in sorted(by_src.items(), key=lambda x: -x[1]["pnl"]):
                        wr_s = (s["w"]/s["n"]*100) if s["n"] else 0
                        lines.append(f"  {src}: n={s['n']} WR={wr_s:.0f}% Rs.{s['pnl']:+,.0f}")
                return "\n".join(lines) + "\n"

            eq = _load(equity_module)
            fno = _load(fno_module)
            msg = ["LIFETIME PERFORMANCE REPORT", ""]
            msg.append(_section("EQUITY", eq))
            msg.append(_section("F&O", fno))
            all_closed = ([t for t in eq if t.get("pnl") is not None and t.get("exit_time")] +
                          [t for t in fno if t.get("pnl") is not None and t.get("exit_time")])
            if all_closed:
                total_combined = sum(t.get("pnl", 0) for t in all_closed)
                msg.append(f"*COMBINED P&L: Rs.{total_combined:+,.0f}* ({len(all_closed)} trades)")
            return "\n".join(msg)
        except Exception as e:
            return f"Lifetime report error: {e}"


    @staticmethod
    def check_lifetime_flag(equity_module, fno_module):
        """PATCH_V73: Polls /home/globalbot/data/.lifetime_pending; if exists,
        sends lifetime report and deletes flag. Called from main scheduler."""
        try:
            flag = DATA_DIR / ".lifetime_pending"
            if flag.exists():
                flag.unlink()
                rpt = EODReporter.generate_lifetime_report(equity_module, fno_module)
                send_telegram(rpt, silent=False)
                log.info("[PATCH_V73] one-shot lifetime report sent")
        except Exception as e:
            log.error(f"[PATCH_V73] flag-trigger failed: {e}")


    @staticmethod
    def generate_weekly_attribution(equity_module, fno_module):
        """
        PATCH_V2_WEEKLY_ATTRIBUTION
        Generate a weekly P&L attribution report grouped by source.
        Reads trade history files (read-only), no orders placed.
        Runs every Sunday at 09:00 IST from main loop.
        """
        try:
            from datetime import datetime, timedelta
            cutoff = now_ist() - timedelta(days=7)

            def _parse_time(t):
                try:
                    return datetime.fromisoformat(t.replace("Z", ""))
                except Exception:
                    return None

            def _load_trades(module):
                try:
                    data = load_json(module._trades_file, {"trades": []})
                    return data.get("trades", [])
                except Exception:
                    return []

            def _filter_week(trades):
                out = []
                for t in trades:
                    et = _parse_time(t.get("exit_time", ""))
                    if et and et.replace(tzinfo=None) >= cutoff.replace(tzinfo=None):
                        out.append(t)
                return out

            eq_trades = _filter_week(_load_trades(equity_module))
            fno_trades = _filter_week(_load_trades(fno_module))

            def _attribute(trades):
                by_source = {}
                for t in trades:
                    src = t.get("source", "UNKNOWN")
                    pnl = t.get("pnl", 0)
                    if src not in by_source:
                        by_source[src] = {"count": 0, "wins": 0, "losses": 0, "pnl": 0.0, "win_pnl": 0.0, "loss_pnl": 0.0}
                    by_source[src]["count"] += 1
                    by_source[src]["pnl"] += pnl
                    if pnl > 0:
                        by_source[src]["wins"] += 1
                        by_source[src]["win_pnl"] += pnl
                    elif pnl < 0:
                        by_source[src]["losses"] += 1
                        by_source[src]["loss_pnl"] += pnl
                return by_source

            eq_attr = _attribute(eq_trades)
            fno_attr = _attribute(fno_trades)

            def _format_section(name, attr, trades):
                if not trades:
                    return f"*{name}*: No trades this week\n"
                total_pnl = sum(t.get("pnl", 0) for t in trades)
                total_count = len(trades)
                wins = sum(1 for t in trades if t.get("pnl", 0) > 0)
                wr = (wins / total_count * 100) if total_count else 0
                lines = [f"*{name}* — {total_count} trades | WR {wr:.0f}% | P&L Rs.{total_pnl:,.0f}"]
                for src in sorted(attr.keys(), key=lambda s: -attr[s]["pnl"]):
                    s = attr[src]
                    src_wr = (s["wins"] / s["count"] * 100) if s["count"] else 0
                    avg_win = (s["win_pnl"] / s["wins"]) if s["wins"] else 0
                    avg_loss = (s["loss_pnl"] / s["losses"]) if s["losses"] else 0
                    pf = (s["win_pnl"] / abs(s["loss_pnl"])) if s["loss_pnl"] != 0 else 0
                    # PATCH_V70_AUDIT_CLEANUP: avg_win/avg_loss were computed but not displayed
                    lines.append(
                        f"`{src:12s}` n={s['count']} WR={src_wr:.0f}% PnL=Rs.{s['pnl']:,.0f} AvgW=Rs.{avg_win:,.0f} AvgL=Rs.{avg_loss:,.0f} PF={pf:.1f}"
                    )
                return "\n".join(lines) + "\n"

            best_eq = max(eq_trades, key=lambda t: t.get("pnl", 0)) if eq_trades else None
            worst_eq = min(eq_trades, key=lambda t: t.get("pnl", 0)) if eq_trades else None
            best_fno = max(fno_trades, key=lambda t: t.get("pnl", 0)) if fno_trades else None
            worst_fno = min(fno_trades, key=lambda t: t.get("pnl", 0)) if fno_trades else None

            msg_lines = ["*WEEKLY ATTRIBUTION REPORT*"]  # PATCH_V70: f-string had no placeholders
            msg_lines.append(f"Period: last 7 days ending {now_ist().strftime('%Y-%m-%d')}")
            msg_lines.append("")
            msg_lines.append(_format_section("EQUITY", eq_attr, eq_trades))
            msg_lines.append(_format_section("F&O", fno_attr, fno_trades))

            if best_eq:
                msg_lines.append(f"Best EQ: {best_eq.get('symbol','?')} Rs.{best_eq.get('pnl',0):,.0f} ({best_eq.get('source','?')})")
            if worst_eq:
                msg_lines.append(f"Worst EQ: {worst_eq.get('symbol','?')} Rs.{worst_eq.get('pnl',0):,.0f} ({worst_eq.get('source','?')})")
            if best_fno:
                msg_lines.append(f"Best FNO: {best_fno.get('tradingsymbol','?')} Rs.{best_fno.get('pnl',0):,.0f} ({best_fno.get('source','?')})")
            if worst_fno:
                msg_lines.append(f"Worst FNO: {worst_fno.get('tradingsymbol','?')} Rs.{worst_fno.get('pnl',0):,.0f} ({worst_fno.get('source','?')})")

            return "\n".join(msg_lines)
        except Exception as e:
            return f"Weekly attribution error: {e}"

    @staticmethod
    def generate_fno_report(fno_module):
        """Generate F&O EOD summary.

        V27 Stage 2 PATCH (Bug #6 follow-up): Filter out phantom positions.
        Only include positions with non-zero qty in Kite. For positions that
        closed today, fetch real exit price from tradebook.
        """
        positions = fno_module.positions
        if not positions:
            return "F&O: No open positions"

        kite = KiteSession.kite()
        if not kite:
            return "F&O: Kite not connected for EOD"

        try:
            # Build Kite truth: which positions have qty != 0
            _kite_qtys = {}
            try:
                for _kp in kite.positions().get("net", []):
                    if _kp.get("exchange") == "NFO":
                        _kite_qtys[_kp.get("tradingsymbol", "")] = _kp.get("quantity", 0)
            except Exception as _ke:
                return f"F&O EOD: Kite positions fetch failed: {_ke}"

            # Build today's tradebook for realized exits
            _tradebook = {}
            try:
                for _t in (kite.trades() or []):
                    if _t.get("exchange") != "NFO":
                        continue
                    _tsym = _t.get("tradingsymbol", "")
                    _ttype = _t.get("transaction_type", "")
                    _qty = _t.get("quantity", 0)
                    _avg = _t.get("average_price", 0)
                    if _tsym not in _tradebook:
                        _tradebook[_tsym] = {"sell_qty": 0, "sell_value": 0}
                    if _ttype == "SELL":
                        _tradebook[_tsym]["sell_qty"] += _qty
                        _tradebook[_tsym]["sell_value"] += _qty * _avg
            except Exception as _te:
                log.warning(f"F&O EOD: tradebook fetch failed: {_te}")

            _open_syms = [t for t, q in _kite_qtys.items() if q != 0]
            _ltp_data = {}
            if _open_syms:
                try:
                    _ltp_data = kite.ltp([f"NFO:{s}" for s in _open_syms])
                except Exception as _le:
                    log.warning(f"F&O EOD: LTP fetch failed: {_le}")

            open_lines, closed_lines = [], []
            total_open_pnl = 0
            total_closed_pnl = 0
            open_count = 0
            closed_count = 0

            for key, pos in positions.items():
                tsym = pos.get("tradingsymbol", "")
                if not tsym:
                    continue
                entry = pos.get("entry_price", 0)
                direction = pos.get("direction", "")
                _kqty = _kite_qtys.get(tsym, 0)

                if _kqty != 0:
                    ltp = _ltp_data.get(f"NFO:{tsym}", {}).get("last_price", 0)
                    pnl = (ltp - entry) * abs(_kqty) if ltp > 0 and entry > 0 else 0
                    total_open_pnl += pnl
                    emoji = "+" if pnl >= 0 else ""
                    open_lines.append(f"`{tsym[:20]:20s}`  {emoji}Rs.{pnl:,.0f} | OPEN | {direction}")
                    open_count += 1
                else:
                    _tb = _tradebook.get(tsym)
                    if _tb and _tb["sell_qty"] > 0:
                        _avg_sell = _tb["sell_value"] / _tb["sell_qty"]
                        _real_pnl = (_avg_sell - entry) * _tb["sell_qty"]
                        total_closed_pnl += _real_pnl
                        emoji = "+" if _real_pnl >= 0 else ""
                        closed_lines.append(f"`{tsym[:20]:20s}`  {emoji}Rs.{_real_pnl:,.0f} | CLOSED at {_avg_sell:,.1f} | {direction}")
                        closed_count += 1
                    else:
                        closed_lines.append(f"`{tsym[:20]:20s}`  CLOSED (no tradebook entry) | {direction}")
                        closed_count += 1

            lines = [f"*F&O EOD: {open_count} open, {closed_count} closed today*\n"]
            if open_lines:
                lines.append("*OPEN POSITIONS:*")
                lines.extend(open_lines)
                lines.append(f"\n*Open Unrealized P&L: Rs.{total_open_pnl:,.0f}*")
            if closed_lines:
                lines.append("\n*CLOSED TODAY:*")
                lines.extend(closed_lines)
                lines.append(f"*Closed Realized P&L: Rs.{total_closed_pnl:,.0f}*")
            lines.append(f"\n*Day Total: Rs.{(total_open_pnl + total_closed_pnl):,.0f}*")
            lines.append(f"F&O Capital: Rs.{fno_module.capital:,.0f}")
            return "\n".join(lines)

        except Exception as e:
            return f"F&O EOD error: {e}"


# ═══════════════════════════════════════════════════════════════════════════════
#  RISK MANAGER — Each module has its own instance
# ═══════════════════════════════════════════════════════════════════════════════

class RiskManager:
    """
    Tracks daily/monthly P&L and enforces risk limits.
    Each module has its own RiskManager — independent tracking.

    Circuit breaker: based on AVERAGE loss per trade (not cumulative).
    """

    def __init__(self, name, capital):
        self.name = name
        self.capital = capital
        # V27 PATCH (Bug #9): Persist breaker state across restarts
        self._state_file = f"/home/globalbot/risk_{name.lower()}.json"
        self._trades_today = []
        self._monthly_pnl = 0.0
        self._halted = False
        self._date = today_ist()
        self._month = today_ist().month
        self._load_state()

    def _load_state(self):
        """V27 PATCH (Bug #9): Reload breaker state on init."""
        try:
            data = load_json(self._state_file, {})
            if not isinstance(data, dict):
                return
            _saved_date = data.get("date", "")
            _saved_month = data.get("month", 0)
            _today = today_ist()
            if _saved_date == _today.isoformat():
                self._trades_today = data.get("trades_today", [])
                self._halted = data.get("halted", False)
                # FIX_V29_AUTO_UNHALT: clear stale halt if loaded date is old
                self._check_new_day()
                if self._halted:
                    log.warning(f"Risk {self.name}: RESTORED HALT STATE from disk (day={_saved_date})")
            if _saved_month == _today.month:
                self._monthly_pnl = data.get("monthly_pnl", 0.0)
        except Exception as e:
            log.error(f"Risk {self.name}: load state failed: {e}")

    def _save_state(self):
        """V27 PATCH (Bug #9): Persist breaker state."""
        try:
            save_json(self._state_file, {
                "date": self._date.isoformat(),
                "month": self._month,
                "trades_today": self._trades_today,
                "monthly_pnl": self._monthly_pnl,
                "halted": self._halted,
            })
        except Exception as e:
            log.error(f"Risk {self.name}: save state failed: {e}")

    def _check_new_day(self):
        d = today_ist()
        if d != self._date:
            self._trades_today = []
            self._halted = False
            self._date = d
        if d.month != self._month:
            self._monthly_pnl = 0.0
            self._month = d.month

    def record_trade(self, pnl):
        """Record a closed trade's P&L."""
        self._check_new_day()
        self._trades_today.append(pnl)
        self._monthly_pnl += pnl
        self._check_limits()
        self._save_state()  # V27 PATCH (Bug #9): persist after every trade

    def _check_limits(self):
        """Check if we should halt trading."""
        if not self._trades_today:
            return

        # Average loss check (not cumulative)
        avg_pnl = sum(self._trades_today) / len(self._trades_today)
        avg_pct = avg_pnl / self.capital if self.capital > 0 else 0

        if avg_pct < -DAILY_LOSS_AVG_LIMIT:
            self._halted = True
            log.warning(f"Risk {self.name}: HALTED — avg loss {avg_pct:.1%} exceeds {DAILY_LOSS_AVG_LIMIT:.0%}")

        # Monthly loss check
        monthly_pct = self._monthly_pnl / self.capital if self.capital > 0 else 0
        if monthly_pct < -MONTHLY_LOSS_LIMIT:
            self._halted = True
            log.warning(f"Risk {self.name}: HALTED — monthly loss {monthly_pct:.1%} exceeds {MONTHLY_LOSS_LIMIT:.0%}")

    def can_trade(self):
        """Can this module place new trades?"""
        self._check_new_day()
        return not self._halted




# ═══════════════════════════════════════════════════════════════════════════════
#  SAFETY FILTERS — Circuit limit, Nifty crash, correlation, bad stock
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
#  EQUITY SMART SCANNERS — Proactive stock discovery
# ═══════════════════════════════════════════════════════════════════════════════

class EquitySmartScanner:
    """
    Proactive equity scanners that find opportunities WITHOUT waiting for news.
    Each scanner runs independently during market hours.
    """

    @staticmethod
    def scan_52week_low_recovery(kite, valid_symbols, trendlyne):
        """FIX_V52: Full universe with tier-aware filters."""
        candidates = []
        try:
            batch_size = 500
            sym_list = list(valid_symbols)

            for i in range(0, len(sym_list), batch_size):
                batch = sym_list[i:i+batch_size]
                try:
                    quotes = kite.quote([f"NSE:{s}" for s in batch])
                    for sym in batch:
                        q = quotes.get(f"NSE:{sym}", {})
                        ltp = q.get("last_price", 0)
                        ohlc = q.get("ohlc", {})
                        low_52 = ohlc.get("low", 0)  # 52-week low
                        high_52 = ohlc.get("high", 0)  # 52-week high

                        if ltp <= 0 or low_52 <= 0 or high_52 <= 0:
                            continue

                        # Stock within 15% of 52-week low
                        pct_from_low = (ltp - low_52) / low_52 if low_52 > 0 else 999
                        if pct_from_low > 0.15:
                            continue  # Not near 52-week low

                        _mcap = trendlyne._data.get(sym, {}).get("market_cap", 0) if trendlyne else 0
                        _min_score = 35 if _mcap > 20000 else (40 if _mcap > 5000 else 45)  # V74: relaxed 60/50/40 -> 45/40/35
                        # V80_CHANGE1: stocks with mcap=0 (not in Trendlyne cache) bypass the score floor.
                        # They still must pass volume/turnover/circuit/panel gates downstream.
                        if _mcap > 0 and trendlyne and trendlyne.score(sym) < _min_score:
                            continue
                        # V80: legacy mcap=0+score=60 check is now dead (V74 made min 45). Kept as no-op.
                        _volume = q.get("volume", 0)
                        # V80_CHANGE3: tier-aware turnover floor
                        _durability = trendlyne._data.get(sym, {}).get("durability", 0) if trendlyne else 0
                        if _mcap == 0:
                            _min_turnover = 10000000   # Rs 1 Cr — no Trendlyne data, extra caution
                        elif _durability >= 80:
                            _min_turnover = 2000000    # Rs 20 L — proven quality stock
                        else:
                            _min_turnover = 5000000    # Rs 50 L — default
                        if _volume * ltp < _min_turnover:
                            continue
                        _uc = q.get("upper_circuit_limit", 0)
                        _lc = q.get("lower_circuit_limit", 0)
                        if _uc > 0 and ltp >= _uc * 0.995: continue
                        if _lc > 0 and ltp <= _lc * 1.005: continue

                        # Must have fallen at least 30% from high (genuine beaten down)
                        pct_from_high = (high_52 - ltp) / high_52 if high_52 > 0 else 0
                        if pct_from_high < 0.30:
                            continue

                        score = 70 + int(pct_from_high * 30)  # Higher score for more beaten down
                        candidates.append({
                            "symbol": sym,
                            "direction": "BULLISH",
                            "score": min(score, 90),
                            "catalyst": f"52WK LOW RECOVERY: {sym} at {pct_from_low:.0%} from low, down {pct_from_high:.0%} from high",
                            "source": "52WK_LOW",
                        })
                except Exception:
                    pass
                time.sleep(0.5)  # Rate limit

        except Exception as e:
            log.debug(f"52wk scanner error: {e}")

        if candidates:
            log.info(f"52wk low scanner: {len(candidates)} recovery candidates")
        return candidates[:10]  # Top 10

    @staticmethod
    def scan_promoter_increase(kite, valid_symbols):
        """
        Find stocks where promoter holding has increased.
        Promoter buying = ultimate insider confidence signal.
        Scan NSE SAST filings for shareholding changes.
        """
        candidates = []
        try:
            session = requests.Session()
            session.headers.update({
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/json",
            })
            session.get("https://www.nseindia.com", timeout=10)
            time.sleep(0.5)

            today = today_ist()
            from_date = (today - timedelta(days=7)).strftime("%d-%m-%Y")
            to_date = today.strftime("%d-%m-%Y")

            url = f"https://www.nseindia.com/api/corporate-announcements?index=equities&from_date={from_date}&to_date={to_date}"
            resp = session.get(url, timeout=15)

            if resp.status_code == 200:
                filings = resp.json()
                if isinstance(filings, list):
                    for f in filings[:200]:
                        subject = f.get("desc", "").lower()
                        symbol = f.get("symbol", "")
                        if not symbol or symbol not in valid_symbols:
                            continue

                        # Look for promoter increase signals
                        if any(kw in subject for kw in [
                            "increase in shareholding", "acquired shares",
                            "promoter buy", "promoter acqui", "sast",
                            "increase in stake", "open market purchase"
                        ]):
                            candidates.append({
                                "symbol": symbol,
                                "direction": "BULLISH",
                                "score": 85,  # High score — promoter buying is strongest signal
                                "catalyst": f"PROMOTER INCREASE: {subject[:80]}",
                                "source": "PROMOTER_BUY",
                            })

        except Exception as e:
            log.debug(f"Promoter scanner error: {e}")

        if candidates:
            log.info(f"Promoter scanner: {len(candidates)} insider buying signals")
        return candidates

    @staticmethod
    def scan_bulk_block_deals(valid_symbols):
        """
        PATCH_V94: Bulk + block deal scout.
        Reads NSE bulk-deals and block-deals feeds, matches client name against
        three lists:
          - Superstar retail investors (Kedia, Damani, Khanna, Kacholia, ...)
          - Foreign institutional investors (Morgan Stanley, GS, Vanguard, ...)
          - Domestic institutional investors (HDFC MF, SBI MF, LIC, ...)
        Emits BULLISH candidates with source SUPERSTAR_BUY / FII_BLOCK_BUY / DII_BLOCK_BUY.
        Daily run after 18:00 IST (NSE publishes T+0 deals around 18:00).
        """
        candidates = []

        # PATCH_V94: hardcoded name lists (case-insensitive substring match on client name)
        SUPERSTAR_NAMES = [
            "rakesh jhunjhunwala", "rekha jhunjhunwala", "rare enterprises",
            "radhakishan damani", "damani", "bright star investments",
            "vijay kedia", "kedia securities", "kedia capital",
            "dolly khanna", "rajiv khanna",
            "mukul agrawal", "mukul mahavir agrawal",
            "ashish kacholia", "lucky securities",
            "porinju veliyath", "equity intelligence",
            "ashish dhawan", "chrys capital",
            "sunil singhania", "abakkus",
            "anil kumar goel", "ashok kumar goel",
            "madhusudan kela", "madhuri kela", "mk ventures",
            "nemish shah", "enam",
            "ramesh damani",
            "bhavook tripathi",
            "hemendra kothari", "dsp blackrock",
            "akash bhanshali", "om kotak", "enam asset",
            "shivanand mankekar", "kalpana mankekar",
        ]
        FII_NAMES = [
            "morgan stanley", "goldman sachs", "merrill lynch", "bank of america",
            "jpmorgan", "j.p. morgan", "citigroup", "credit suisse", "ubs",
            "deutsche bank", "barclays", "hsbc",
            "vanguard", "blackrock", "fidelity", "capital group", "capital research",
            "t. rowe price", "wellington management", "invesco",
            "gmo llc", "gmo", "schroder", "schroders", "aberdeen", "abrdn",
            "matthews asia", "mathews india", "amundi",
            "government of singapore", "gic private", "monetary authority of singapore",
            "abu dhabi investment", "adia", "kuwait investment authority",
            "norges bank", "norway pension", "government pension fund global",
            "saudi arabia monetary", "qatar investment",
            "smallcap world fund", "emerging markets growth",
        ]
        DII_NAMES = [
            "hdfc mutual fund", "hdfc mf", "hdfc trustee",
            "sbi mutual fund", "sbi mf", "sbi funds management",
            "icici prudential", "icici pru",
            "kotak mahindra mutual", "kotak mf", "kotak mahindra mf",
            "axis mutual fund", "axis mf",
            "nippon life", "nippon india",
            "aditya birla sun life", "absl mutual fund", "absl mf",
            "uti mutual fund", "uti mf",
            "dsp mutual fund", "dsp mf", "dsp blackrock",
            "tata mutual fund", "tata mf",
            "mirae asset", "mirae mf",
            "franklin templeton", "franklin india",
            "life insurance corporation", "lic of india", "lic mf",
            "general insurance corporation", "gic re",
            "new india assurance", "national insurance",
            "ppfas mutual fund", "parag parikh",
            "quantum mutual fund", "quantum mf",
            "edelweiss mutual fund", "edelweiss mf",
        ]

        try:
            session = requests.Session()
            session.headers.update({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Accept": "application/json",
                "Accept-Language": "en-US,en;q=0.9",
                "Referer": "https://www.nseindia.com/",
            })
            session.get("https://www.nseindia.com", timeout=10)
            time.sleep(0.6)

            for feed_name, url in [
                ("BULK", "https://www.nseindia.com/api/historical/cm/bulkdeals"),
                ("BLOCK", "https://www.nseindia.com/api/historical/cm/blockdeals"),
            ]:
                try:
                    today = today_ist()
                    from_date = (today - timedelta(days=2)).strftime("%d-%m-%Y")
                    to_date = today.strftime("%d-%m-%Y")
                    params_url = f"{url}?from={from_date}&to={to_date}"
                    resp = session.get(params_url, timeout=15)
                    if resp.status_code != 200:
                        log.debug(f"[V94_{feed_name}] HTTP {resp.status_code}")
                        continue
                    data = resp.json() if resp.content else {}
                    rows = data.get("data", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
                    log.info(f"[V94_{feed_name}] {len(rows)} {feed_name.lower()}-deal rows fetched")

                    for row in rows:
                        sym = (row.get("symbol") or row.get("BD_SYMBOL") or "").strip()
                        if not sym or sym not in valid_symbols:
                            continue
                        client = (row.get("clientName") or row.get("BD_CLIENT_NAME") or "").lower().strip()
                        deal_type = (row.get("buySell") or row.get("BD_BUY_SELL") or "").upper().strip()
                        # Only BUY-side activity is bullish
                        if "BUY" not in deal_type:
                            continue
                        if not client:
                            continue

                        matched_source = None
                        matched_score = 0
                        if any(name in client for name in SUPERSTAR_NAMES):
                            matched_source = "SUPERSTAR_BUY"
                            matched_score = 90
                        elif any(name in client for name in FII_NAMES):
                            matched_source = "FII_BLOCK_BUY"
                            matched_score = 85
                        elif any(name in client for name in DII_NAMES):
                            matched_source = "DII_BLOCK_BUY"
                            matched_score = 80

                        if matched_source:
                            qty = row.get("quantity") or row.get("BD_QTY_TRD") or "?"
                            price = row.get("tradePrice") or row.get("BD_TP_WATP") or "?"
                            candidates.append({
                                "symbol": sym,
                                "direction": "BULLISH",
                                "score": matched_score,
                                "catalyst": f"{matched_source} via {feed_name}: {client[:60]} bought {qty} @ {price}",
                                "source": matched_source,
                            })
                            log.info(f"[V94_{feed_name}] MATCH {sym} <- {matched_source} ({client[:40]})")
                except Exception as _fe:
                    log.warning(f"[V94_{feed_name}] feed error: {_fe}")

        except Exception as e:
            log.warning(f"[V94] bulk/block scout outer error: {e}")

        if candidates:
            log.info(f"[V94] bulk/block scout: {len(candidates)} institutional buy signals")
        return candidates

    @staticmethod
    def scan_sector_rotation(kite):
        """
        Detect which sectors money is flowing INTO and OUT OF.
        Buy stocks in strengthening sectors, avoid weakening sectors.
        Recovery cycle: Banks → Infra → Auto → Consumer → IT
        """
        candidates = []
        SECTOR_INDICES = {
            "BANKING": "NSE:NIFTY BANK",
            "IT": "NSE:NIFTY IT",
            "PHARMA": "NSE:NIFTY PHARMA",
            "FMCG": "NSE:NIFTY FMCG CONSUMPTION",
            "AUTO": "NSE:NIFTY AUTO",
            "METALS": "NSE:NIFTY METAL",
            "ENERGY": "NSE:NIFTY ENERGY",
            "INFRA": "NSE:NIFTY INFRA",
        }

        SECTOR_STOCKS = {
            "BANKING": ["HDFCBANK", "ICICIBANK", "SBIN", "KOTAKBANK", "AXISBANK"],
            "IT": ["TCS", "INFY", "WIPRO", "HCLTECH", "TECHM"],
            "PHARMA": ["SUNPHARMA", "DRREDDY", "CIPLA", "DIVISLAB"],
            "FMCG": ["HINDUNILVR", "ITC", "NESTLEIND", "BRITANNIA"],
            "AUTO": ["MARUTI", "TATAMOTORS", "M&M", "BAJAJ-AUTO"],
            "METALS": ["TATASTEEL", "JSWSTEEL", "HINDALCO", "VEDL"],
            "ENERGY": ["RELIANCE", "ONGC", "BPCL", "NTPC"],
            "INFRA": ["LT", "ADANIENT", "ADANIPORTS"],
        }

        try:
            # Get sector index performance
            idx_list = list(SECTOR_INDICES.values())
            idx_list.append("NSE:NIFTY 50")  # Benchmark
            quotes = kite.quote(idx_list)

            nifty_q = quotes.get("NSE:NIFTY 50", {})
            nifty_change = 0
            nifty_prev = nifty_q.get("ohlc", {}).get("close", 0)
            nifty_ltp = nifty_q.get("last_price", 0)
            if nifty_prev > 0 and nifty_ltp > 0:
                nifty_change = (nifty_ltp - nifty_prev) / nifty_prev

            strong_sectors = []
            weak_sectors = []

            for sector, idx_sym in SECTOR_INDICES.items():
                q = quotes.get(idx_sym, {})
                prev = q.get("ohlc", {}).get("close", 0)
                ltp = q.get("last_price", 0)
                if prev <= 0 or ltp <= 0:
                    continue

                change = (ltp - prev) / prev
                relative = change - nifty_change  # Performance vs Nifty

                if relative > 0.005:  # Outperforming Nifty by 0.5%+
                    strong_sectors.append((sector, relative))
                elif relative < -0.005:  # Underperforming by 0.5%+
                    weak_sectors.append((sector, relative))

            # Buy stocks in strong sectors
            for sector, rel_perf in sorted(strong_sectors, key=lambda x: x[1], reverse=True)[:3]:
                stocks = SECTOR_STOCKS.get(sector, [])
                for sym in stocks[:2]:  # Top 2 per sector
                    candidates.append({
                        "symbol": sym,
                        "direction": "BULLISH",
                        "score": 65 + int(rel_perf * 1000),  # Higher score for stronger rotation
                        "catalyst": f"SECTOR ROTATION: {sector} outperforming Nifty by {rel_perf:.1%}",
                        "source": "SECTOR_ROTATION",
                    })

        except Exception as e:
            log.debug(f"Sector rotation error: {e}")

        if candidates:
            log.info(f"Sector rotation: {len(candidates)} rotation candidates from {len(strong_sectors)} strong sectors")
        return candidates

    @staticmethod
    def scan_breakout(kite, valid_symbols, avg_vol_map=None, trendlyne=None):
        """FIX_V52: Full universe breakout with pre-market avg volume cache."""
        candidates = []
        try:
            sym_list = list(valid_symbols)
            batch_size = 500

            for i in range(0, len(sym_list), batch_size):
                batch = sym_list[i:i+batch_size]
                try:
                    quotes = kite.quote([f"NSE:{s}" for s in batch])
                    for sym in batch:
                        q = quotes.get(f"NSE:{sym}", {})
                        ltp = q.get("last_price", 0)
                        ohlc = q.get("ohlc", {})
                        high_52 = ohlc.get("high", 0)
                        volume = q.get("volume", 0)
                        avg_volume = avg_vol_map.get(sym, 0) if avg_vol_map else 0

                        if ltp <= 0 or high_52 <= 0:
                            continue
                        pct_from_high = (ltp - high_52) / high_52 if high_52 > 0 else -1
                        if pct_from_high < -0.02:
                            continue
                        if avg_volume > 0 and volume > 0:
                            vol_ratio = volume / avg_volume
                            if vol_ratio < 1.5:
                                continue
                        else:
                            # FIX_V53_AVGVOL_FALLBACK: avg cache not ready (first 18 min of
                            # fresh deploy). Fall back to absolute volume gate so we don't
                            # produce zero candidates during cold start.
                            if volume <= 0 or volume * ltp < 5000000:  # V75: 1cr -> 50L  # Rs 1 Cr absolute turnover
                                continue
                            vol_ratio = 1.5  # neutral assumption for scoring
                        _mcap = trendlyne._data.get(sym, {}).get("market_cap", 0) if trendlyne else 0
                        _min_score = 35 if _mcap > 20000 else (40 if _mcap > 5000 else 45)  # V74: relaxed 60/50/40 -> 45/40/35
                        # V80_CHANGE1: stocks with mcap=0 (not in Trendlyne cache) bypass the score floor.
                        # They still must pass volume/turnover/circuit/panel gates downstream.
                        if _mcap > 0 and trendlyne and trendlyne.score(sym) < _min_score:
                            continue
                        # V80: legacy mcap=0+score=60 check is now dead (V74 made min 45). Kept as no-op.
                        if volume * ltp < 5000000:
                            continue
                        _uc = q.get("upper_circuit_limit", 0)
                        _lc = q.get("lower_circuit_limit", 0)
                        if _uc > 0 and ltp >= _uc * 0.995: continue
                        if _lc > 0 and ltp <= _lc * 1.005: continue

                        score = 75 + int(min(vol_ratio, 5) * 3)
                        candidates.append({
                            "symbol": sym,
                            "direction": "BULLISH",
                            "score": min(score, 90),
                            "catalyst": f"BREAKOUT: {sym} at 52wk high, volume {vol_ratio:.1f}x average",
                            "source": "BREAKOUT",
                        })
                except Exception:
                    pass
                time.sleep(0.5)

        except Exception as e:
            log.debug(f"Breakout scanner error: {e}")

        if candidates:
            log.info(f"Breakout scanner: {len(candidates)} breakout candidates")
        return candidates[:10]

    @staticmethod
    def scan_intraday_momentum(kite, valid_symbols, avg_vol_map=None, trendlyne=None):
        """FIX_V48/V52: Full universe intraday momentum hunter."""
        candidates = []
        try:
            sym_list = list(valid_symbols)
            batch_size = 500
            for i in range(0, len(sym_list), batch_size):
                batch = sym_list[i:i+batch_size]
                try:
                    quotes = kite.quote([f"NSE:{s}" for s in batch])
                    for sym in batch:
                        q = quotes.get(f"NSE:{sym}", {})
                        ltp = q.get("last_price", 0)
                        ohlc = q.get("ohlc", {})
                        prev_close = ohlc.get("close", 0)
                        volume = q.get("volume", 0)
                        if ltp <= 0 or prev_close <= 0: continue
                        pct_change = (ltp - prev_close) / prev_close
                        # V80_CHANGE4: tiered momentum threshold — earlier entry but stricter volume confirmation
                        if pct_change < 0.025: continue
                        _is_strong_move = pct_change >= 0.04
                        avg_volume = avg_vol_map.get(sym, 0) if avg_vol_map else 0
                        # FIX_V53_AVGVOL_FALLBACK: allow strong momentum to enter
                        # even if avg_vol cache not ready (first 18 min after deploy).
                        if avg_volume > 0 and volume > 0:
                            vol_ratio = volume / avg_volume
                            # V80_CHANGE4: weaker move (2.5-4%) requires 2.5x vol; strong move (4%+) keeps 1.8x
                            _min_vol_ratio = 1.8 if _is_strong_move else 2.5
                            if vol_ratio < _min_vol_ratio: continue
                        else:
                            # V80_CHANGE4: weaker move requires Rs 1 Cr turnover; strong move keeps existing logic
                            _min_turnover_momentum = 10000000 if _is_strong_move else 20000000
                            if volume <= 0 or volume * ltp < _min_turnover_momentum:
                                continue
                            vol_ratio = 1.8 if _is_strong_move else 2.5  # neutral assumption for scoring
                        upper_circuit = q.get("upper_circuit_limit", 0)
                        lower_circuit = q.get("lower_circuit_limit", 0)
                        if upper_circuit > 0 and ltp >= upper_circuit * 0.995: continue
                        if lower_circuit > 0 and ltp <= lower_circuit * 1.005: continue
                        _mcap = trendlyne._data.get(sym, {}).get("market_cap", 0) if trendlyne else 0
                        _min_score = 35 if _mcap > 20000 else (40 if _mcap > 5000 else 45)  # V74: relaxed 60/50/40 -> 45/40/35
                        # V80_CHANGE1: stocks with mcap=0 (not in Trendlyne cache) bypass the score floor.
                        # They still must pass volume/turnover/circuit/panel gates downstream.
                        if _mcap > 0 and trendlyne and trendlyne.score(sym) < _min_score:
                            continue
                        # V80: legacy mcap=0+score=60 check is now dead (V74 made min 45). Kept as no-op.
                        # V80_CHANGE3: tier-aware turnover floor (same logic as Falling Knife scanner)
                        _durability = trendlyne._data.get(sym, {}).get("durability", 0) if trendlyne else 0
                        if _mcap == 0:
                            _min_turnover = 10000000
                        elif _durability >= 80:
                            _min_turnover = 2000000
                        else:
                            _min_turnover = 5000000
                        if volume * ltp < _min_turnover:
                            continue
                        score = 70 + int(min(pct_change * 100, 15) * 1.2) + int(min(vol_ratio, 5))
                        candidates.append({"symbol": sym, "direction": "BULLISH", "score": min(score, 90), "catalyst": f"MOMENTUM: {sym} +{pct_change*100:.1f}% today, volume {vol_ratio:.1f}x avg", "source": "MOMENTUM"})
                except Exception: pass
                import time; time.sleep(0.5)
        except Exception as e:
            import logging; logging.getLogger("GIR").debug(f"[FIX_V48] Intraday momentum scanner error: {e}")
        if candidates:
            import logging; logging.getLogger("GIR").info(f"[FIX_V48] Intraday momentum scanner: {len(candidates)} candidates")
        return candidates[:10]
    @staticmethod
    def scan_hidden_gem(kite, valid_symbols, trendlyne, avg_vol_map=None):
        """FIX_V52: Smallcap alpha hunter"""
        candidates = []
        try:
            sym_list = list(valid_symbols)
            for i in range(0, len(sym_list), 500):
                batch = sym_list[i:i+500]
                try:
                    quotes = kite.quote([f"NSE:{s}" for s in batch])
                    for sym in batch:
                        d = trendlyne._data.get(sym, {}) if trendlyne else {}
                        mcap = d.get("market_cap", 0)
                        if not (300 <= mcap <= 200000): continue  # V80_CHANGE2: widened upper to 2L Cr (was 50K Cr); excludes only mega-caps where gem signal is noise
                        if d.get("durability", 0) < 50: continue  # V74: relaxed from 60
                        pc = d.get("promoter_chg_qoq", 0) + d.get("promoter_chg_4q", 0) * 0.5
                        fc = d.get("fii_chg_qoq", 0) + d.get("fii_chg_4q", 0) * 0.5
                        mc = d.get("mf_chg_qoq", 0) + d.get("mf_chg_1m", 0) * 0.5
                        if sum(1 for x in [pc, fc, mc] if x >= 0.1) < 1: continue  # V74: relaxed 0.3/2 -> 0.1/1
                        q = quotes.get(f"NSE:{sym}", {})
                        ltp = q.get("last_price", 0)
                        volume = q.get("volume", 0)
                        ohlc = q.get("ohlc", {})
                        prev_close = ohlc.get("close", 0)
                        if ltp <= 0 or prev_close <= 0 or volume <= 0: continue
                        if volume * ltp < 5000000: continue
                        pct_change = (ltp - prev_close) / prev_close
                        if pct_change < 0.02: continue
                        avg_volume = avg_vol_map.get(sym, 0) if avg_vol_map else 0
                        vol_ratio = (volume / avg_volume) if avg_volume > 0 else 0
                        if vol_ratio < 1.3: continue  # V74: relaxed from 1.5
                        _uc = q.get("upper_circuit_limit", 0)
                        _lc = q.get("lower_circuit_limit", 0)
                        if _uc > 0 and ltp >= _uc * 0.995: continue
                        if _lc > 0 and ltp <= _lc * 1.005: continue
                        score = 75 + int(pc + fc + mc) + int(min(pct_change * 100, 10))
                        candidates.append({
                            "symbol": sym, "direction": "BULLISH",
                            "score": min(score, 95),
                            "catalyst": f"HIDDEN GEM: {sym} mcap={mcap:.0f}Cr +{pct_change*100:.1f}% {vol_ratio:.1f}x",
                            "source": "HIDDEN_GEM",
                        })
                except Exception as _be:
                    import logging; logging.getLogger("GIR").debug(f"[FIX_V52 GEM] {_be}")
                time.sleep(0.15)
        except Exception as _e:
            import logging; logging.getLogger("GIR").debug(f"[FIX_V52 GEM] {_e}")
        if candidates:
            candidates.sort(key=lambda x: x["score"], reverse=True)
            import logging; logging.getLogger("GIR").info(f"[FIX_V52 GEM] {len(candidates)} hidden gems")
        return candidates[:5]




class SafetyFilters:
    _circuit_stocks = set()

    @classmethod
    def check_circuit(cls, kite, symbol):
        try:
            q = kite.quote([f"NSE:{symbol}"]).get(f"NSE:{symbol}", {})
            ltp = q.get("last_price", 0)
            lower = q.get("lower_circuit_limit", 0)
            upper = q.get("upper_circuit_limit", 0)
            if ltp <= 0: return True
            if upper > 0 and ltp >= upper * 0.995:
                log.warning(f"Safety: {symbol} at UPPER circuit")
                return False
            if lower > 0 and ltp <= lower * 1.005:
                log.warning(f"Safety: {symbol} at LOWER circuit")
                cls._circuit_stocks.add(symbol)
                return False
            cls._circuit_stocks.discard(symbol)
            return True
        except Exception: return True  # V28 A8

    @classmethod
    def check_nifty_crash(cls, kite):
        try:
            q = kite.quote(["NSE:NIFTY 50"]).get("NSE:NIFTY 50", {})
            ltp = q.get("last_price", 0)
            prev = q.get("ohlc", {}).get("close", 0)
            if ltp <= 0 or prev <= 0: return True
            if (ltp - prev) / prev < -CRASH_NIFTY_DROP:
                log.warning("Safety: NIFTY CRASH! Blocking ALL entries")
                return False
            return True
        except Exception: return True  # V28 A8

    CORRELATED = {
        "PVT_BANK": {"HDFCBANK","ICICIBANK","KOTAKBANK","AXISBANK","INDUSINDBK","FEDERALBNK","BANDHANBNK"},
        "PSU_BANK": {"SBIN","PNB","BANKBARODA","CANBK","UNIONBANK","IOB"},
        "IT": {"TCS","INFY","WIPRO","HCLTECH","TECHM","LTIM"},
        "PHARMA": {"SUNPHARMA","DRREDDY","CIPLA","DIVISLAB","BIOCON","AUROPHARMA","LUPIN"},
        "OIL": {"RELIANCE","ONGC","BPCL","IOC","GAIL","HINDPETRO"},
        "METALS": {"TATASTEEL","JSWSTEEL","HINDALCO","VEDL","NMDC","COALINDIA"},
        "ADANI": {"ADANIENT","ADANIPORTS","ADANIGREEN","ADANIPOWER","ATGL"},
        "TATA": {"TCS","TATAMOTORS","TATASTEEL","TATAPOWER","TATACONSUM","TITAN"},
        "AUTO": {"MARUTI","TATAMOTORS","M&M","BAJAJ-AUTO","HEROMOTOCO","ASHOKLEY"},
        "FMCG": {"HINDUNILVR","ITC","NESTLEIND","BRITANNIA","DABUR","MARICO"},
        "INFRA": {"LT","ADANIENT","ADANIPORTS","NCC","NBCC"},
        "INSURANCE": {"HDFCLIFE","SBILIFE","ICICIPRULI","MAXHEALTH","STARHEALTH"},
    }

    @classmethod
    def check_correlation(cls, symbol, positions):
        existing = set(positions.keys()) if isinstance(positions, dict) else set()
        for grp, stocks in cls.CORRELATED.items():
            if symbol in stocks:
                overlap = existing & stocks
                if len(overlap) >= 2:
                    log.info(f"Safety: {symbol} blocked — {len(overlap)} in {grp}")
                    return False
        return True

    @classmethod
    def is_bad_stock(cls, kite, symbol):
        if symbol in cls._circuit_stocks or symbol in BLACKLIST:
            return True
        try:
            q = kite.quote([f"NSE:{symbol}"]).get(f"NSE:{symbol}", {})
            if 0 < q.get("last_price", 0) < 10: return True
            if q.get("volume", 0) == 0: return True
            return False
        except Exception: return False  # V28 A8


class OrderGuard:
    """
    Prevents duplicate orders and enforces SEBI compliance.
    - No duplicate BUY for same symbol within 30 minutes
    - No duplicate F&O contract (same tradingsymbol)
    - SEBI rate limit: max 10 orders/second, max 200 orders/day
    - Market hours enforcement
    """
    _recent_orders = {}  # symbol -> timestamp of last order
    _daily_count = 0
    _daily_date = None
    _second_count = 0
    _second_ts = 0

    @classmethod
    def _reset_daily(cls):
        today = today_ist()
        if cls._daily_date != today:
            cls._daily_count = 0
            cls._daily_date = today

    @classmethod
    def can_place_order(cls, symbol, is_fno=False):
        """
        Check all order guards before placing.
        Returns (True, "OK") or (False, "reason").
        """
        cls._reset_daily()
        now = time.time()

        # SEBI: max 200 orders/day
        if cls._daily_count >= 200:
            return False, "SEBI_DAILY_LIMIT_200"

        # SEBI: max 10 orders/second
        # FIX_V53_RATE: sliding 1-sec window. Prior logic never reset _second_ts
        # mid-stream, so continuous orders at >1/sec would stay BLOCKED forever
        # even when actual rate fell below 10/sec.
        if now - cls._second_ts >= 1:
            # Window expired — start fresh
            cls._second_ts = now
            cls._second_count = 1
        else:
            cls._second_count += 1
            if cls._second_count > 10:
                return False, "SEBI_RATE_LIMIT_10/sec"

        # Market hours check (skip for AMO)
        if not is_market_hours() and not is_amo_window():
            return False, "MARKET_CLOSED"

        # Duplicate check: same symbol within 30 minutes
        last_time = cls._recent_orders.get(symbol, 0)
        if now - last_time < 1800:  # 30 minutes
            mins_ago = int((now - last_time) / 60)
            return False, f"DUPLICATE_ORDER_{mins_ago}min_ago"

        return True, "OK"

    @classmethod
    def record_order(cls, symbol):
        """Record that an order was placed for this symbol."""
        cls._recent_orders[symbol] = time.time()
        cls._daily_count += 1
        # Clean old entries (older than 2 hours)
        now = time.time()
        cls._recent_orders = {s: t for s, t in cls._recent_orders.items() if now - t < 7200}

    @classmethod
    def check_pending_orders(cls, kite, symbol, is_fno=False):
        """
        Check if there's already a pending BUY order for this symbol on Kite.
        V28 A7: Exact match for equity; F&O matches contracts starting with underlying.
        Equity path no longer false-blocks when an F&O contract of same underlying is pending.
        """
        try:
            orders = kite.orders()
            for o in orders:
                ts = o.get("tradingsymbol", "")
                exch = o.get("exchange", "")
                if o.get("transaction_type") != "BUY":
                    continue
                if o.get("status") not in ("OPEN", "TRIGGER PENDING", "AMO REQ RECEIVED"):
                    continue
                # V28 A7: correct matching per exchange
                if is_fno:
                    # F&O contracts live on NFO and are like SBIN26APR820CE
                    if exch == "NFO" and ts.startswith(symbol):
                        log.info(f"OrderGuard: {symbol} already has pending F&O BUY {o.get('order_id')} ({ts})")
                        return True
                else:
                    # Equity: exact match on NSE only
                    if exch == "NSE" and ts == symbol:
                        log.info(f"OrderGuard: {symbol} already has pending equity BUY {o.get('order_id')}")
                        return True
            return False
        except Exception:
            return False


# FIX_V46: Expert Panel
class LossHistoryTracker:
    def __init__(self):
        self.file = Path("/home/globalbot/data/loss_history.json")
        self.data = self._load()
    def _load(self):
        try:
            if self.file.exists():
                with open(self.file, "r") as f:
                    return json.load(f)
        except Exception as e:
            log.warning(f"[FIX_V46_LOSS] load failed: {e}")
        return {}
    def _save(self):
        try:
            self.file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.file, "w") as f:
                json.dump(self.data, f, indent=2)
        except Exception as e:
            log.error(f"[FIX_V46_LOSS] save failed: {e}")
    def record_loss(self, symbol, pnl_pct=None):
        today = today_ist().isoformat()
        if symbol not in self.data:
            self.data[symbol] = []
        self.data[symbol].append({"date": today, "pnl_pct": pnl_pct or 0})
        cutoff = today_ist() - timedelta(days=30)
        self.data[symbol] = [e for e in self.data[symbol] if datetime.fromisoformat(e["date"]).date() >= cutoff]
        self._save()
    def days_since_last_loss(self, symbol):
        entries = self.data.get(symbol, [])
        if not entries:
            return None
        try:
            last = max(datetime.fromisoformat(e["date"]).date() for e in entries)
            return (today_ist() - last).days
        except Exception:
            return None
    def count_recent_losses(self, symbol, days=7):
        entries = self.data.get(symbol, [])
        if not entries:
            return 0
        cutoff = today_ist() - timedelta(days=days)
        return sum(1 for e in entries if datetime.fromisoformat(e["date"]).date() >= cutoff)

_LOSS_TRACKER = None


# ─── V83_EXIT_COOLDOWN_24H: tracks ANY exit (loss/be/win/external) per symbol ───
# Distinct from LossHistoryTracker (which only records losses with day granularity).
# This tracker uses ISO timestamps (minute precision) for the 24-hour hard cooldown.
class ExitHistoryTracker:
    def __init__(self):
        self.file = Path("/home/globalbot/data/exit_history.json")
        self.data = self._load()
    def _load(self):
        try:
            if self.file.exists():
                with open(self.file, "r") as f:
                    return json.load(f)
        except Exception as e:
            log.warning(f"[V83_EXIT_TRACKER] load failed: {e}")
        return {}
    def _save(self):
        try:
            self.file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.file, "w") as f:
                json.dump(self.data, f, indent=2)
        except Exception as e:
            log.error(f"[V83_EXIT_TRACKER] save failed: {e}")
    def record_exit(self, symbol, reason="UNKNOWN"):
        ts = now_ist().isoformat()
        if symbol not in self.data:
            self.data[symbol] = []
        self.data[symbol].append({"ts": ts, "reason": reason})
        cutoff = now_ist() - timedelta(days=7)
        self.data[symbol] = [
            e for e in self.data[symbol]
            if datetime.fromisoformat(e["ts"]) >= cutoff
        ]
        self._save()
    def hours_since_last_exit(self, symbol):
        entries = self.data.get(symbol, [])
        if not entries:
            return None
        try:
            last = max(datetime.fromisoformat(e["ts"]) for e in entries)
            delta = now_ist() - last
            return delta.total_seconds() / 3600.0
        except Exception:
            return None

_EXIT_TRACKER = None
def get_exit_tracker():
    global _EXIT_TRACKER
    if _EXIT_TRACKER is None:
        _EXIT_TRACKER = ExitHistoryTracker()
    return _EXIT_TRACKER

def get_loss_tracker():
    global _LOSS_TRACKER
    if _LOSS_TRACKER is None:
        _LOSS_TRACKER = LossHistoryTracker()
    return _LOSS_TRACKER


class ExpertPanel:
    # FIX_V58_PANEL_EVERYWHERE + FIX_V59_THRESHOLD_TUNE: calibrated to 60% quality bar
    # EQ 60/100 = 60%, FNO 72/120 = 60%. Same standard for both modules.
    # V58 (EQ 55, FNO 70) was too loose — let mediocre scores (RELIANCE 51, SUNPHARMA 51) pass.
    # V59 60/72 blocks mediocre, passes ONGC(60)-class or better.
    EQUITY_THRESHOLD = 55  # V74: relaxed from 60
    FNO_THRESHOLD = 65  # V74: relaxed from 72
    def __init__(self, trendlyne, sector_rotation, news_brain):
        self.trendlyne = trendlyne
        self.sector_rotation = sector_rotation
        self.news_brain = news_brain
        self.loss_tracker = get_loss_tracker()
    def score_news(self, symbol, direction, scout_data):
        try:
            if scout_data.get("newsbrain_rejected"): return (0, True)
            _src = (scout_data.get("source") or "").upper()
            if "INDEX" in _src:
                _s = scout_data.get("score", 60)
                return (20 if _s>=75 else 15 if _s>=65 else 10, False)
            base = 10
            if scout_data.get("validated"):
                base = 20
                c = int(scout_data.get("score", 0))
                if self.news_brain and hasattr(self.news_brain, "is_validated"):
                    try:
                        f, nc = self.news_brain.is_validated(symbol, direction)
                        if f and nc > c: c = nc
                    except Exception: pass  # V77_FIX1
                if c >= 85: base += 25
                elif c >= 70: base += 15
                cat = (scout_data.get("catalyst") or "").upper()
                if any(k in cat for k in ("EARNINGS","RESULTS","QUARTER","PROFIT","REVENUE","BEAT","MISS","EBITDA")):
                    if any(k in cat for k in ("BEAT","ABOVE ESTIMATE","STRONG RESULT","RECORD PROFIT")) and direction == "BULLISH": base += 15
                    elif any(k in cat for k in ("MISS","BELOW ESTIMATE","WEAK RESULT","LOSS WIDEN")) and direction == "BEARISH": base += 20
            return (min(100, base), False)
        except Exception: return (10, False)  # V77_FIX1

    def score_fundamental(self, symbol):
        try:
            dvm = self.trendlyne.score(symbol)
            return max(0, min(20, int(dvm / 5)))
        except Exception:
            return 10
    def score_technical(self, kite, symbol, direction):
        if not kite:
            return 10
        try:
            tok = 0
            try:
                q = kite.ltp(f"NSE:{symbol}")
                tok = q.get(f"NSE:{symbol}", {}).get("instrument_token", 0)
            except Exception:
                return 10
            if not tok:
                return 10
            hist = kite.historical_data(instrument_token=tok, from_date=(now_ist() - timedelta(days=80)).date(), to_date=today_ist(), interval="day")
            if not hist or len(hist) < 50:
                return 10
            closes = [h["close"] for h in hist]
            volumes = [h["volume"] for h in hist]
            ltp = closes[-1]
            sma50 = sum(closes[-50:]) / 50
            avg_vol = sum(volumes[-20:-1]) / 19 if len(volumes) >= 21 else 0
            today_vol = volumes[-1]
            gains = sum(max(0, closes[i] - closes[i-1]) for i in range(-14, 0))
            losses = sum(max(0, closes[i-1] - closes[i]) for i in range(-14, 0))
            rs = gains / losses if losses > 0 else 100
            rsi = 100 - (100 / (1 + rs))
            score = 0
            if direction == "BULLISH" and ltp > sma50:
                score += 7
            elif direction == "BEARISH" and ltp < sma50:
                score += 7
            if direction == "BULLISH" and 40 <= rsi <= 70:
                score += 7
            elif direction == "BEARISH" and 30 <= rsi <= 60:
                score += 7
            if avg_vol > 0 and today_vol >= 1.5 * avg_vol:
                score += 6
            return score
        except Exception:
            return 10
    def score_sector(self, symbol, direction):
        if not self.sector_rotation:
            return 10
        try:
            bias, _sec, _pct = self.sector_rotation.get_sector_bias(symbol)
            if bias is None:
                return 10
            if direction == "BULLISH":
                return {"TOP": 20, "MIDDLE": 10, "BOTTOM": 0}.get(bias, 10)
            else:
                return {"BOTTOM": 20, "MIDDLE": 10, "TOP": 0}.get(bias, 10)
        except Exception:
            return 10
    def score_history(self, symbol):
        try:
            if self.loss_tracker.count_recent_losses(symbol, days=7) >= 2:
                return 0
            days = self.loss_tracker.days_since_last_loss(symbol)
            if days is None:
                return 15
            if days >= 7:
                return 20
            if days >= 3:
                return 10
            return 0
        except Exception:
            return 15
    def score_option(self, option_data):
        score = 0
        try:
            dte = option_data.get("dte", 0)
            if 14 <= dte <= 40:
                score += 7
            delta = abs(option_data.get("delta", 0.45))
            if 0.30 <= delta <= 0.60:
                score += 7
            iv = option_data.get("iv", 30)
            if iv < 60:
                score += 6
            elif iv < 80:
                score += 3
            return score
        except Exception:
            return 10
    def evaluate_equity(self, kite, symbol, direction, scout_data):
        e1, veto = self.score_news(symbol, direction, scout_data)
        e2 = self.score_fundamental(symbol)
        e3 = self.score_technical(kite, symbol, direction)
        e4 = self.score_sector(symbol, direction)
        e5 = self.score_history(symbol)
        total = e1 + e2 + e3 + e4 + e5
        result = {"symbol": symbol, "direction": direction, "news": e1, "fundamental": e2, "technical": e3, "sector": e4, "history": e5, "total": total, "threshold": self.EQUITY_THRESHOLD, "veto": veto, "allow": (not veto) and (total >= self.EQUITY_THRESHOLD)}
        # V79: replay-harness decision tap (separate from V63 legacy log)
        try:
            if _TRADE_RECORDER_AVAILABLE:
                TradeRecorder.record_decision("EQUITY", result, scout_data, extras={
                    "components": {"news": e1, "fundamental": e2, "technical": e3, "sector": e4, "history": e5},
                    "veto": veto,
                })
        except Exception:
            pass
        # PATCH_V63_PANEL_TRACKER: persist every equity panel decision for Friday review
        try:
            self._record_panel_decision("EQUITY", result, scout_data)
        except Exception:
            pass
        return result
    def evaluate_fno(self, kite, symbol, direction, scout_data, option_data):
        e1, veto = self.score_news(symbol, direction, scout_data)
        e2 = 10
        try:
            cat = scout_data.get("catalyst", "")
            src = scout_data.get("source", "")
            if "OI" in src.upper() or "CALL_WRITING" in cat or "PUT_WRITING" in cat:
                e2 = 20
        except Exception:
            pass
        e3 = self.score_technical(kite, symbol, direction)
        e4 = self.score_sector(symbol, direction)
        e5 = self.score_history(symbol)
        e6 = self.score_option(option_data or {})
        total = e1 + e2 + e3 + e4 + e5 + e6
        result = {"symbol": symbol, "direction": direction, "news": e1, "oi": e2, "technical": e3, "sector": e4, "history": e5, "option": e6, "total": total, "threshold": self.FNO_THRESHOLD, "veto": veto, "allow": (not veto) and (total >= self.FNO_THRESHOLD)}
        # V79: replay-harness decision tap (separate from V63 legacy log)
        try:
            if _TRADE_RECORDER_AVAILABLE:
                TradeRecorder.record_decision("FNO", result, scout_data, extras={
                    "components": {"news": e1, "oi": e2, "technical": e3, "sector": e4, "history": e5, "option": e6},
                    "veto": veto,
                    "option_data": option_data or {},
                })
        except Exception:
            pass
        # PATCH_V63_PANEL_TRACKER: persist every F&O panel decision for Friday review
        try:
            self._record_panel_decision("FNO", result, scout_data)
        except Exception:
            pass
        return result

    # ═══════════════════════════════════════════════════════════════════════
    # PATCH_V63_PANEL_TRACKER — daily rollup of every panel decision
    # File: panel_decisions_YYYYMMDD.jsonl  (one JSON per line, easy grep)
    # Use:  cat data/panel_decisions_$(date +%Y%m%d).jsonl | python3 -c "
    #          import sys, json
    #          rows = [json.loads(l) for l in sys.stdin]
    #          print(f'Total: {len(rows)}')
    #          print(f'TRADE: {sum(1 for r in rows if r[\"allow\"])}')
    #          print(f'BLOCK: {sum(1 for r in rows if not r[\"allow\"])}')
    #          for r in sorted(rows, key=lambda x: -x['total'])[:10]:
    #              print(r)
    #       "
    # ═══════════════════════════════════════════════════════════════════════
    def _record_panel_decision(self, kind, result, scout_data):
        # PATCH_V70_AUDIT_CLEANUP: removed unused 'from pathlib import Path'
        import json as _json
        try:
            today = today_ist().strftime("%Y%m%d")
            path = DATA_DIR / f"panel_decisions_{today}.jsonl"
            row = {
                "ts": now_ist().isoformat(),
                "kind": kind,  # EQUITY or FNO
                "symbol": result.get("symbol"),
                "direction": result.get("direction"),
                "total": result.get("total"),
                "threshold": result.get("threshold"),
                "allow": result.get("allow"),
                "veto": result.get("veto"),
                "news": result.get("news"),
                "fundamental": result.get("fundamental"),
                "oi": result.get("oi"),
                "technical": result.get("technical"),
                "sector": result.get("sector"),
                "history": result.get("history"),
                "option": result.get("option"),
                "catalyst": (scout_data or {}).get("catalyst", "")[:200],
                "source": (scout_data or {}).get("source", ""),
                "scout_score": (scout_data or {}).get("score"),
            }
            with open(path, "a") as f:
                f.write(_json.dumps(row) + "\n")
        except Exception:
            # NEVER let logging break the panel
            pass

    @staticmethod
    def format_equity_log(r):
        v = " VETO" if r["veto"] else ""
        d = "TRADE" if r["allow"] else "BLOCK"
        return f"[FIX_V46_PANEL_EQ] {r['symbol']} {r['direction']}: news={r['news']} fund={r['fundamental']} tech={r['technical']} sector={r['sector']} hist={r['history']} -> TOTAL={r['total']}/100 thr={r['threshold']} -> {d}{v}"
    @staticmethod
    def format_fno_log(r):
        v = " VETO" if r["veto"] else ""
        d = "TRADE" if r["allow"] else "BLOCK"
        return f"[FIX_V46_PANEL_FNO] {r['symbol']} {r['direction']}: news={r['news']} oi={r['oi']} tech={r['technical']} sector={r['sector']} hist={r['history']} opt={r['option']} -> TOTAL={r['total']}/120 thr={r['threshold']} -> {d}{v}"


# ═══════════════════════════════════════════════════════════════════════════════
#  EQUITY MODULE — Completely independent. No F&O awareness.
# ═══════════════════════════════════════════════════════════════════════════════

class EquityModule:
    """
    CNC delivery equity trading.

    Entry: News/filing catalyst + RSI not overbought + Volume above avg
    Exit:  Trailing SL (2.5% below LTP, only moves UP) + Dead money (15 days)
    Protection: GTT on every position, checked/modified every 5 min

    Own news scanner, own GTT manager, own risk manager, own capital.
    Zero awareness of F&O module.
    """

    def __init__(self):
        self.positions = {}     # symbol -> position dict
        self.gtt = GTTManager("EQUITY")
        self.risk = None        # Initialized after capital is known
        self.telegram = TelegramThrottle("EQUITY", max_per_hour=15)
        self.news = None        # Initialized after symbols are loaded
        self.filings = None     # FilingMonitor
        self._sector_rotation = None  # FIX_V49b: needed by ExpertPanel
        self.brokerage = None   # BrokerageMonitor
        self.macro = None       # MacroDetector
        self.trendlyne = TrendlyneScorer()
        self.amo = AMOEngine("EQUITY")
        self.news_brain = None  # FIX_V40_EQUITY_LLM
        self.capital = 0
        self.available = 0
        self.all_symbols = set()
        self.fno_stocks = set()  # Loaded from Kite
        self._pos_file = DATA_DIR / "equity_positions.json"
        self._trades_file = DATA_DIR / "equity_trades.json"
        self._last_scan = None
        self._last_trail_check = None

    def init(self):
        """Initialize after Kite login."""
        kite = KiteSession.kite()
        if not kite:
            log.error("Equity: cannot init without Kite")
            return False

        # Load all NSE symbols
        try:
            instruments = kite.instruments("NSE")
            self.all_symbols = set()
            for inst in instruments:
                sym = inst.get("tradingsymbol", "")
                if is_valid_symbol(sym) and inst.get("instrument_type") == "EQ":
                    price = inst.get("last_price", 0)
                    # Skip penny stocks (< Rs.5)
                    if price and price < 5:
                        continue
                    self.all_symbols.add(sym)
            log.info(f"Equity: {len(self.all_symbols)} valid stocks loaded")
            # FIX_V52: Initialize avg volume loader AFTER all_symbols is populated
            try:
                self._avg_vol_loader = AvgVolumeLoader()
                self._avg_vol_loader.load_from_disk()
                if not self._avg_vol_loader.is_fresh():
                    self._avg_vol_loader.refresh_async(kite, self.all_symbols)
            except Exception as _ae:
                log.warning(f"[FIX_V52 AVGVOL] startup kick failed: {_ae}")
                self._avg_vol_loader = AvgVolumeLoader()  # empty fallback
            # FIX_V52: Rank universe by market cap for tier-aware filtering
            try:
                self.all_symbols_ranked = sorted(
                    self.all_symbols,
                    key=lambda s: self.trendlyne._data.get(s, {}).get("market_cap", 0),
                    reverse=True
                )
                log.info(f"[FIX_V52] Ranked {len(self.all_symbols_ranked)} stocks by market cap")
            except Exception as _se:
                log.warning(f"[FIX_V52] market cap ranking failed: {_se}")
                self.all_symbols_ranked = list(self.all_symbols)
        except Exception as e:
            log.error(f"Equity: instrument load failed: {e}")
            return False

        # Get capital
        self._refresh_capital()

        # Init risk manager
        self.risk = RiskManager("EQUITY", self.capital)

        # PATCH_V28_SYMBOL_REGEX: build company name -> symbol map from instruments
        self._name_map = {}
        try:
            for inst in instruments:
                _n = inst.get("name", "").upper().strip()
                _s = inst.get("tradingsymbol", "")
                if _n and _s and inst.get("instrument_type") == "EQ":
                    self._name_map[_n] = _s
                    # Also add first 2 words as key (e.g. "TATA CONSULTANCY" -> TCS)
                    _words = _n.split()
                    if len(_words) >= 2:
                        self._name_map[" ".join(_words[:2])] = _s
            log.info(f"[PATCH_V28_SYMBOL_REGEX] built name_map with {len(self._name_map)} entries")
        except Exception as _nme:
            log.warning(f"[PATCH_V28_SYMBOL_REGEX] name_map build failed: {_nme}")

        # FIX_V48: Build Trendlyne Stock Name -> NSE Code map
        self._trendlyne_name_map = {}
        try:
            _tpath = DATA_DIR / "trendlyne_data.xlsx"
            if _tpath.exists():
                import openpyxl as _opx
                _wb = _opx.load_workbook(_tpath, read_only=True)
                _ws = _wb.active
                _rows = list(_ws.iter_rows(values_only=True))
                if _rows:
                    _hdrs = [str(h).strip() if h else "" for h in _rows[0]]
                    _nc = _cc = None
                    for _ci, _h in enumerate(_hdrs):
                        if _h == "Stock Name": _nc = _ci
                        elif _h in ("NSE Code", "Stock Code"): _cc = _ci
                    if _nc is not None and _cc is not None:
                        for _r in _rows[1:]:
                            if len(_r) > max(_nc, _cc):
                                _sn = str(_r[_nc] or "").strip().upper()
                                _sc = str(_r[_cc] or "").strip().upper()
                                if _sn and _sc and len(_sn) >= 3:
                                    self._trendlyne_name_map[_sn] = _sc
                _wb.close()
                log.info(f"[FIX_V48] Trendlyne name_map: {len(self._trendlyne_name_map)} entries")
        except Exception as _te:
            log.warning(f"[FIX_V48] Trendlyne name_map build failed: {_te}")
        # Init news scanner with equity symbols + name mapping
        self.news = NewsScanner("EQUITY", self.all_symbols, name_map=self._name_map, trendlyne_name_map=self._trendlyne_name_map)

        # Init filing monitor
        self.filings = FilingMonitor("EQUITY", self.all_symbols)

        # Init brokerage monitor
        self.brokerage = BrokerageMonitor("EQUITY", self.all_symbols)

        # Init macro detector
        self.macro = MacroDetector("EQUITY", self.all_symbols, trendlyne_scorer=self.trendlyne)  # FIX_V4_DYNAMIC_MACRO

        # FIX_V40_EQUITY_LLM: independent NewsBrain for equity
        self.news_brain = None
        if USE_LLM_NEWS and _NEWS_BRAIN_AVAILABLE and NewsBrain:
            try:
                try:
                    # FIX_NB_G2: relaxed filters — conv 70->60, cross 2->1, age 24h->72h
                    self.news_brain = NewsBrain(
                        universe_stocks=self.all_symbols,
                        news_feeds=NEWS_FEEDS,
                        module_name="EQUITY",
                        direction_mode="BULLISH_ONLY",
                        min_conviction=60,
                        require_cross_sources=1,
                        require_sentence_citation=True,
                        max_article_age_hours=72,
                    )
                    log.info("EQUITY: NewsBrain v4 [FIX_NB_G2] - BULLISH_ONLY, min_conv=60, cross>=1, cite=yes, max_age=72h")
                except TypeError:
                    self.news_brain = NewsBrain(fno_stocks=self.all_symbols)
                    log.warning("EQUITY: NewsBrain v3 fallback - safety layers NOT active")
            except Exception as _nbe:
                log.error(f"EQUITY: NewsBrain init failed: {_nbe}")
        else:
            log.info(f"EQUITY: NewsBrain disabled (USE_LLM_NEWS={USE_LLM_NEWS})")

        # PATCH_V28_KITE_ONLY_EQUITY: DO NOT load from JSON. Kite is the only source of truth.
        self.positions = {}

        # FIX_V56_TICK_SIZE: load tick-size cache BEFORE any GTT placement
        # Prevents TRENT-class (Apr 23 2026) rejection where hardcoded 0.05 used for 0.10-tick stock
        try:
            _loaded = TickSizeCache.load()
            log.info(f"[FIX_V56_TICK_SIZE] tick cache loaded: {_loaded} entries")
        except Exception as _tse:
            log.error(f"[FIX_V56_TICK_SIZE] cache load failed: {_tse} (using fallback buckets)")

        # FIX_V55_AMO_RECOVERY: track pending AMO BUY orders across restarts
        self.pending_amo_fills = {}

        # Sync with Kite holdings — this is now the ONLY source of positions
        self._sync_holdings()

        # FIX_V55_AMO_RECOVERY: discover pending AMO BUYs from dead previous process
        # MUST run BEFORE _gtt_audit so any fills already happened during restart
        # get tracked, placed in positions, and GTT-protected.
        self._recover_pending_amo_orders()

        # Run GTT audit
        self._gtt_audit()

        log.info(f"Equity: init complete. Capital=Rs.{self.capital:,.0f} Positions={len(self.positions)}")
        self.telegram.send(
            f"Equity Module started\n"
            f"Capital: Rs.{self.capital:,.0f}\n"
            f"Positions: {len(self.positions)}\n"
            f"Stocks scanned: {len(self.all_symbols)}",
            force=True,
        )
        return True

    def _refresh_capital(self):
        """Get real capital from Kite.
        PATCH_V10_CAPITAL_RETRY: retry up to 3 times with 1s delay if Kite returns 0.
        """
        kite = KiteSession.kite()
        if not kite:
            return
        import time as _t
        total = 0
        for _attempt in range(1, 4):
            try:
                margins = kite.margins("equity")
                _t1 = margins.get("net", 0) or margins.get("available", {}).get("live_balance", 0)
                if not _t1:
                    _t1 = margins.get("available", {}).get("cash", 0)
                if _t1 and _t1 > 0:
                    total = _t1
                    break
                log.warning(f"[PATCH_V10_CAPITAL_RETRY] Equity attempt {_attempt}/3: Kite returned 0, retrying")
            except Exception as _e:
                log.warning(f"[PATCH_V10_CAPITAL_RETRY] Equity attempt {_attempt}/3: {type(_e).__name__}: {_e}")
            if _attempt < 3:
                _t.sleep(1)
        if total > 0:
            self.capital = round(total * EQUITY_PCT, 2)
            invested = sum(
                p.get("qty", 0) * p.get("entry_price", 0)
                for p in self.positions.values()
            )
            self.available = max(0, self.capital - invested)
            log.info(f"[PATCH_V10_CAPITAL_RETRY] Equity refreshed: total={total} capital={self.capital} available={self.available}")
        else:
            log.warning(f"[PATCH_V10_CAPITAL_RETRY] Equity: all 3 attempts returned 0, preserving previous capital={self.capital}")


    def _save_positions(self):
        """Save positions to file."""
        save_json(self._pos_file, {"positions": self.positions})

    def _sync_holdings(self):
        """
        Sync with Kite holdings. Every stock in Kite = bot owned.
        No adoption concept — just sync.
        """
        kite = KiteSession.kite()
        if not kite:
            return

        try:
            holdings = kite.holdings()
            kite_syms = {}
            for h in holdings:
                sym = h.get("tradingsymbol", "")
                qty = h.get("quantity", 0) + h.get("t1_quantity", 0)
                avg = h.get("average_price", 0)
                if qty > 0 and sym and sym not in BLACKLIST:
                    kite_syms[sym] = {"qty": qty, "avg": avg}
            
            # Also check positions (day trades not yet settled)
            positions = kite.positions().get("net", [])
            for p in positions:
                sym = p.get("tradingsymbol", "")
                qty = p.get("quantity", 0)
                avg = p.get("average_price", 0)
                exch = p.get("exchange", "")
                if qty > 0 and sym and sym not in BLACKLIST and exch == "NSE":
                    if sym not in kite_syms:
                        kite_syms[sym] = {"qty": qty, "avg": avg}

            # Add Kite holdings not in our positions
            for sym, info in kite_syms.items():
                if sym not in self.positions:
                    # PATCH_V1_RESTART_PEAK
                    _live_sl = _snap_sl_tick_up(info["avg"] * (1 - EQ_INITIAL_SL_PCT), symbol=sym, exchange="NSE")
                    _live_peak = info["avg"]
                    try:
                        _existing_gtts = kite.get_gtts() or []
                        for _g in _existing_gtts:
                            if _g.get("status") != "active":
                                continue
                            _cond = _g.get("condition", {})
                            if _cond.get("tradingsymbol") == sym and _cond.get("exchange") == "NSE":
                                _triggers = _cond.get("trigger_values", [])
                                if _triggers and _triggers[0] > 0:
                                    _live_sl = max(_live_sl, float(_triggers[0]))
                                    log.info(f"Equity sync: {sym} using live GTT SL={_live_sl}")
                                    break
                    except Exception as _e:
                        log.warning(f"Equity sync: GTT lookup failed for {sym}: {_e}")
                    try:
                        _ltp_data = kite.ltp([f"NSE:{sym}"])
                        _cur_ltp = _ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                        if _cur_ltp > 0:
                            _live_peak = max(_live_peak, _cur_ltp)
                    except Exception as _e:
                        log.warning(f"Equity sync: LTP lookup failed for {sym}: {_e}")
                    # END_PATCH_V1_RESTART_PEAK
                    self.positions[sym] = {
                        "symbol": sym,
                        "qty": info["qty"],
                        "entry_price": info["avg"],
                        "entry_time": now_ist().isoformat(),
                        "sl": _live_sl,
                        "target": 0,
                        "peak_price": _live_peak,
                        "source": "KITE_SYNC",
                    }
                    log.info(f"Equity: synced {sym} qty={info['qty']} avg={info['avg']} sl={_live_sl} peak={_live_peak}")

            # V27 PATCH (Bug #6): Remove positions not in Kite, fetch REAL exit price
            # from tradebook (was recording 0, polluting all P&L reports).
            for sym in list(self.positions.keys()):
                if sym not in kite_syms:
                    pos = self.positions[sym]
                    _exit_price = 0
                    _reason = "SOLD_EXTERNAL"
                    try:
                        _trades = kite.trades() or []
                        _sym_sells = []
                        for _t in _trades:
                            if (_t.get("tradingsymbol") == sym
                                and _t.get("exchange") == "NSE"
                                and _t.get("transaction_type") == "SELL"):
                                _sym_sells.append(_t)
                        if _sym_sells:
                            _total_qty = sum(_t.get("quantity", 0) for _t in _sym_sells)
                            _total_val = sum(_t.get("quantity", 0) * _t.get("average_price", 0) for _t in _sym_sells)
                            if _total_qty > 0:
                                _exit_price = round(_total_val / _total_qty, 2)
                                _reason = "MANUAL_SELL"
                                log.info(f"Equity: {sym} external sell detected - exit_price={_exit_price} from tradebook")
                    except Exception as _te:
                        log.warning(f"Equity: tradebook lookup for {sym} failed: {_te}, recording exit=0")
                    self._record_trade(sym, pos, _exit_price, _reason)
                    del self.positions[sym]
                    log.info(f"Equity: removed {sym} (exit_price={_exit_price} reason={_reason})")

            self._save_positions()

        except Exception as e:
            log.error(f"Equity: sync holdings failed: {e}")

    def _estimate_atr(self, symbol):
        """FIX_V55 helper: quick ATR estimate from last 14 daily bars via Kite historical.
        Returns 0 if unavailable. Non-fatal — callers handle gracefully."""
        kite = KiteSession.kite()
        if not kite:
            return 0
        try:
            # Find instrument token
            instr = kite.instruments("NSE") or []
            token = None
            for i in instr:
                if i.get("tradingsymbol") == symbol:
                    token = i.get("instrument_token")
                    break
            if not token:
                return 0
            from datetime import timedelta
            to_date = now_ist().date()
            from_date = to_date - timedelta(days=25)
            bars = kite.historical_data(token, from_date, to_date, "day") or []
            if len(bars) < 14:
                return 0
            tr_list = []
            for i in range(1, len(bars)):
                h = bars[i]["high"]
                l = bars[i]["low"]
                c_prev = bars[i-1]["close"]
                tr = max(h - l, abs(h - c_prev), abs(l - c_prev))
                tr_list.append(tr)
            if not tr_list:
                return 0
            return sum(tr_list[-14:]) / min(14, len(tr_list))
        except Exception:
            return 0

    def _recover_pending_amo_orders(self):
        """FIX_V55_AMO_RECOVERY: Scan Kite for pending AMO BUY orders at startup / tick.
        
        On restart, bot may miss AMO orders placed by previous process that haven't
        filled yet. This method:
        1. Reads kite.orders() for AMO BUYs still pending (OPEN/TRIGGER_PENDING/AMO_RECEIVED)
        2. Registers them in self.pending_amo_fills[order_id] = {symbol, qty, price}
        3. On next tick, if order_id now shows COMPLETE -> place GTT immediately
        4. Drops tracking if REJECTED/CANCELLED
        
        Called from:
        - init() at startup (recovers orders from dead previous process)
        - tick() every minute during market hours 09:00-09:30 (catches fills at open)
        """
        kite = KiteSession.kite()
        if not kite:
            return
        
        try:
            orders = kite.orders() or []
        except Exception as e:
            log.warning(f"[FIX_V55_AMO_RECOVERY] kite.orders() failed: {e}")
            return
        
        # Build current status map: {order_id: (status, tradingsymbol, qty, avg_price, transaction)}
        live_orders = {}
        for o in orders:
            oid = str(o.get("order_id", ""))
            if not oid:
                continue
            live_orders[oid] = {
                "status": o.get("status", ""),
                "symbol": o.get("tradingsymbol", ""),
                "qty": o.get("quantity", 0),
                "avg_price": o.get("average_price", 0),
                "txn": o.get("transaction_type", ""),
                "variety": o.get("variety", ""),
                "exchange": o.get("exchange", ""),
                "product": o.get("product", ""),
            }
        
        # Step 1: Discover new pending AMO BUYs not yet tracked
        PENDING_STATUSES = ("OPEN", "TRIGGER PENDING", "AMO REQ RECEIVED", "AMO REQUEST RECEIVED", "PUT ORDER REQ RECEIVED")
        
        for oid, info in live_orders.items():
            if oid in self.pending_amo_fills:
                continue  # already tracked
            
            is_amo_buy = (
                info["variety"] == "amo"
                and info["txn"] == "BUY"
                and info["exchange"] == "NSE"
                and info["product"] == "CNC"
                and info["status"] in PENDING_STATUSES
            )
            
            if is_amo_buy:
                sym = info["symbol"]
                # Skip if already in positions (shouldn't happen but defensive)
                if sym in self.positions:
                    continue
                self.pending_amo_fills[oid] = {
                    "symbol": sym,
                    "qty": info["qty"],
                    "registered_at": now_ist().isoformat(),
                }
                log.info(f"[FIX_V55_AMO_RECOVERY] registered pending AMO BUY: {sym} qty={info['qty']} order_id={oid}")
        
        # Step 2: Check tracked orders for state transitions
        fills_to_process = []
        for oid in list(self.pending_amo_fills.keys()):
            tracked = self.pending_amo_fills[oid]
            sym = tracked["symbol"]
            
            # If order vanished from orders() — could be very old. Clean up.
            if oid not in live_orders:
                # Check if symbol now in holdings (fill happened, order aged out)
                try:
                    _h = kite.holdings() or []
                    _held = any(h.get("tradingsymbol") == sym and (h.get("quantity", 0) + h.get("t1_quantity", 0)) > 0 for h in _h)
                    if _held:
                        log.info(f"[FIX_V55_AMO_RECOVERY] {sym} order_id={oid} not in orders() but IS in holdings -> treat as filled")
                        fills_to_process.append((oid, sym, tracked["qty"], 0))
                    else:
                        log.info(f"[FIX_V55_AMO_RECOVERY] {sym} order_id={oid} vanished and not held -> dropping")
                except Exception:
                    pass
                del self.pending_amo_fills[oid]
                continue
            
            status = live_orders[oid]["status"]
            
            if status == "COMPLETE":
                avg = live_orders[oid]["avg_price"] or 0
                qty = live_orders[oid]["qty"] or tracked["qty"]
                log.info(f"[FIX_V55_AMO_RECOVERY] FILL DETECTED: {sym} qty={qty} avg={avg} order_id={oid}")
                fills_to_process.append((oid, sym, qty, avg))
                del self.pending_amo_fills[oid]
            elif status in ("REJECTED", "CANCELLED"):
                log.warning(f"[FIX_V55_AMO_RECOVERY] {sym} order_id={oid} status={status} -> dropping tracking")
                del self.pending_amo_fills[oid]
            # else: still pending, keep tracking
        
        # Step 3: For each fill, register position + place GTT immediately
        for oid, sym, qty, fill_avg in fills_to_process:
            try:
                # If avg not in order (edge case), fetch from holdings
                if fill_avg <= 0:
                    try:
                        for h in (kite.holdings() or []):
                            if h.get("tradingsymbol") == sym:
                                fill_avg = h.get("average_price", 0)
                                break
                    except Exception:
                        pass
                
                if fill_avg <= 0:
                    log.error(f"[FIX_V55_AMO_RECOVERY] {sym} fill detected but avg=0 — skipping GTT (hourly audit will retry)")
                    continue
                
                # Get ATR for SL calc (same as regular entry path)
                try:
                    _atr = self._estimate_atr(sym) or 0
                except Exception:
                    _atr = 0
                
                sl_price = _snap_sl_tick_up(fill_avg * (1 - EQ_INITIAL_SL_PCT), symbol=sym, exchange="NSE")
                if _atr > 0:
                    sl_price = _snap_sl_tick_up(max(sl_price, fill_avg - 1.5 * _atr), symbol=sym, exchange="NSE")
                target = round(fill_avg + EQ_T2_ATR_MULT * _atr, 2) if _atr > 0 else 0
                
                # Register in positions
                self.positions[sym] = {
                    "symbol": sym,
                    "qty": qty,
                    "entry_price": fill_avg,
                    "entry_time": now_ist().isoformat(),
                    "sl": sl_price,
                    "target": target,
                    "peak_price": fill_avg,
                    "atr": _atr,
                    "catalyst": "AMO_RECOVERED",
                    "order_id": str(oid),
                    "source": "AMO_RECOVERED_V55",
                }
                self._save_positions()
                log.info(f"[FIX_V55_AMO_RECOVERY] registered position: {sym} qty={qty} avg={fill_avg} sl={sl_price}")
                
                # Place GTT immediately
                try:
                    gtt_id = self.gtt.ensure_gtt(sym, "NSE", qty, sl_price, target)
                    if gtt_id:
                        self.positions[sym]["gtt_id"] = gtt_id
                        self._save_positions()
                        log.info(f"[FIX_V55_AMO_RECOVERY] GTT PLACED: {sym} SL={sl_price} id={gtt_id}")
                        try:
                            self.telegram.send(f"AMO FILL RECOVERED\n{sym} qty={qty} @ {fill_avg}\nSL GTT placed @ {sl_price}")
                        except Exception:
                            pass
                    else:
                        log.error(f"[FIX_V55_AMO_RECOVERY] GTT placement returned None for {sym} — next audit will retry")
                except Exception as _ge:
                    log.error(f"[FIX_V55_AMO_RECOVERY] GTT placement raised for {sym}: {_ge} — next audit will retry")
                
            except Exception as _fe:
                log.error(f"[FIX_V55_AMO_RECOVERY] fill processing failed for {sym}: {_fe}")

    def _gtt_audit(self):
        """
        Ensure every holding has exactly 1 GTT.
        Source of truth: Kite holdings (NOT JSON file).
        """
        kite = KiteSession.kite()
        if not kite:
            return
        
        try:
            # Step 1: What do I ACTUALLY own? (Kite = truth)
            # Check BOTH holdings AND positions — covers T1, T2, and day positions
            kite_holdings = {}
            
            # Holdings (settled + T1)
            holdings = kite.holdings()
            for h in holdings:
                sym = h.get("tradingsymbol", "")
                qty = h.get("quantity", 0) + h.get("t1_quantity", 0)
                avg = h.get("average_price", 0)
                if qty > 0 and sym and sym not in BLACKLIST:
                    kite_holdings[sym] = {"qty": qty, "avg": avg}
            
            # Positions (day positions, not yet in holdings)
            positions = kite.positions().get("net", [])
            for p in positions:
                sym = p.get("tradingsymbol", "")
                qty = p.get("quantity", 0)
                avg = p.get("average_price", 0)
                exch = p.get("exchange", "")
                if qty > 0 and sym and sym not in BLACKLIST and exch == "NSE":
                    if sym not in kite_holdings:
                        kite_holdings[sym] = {"qty": qty, "avg": avg}
                    else:
                        kite_holdings[sym]["qty"] += qty
            
            # Step 2: What GTTs exist? (Kite = truth)
            all_gtts = kite.get_gtts() or []
            gtt_by_sym = defaultdict(list)
            for g in all_gtts:
                if g.get("status") == "active":
                    sym = g.get("condition", {}).get("tradingsymbol", "")
                    exch = g.get("condition", {}).get("exchange", "")
                    if sym and exch == "NSE":
                        gtt_by_sym[sym].append(g)
            
            placed, modified, deleted = 0, 0, 0
            
            # FIX_V53_EQ_DEDUP: Step 3 — keep HIGHEST trigger, not newest id.
            # Prior logic sorted by id and kept newest, which could silently
            # regress a profit-locking SL when a newer GTT had lower trigger.
            # Now matches F&O V30 no-regress pattern.
            def _trig_of_eq(g):
                tv = g.get("condition", {}).get("trigger_values") or [0]
                try:
                    return float(tv[0])
                except Exception:
                    return 0.0
            
            for sym, gtts in gtt_by_sym.items():
                if len(gtts) > 1:
                    gtts.sort(key=_trig_of_eq, reverse=True)
                    keeper = gtts[0]
                    for dup in gtts[1:]:
                        try:
                            kite.delete_gtt(dup["id"])
                            deleted += 1
                            log.warning(f"GTT EQUITY: deleted dup {sym} id={dup['id']} trig={_trig_of_eq(dup)} (kept id={keeper['id']} trig={_trig_of_eq(keeper)}) [FIX_V53_EQ_DEDUP]")
                        except Exception as _e71:
                            log.warning(f"GTT EQUITY: failed to delete dup {sym} id={dup['id']}: {_e71} [PATCH_V71_RESILIENCE]")
                    gtt_by_sym[sym] = [keeper]
            
            # Step 4: Delete orphans (GTT exists but stock not in holdings)
            for sym in list(gtt_by_sym.keys()):
                if sym not in kite_holdings:
                    for g in gtt_by_sym[sym]:
                        try:
                            kite.delete_gtt(g["id"])
                            deleted += 1
                            log.info(f"GTT EQUITY: deleted orphan {sym} id={g['id']}")
                        except Exception as _e71:
                            log.warning(f"GTT EQUITY: failed to delete orphan {sym} id={g['id']}: {_e71} [PATCH_V71_RESILIENCE]")
            
            # Step 5: Ensure every holding has GTT with valid triggers
            ltp_data = kite.ltp([f"NSE:{s}" for s in kite_holdings]) if kite_holdings else {}
            
            for sym, info in kite_holdings.items():
                qty = info["qty"]
                ltp = ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                if ltp <= 0:
                    ltp = info["avg"]
                
                # FIX_V29_SINGLE_ONLY: all GTTs are SINGLE (SL-only). Software handles target exit.
                # FIX_V29_SL_PROTECT: use max(fresh_sl, position_sl) to never lower SL on new placement.
                # FIX_V53_EQ_FLOOR: full floor rule — max(formula, existing_gtt_trigger, pos_sl).
                # Mirrors F&O _fno_gtt_audit V30 no-regress logic.
                formula_sl = _snap_sl_tick_up(ltp * (1 - EQ_TRAIL_SL_PCT), symbol=sym, exchange="NSE")
                _pos_sl = 0.0
                try:
                    _pos_sl = float(self.positions.get(sym, {}).get("sl", 0) or 0)
                except Exception:
                    _pos_sl = 0.0
                
                existing_gtts = gtt_by_sym.get(sym, [])
                existing_trigger = _trig_of_eq(existing_gtts[0]) if existing_gtts else 0.0
                
                sl = max(formula_sl, _snap_sl_tick_up(_pos_sl, symbol=sym, exchange="NSE") if _pos_sl > 0 else 0.0, existing_trigger)
                
                # FIX_V53_EQ_CAP + FIX_V57_EQ_TIGHT_CAP: safety cap — SL must be BELOW LTP.
                # V53 used 15% (too wide). V57 tightens to 2.5% matching standard equity SL.
                # Still wide enough to avoid Kite 0.25% gap rejection.
                if ltp > 0 and sl >= ltp:
                    capped = _snap_tick(ltp * 0.975, symbol=sym, exchange="NSE")
                    log.warning(f"GTT EQUITY: [CAP] {sym} sl={sl} >= ltp={ltp}, capping to {capped} [FIX_V57_EQ_TIGHT_CAP]")
                    sl = capped
                
                # FIX_V53_GAP: enforce Kite's 0.25% minimum trigger gap
                sl = _gtt_safe_trigger(sl, ltp, side="SELL")
                
                if sl <= 0:
                    log.warning(f"GTT EQUITY: {sym} computed sl={sl}, skip")
                    continue
                
                if not existing_gtts:
                    # No GTT — place new SINGLE SL-only
                    try:
                        gtt_id = _safe_place_gtt(kite, 
                            trigger_type=kite.GTT_TYPE_SINGLE,
                            tradingsymbol=sym, exchange="NSE",
                            trigger_values=[sl], last_price=ltp,
                            orders=[
                                {"transaction_type": "SELL", "quantity": qty, "price": round(sl * 0.98, 1), "order_type": "LIMIT", "product": "CNC"},
                            ],
                        )
                        placed += 1
                        log.info(f"GTT EQUITY: placed {sym} SL={sl} (SINGLE) id={gtt_id}")
                    except Exception as e:
                        log.error(f"GTT EQUITY: place {sym} failed: {e}")
                else:
                    # GTT exists — check if triggers are valid
                    gtt = existing_gtts[0]
                    triggers = gtt.get("trigger_values", [])
                    if not triggers or len(triggers) == 0:
                        # Empty triggers — delete and re-place as SINGLE
                        try:
                            kite.delete_gtt(gtt["id"])
                            gtt_id = _safe_place_gtt(kite, 
                                trigger_type=kite.GTT_TYPE_SINGLE,
                                tradingsymbol=sym, exchange="NSE",
                                trigger_values=[sl], last_price=ltp,
                                orders=[
                                    {"transaction_type": "SELL", "quantity": qty, "price": round(sl * 0.98, 1), "order_type": "LIMIT", "product": "CNC"},
                                ],
                            )
                            placed += 1
                            log.info(f"GTT EQUITY: re-placed {sym} (empty triggers) SL={sl} (SINGLE)")
                        except Exception as e:
                            log.error(f"GTT EQUITY: re-place {sym} failed: {e}")
                    else:
                        # Valid triggers — check if SL should be trailed up
                        old_sl = triggers[0] if triggers else 0
                        if sl > old_sl:
                            # FIX_V29_SINGLE_ONLY: always modify as SINGLE
                            try:
                                _safe_modify_gtt(kite, 
                                    trigger_id=gtt["id"], trigger_type=kite.GTT_TYPE_SINGLE,
                                    tradingsymbol=sym, exchange="NSE",
                                    trigger_values=[sl], last_price=ltp,
                                    orders=[
                                        {"transaction_type": "SELL", "quantity": qty, "price": round(sl * 0.98, 1), "order_type": "LIMIT", "product": "CNC"},
                                    ],
                                )
                                modified += 1
                                log.info(f"GTT EQUITY: modified {sym} SL {old_sl}->{sl}")
                            except Exception:
                                # Modify failed (type mismatch etc) — delete and re-place as SINGLE
                                try:
                                    kite.delete_gtt(gtt["id"])
                                    _safe_place_gtt(kite, 
                                        trigger_type=kite.GTT_TYPE_SINGLE,
                                        tradingsymbol=sym, exchange="NSE",
                                        trigger_values=[sl], last_price=ltp,
                                        orders=[
                                            {"transaction_type": "SELL", "quantity": qty, "price": round(sl * 0.98, 1), "order_type": "LIMIT", "product": "CNC"},
                                        ],
                                    )
                                    placed += 1
                                    log.info(f"GTT EQUITY: re-placed {sym} (modify failed) SL={sl} (SINGLE)")
                                except Exception as e2:
                                    log.error(f"GTT EQUITY: re-place {sym} failed: {e2}")
                        else:
                            modified += 1
            
            log.info(f"GTT EQUITY AUDIT: placed={placed} modified={modified} deleted={deleted}")
        
        except Exception as e:
            log.error(f"GTT EQUITY audit error: {e}")

    def _record_trade(self, symbol, pos, exit_price, reason):
        """Record a completed trade."""
        entry = pos.get("entry_price", 0)
        qty = pos.get("qty", 0)
        pnl = (exit_price - entry) * qty if exit_price > 0 else 0
        # V83_EXIT_HOOK: record EVERY exit (loss/breakeven/win/external) for 24h cooldown.
        # Distinct from V79A_LOSS_HOOK below (which only fires on real losses).
        # Critical: must fire on SOLD_EXTERNAL/breakeven too — VENUSREM 1st re-entry
        # was after a pnl=0 exit, would have slipped through a loss-only hook.
        try:
            get_exit_tracker().record_exit(symbol, reason=reason or "UNKNOWN")
            log.info(f"[V83_EXIT_HOOK] recorded exit: {symbol} reason={reason} pnl={pnl:.2f}")
        except Exception as _v83eh:
            log.warning(f"[V83_EXIT_HOOK] failed for {symbol}: {_v83eh}")
        trade = {
            "symbol": symbol,
            "entry_price": entry,
            "exit_price": exit_price,
            "qty": qty,
            "pnl": round(pnl, 2),
            "reason": reason,
            "entry_time": pos.get("entry_time", ""),
            "exit_time": now_ist().isoformat(),
            "source": pos.get("source", "UNKNOWN"),
        }
        # Append to trades file
        trades = load_json(self._trades_file, {"trades": []})
        trades["trades"].append(trade)
        save_json(self._trades_file, trades)

        if self.risk and pnl != 0:
            self.risk.record_trade(pnl)

        # V79A_LOSS_HOOK: feed loss_tracker so cooldown can see history
        try:
            if pnl < 0:
                _m = re.match(r'^([A-Z&]+?)\d{2}', symbol or "")
                _stock = _m.group(1) if _m else (symbol or "")
                if _stock:
                    _entry_px = entry if entry else 0
                    _pnl_pct = (pnl / (_entry_px * qty) * 100) if (_entry_px and qty) else None
                    get_loss_tracker().record_loss(_stock, _pnl_pct)
                    log.info(f"[V79A_LOSS_HOOK] recorded loss: {_stock} pnl={pnl:.2f} pct={_pnl_pct}")
        except Exception as _v79ae:
            log.warning(f"[V79A_LOSS_HOOK] failed: {_v79ae}")

    # ── ENTRY LOGIC ──

    # V28 A1: Kite-holdings cache (30s) to avoid hammering API
    _kite_holdings_cache = {"ts": 0, "syms": set()}

    @classmethod
    def _kite_held_symbols(cls):
        """Get currently held equity symbols from Kite (cached 30s).
        Returns None if Kite unreachable so caller can fall back to local JSON.
        """
        import time as _t
        now = _t.time()
        if now - cls._kite_holdings_cache["ts"] < 30:
            return cls._kite_holdings_cache["syms"]
        kite = KiteSession.kite()
        if not kite:
            return None
        try:
            held = set()
            for h in (kite.holdings() or []):
                sym = h.get("tradingsymbol", "")
                qty = h.get("quantity", 0) + h.get("t1_quantity", 0)
                if sym and qty > 0:
                    held.add(sym)
            # Also include same-day positions on NSE (not yet holdings)
            for p in (kite.positions().get("net", []) or []):
                if p.get("exchange") == "NSE" and p.get("quantity", 0) > 0:
                    ts = p.get("tradingsymbol", "")
                    if ts:
                        held.add(ts)
            cls._kite_holdings_cache = {"ts": now, "syms": held}
            return held
        except Exception as e:
            log.warning(f"_kite_held_symbols: kite.holdings() failed: {e}")
            return None

    def _can_enter(self, symbol):
        """Pre-entry checks.
        V28 A1: Authoritative position count comes from Kite, not stale local JSON.
        """
        if symbol in BLACKLIST:
            return False, "BLACKLISTED"
        if symbol in self.positions:
            return False, "ALREADY_HELD"

        # V28 A1: check Kite live
        held = EquityModule._kite_held_symbols()
        if held is not None:
            if symbol in held:
                return False, "ALREADY_HELD_ON_KITE"
            if len(held) >= EQ_MAX_POS:
                return False, f"KITE_MAX_POSITIONS_{len(held)}"
            # Sector limit based on LIVE Kite holdings
            sector = get_sector(symbol)
            sector_count = sum(1 for s in held if get_sector(s) == sector)
            if sector_count >= EQ_MAX_PER_SECTOR:
                return False, f"SECTOR_LIMIT_KITE_{sector}"
        else:
            # Fallback to local JSON if Kite unreachable (safer to under-trade)
            if len(self.positions) >= EQ_MAX_POS:
                return False, "MAX_POSITIONS"
            sector = get_sector(symbol)
            sector_count = sum(1 for s in self.positions if get_sector(s) == sector)
            if sector_count >= EQ_MAX_PER_SECTOR:
                return False, "SECTOR_LIMIT"

        if not self.risk or not self.risk.can_trade():
            return False, "RISK_HALTED"
        return True, "OK"

    def _calc_qty(self, price, sl_price, score=60):
        """Calculate quantity using risk-per-trade rule.
        V28 C10: Edge-weighted — higher score = bigger bet.
                 multiplier = score / 60  (score 60 = 1x baseline)
                 Capped at 3% risk (1.5x of 2% baseline).
        V28 D1:  With EQ_MAX_POS=5, allow up to 25% per position (was 15% for 10-pos).
        V28 P2:  DRAWDOWN-BASED SIZING — reduce size after losing streaks.
                 Looks at last 5 trades from trades_equity.json.
                 4+ losses out of 5 -> 50% size. 3 losses -> 75% size.
                 Only REDUCES size, never increases beyond normal.
        """
        if price <= 0 or sl_price <= 0 or sl_price >= price:
            return 0
        risk_per_share = price - sl_price
        if risk_per_share <= 0:
            return 0

        # V28 C10: edge-weighted risk. Baseline 2%, max 3%, min 1%
        try:
            sc = float(score) if score else 60.0
        except Exception:
            sc = 60.0
        score_mult = max(0.5, min(1.5, sc / 60.0))   # 0.5x..1.5x
        effective_risk_pct = EQ_RISK_PER_TRADE * score_mult
        effective_risk_pct = min(effective_risk_pct, 0.03)  # hard cap 3%

        # V28 P2: Drawdown-based size multiplier
        _drawdown_mult = 1.0
        try:
            trades_data = load_json(self._trades_file, {"trades": []})
            last_trades = trades_data.get("trades", [])[-5:]
            if len(last_trades) >= 5:
                losses = sum(1 for t in last_trades if t.get("pnl", 0) < 0)
                if losses >= 4:
                    _drawdown_mult = 0.5
                    log.info(f"[V28 P2] {losses}/5 losses in recent trades — sizing at 50%")
                elif losses == 3:
                    _drawdown_mult = 0.75
                    log.info(f"[V28 P2] {losses}/5 losses in recent trades — sizing at 75%")
        except Exception as _pe:
            log.debug(f"[V28 P2] drawdown check failed: {_pe}, using full size")
        effective_risk_pct = effective_risk_pct * _drawdown_mult

        risk_amount = self.available * effective_risk_pct
        qty = int(risk_amount / risk_per_share)

        # V28 D1: concentration cap raised from 15% to 25% (matches 5-position max)
        max_cost = self.available * 0.25
        if qty * price > max_cost:
            qty = int(max_cost / price)
        return max(0, qty)

    # V80_FIX1: Score-delta replacement
    # Rule: if at EQ_MAX_POS cap and a NEW high-score candidate appears,
    # exit the worst-performing flat/losing holding to make room.
    # Guards: score>=75, worst holding PnL<=+0.5%, held>=3d, max 1 swap/day.
    # Why: prevents missing major catalysts (e.g. HAL bags 20kCr order, score 90)
    # when 10 positions are full of mediocre flat trades.
    def try_replace_weak_holding(self, new_symbol, new_score):
        """V80_FIX1: Try to free a slot by exiting worst flat/losing holding.
        Returns True if a replacement happened, False otherwise."""
        try:
            # Guard 1: score floor
            if new_score < 75:
                log.debug(f"V80_FIX1 REPLACE: {new_symbol} score={new_score} < 75, skip")
                return False
            # Guard 2: one swap per day
            today = today_ist().isoformat()
            if getattr(self, "_last_replace_day", None) == today:
                log.debug(f"V80_FIX1 REPLACE: already swapped today ({self._last_replace_day}), skip")
                return False
            # Get Kite holdings for live PnL
            kite = KiteSession.kite()
            if not kite:
                return False
            try:
                holdings = kite.holdings() or []
            except Exception as _eh:
                log.warning(f"V80_FIX1 REPLACE: kite.holdings() failed: {_eh}")
                return False
            # Find worst-PnL holding that meets guards
            from datetime import datetime as _dt
            candidates = []
            for h in holdings:
                sym = h.get("tradingsymbol", "")
                qty = (h.get("quantity", 0) or 0) + (h.get("t1_quantity", 0) or 0)
                if qty <= 0 or not sym:
                    continue
                avg = float(h.get("average_price", 0) or 0)
                ltp = float(h.get("last_price", 0) or 0)
                if avg <= 0 or ltp <= 0:
                    continue
                pnl_pct = (ltp - avg) / avg
                # Guard 3: must be flat or losing (PnL <= +0.5%)
                if pnl_pct > 0.005:
                    continue
                # Guard 4: must be held >= 3 days
                pos = self.positions.get(sym, {})
                entry_time = pos.get("entry_time", "")
                if not entry_time:
                    # Unknown entry time -> assume eligible (legacy holding)
                    days_held = 999
                else:
                    try:
                        entry_dt = _dt.fromisoformat(entry_time.replace("Z", "+00:00"))
                        days_held = (now_ist() - entry_dt).days
                    except Exception:
                        days_held = 999
                if days_held < 3:
                    continue
                candidates.append((sym, pnl_pct, days_held, avg, ltp))
            if not candidates:
                log.info(f"V80_FIX1 REPLACE: {new_symbol} score={new_score} but no eligible weak holdings")
                return False
            # Pick worst PnL
            candidates.sort(key=lambda x: x[1])  # ascending PnL
            worst_sym, worst_pnl, worst_days, worst_avg, worst_ltp = candidates[0]
            log.info(
                f"V80_FIX1 REPLACE_TRIGGER: new={new_symbol} score={new_score} "
                f"-> exit worst={worst_sym} pnl={worst_pnl*100:.2f}% days={worst_days} "
                f"avg={worst_avg:.2f} ltp={worst_ltp:.2f}"
            )
            # Execute exit
            try:
                self._exit_position(worst_sym, f"V80_REPLACE_FOR_{new_symbol}")
                self._last_replace_day = today
                log.info(f"V80_FIX1 REPLACE: {worst_sym} exited successfully, slot freed for {new_symbol}")
                return True
            except Exception as _ee:
                log.error(f"V80_FIX1 REPLACE: exit {worst_sym} failed: {_ee}")
                return False
        except Exception as _e_outer:
            log.warning(f"V80_FIX1 try_replace_weak_holding failed: {_e_outer}")
            return False

    def enter(self, symbol, catalyst, direction, score, source="UNKNOWN"):
        """
        Place a real BUY order for equity.
        Returns True on success.
        """
        # ═══ V83 ENTRY GUARDS (executed in order; first failure returns False) ═══
        # V83_KILL_MACRO_EQ: V68_MACRO is structurally low-quality for equity entries.
        # 91% of equity candidates came from V68_MACRO (avg score 68.6) and produced
        # the bulk of the same-day SL bleed. F&O still uses V68_MACRO via separate path.
        if (source or "").upper() == "V68_MACRO":
            log.info(f"[V83_KILL_MACRO_EQ] BLOCKED {symbol}: V68_MACRO not allowed for equity (use VALIDATED_FILING/BROKERAGE/NEWS)")
            return False

        # V83_EXIT_COOLDOWN_24H: hard block re-entry within 24h of any exit.
        # Stops the VENUSREM-3x and CAMS-2x re-entry death spirals.
        try:
            _et = get_exit_tracker()
            _hrs = _et.hours_since_last_exit(symbol)
            if _hrs is not None and _hrs < 24.0:
                log.info(f"[V83_EXIT_COOLDOWN_24H] BLOCKED {symbol}: last exit {_hrs:.1f}h ago (need 24h)")
                return False
        except Exception as _v83ee:
            log.warning(f"[V83_EXIT_COOLDOWN_24H] check failed for {symbol}: {_v83ee}, allowing")

        # V83_CONTRADICT: block if same symbol has both BULLISH and BEARISH
        # scout entries (score>=65) within last 4 hours. Indicates split conviction.
        try:
            _nb = getattr(self, "news_brain", None)
            if _nb and hasattr(_nb, "get_recent_decisions"):
                _recent = _nb.get_recent_decisions(symbol, hours=4)
                _bull = any(c.get("direction") == "BULLISH" and c.get("score", 0) >= 65 for c in _recent)
                _bear = any(c.get("direction") == "BEARISH" and c.get("score", 0) >= 65 for c in _recent)
                if _bull and _bear:
                    log.info(f"[V83_CONTRADICT] BLOCKED {symbol}: BULLISH+BEARISH both flagged within 4h (split conviction)")
                    return False
        except Exception as _v83ce:
            log.warning(f"[V83_CONTRADICT] check failed for {symbol}: {_v83ce}, allowing")

        # V83_REFLEX: if the same article triggered >=3 candidates simultaneously,
        # this is a basket/reflex signal. Require this specific symbol to also be
        # confirming with price action (>=0.5% move in predicted direction).
        try:
            _nb = getattr(self, "news_brain", None)
            if _nb and hasattr(_nb, "get_candidates_by_catalyst") and catalyst:
                _basket = _nb.get_candidates_by_catalyst(catalyst, hours=4)
                _basket_syms = set(c.get("symbol") for c in _basket if c.get("symbol"))
                if len(_basket_syms) >= 3:
                    try:
                        _kite = KiteSession.kite()
                        if _kite:
                            _q = _kite.quote([f"NSE:{symbol}"]).get(f"NSE:{symbol}", {})
                            _ltp = _q.get("last_price", 0)
                            _pclose = _q.get("ohlc", {}).get("close", 0)
                            if _ltp > 0 and _pclose > 0:
                                _pct = (_ltp - _pclose) / _pclose * 100
                                _confirm_dir = (direction == "BULLISH" and _pct >= 0.5) or \
                                               (direction == "BEARISH" and _pct <= -0.5)
                                if not _confirm_dir:
                                    log.info(f"[V83_REFLEX] BLOCKED {symbol}: basket signal ({len(_basket_syms)} stocks share catalyst) but price not confirming ({_pct:+.2f}% vs {direction})")
                                    return False
                                log.info(f"[V83_REFLEX] {symbol} basket signal CONFIRMED by price ({_pct:+.2f}%)")
                    except Exception as _v83re:
                        log.warning(f"[V83_REFLEX] price check failed for {symbol}: {_v83re}, allowing")
        except Exception as _v83re2:
            log.warning(f"[V83_REFLEX] outer check failed for {symbol}: {_v83re2}, allowing")

        # V83_HOURLY: tiered filter by IST entry hour. Morning entries face stricter
        # thresholds — the 13:00 cluster produced the only profitable equity trades.
        try:
            _now = now_ist()
            _hm = _now.hour * 60 + _now.minute
            _src_upper = (source or "").upper()
            _is_validated = _src_upper.startswith("VALIDATED_")
            # 9:15->555, 10:30->630, 12:30->750, 14:30->870, 15:15->915
            if 555 <= _hm < 630:
                if score < 75:
                    log.info(f"[V83_HOURLY] BLOCKED {symbol}: morning slot needs score>=75, got {score}")
                    return False
                if not _is_validated:
                    log.info(f"[V83_HOURLY] BLOCKED {symbol}: morning slot needs VALIDATED_*, got {source}")
                    return False
            elif 630 <= _hm < 750:
                if score < 70:
                    log.info(f"[V83_HOURLY] BLOCKED {symbol}: mid-morning slot needs score>=70, got {score}")
                    return False
            elif 870 <= _hm < 915:
                if score < 70:
                    log.info(f"[V83_HOURLY] BLOCKED {symbol}: late slot needs score>=70 (overnight gap risk), got {score}")
                    return False
        except Exception as _v83he:
            log.warning(f"[V83_HOURLY] check failed for {symbol}: {_v83he}, allowing")
        # ═══ END V83 GUARDS ═══

        can, reason = self._can_enter(symbol)
        if not can:
            # V80_FIX1: if blocked due to position cap, try replacement
            if reason and (reason.startswith("KITE_MAX_POSITIONS_") or reason == "MAX_POSITIONS"):
                if self.try_replace_weak_holding(symbol, score):
                    # Re-check after replacement
                    can, reason = self._can_enter(symbol)
                    if can:
                        log.info(f"V80_FIX1 REPLACE: re-check passed for {symbol}, proceeding with entry")
                    else:
                        log.info(f"V80_FIX1 REPLACE: re-check still blocked for {symbol}: {reason}")
                        return False
                else:
                    log.debug(f"Equity entry blocked {symbol}: {reason}")
                    return False
            else:
                log.debug(f"Equity entry blocked {symbol}: {reason}")
                return False

        kite = KiteSession.kite()
        if not kite:
            return False

        # Safety: crash, circuit, correlation, bad stock
        if not SafetyFilters.check_nifty_crash(kite): return False
        if not SafetyFilters.check_circuit(kite, symbol): return False
        if not SafetyFilters.check_correlation(symbol, self.positions): return False
        if SafetyFilters.is_bad_stock(kite, symbol): return False


        # FIX_V57_BREAKOUT_PRICE_ACTION START
        # 52wk-high breakouts: real breakouts hold near high with green close.
        # Distribution traps spike then reverse with long upper wick + red close.
        # Same scanner signal, opposite outcomes. Verify intraday before buy.
        if source == "BREAKOUT":
            try:
                _q = kite.quote([f"NSE:{symbol}"]).get(f"NSE:{symbol}", {})
                _ltp = _q.get("last_price", 0)
                _ohlc = _q.get("ohlc", {})
                _open = _ohlc.get("open", 0)
                _high = _ohlc.get("high", 0)
                _low = _ohlc.get("low", 0)
                if _ltp > 0 and _high > 0 and _low > 0 and _open > 0:
                    _pct_off_high = (_high - _ltp) / _high * 100 if _high else 0
                    _day_range = _high - _low
                    _upper_wick_pct = ((_high - _ltp) / _day_range * 100) if _day_range > 0 else 0
                    _is_green = _ltp >= _open
                    _reasons = []
                    if _pct_off_high > 3.0:
                        _reasons.append(f"{_pct_off_high:.1f}% off high (reversed)")
                    if _upper_wick_pct > 50.0:
                        _reasons.append(f"{_upper_wick_pct:.0f}% upper wick (failed BO)")
                    if not _is_green:
                        _reasons.append(f"red candle (open={_open:.2f} ltp={_ltp:.2f})")
                    if _reasons:
                        log.info(f"[V57] {symbol} BREAKOUT BLOCKED: {' | '.join(_reasons)}")
                        return False
                    log.info(f"[V57] {symbol} BREAKOUT confirmed: ltp={_ltp:.2f} high={_high:.2f} ({_pct_off_high:.1f}% off) wick={_upper_wick_pct:.0f}% green={_is_green}")
            except Exception as e:
                log.warning(f"[V57] price action check error for {symbol}: {e} - allowing through")
        # FIX_V57_BREAKOUT_PRICE_ACTION END

        # V28 A10: EQUITY REGIME FILTER
        # Prevents two failure modes that have historically cost us money:
        #   (a) Buying on bullish news when the MARKET is falling — you catch the tide
        #   (b) Chasing a stock that has already jumped >3% since the signal fired
        # Exception: very high-conviction catalysts (score >= 85, e.g. BUYBACK) allowed
        #            with a looser relative-strength threshold.
        try:
            _nifty_q = kite.quote(["NSE:NIFTY 50"]).get("NSE:NIFTY 50", {})
            _n_ltp = _nifty_q.get("last_price", 0)
            _n_prev = _nifty_q.get("ohlc", {}).get("close", 0)
            _stock_q = kite.quote([f"NSE:{symbol}"]).get(f"NSE:{symbol}", {})
            _s_ltp = _stock_q.get("last_price", 0)
            _s_prev = _stock_q.get("ohlc", {}).get("close", 0)

            if _n_ltp > 0 and _n_prev > 0 and _s_ltp > 0 and _s_prev > 0:
                _nifty_pct = (_n_ltp - _n_prev) / _n_prev * 100
                _stock_pct = (_s_ltp - _s_prev) / _s_prev * 100
                _rel_strength = _stock_pct - _nifty_pct

                # Threshold varies by catalyst conviction
                _high_conv = (score or 0) >= 85
                _rel_threshold = 0.2 if _high_conv else 0.5

                # Rule 1: falling market + stock not outperforming → BLOCK
                if _nifty_pct <= -1.0 and _rel_strength < _rel_threshold:
                    log.info(
                        f"[V28 A10] {symbol} BLOCKED: NIFTY {_nifty_pct:+.2f}%, "
                        f"stock {_stock_pct:+.2f}%, rel_strength {_rel_strength:+.2f}% "
                        f"< {_rel_threshold:+.1f}% threshold (conv={'HIGH' if _high_conv else 'NORMAL'})"
                    )
                    return False

                # Rule 2: chasing — stock already ran >3% today → BLOCK
                # (This is the v24.1.3 chase protection we used to have.)
                # High-conviction catalysts get 5% tolerance.
                _chase_limit = 10.0 if _high_conv else 6.0  # V75: relaxed 5/3 -> 10/6
                if _stock_pct > _chase_limit:
                    log.info(
                        f"[V28 A10] {symbol} BLOCKED: already +{_stock_pct:.2f}% today "
                        f"> chase_limit {_chase_limit}% (signal arrived late)"
                    )
                    return False

                log.debug(
                    f"[V28 A10] {symbol} regime OK: NIFTY {_nifty_pct:+.2f}%, "
                    f"stock {_stock_pct:+.2f}%, rel {_rel_strength:+.2f}%"
                )
        except Exception as _ae:
            # If regime check fails for any reason (no market data, API error),
            # allow entry rather than block — other safety filters still apply.
            log.warning(f"[V28 A10] regime check failed for {symbol}: {_ae}, allowing entry")

        # Order guard: duplicate check + SEBI compliance
        can_order, guard_reason = OrderGuard.can_place_order(symbol)
        if not can_order:
            log.info(f"Equity: {symbol} blocked by OrderGuard: {guard_reason}")
            return False
        if OrderGuard.check_pending_orders(kite, symbol):
            return False

        try:
            # Get live price
            ltp_data = kite.ltp([f"NSE:{symbol}"])
            price = ltp_data.get(f"NSE:{symbol}", {}).get("last_price", 0)
            if price <= 0:
                return False

            # V28 C3: Freak trade protection — reject LTP more than 20% off prev close
            try:
                q = kite.quote([f"NSE:{symbol}"]).get(f"NSE:{symbol}", {})
                prev_close = q.get("ohlc", {}).get("close", 0)
                if prev_close > 0 and abs(price - prev_close) / prev_close > 0.20:
                    log.warning(f"Equity: {symbol} FREAK PRICE — LTP={price}, prev_close={prev_close}, diff={(price-prev_close)/prev_close*100:.1f}%, REJECTING")
                    return False
            except Exception as _fe:
                log.debug(f"Equity: {symbol} freak-check failed: {_fe}")

            # Get historical data for ATR
            today = today_ist()
            hist = kite.historical_data(
                instrument_token=self._get_instrument_token(kite, symbol),
                from_date=(today - timedelta(days=30)).isoformat(),
                to_date=today.isoformat(),
                interval="day",
            )
            if not hist or len(hist) < 5:
                atr = price * 0.02  # Default 2%
                rsi = 50
                vol_ratio = 1.0
            else:
                closes = [c["close"] for c in hist]
                highs = [c["high"] for c in hist]
                lows = [c["low"] for c in hist]
                volumes = [c["volume"] for c in hist]
                atr = calc_atr(highs, lows, closes)
                rsi = calc_rsi(closes)
                vol_ratio = calc_volume_ratio(volumes)

            # ── Technical confirmation ──
            # RSI: don't buy overbought
            if rsi > 75:
                log.info(f"Equity: {symbol} RSI={rsi:.0f} too high, skipping")
                return False

            # Volume: need above average for conviction
            if vol_ratio < 0.8:
                log.info(f"Equity: {symbol} volume ratio={vol_ratio:.1f} too low, skipping")
                return False

            # ── Calculate SL and target ──
            sl_price = round(price - max(1.5 * atr, price * EQ_INITIAL_SL_PCT), 2)
            # SL must be below current price
            sl_price = min(sl_price, _snap_sl_tick_up(price * (1 - EQ_INITIAL_SL_PCT), symbol=symbol, exchange="NSE"))
            target = round(price + EQ_T2_ATR_MULT * atr, 2)

            # ── Position sizing ──
            qty = self._calc_qty(price, sl_price, score=score)  # V28 C10: edge-weighted
            if qty <= 0:
                log.info(f"Equity: {symbol} qty=0, insufficient capital")
                return False

            cost = qty * price
            if cost > self.available:
                log.info(f"Equity: {symbol} cost={cost:.0f} > available={self.available:.0f}")
                return False

            # ── Place order ──
            order_id = kite.place_order(
                variety=kite.VARIETY_REGULAR,
                exchange="NSE",
                tradingsymbol=symbol,
                transaction_type="BUY",
                quantity=qty,
                order_type="LIMIT",
                price=round(price * 1.005, 1),  # 0.5% buffer above LTP
                product="CNC",
                validity="DAY",
                market_protection=5,  # SEBI mandatory — 5% market protection
            )
            log.info(f"Equity: ORDER PLACED {symbol} qty={qty} price={price} order_id={order_id}")
            OrderGuard.record_order(symbol)

            # FIX_V53_EQ_POLL: 15-sec poll for actual fill (not 2-sec assumption).
            # Prior code checked orders() once after 2s — if fill slower, GTT never
            # placed → position unprotected. Root cause of 19x recurring bug.
            # Now matches F&O poll pattern: 15s max, 1.5s interval, cancel if timeout.
            fill_price = 0.0
            order_status = None
            _poll_max_sec = 15
            _poll_interval = 1.5
            _elapsed = 0.0
            while _elapsed < _poll_max_sec:
                time.sleep(_poll_interval)
                _elapsed += _poll_interval
                try:
                    orders = kite.orders()
                    for o in orders:
                        if str(o.get("order_id")) == str(order_id):
                            order_status = o.get("status")
                            if o.get("average_price", 0) > 0:
                                fill_price = o["average_price"]
                            break
                except Exception as _oe:
                    log.warning(f"Equity: {symbol} order poll error: {_oe}")
                    continue

                if order_status == "COMPLETE":
                    log.info(f"Equity: {symbol} FILLED at {fill_price} after {_elapsed:.1f}s")
                    break
                if order_status == "REJECTED":
                    log.error(f"Equity: {symbol} order REJECTED")
                    try:
                        health.counters["orders_rejected"] += 1
                        health.fire(f"EQUITY REJECTED: {symbol} qty={qty}")
                    except Exception as _e71:
                        log.warning(f"Equity: telegram fire failed for REJECTED {symbol}: {_e71} [PATCH_V71_RESILIENCE]")
                    return False
                if order_status == "CANCELLED":
                    log.warning(f"Equity: {symbol} order CANCELLED externally")
                    try:
                        health.counters["orders_cancelled"] += 1
                        health.fire(f"EQUITY CANCELLED: {symbol}")
                    except Exception as _e71:
                        log.warning(f"Equity: telegram fire failed for CANCELLED {symbol}: {_e71} [PATCH_V71_RESILIENCE]")
                    return False

            if order_status != "COMPLETE":
                log.warning(f"Equity: {symbol} status={order_status} after {_poll_max_sec}s, CANCELLING to avoid phantom position [FIX_V53_EQ_POLL]")
                try:
                    kite.cancel_order(variety=kite.VARIETY_REGULAR, order_id=order_id)
                except Exception as _ce:
                    log.error(f"Equity: {symbol} cancel failed: {_ce}")
                return False

            if fill_price <= 0:
                log.error(f"Equity: {symbol} COMPLETE but fill_price=0, ABORTING (will not save phantom position) [FIX_V53_EQ_POLL]")
                return False

            # Update price/sl/target with actual fill
            price = fill_price
            sl_price = round(price - max(1.5 * atr, price * EQ_INITIAL_SL_PCT), 2)
            sl_price = min(sl_price, _snap_sl_tick_up(price * (1 - EQ_INITIAL_SL_PCT), symbol=symbol, exchange="NSE"))  # PATCH_V70_AUDIT_CLEANUP: was sym (undefined)
            target = round(price + EQ_T2_ATR_MULT * atr, 2)

            # ── Save position (fill confirmed) ──
            self.positions[symbol] = {
                "symbol": symbol,
                "qty": qty,
                "entry_price": price,
                "entry_time": now_ist().isoformat(),
                "sl": sl_price,
                "target": target,
                "peak_price": price,
                "atr": atr,
                "catalyst": catalyst,
                "order_id": str(order_id),
                "source": source,
                "score": score,
                "v85_partial_taken": False,
            }
            self.available -= cost
            self._save_positions()

            # FIX_V53_EQ_POLL: COMPLETE guaranteed by poll above — place GTT immediately
            try:
                gtt_id = self.gtt.ensure_gtt(symbol, "NSE", qty, sl_price, target)
                if gtt_id:
                    self.positions[symbol]["gtt_id"] = gtt_id
                    self._save_positions()
                    log.info(f"[FIX_V53_EQ_POLL] GTT placed at fill: {symbol} SL={sl_price} id={gtt_id}")
                else:
                    log.error(f"[FIX_V53_EQ_POLL] GTT placement returned None for {symbol} — hourly audit will retry")
            except Exception as _ge:
                log.error(f"[FIX_V53_EQ_POLL] GTT placement raised for {symbol}: {_ge} — hourly audit will retry")

            self.telegram.send(
                f"BUY {symbol} x{qty} @ Rs.{price:,.1f}\n"
                f"SL: Rs.{sl_price:,.1f} | Target: Rs.{target:,.1f}\n"
                f"Cost: Rs.{cost:,.0f} | RSI: {rsi:.0f} | Vol: {vol_ratio:.1f}x\n"
                f"Catalyst: {catalyst[:60]}",
                force=True,
            )
            return True

        except Exception as e:
            log.error(f"Equity: entry {symbol} failed: {e}")
            return False

    def _get_instrument_token(self, kite, symbol):
        """Get instrument token for historical data."""
        try:
            instruments = kite.instruments("NSE")
            for inst in instruments:
                if inst.get("tradingsymbol") == symbol and inst.get("instrument_type") == "EQ":
                    return inst["instrument_token"]
        except Exception as _e72:
            log.warning(f"_get_instrument_token({symbol}) failed: {_e72} [PATCH_V72_COMPREHENSIVE]")
        return None

    # ── TRAILING SL ──

    def check_trailing_sl(self):
        """
        Check all positions. If price went up, move SL up.
        SL = 2.5% below current LTP.
        SL ONLY moves UP, NEVER down.
        SL must ALWAYS be BELOW current LTP.

        Runs every 5 minutes during market hours.
        """
        if not self.positions:
            return

        kite = KiteSession.kite()
        if not kite:
            return

        try:
            # Batch LTP fetch
            sym_list = [f"NSE:{s}" for s in self.positions]
            ltp_data = kite.ltp(sym_list)

            for sym, pos in list(self.positions.items()):
                ltp = ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                if ltp <= 0:
                    continue

                old_sl = pos.get("sl", 0)
                peak = pos.get("peak_price", pos.get("entry_price", 0))

                # Update peak
                if ltp > peak:
                    pos["peak_price"] = ltp

                # V84_EQUITY_GRACE_PERIOD: do not trail SL for first 30 min after entry.
                _v84b_grace_skip = False
                try:
                    _v84b_et_str = pos.get("entry_time", "")
                    if _v84b_et_str:
                        _v84b_et = datetime.fromisoformat(_v84b_et_str)
                        _v84b_age_min = (now_ist() - _v84b_et).total_seconds() / 60.0
                        if _v84b_age_min < 30.0:
                            _v84b_grace_skip = True
                            log.debug(f"[V84_EQUITY_GRACE_PERIOD] {sym}: age {_v84b_age_min:.1f}min < 30min, SL trail deferred")
                except Exception as _v84be:
                    log.warning(f"[V84_EQUITY_GRACE_PERIOD] {sym}: age calc failed ({_v84be}), allowing trail")

                # V85_TIERED_TRAIL: conviction- and gain-tiered trail width.
                _v85_score = pos.get("score", 60) or 60
                _v85_entry = pos.get("entry_price", 0)
                _v85_unreal_pct = ((ltp - _v85_entry) / _v85_entry * 100) if _v85_entry > 0 else 0
                if _v85_unreal_pct >= 5.0:
                    _v85_trail_pct = 0.040
                    _v85_tier = "WINNER>5%"
                elif _v85_score >= 80:
                    _v85_trail_pct = 0.040
                    _v85_tier = "CONV>=80"
                elif _v85_score >= 65:
                    _v85_trail_pct = 0.030
                    _v85_tier = "CONV>=65"
                else:
                    _v85_trail_pct = EQ_TRAIL_SL_PCT
                    _v85_tier = "DEFAULT"

                # V85_PARTIAL_TAKE: at +7%, sell 50%, ride 50%.
                _v85_partial_done = pos.get("v85_partial_taken", False)
                if (not _v85_partial_done) and _v85_unreal_pct >= 7.0 and not _v84b_grace_skip:
                    _v85_qty = pos.get("qty", 0)
                    _v85_half = _v85_qty // 2
                    if _v85_half > 0:
                        try:
                            _v85_partial_oid = kite.place_order(
                                variety=kite.VARIETY_REGULAR,
                                exchange="NSE",
                                tradingsymbol=sym,
                                transaction_type="SELL",
                                quantity=_v85_half,
                                order_type="MARKET",
                                product="CNC",
                                validity="DAY",
                                market_protection=5,
                            )
                            log.info(f"[V85_PARTIAL_TAKE] {sym}: +{_v85_unreal_pct:.1f}%, selling {_v85_half}/{_v85_qty} @ market")
                            pos["qty"] = _v85_qty - _v85_half
                            pos["v85_partial_taken"] = True
                            self._save_positions()
                            try:
                                self.telegram.send(f"PARTIAL TAKE {sym}\nSold {_v85_half}/{_v85_qty} @ ~Rs.{ltp} (+{_v85_unreal_pct:.1f}%)\nRiding {_v85_qty - _v85_half}", force=True)
                            except Exception:
                                pass
                        except Exception as _v85_pte:
                            log.warning(f"[V85_PARTIAL_TAKE] {sym}: partial sell failed: {_v85_pte}")

                # Calculate new trailing SL using tiered width
                new_sl = _snap_sl_tick_up(ltp * (1 - _v85_trail_pct), symbol=sym, exchange="NSE")

                # CRITICAL SAFETY CHECKS:
                # 1. New SL must be HIGHER than old SL (only moves UP)
                # 2. New SL must be BELOW current LTP (never above)
                # 3. V84: skip if within 30-min grace period
                if not _v84b_grace_skip and new_sl > old_sl and new_sl < ltp:
                    pos["sl"] = new_sl
                    self._save_positions()

                    # Update GTT
                    self.gtt.ensure_gtt(sym, "NSE", pos.get("qty", 0), new_sl, pos.get("target", 0))
                    log.info(f"Equity trail: {sym} SL {old_sl}→{new_sl} (LTP={ltp}) [V85_TIER={_v85_tier} trail={_v85_trail_pct:.3f} unreal={_v85_unreal_pct:+.2f}%]")

                # FIX_V29_SINGLE_ONLY: software target exit (replaces OCO target trigger)
                _target = pos.get("target", 0)
                if _target > 0 and ltp >= _target:
                    log.info(f"Equity: {sym} TARGET HIT LTP={ltp} >= target={_target}")
                    self._exit_position(sym, "TARGET_HIT")
                    continue

                # Check if SL breached
                if ltp <= old_sl and old_sl > 0:
                    log.info(f"Equity: {sym} SL breached LTP={ltp} <= SL={old_sl}")
                    self._exit_position(sym, "SL_BREACHED")

        except Exception as e:
            log.error(f"Equity trailing SL error: {e}")

    def _exit_position(self, symbol, reason):
        """Exit a position by placing a sell order."""
        pos = self.positions.get(symbol)
        if not pos:
            return

        kite = KiteSession.kite()
        if not kite:
            return

        # Don't sell at lower circuit — will be rejected
        if not SafetyFilters.check_circuit(kite, symbol):
            log.warning(f"Equity: {symbol} at circuit, cannot exit. Retry next cycle.")
            return

        try:
            qty = pos.get("qty", 0)
            if qty <= 0:
                return

            # V85_PRE_SELL_CHECK: verify Kite holds this stock before placing SELL.
            try:
                _v85_holdings = kite.holdings() or []
                _v85_held_qty = 0
                for _v85_h in _v85_holdings:
                    if _v85_h.get("tradingsymbol") == symbol:
                        _v85_held_qty = (_v85_h.get("quantity", 0) or 0) + (_v85_h.get("t1_quantity", 0) or 0)
                        break
                if _v85_held_qty < qty:
                    log.warning(f"[V85_PRE_SELL_CHECK] {symbol}: bot wants {qty} but Kite holds {_v85_held_qty} — clearing bot state")
                    self.positions.pop(symbol, None)
                    self._save_positions()
                    return
            except Exception as _v85_pse:
                log.warning(f"[V85_PRE_SELL_CHECK] {symbol}: holdings check failed ({_v85_pse}), proceeding cautiously")

            order_id = kite.place_order(
                variety=kite.VARIETY_REGULAR,
                exchange="NSE",
                tradingsymbol=symbol,
                transaction_type="SELL",
                quantity=qty,
                order_type="MARKET",
                product="CNC",
                validity="DAY",
                market_protection=5,  # SEBI mandatory
            )
            log.info(f"Equity: SELL {symbol} x{qty} reason={reason} order_id={order_id}")

            # V85_EXIT_POLL: poll up to 30s for COMPLETE status
            exit_price = 0
            _v85_status = "UNKNOWN"
            for _v85_attempt in range(15):
                time.sleep(2)
                try:
                    _v85_orders = kite.orders()
                    for _v85_o in _v85_orders:
                        if str(_v85_o.get("order_id")) == str(order_id):
                            _v85_status = _v85_o.get("status", "UNKNOWN")
                            if _v85_status == "COMPLETE":
                                exit_price = _v85_o.get("average_price", 0)
                            break
                except Exception as _v85_e:
                    log.warning(f"[V85_EXIT_POLL] {symbol}: kite.orders() failed: {_v85_e}")
                if _v85_status == "COMPLETE" and exit_price > 0:
                    log.info(f"[V85_EXIT_POLL] {symbol}: filled @ Rs.{exit_price} after {(_v85_attempt+1)*2}s")
                    break
                if _v85_status in ("REJECTED", "CANCELLED"):
                    log.warning(f"[V85_EXIT_POLL] {symbol}: {_v85_status} after {(_v85_attempt+1)*2}s")
                    break
            else:
                log.warning(f"[V85_EXIT_POLL] {symbol}: no fill after 30s, last_status={_v85_status}")

            self._record_trade(symbol, pos, exit_price, reason)
            del self.positions[symbol]
            self._save_positions()

            pnl = (exit_price - pos.get("entry_price", 0)) * qty if exit_price > 0 else 0
            self.telegram.send(
                f"SELL {symbol} x{qty} @ Rs.{exit_price:,.1f}\n"
                f"P&L: Rs.{pnl:,.0f} | Reason: {reason}",
                force=True,
            )

        except Exception as e:
            # FIX_V29_STATE_DESYNC: auto-clear position state when Kite says already sold
            err_str = str(e)
            if "Insufficient" in err_str or "pending sell" in err_str or "Holding quantity: 0" in err_str:
                self.positions.pop(symbol, None)
                log.warning(f"Equity: {symbol} already sold on Kite, removed from bot state")
            else:
                log.error(f"Equity: exit {symbol} failed: {e}")

    # ── DEAD MONEY CHECK ──

    # V28 D3: Catalyst-specific dead-money windows. Slow catalysts (buyback, promoter buy)
    # need weeks-to-months to play out. Fast catalysts (results beat, momentum) should
    # move quickly or not at all. Single 15d rule was cutting winners on slow catalysts.
    _DEAD_MONEY_BY_CATALYST = {
        "BUYBACK":       60,   # buybacks run weeks-months
        "PROMOTER_BUY":  45,   # insider conviction signals
        "BONUS":         45,
        "STOCK_SPLIT":   30,
        "ACQUISITION":   45,
        "EXPANSION":     30,
        "ORDER_WIN":     30,
        "DIVIDEND":      25,
        "RESULTS_BEAT":  15,   # fast move expected
        "BREAKOUT":      10,
        "MOMENTUM":      10,
    }

    @classmethod
    def _dead_money_days_for(cls, catalyst_text, default=15):
        """Pick a dead-money window based on catalyst type keywords."""
        if not catalyst_text:
            return default
        up = str(catalyst_text).upper()
        # Check FILING_SIGNALS keys first (explicit)
        for key, days in cls._DEAD_MONEY_BY_CATALYST.items():
            if key in up:
                return days
        # Heuristic fallbacks
        if "BUYBACK" in up:    return 60
        if "ORDER" in up:      return 30
        if "BONUS" in up:      return 45
        if "DIVIDEND" in up:   return 25
        if "RESULT" in up or "EARNINGS" in up: return 15
        return default

    # V28 D7: Press winners — when capital sits idle, add to strongest existing winner
    # Rule: if idle capital > 15% of equity capital AND a position is up > 10%,
    #       top up that winner with up to 50% of its current position value.
    # Only runs once per trading day (tracked via self._last_press_day).
    # Respects all safety guards (circuit, crash, correlation, OrderGuard).
    def press_winners(self):
        try:
            if not is_market_hours():
                return
            today = today_ist()
            if getattr(self, "_last_press_day", None) == today:
                return
            if self.capital <= 0:
                return
            idle_pct = self.available / self.capital if self.capital > 0 else 0
            if idle_pct < 0.10:  # V77_FIX10: 0.15 -> 0.10
                return
            if not self.positions:
                return
            kite = KiteSession.kite()
            if not kite:
                return
            # Find best winner (highest unrealized % gain)
            best_sym = None
            best_gain = 0
            best_ltp = 0
            sym_list = [f"NSE:{s}" for s in self.positions]
            try:
                ltp_data = kite.ltp(sym_list)
            except Exception:
                return
            for sym, pos in self.positions.items():
                entry = pos.get("entry_price", 0)
                ltp = ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                if entry <= 0 or ltp <= 0:
                    continue
                gain = (ltp - entry) / entry
                if gain > best_gain:
                    best_gain = gain
                    best_sym = sym
                    best_ltp = ltp
            if not best_sym or best_gain < 0.05:  # V77_FIX10: 0.10 -> 0.05
                log.debug(f"Equity press: idle={idle_pct*100:.0f}% but no winner >5%")
                self._last_press_day = today
                return
            # Safety guards
            if not SafetyFilters.check_nifty_crash(kite): return
            if not SafetyFilters.check_circuit(kite, best_sym): return
            pos = self.positions[best_sym]
            current_value = pos.get("qty", 0) * best_ltp
            add_budget = min(self.available * 0.5, current_value * 0.5)
            if add_budget < best_ltp:
                return
            add_qty = int(add_budget / best_ltp)
            if add_qty <= 0:
                return
            # OrderGuard: skip if duplicate/pending
            can_order, reason = OrderGuard.can_place_order(best_sym)
            if not can_order:
                log.info(f"Equity press: {best_sym} guard blocked: {reason}")
                return
            if OrderGuard.check_pending_orders(kite, best_sym):
                return
            try:
                order_id = kite.place_order(
                    variety=kite.VARIETY_REGULAR,
                    exchange="NSE",
                    tradingsymbol=best_sym,
                    transaction_type="BUY",
                    quantity=add_qty,
                    order_type="LIMIT",
                    price=round(best_ltp * 1.005, 1),
                    product="CNC",
                    validity="DAY",
                    market_protection=5,
                )
                log.info(f"Equity PRESS WINNER: {best_sym} +{add_qty} @ {best_ltp} (was up {best_gain*100:.1f}%, idle {idle_pct*100:.0f}%) order={order_id}")
                OrderGuard.record_order(best_sym)

                # V28 D7: Poll for fill before touching local state to avoid drift
                # if order doesn't fill. Don't leak state to self.positions until COMPLETE.
                fill_price = 0.0
                order_status = None
                _elapsed = 0.0
                while _elapsed < 15:
                    time.sleep(1.5)
                    _elapsed += 1.5
                    try:
                        for o in (kite.orders() or []):
                            if str(o.get("order_id")) == str(order_id):
                                order_status = o.get("status")
                                if o.get("average_price", 0) > 0:
                                    fill_price = o["average_price"]
                                break
                    except Exception:
                        continue
                    if order_status in ("COMPLETE", "REJECTED", "CANCELLED"):
                        break

                if order_status != "COMPLETE" or fill_price <= 0:
                    log.warning(f"Equity press: {best_sym} not filled (status={order_status}), cancelling and skipping state update")
                    try:
                        kite.cancel_order(variety=kite.VARIETY_REGULAR, order_id=order_id)
                    except Exception as _e72:
                        log.warning(f"Equity press: cancel_order failed for {best_sym} order_id={order_id}: {_e72} [PATCH_V72_COMPREHENSIVE]")
                    self._last_press_day = today
                    return

                self.telegram.send(
                    f"PRESS WINNER {best_sym} +{add_qty} @ Rs.{fill_price:,.1f}\n"
                    f"Already up {best_gain*100:.1f}% | Idle was {idle_pct*100:.0f}%",
                    silent=True,
                )
                # Update position with ACTUAL fill price (blended average)
                pos["qty"] = pos.get("qty", 0) + add_qty
                pos["entry_price"] = round(
                    (pos.get("entry_price", fill_price) * (pos["qty"] - add_qty) + fill_price * add_qty) / pos["qty"], 2
                )
                self.available -= add_qty * fill_price
                self._save_positions()
            except Exception as pe:
                log.error(f"Equity press: order failed for {best_sym}: {pe}")
            self._last_press_day = today
        except Exception as e:
            log.error(f"press_winners error: {e}")

    def check_thesis_invalidation(self):
        """V28 P1: Re-scan news for held positions. Exit if bearish signal appears.
        Runs every EQ_SCAN_INTERVAL_MIN during market hours.
        Logic:
          For each held equity position (all are BULLISH entries):
            Scan current news candidates for this symbol
            If ANY BEARISH candidate with score >= 70 appears -> exit position
        Threshold 70 = strong signal only, filters out noise.
        """
        if not self.positions or not is_market_hours():
            return
        if not self.news:
            return
        try:
            candidates = self.news.scan() or []
        except Exception as _e:
            log.debug(f"[V28 P1] news scan failed: {_e}")
            return

        bearish_map = {}
        for c in candidates:
            if c.get("direction") != "BEARISH":
                continue
            sym = c.get("symbol")
            score = c.get("score", 0)
            if not sym or score < 70:
                continue
            if sym not in bearish_map or score > bearish_map[sym].get("score", 0):
                bearish_map[sym] = c

        for sym in list(self.positions.keys()):
            if sym in bearish_map:
                bad = bearish_map[sym]
                cat = str(bad.get("catalyst", ""))[:60]
                log.warning(f"[V28 P1] THESIS INVALIDATED {sym}: bearish news score={bad.get('score')} catalyst='{cat}' — EXITING")
                try:
                    self.telegram.send(
                        f"THESIS DEAD {sym}\nBearish news score {bad.get('score')}: {cat}\nExiting to prevent loss",
                        silent=False,
                    )
                except Exception as _e72:
                    log.warning(f"check_thesis_invalidation: telegram failed for {sym}: {_e72} [PATCH_V72_COMPREHENSIVE]")
                try:
                    self._exit_position(sym, f"THESIS_DEAD_news_score_{bad.get('score')}")
                except Exception as _ee:
                    log.error(f"[V28 P1] exit failed for {sym}: {_ee}")

    def check_dead_money(self):
        """Exit positions with no meaningful profit after catalyst-specific window.
        V28 D3: Window is catalyst-aware (was a single 15-day rule).
        """
        for sym in list(self.positions.keys()):
            pos = self.positions[sym]
            entry_time = pos.get("entry_time", "")
            if not entry_time:
                continue
            try:
                entry_dt = datetime.fromisoformat(entry_time.replace("Z", "+00:00"))
                days_held = (now_ist() - entry_dt).days
                # V28 D3: catalyst-aware window
                _cat = pos.get("catalyst", "") or ""
                window = EquityModule._dead_money_days_for(_cat, default=EQ_DEAD_MONEY_DAYS)
                if days_held >= window:
                    entry_price = pos.get("entry_price", 0)
                    kite = KiteSession.kite()
                    if kite:
                        ltp_data = kite.ltp([f"NSE:{sym}"])
                        ltp = ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                        if ltp > 0 and ltp <= entry_price * 1.015:  # Less than 1.5% profit
                            log.info(f"Equity: {sym} dead money {days_held}d >= window {window}d (catalyst='{_cat[:40]}'), exiting")
                            self._exit_position(sym, f"DEAD_MONEY_{days_held}d_w{window}")
            except Exception as _e72:
                log.warning(f"check_dead_money loop iteration failed: {_e72} [PATCH_V72_COMPREHENSIVE]")

    # V79_FIX4: Time-stagnation exit (7 days for room to breathe)
    # Rule: exit positions that have not moved 1xATR(20) after 7 trading days,
    # AND are flat or losing. Faster than check_dead_money (15-60 day windows).
    # Catches "stuck in dead air" trades before they bleed into SL.
    # Source: Connors/Alvarez swing-trade research, Minervini SEPA stagnation rule.
    def check_time_stagnation(self):
        """V79_FIX4: Exit positions stagnant after 7 days with movement < 1xATR.
        Runs daily in main scan loop alongside check_dead_money."""
        try:
            from datetime import datetime as _dt
            for sym in list(self.positions.keys()):
                pos = self.positions[sym]
                entry_time = pos.get("entry_time", "")
                if not entry_time:
                    continue
                try:
                    entry_dt = _dt.fromisoformat(entry_time.replace("Z", "+00:00"))
                    days_held = (now_ist() - entry_dt).days
                    if days_held < 7:
                        continue  # 7 days room to breathe
                    entry_price = float(pos.get("entry_price", 0) or 0)
                    if entry_price <= 0:
                        continue
                    kite = KiteSession.kite()
                    if not kite:
                        continue
                    ltp_data = kite.ltp([f"NSE:{sym}"])
                    ltp = float(ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0) or 0)
                    if ltp <= 0:
                        continue
                    # Only exit if flat or losing (don't kill winners)
                    if ltp > entry_price * 1.005:  # >0.5% in profit -> let winners run
                        continue
                    # Get ATR; if unavailable, fall back to 2% threshold
                    atr_val = self._estimate_atr(sym) or (entry_price * 0.02)
                    move = abs(ltp - entry_price)
                    if move < atr_val:
                        log.info(
                            f"V79_FIX4 TIME_STAG_7D: {sym} held={days_held}d "
                            f"move=Rs.{move:.2f} < ATR=Rs.{atr_val:.2f} "
                            f"entry={entry_price:.2f} ltp={ltp:.2f} -> EXIT"
                        )
                        self._exit_position(sym, f"TIME_STAG_{days_held}d_move<ATR")
                except Exception as _e_ts:
                    log.warning(f"V79_FIX4 check_time_stagnation iter failed for {sym}: {_e_ts}")
        except Exception as _e_ts_outer:
            log.warning(f"V79_FIX4 check_time_stagnation failed: {_e_ts_outer}")

    # V79_FIX5: Consecutive upper-circuit exit
    # Rule: exit position if stock has hit upper circuit 5 consecutive trading days.
    # Why: 5 consecutive UCs = parabolic move (25-100% gain), high reversal risk,
    # and risk of getting trapped if stock locks lower circuit.
    # Source: Livermore "sell into euphoria"; standard Indian retail risk-mgmt rule.
    # State: per-position uc_streak counter (resets on non-UC day).
    def check_consecutive_uc(self):
        """V79_FIX5: Exit position if 5 consecutive upper-circuit days.
        Runs daily in main scan loop after EOD or near close."""
        try:
            kite = KiteSession.kite()
            if not kite:
                return
            for sym in list(self.positions.keys()):
                pos = self.positions[sym]
                try:
                    quote_data = kite.quote([f"NSE:{sym}"])
                    q = quote_data.get(f"NSE:{sym}", {})
                    ltp = float(q.get("last_price", 0) or 0)
                    upper_circuit = float(q.get("upper_circuit_limit", 0) or 0)
                    if ltp <= 0 or upper_circuit <= 0:
                        continue
                    # Same threshold pattern as elsewhere in code (0.995)
                    at_uc = (ltp >= upper_circuit * 0.995)
                    streak = int(pos.get("uc_streak", 0) or 0)
                    last_uc_date = pos.get("uc_last_date", "")
                    today_str = today_ist().isoformat()
                    if at_uc:
                        # Only increment once per day (idempotent if check runs multiple times)
                        if last_uc_date != today_str:
                            streak += 1
                            pos["uc_streak"] = streak
                            pos["uc_last_date"] = today_str
                            log.info(
                                f"V79_FIX5 UC_STREAK: {sym} ltp={ltp:.2f} "
                                f"uc={upper_circuit:.2f} streak={streak}/5"
                            )
                            if streak >= 5:
                                log.info(
                                    f"V79_FIX5 UC_EXIT: {sym} hit 5 consecutive UCs "
                                    f"-> EXIT (sell into euphoria)"
                                )
                                self._exit_position(sym, f"UC_STREAK_5d")
                    else:
                        # Reset streak if not at UC
                        if streak > 0:
                            log.info(
                                f"V79_FIX5 UC_RESET: {sym} streak {streak}->0 "
                                f"(ltp={ltp:.2f} not at UC {upper_circuit:.2f})"
                            )
                            pos["uc_streak"] = 0
                            pos["uc_last_date"] = ""
                except Exception as _e_uc:
                    log.warning(f"V79_FIX5 check_consecutive_uc iter failed for {sym}: {_e_uc}")
        except Exception as _e_uc_outer:
            log.warning(f"V79_FIX5 check_consecutive_uc failed: {_e_uc_outer}")

    # ── SCAN & TRADE ──

    def scan_and_trade(self):
        """
        Main equity scan loop.
        1. Scan ALL sources: news, filings, brokerage, macro
        2. Filter by Trendlyne fundamentals
        3. Confirm with technicals (RSI, Volume)
        4. Enter best candidates
        """
        if not self.risk or not self.risk.can_trade():
            return

        # FIX_V40_EQUITY_LLM: news scouts -> LLM; technical scanners pass through
        news_scout_candidates = []
        if self.news:
            news_scout_candidates.extend(self.news.scan())
        if self.filings:
            news_scout_candidates.extend(self.filings.scan())
        if self.brokerage:
            news_scout_candidates.extend(self.brokerage.scan())
        if self.macro:
            news_scout_candidates.extend(self.macro.scan())

        # PATCH_V68_LLM_FIRST_DECISION: split V68 article-items from legacy scouts
        _v68_articles = [c for c in news_scout_candidates if c.get("v68_article_item")]
        _legacy_scouts = [c for c in news_scout_candidates if not c.get("v68_article_item")]
        news_scout_candidates = _legacy_scouts
        _v68_decided = []
        if _v68_articles and getattr(self, "news_brain", None) and hasattr(self.news_brain, "read_and_decide_parallel"):
            _v68_pairs = [
                ({"title": a.get("title",""), "summary": a.get("summary",""),
                  "link": a.get("link",""), "source_feed": a.get("source_feed", a.get("source",""))},
                 a.get("hints", []),
                 a.get("source", "MACRO"))
                for a in _v68_articles
            ]
            try:
                _v68_decided = self.news_brain.read_and_decide_parallel(_v68_pairs, max_workers=5)
                log.info(f"[PATCH_V68] equity LLM-decide: {len(_v68_articles)} articles -> {len(_v68_decided)} candidates")
            except Exception as _ve:
                log.error(f"[PATCH_V68] equity read_and_decide_parallel failed: {_ve}")

        candidates = []
        _use_v4_eq = os.getenv("USE_V4_EQUITY_VALIDATION", "True").lower() in ("true", "1", "yes")
        if _use_v4_eq and getattr(self, "news_brain", None) and hasattr(self.news_brain, "validate_scout_candidate"):
            # PATCH_V63_PARALLEL: replaced sequential for-loop with parallel batch call
            # Same counters, same appends, same logs - just 5x faster due to ThreadPoolExecutor
            _pairs = []
            for sc in news_scout_candidates:
                _pairs.append((sc, {
                    "title": sc.get("title", ""),
                    "summary": sc.get("summary", ""),
                    "link": sc.get("link", ""),
                    "source_feed": sc.get("source_feed", sc.get("source", "")),
                }))
            try:
                _all_validated = self.news_brain.validate_scout_candidates_parallel(_pairs, max_workers=5)
            except Exception as _pe:
                log.error(f"PATCH_V63: parallel validation failed, falling back sequential: {_pe}")
                _all_validated = []
                for sc, ni in _pairs:
                    try:
                        _all_validated.extend(self.news_brain.validate_scout_candidate(sc, ni) or [])
                    except Exception as _ve:
                        log.error(f"FIX_V40: equity validator exception for {sc.get('symbol','?')}: {_ve}")
            _val = 0
            _cor = 0
            # PATCH_V68_LLM_FIRST_DECISION: prepend LLM-decided candidates
            for _vc in _v68_decided:
                candidates.append(_vc)
                _val += 1
            for v in _all_validated:
                if v.get("corrected"):
                    _cor += 1
                candidates.append(v)
                _val += 1
            _rej = len(news_scout_candidates) - (_val - _cor) - _cor  # approx; parallel loses per-scout mapping
            _rej = max(0, len(news_scout_candidates) - len(set(v.get("scout_symbol","") for v in _all_validated)))
            log.info(f"FIX_V40 + PATCH_V63 EQUITY news scouts: total={len(news_scout_candidates)} -> validated={_val}, rejected={_rej}, corrected={_cor}")
        else:
            log.warning("FIX_V40: equity v4 validation UNAVAILABLE - fallback to raw scouts")
            candidates = news_scout_candidates

        if not candidates:
            candidates = []

        # Source 5: 52-week low recovery (Kedia/Damani style)
        if is_market_hours():
            kite = KiteSession.kite()
            if kite:
                try:
                    candidates.extend(EquitySmartScanner.scan_52week_low_recovery(kite, self.all_symbols, self.trendlyne))
                except Exception as e:
                    log.debug(f"52wk scanner error: {e}")

                # Source 6: Promoter buying (strongest insider signal)
                try:
                    candidates.extend(EquitySmartScanner.scan_promoter_increase(kite, self.all_symbols))
                except Exception as e:
                    log.debug(f"Promoter scanner error: {e}")

                # Source 7: Sector rotation (follow the money)
                try:
                    candidates.extend(EquitySmartScanner.scan_sector_rotation(kite))
                except Exception as e:
                    log.debug(f"Sector rotation error: {e}")

                # Source 8: Breakout (52-week high on volume)
                try:
                    candidates.extend(EquitySmartScanner.scan_breakout(kite, self.all_symbols, self._avg_vol_loader._avg_vol, self.trendlyne))  # FIX_V52
                except Exception as e:
                    log.debug(f"Breakout scanner error: {e}")

                # Source 10: Hidden Gem smallcap alpha (FIX_V52)
                try:
                    candidates.extend(EquitySmartScanner.scan_hidden_gem(kite, self.all_symbols, self.trendlyne, self._avg_vol_loader._avg_vol))
                except Exception as e:
                    log.debug(f"[FIX_V52 GEM] {e}")

                # Source 9: Intraday momentum (FIX_V48)
                try:
                    candidates.extend(EquitySmartScanner.scan_intraday_momentum(kite, self.all_symbols, self._avg_vol_loader._avg_vol, self.trendlyne))  # FIX_V52
                except Exception as e:
                    log.debug(f"[FIX_V48] Momentum scanner error: {e}")

                # Source 11: PATCH_V94 bulk + block deal scout (superstars + FII + DII)
                # Run after 17:30 IST so NSE has had time to publish T+0 deals.
                try:
                    _v94_now = now_ist()
                    if (_v94_now.hour, _v94_now.minute) >= (17, 30):
                        candidates.extend(EquitySmartScanner.scan_bulk_block_deals(self.all_symbols))
                except Exception as _v94e:
                    log.debug(f"[V94] bulk/block scanner error: {_v94e}")

        # PATCH_V28_SOURCE_WEIGHT: boost scores based on source reliability
        # Filing/Brokerage = highest conviction, News/Momentum = lowest
        _SOURCE_WEIGHTS = {
            "FILING": 1.15,        # NSE filings — hard data, highest conviction
            "BROKERAGE": 1.12,     # Analyst upgrades/downgrades — professional research
            "PROMOTER_BUY": 1.10,  # Insider buying — strong signal
            "MACRO": 1.05,         # RBI/Fed/GDP — affects entire market
            "52WK_LOW": 1.0,       # Technical — no boost
            "BREAKOUT": 1.0,       # Technical — no boost
            "SECTOR_ROTATION": 1.0,# Sector play — no boost
            "NEWS": 0.95,          # RSS headlines — noisy, slight penalty
            "MOMENTUM": 1.00,      # Price momentum (V93: 0.90->1.00, V85+V83_HOURLY already filter noise)
            "SUPERSTAR_BUY": 1.20, # V94: superstar investor bulk/block buy — highest predictive signal
            "FII_BLOCK_BUY": 1.15, # V94: FII institutional bulk/block buy
            "DII_BLOCK_BUY": 1.10, # V94: DII (MF/insurance) bulk/block buy
        }
        for c in candidates:
            _src = c.get("source", "")
            _w = _SOURCE_WEIGHTS.get(_src, 1.0)
            if _w != 1.0:
                _old_score = c.get("score", 0)
                c["score"] = min(95, int(_old_score * _w))
                if c["score"] != _old_score:
                    log.debug(f"[PATCH_V28_SOURCE_WEIGHT] {c.get('symbol')} {_src}: {_old_score} -> {c['score']} (x{_w})")

        # Deduplicate by symbol (keep highest score)
        best_by_sym = {}
        for c in candidates:
            sym = c["symbol"]
            if sym not in best_by_sym or c.get("score", 0) > best_by_sym[sym].get("score", 0):
                best_by_sym[sym] = c
        candidates = list(best_by_sym.values())

        # PATCH_V92_TRENDLYNE_REWORK: Trendlyne becomes a CONVICTION BOOSTER, not a GATE.
        # Predictive accumulation — pre-position stocks BEFORE catalysts fire,
        # including stocks Trendlyne hasn't scored yet (no_data symbols pass through).
        # DVM>=70: +5 score boost. DVM 40-70: no change. DVM<40: -5 (deprioritize).
        # No-data: no change (TrendlyneScorer.score returns 50 for no-data).
        qualified = []
        for c in candidates:
            sym = c["symbol"]
            try:
                _dvm = self.trendlyne.score(sym)
                _old = c.get("score", 0)
                if _dvm >= 70:
                    c["score"] = min(95, _old + 5)
                    if c["score"] != _old:
                        log.debug(f"[V92_TL_BOOST] {sym} DVM={_dvm:.0f} score {_old}->{c['score']}")
                elif _dvm < 40 and _dvm > 0:
                    c["score"] = max(0, _old - 5)
                    if c["score"] != _old:
                        log.debug(f"[V92_TL_DEPRIO] {sym} DVM={_dvm:.0f} score {_old}->{c['score']}")
                # else: 40<=DVM<70 OR no-data (50): no change
            except Exception as _v92e:
                log.debug(f"[V92_TL_REWORK] {sym} score lookup failed: {_v92e}, allowing")
            qualified.append(c)

        # Sort by score (highest first)
        qualified.sort(key=lambda c: c.get("score", 0), reverse=True)

        # Only take BULLISH for equity (we buy stocks, not short)
        bullish = [c for c in qualified if c.get("direction") == "BULLISH"]

        entered = 0
        for cand in bullish[:5]:  # Try top 5 candidates
            sym = cand["symbol"]
            if sym not in self.all_symbols:
                continue
            if not is_valid_symbol(sym):
                continue

            try:
                if not hasattr(self, "_expert_panel"):
                    self._expert_panel = ExpertPanel(self.trendlyne, getattr(self, "_sector_rotation", None), getattr(self, "news_brain", None))
                kite_p = KiteSession.kite()
                pr = self._expert_panel.evaluate_equity(kite_p, sym, "BULLISH", cand)
                log.info(ExpertPanel.format_equity_log(pr))
                if not pr["allow"]:
                    continue
            except Exception as _pe:
                log.error(f"[FIX_V46_PANEL_EQ] {sym} error: {_pe}")
                continue  # FIX_V49: panel crash = BLOCK, never fall through to enter
            success = self.enter(sym, cand.get("catalyst", ""), "BULLISH", cand.get("score", 0), cand.get("source", "UNKNOWN"))
            if success:
                entered += 1
                if entered >= 3:
                    break

        if entered > 0:
            log.info(f"Equity: entered {entered} positions this cycle")

    # ── MAIN TICK ──

    def tick(self):
        """Called every minute from main loop."""
        n = now_ist()

        # Morning sync at 08:00
        if n.hour == 8 and n.minute == 0:
            self._refresh_capital()
            self._sync_holdings()
            self._gtt_audit()

        # FIX_V55_AMO_RECOVERY: check for AMO fills every minute from 09:10-09:30 IST
        # Market opens 09:15, AMO orders fill in first few seconds.
        # Tight window (20 min) avoids log spam but catches every possible fill.
        if n.hour == 9 and 10 <= n.minute <= 30:
            try:
                self._recover_pending_amo_orders()
            except Exception as _ve:
                log.error(f"[FIX_V55_AMO_RECOVERY] tick call failed: {_ve}")

        # FIX_V55_AMO_RECOVERY: also register new AMO BUYs placed during _evening_amo_scan / premarket
        # at 08:35 and 16:05 IST (30 seconds after AMO placement completes).
        if (n.hour == 8 and n.minute == 35) or (n.hour == 16 and n.minute == 5):
            try:
                self._recover_pending_amo_orders()
                log.info(f"[FIX_V55_AMO_RECOVERY] post-AMO-placement registration at {n.hour:02d}:{n.minute:02d}")
            except Exception as _ve:
                log.error(f"[FIX_V55_AMO_RECOVERY] post-placement call failed: {_ve}")

        # FIX_V41_PREMARKET_AMO: catch overnight/early-morning catalysts before AMO cutoff
        if n.hour == 8 and n.minute == 30 and is_trading_day():
            try:
                # PATCH_V54: equity GIFT awareness (log-only, non-blocking)
                try:
                    from kite_session import KiteSession
                    _kite_v54 = KiteSession.kite()
                    if _kite_v54:
                        _gd = _kite_v54.ltp(["NSEIX:GIFT NIFTY"])
                        _gltp = _gd.get("NSEIX:GIFT NIFTY", {}).get("last_price", 0)
                        _no = _kite_v54.ohlc(["NSE:NIFTY 50"]).get("NSE:NIFTY 50", {}).get("ohlc", {})
                        _yc = _no.get("close", 0)
                        if _gltp > 0 and _yc > 0:
                            _gp = ((_gltp - _yc) / _yc) * 100
                            if _gp <= -0.5:
                                log.warning(f"[PATCH_V54_EQUITY] GIFT {_gp:+.2f}% bearish premarket; equity AMO proceeding")
                            elif _gp >= 0.5:
                                log.info(f"[PATCH_V54_EQUITY] GIFT {_gp:+.2f}% bullish premarket")
                            else:
                                log.info(f"[PATCH_V54_EQUITY] GIFT {_gp:+.2f}% neutral premarket")
                except Exception as _v54ee:
                    log.debug(f"[PATCH_V54_EQUITY] GIFT read skipped: {_v54ee}")
                log.info("[FIX_V41_PREMARKET_AMO] running pre-market AMO refresh")
                self._evening_amo_scan()
            except Exception as _pe:
                log.error(f"[FIX_V41_PREMARKET_AMO] failed: {_pe}")

        # SL re-place at 09:16 (SL orders expire daily)
        if n.hour == 9 and n.minute == 16:
            self._gtt_audit()

        # FIX_V41_FAST_SCAN: 10-min cadence in first hour (news impact peak), 30-min after
        # PATCH_V70_AUDIT_CLEANUP: applied 2026-04-28
        # PATCH_V72_COMPREHENSIVE: applied 2026-04-28 - JSON guard + 19 except hardening + AMO V68 + hook integration
        # PATCH_V71_RESILIENCE: applied 2026-04-28 - JSON guard + 11 critical except hardening — see fix list in patch script
        # PATCH_V69_RESTART_SCAN: also fire on first tick after a mid-market restart,
        # else equity sits idle up to 30 minutes after restart.
        _scan_interval = 10 if (n.hour == 9 or (n.hour == 10 and n.minute <= 15)) else EQ_SCAN_INTERVAL_MIN
        _eq_first_tick = (self._last_scan is None)
        if is_market_hours() and (n.minute % _scan_interval == 0 or _eq_first_tick):
            if self._last_scan != n.hour * 60 + n.minute:
                self._last_scan = n.hour * 60 + n.minute
                _trig = "first-tick-after-restart" if _eq_first_tick else f"interval={_scan_interval}m"
                log.info(f"[PATCH_V69_RESTART_SCAN] equity scan fired at {n.hour:02d}:{n.minute:02d} ({_trig})")
                self.scan_and_trade()

        # Trailing SL check every 5 minutes during market hours
        if is_market_hours() and n.minute % 5 == 0:
            if self._last_trail_check != n.hour * 60 + n.minute:
                self._last_trail_check = n.hour * 60 + n.minute
                self.check_trailing_sl()

        # FIX_V29_MIDDAY_AUDIT: midday GTT safety check at 12:30
        if n.hour == 12 and n.minute == 30:
            self._gtt_audit()

        # Dead money check at 15:00
        if n.hour == 15 and n.minute == 0:
            self.check_dead_money()
            self.check_time_stagnation()  # V79_FIX4
            self.check_consecutive_uc()  # V79_FIX5

        # V28 P1: Thesis invalidation check every EQ_SCAN_INTERVAL_MIN during market hours
        if is_market_hours() and n.minute % EQ_SCAN_INTERVAL_MIN == 0:
            try:
                _ts_key = n.hour * 60 + n.minute
                if getattr(self, "_last_thesis_check", None) != _ts_key:
                    self._last_thesis_check = _ts_key
                    self.check_thesis_invalidation()
            except Exception as _pe:
                log.error(f"[V28 P1] thesis check error: {_pe}")

        # V28 D7 + V77_FIX10: Press winners every hour 11:00-15:00 IST when capital idle
        # Was once at 14:00. Now hourly to catch idle capital throughout the day.
        # press_winners() internal _last_press_day prevents same-day duplicate execution.
        if is_trading_day() and 11 <= n.hour <= 15 and n.minute == 0:
            try:
                self.press_winners()
            except Exception as _pe:
                log.error(f"press_winners tick error: {_pe}")

        # Cancel unfilled AMOs — PATCH_V28_STALE_WINDOW: 09:20-09:30 window, once per day
        # Previously: fired only at 09:20:00 exactly. If main loop was busy that minute,
        # the cancel was skipped entirely until next day (SURAJEST-class miss).
        # Now: tries every minute in the window until successful, then flags done for the day.
        if n.hour == 9 and 20 <= n.minute <= 30:
            try:
                _stale_marker = DATA_DIR / "equity_stale_cancel_today.json"
                _stale_existing = load_json(_stale_marker, {})
                if _stale_existing.get("date") != str(today_ist()):
                    log.info(f"[PATCH_V28_STALE_WINDOW] running equity stale-cancel at {n.hour:02d}:{n.minute:02d} IST")
                    # FIX_V29_DEDUP_AMO: removed duplicate _cancel_stale_amos(), amo.cancel_stale() does same + clears internal list
                    self.amo.cancel_stale()
                    save_json(_stale_marker, {"date": str(today_ist()), "ran_at": f"{n.hour:02d}:{n.minute:02d}"})
            except Exception as _stale_err:
                log.error(f"[PATCH_V28_STALE_WINDOW] failed: {_stale_err}")

        # EOD report at 15:35
        if n.hour == 15 and n.minute == 35 and is_trading_day():
            report = EODReporter.generate_equity_report(self)
            self.telegram.send(report, force=True)

        # V28 A2: Evening AMO scan runs ONCE per evening (20:00 IST), not twice.
        # Previously ran at 16:00 AND 20:00 causing multiple AMO batches per night.
        # 20:00 gives full evening window for news/filings to surface.
        if n.hour == 20 and n.minute == 0 and is_trading_day():
            self._evening_amo_scan()


    def _evening_amo_scan(self):
        """
        Evening scan for next-day AMO orders — runs ONCE at 20:00 IST (V28 A2).
        Checks LIVE Kite holdings before sizing slots (not stale local JSON).
        Counts pending AMOs against the slot budget.
        Total AMO cost must NOT exceed available capital.
        """
        if not is_amo_window():
            return

        # V28 A2: Pull LIVE Kite holdings. If Kite unreachable, abort rather than guess.
        kite = KiteSession.kite()
        if not kite:
            log.error("Equity AMO: no Kite session, aborting scan")
            return

        try:
            held_syms = set()
            for h in (kite.holdings() or []):
                if h.get("tradingsymbol") and (h.get("quantity", 0) + h.get("t1_quantity", 0)) > 0:
                    held_syms.add(h["tradingsymbol"])
            for p in (kite.positions().get("net", []) or []):
                if p.get("exchange") == "NSE" and p.get("quantity", 0) > 0 and p.get("tradingsymbol"):
                    held_syms.add(p["tradingsymbol"])
        except Exception as e:
            log.error(f"Equity AMO: kite.holdings() failed — aborting scan to avoid over-buying: {e}")
            return

        current_count = len(held_syms)
        log.info(f"Equity AMO: kite shows {current_count} live holdings before AMO scan")

        if current_count >= EQ_MAX_POS:
            log.info(f"Equity AMO: kite has {current_count} >= MAX {EQ_MAX_POS}, skipping AMO scan entirely")
            return

        free_slots = EQ_MAX_POS - current_count

        # Also subtract pending AMO BUY orders on Kite (already queued for tomorrow)
        pending_amo_syms = set()
        try:
            for o in (kite.orders() or []):
                if (o.get("variety") == "amo"
                    and o.get("transaction_type") == "BUY"
                    and o.get("status") in ("OPEN", "TRIGGER PENDING", "AMO REQ RECEIVED")
                    and o.get("exchange") == "NSE"):
                    ts = o.get("tradingsymbol")
                    if ts:
                        pending_amo_syms.add(ts)
        except Exception as pe:
            log.warning(f"Equity AMO: pending-order check failed: {pe}")

        free_slots = max(0, free_slots - len(pending_amo_syms))
        log.info(f"Equity AMO: {free_slots} free slots ({current_count}/{EQ_MAX_POS} used + {len(pending_amo_syms)} pending)")

        if free_slots <= 0:
            return

        # FIX_V41_EVENING_LLM: evening AMO now uses same LLM validation as intraday
        news_scout_candidates = []
        if self.filings:
            news_scout_candidates.extend(self.filings.scan())
        if self.news:
            news_scout_candidates.extend(self.news.scan())
        if self.brokerage:
            news_scout_candidates.extend(self.brokerage.scan())
        if self.macro:
            news_scout_candidates.extend(self.macro.scan())

        candidates = []
        _use_v4_eq = os.getenv("USE_V4_EQUITY_VALIDATION", "True").lower() in ("true", "1", "yes")
        if _use_v4_eq and getattr(self, "news_brain", None) and hasattr(self.news_brain, "validate_scout_candidate"):
            # PATCH_V63_PARALLEL: evening AMO parallel batch (was sequential)
            _pairs2 = []
            for sc in news_scout_candidates:
                _pairs2.append((sc, {
                    "title": sc.get("title", ""),
                    "summary": sc.get("summary", ""),
                    "link": sc.get("link", ""),
                    "source_feed": sc.get("source_feed", sc.get("source", "")),
                }))
            try:
                _all_validated2 = self.news_brain.validate_scout_candidates_parallel(_pairs2, max_workers=5)
            except Exception as _pe:
                log.error(f"PATCH_V63: AMO parallel validation failed, falling back sequential: {_pe}")
                _all_validated2 = []
                for sc, ni in _pairs2:
                    try:
                        _all_validated2.extend(self.news_brain.validate_scout_candidate(sc, ni) or [])
                    except Exception as _ve:
                        log.error(f"[FIX_V41_EVENING_LLM] validator exception for {sc.get('symbol','?')}: {_ve}")
            _val = _cor = 0
            for v in _all_validated2:
                if v.get("corrected"):
                    _cor += 1
                candidates.append(v)
                _val += 1
            _rej = max(0, len(news_scout_candidates) - len(set(v.get("scout_symbol","") for v in _all_validated2)))
            log.info(f"[FIX_V41_EVENING_LLM + PATCH_V63] evening AMO scouts: total={len(news_scout_candidates)} validated={_val} rejected={_rej} corrected={_cor}")
        else:
            log.warning("[FIX_V41_EVENING_LLM] v4 validation UNAVAILABLE - fallback to raw scouts")
            candidates = news_scout_candidates

        # Only strong catalysts for AMO (score >= 75)
        # V28 A2: exclude symbols held on Kite AND pending AMOs (not just local positions)
        strong = [c for c in candidates
                  if c.get("direction") == "BULLISH"
                  and c.get("score", 0) >= 75
                  and c.get("symbol") in self.all_symbols
                  and is_valid_symbol(c.get("symbol", ""))
                  and c.get("symbol") not in self.positions
                  and c.get("symbol") not in held_syms
                  and c.get("symbol") not in pending_amo_syms]

        if not strong:
            log.info("Equity AMO: no new strong candidates after holdings/pending filter")
            return

        strong.sort(key=lambda c: c.get("score", 0), reverse=True)
        # V28 A2: cap by actual free_slots (dynamic), and never more than 3 per evening
        strong = strong[: min(free_slots, 3)]

        # Check available capital (minus pending AMO cost)
        pending_cost = self.amo.get_total_pending_cost()
        amo_available = self.available - pending_cost

        for cand in strong:  # V28 A2: strong already capped by free_slots
            sym = cand["symbol"]
            try:
                ltp_data = kite.ltp([f"NSE:{sym}"])
                price = ltp_data.get(f"NSE:{sym}", {}).get("last_price", 0)
                if price <= 0:
                    continue

                # V28 C3: freak-trade protection — reject LTP >20% off prev close
                try:
                    q = kite.quote([f"NSE:{sym}"]).get(f"NSE:{sym}", {})
                    prev_close = q.get("ohlc", {}).get("close", 0)
                    if prev_close > 0 and abs(price - prev_close) / prev_close > 0.20:
                        log.warning(f"Equity AMO: {sym} FREAK LTP={price} vs prev_close={prev_close}, skipping")
                        continue
                except Exception as _e72:
                    log.warning(f"Equity AMO: freak-LTP check failed for {sym}: {_e72} [PATCH_V72_COMPREHENSIVE]")

                # Simple qty: 2% risk, edge-weighted (V28 C10)
                sl_price = _snap_sl_tick_up(price * (1 - EQ_INITIAL_SL_PCT), symbol=sym, exchange="NSE")
                qty = self._calc_qty(price, sl_price, score=cand.get("score", 60))
                if qty <= 0:
                    continue

                cost = qty * price * 1.01  # Include 1% buffer
                if cost > amo_available:
                    log.info(f"Equity AMO: {sym} cost={cost:.0f} exceeds available={amo_available:.0f}")
                    continue

                order_id = self.amo.place_amo(sym, price, qty)
                if order_id:
                    amo_available -= cost
                    self.telegram.send(
                        f"AMO PLACED {sym} x{qty} @ ~Rs.{price:,.1f}\n"
                        f"Catalyst: {cand.get('catalyst', '')[:60]}",
                        silent=True,
                    )
            except Exception as e:
                log.error(f"Equity AMO: {sym} failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  F&O MODULE — Completely independent. No equity awareness.
# ═══════════════════════════════════════════════════════════════════════════════

class SectorRotation:
    """
    PATCH_V8_SECTOR_ROTATION
    Tracks sectoral indices and ranks them by today's intraday % move.
    Refreshes every 15 minutes during market hours.
    Top-3 sectors get score boost, bottom-2 sectors block new entries.
    """

    def __init__(self, fno_module):
        self.fno = fno_module
        self.ranking_file = DATA_DIR / "sector_ranking.json"
        # PATCH_V72_SECTOR_INIT: restore _last_refresh_minute that V70 wrongly removed.
        # Used by FnoModule.tick() at gir.py L9289 to throttle refresh to once per 15min.
        self._last_refresh_minute = -1
        self._ranking = {}  # sector_name -> {"pct": X, "rank": N}

    def refresh(self):
        """Fetch all sectoral indices, compute today's % move, rank them."""
        kite = KiteSession.kite()
        if not kite:
            return {}
        try:
            keys = list(SECTOR_INDICES_MAP.values())
            ohlc_data = kite.ohlc(keys)
            sector_moves = []
            for sector, key in SECTOR_INDICES_MAP.items():
                d = ohlc_data.get(key, {})
                ltp = d.get("last_price", 0)
                ohlc = d.get("ohlc", {})
                prev_close = ohlc.get("close", 0)
                if ltp > 0 and prev_close > 0:
                    pct = ((ltp - prev_close) / prev_close) * 100
                    sector_moves.append((sector, pct))
            # Sort descending by % move
            sector_moves.sort(key=lambda x: x[1], reverse=True)
            ranking = {}
            for rank, (sector, pct) in enumerate(sector_moves, start=1):
                ranking[sector] = {"pct": round(pct, 2), "rank": rank}
            self._ranking = ranking
            try:
                save_json(self.ranking_file, {
                    "timestamp": now_ist().isoformat(),
                    "ranking": ranking,
                })
            except Exception:
                pass
            top3 = [s for s, d in ranking.items() if d["rank"] <= 3]
            bot2 = [s for s, d in ranking.items() if d["rank"] >= len(ranking) - 1]
            log.info(f"[PATCH_V8_SECTOR] ranking refreshed: TOP3={top3} BOTTOM2={bot2}")
            return ranking
        except Exception as _e:
            log.error(f"[PATCH_V8_SECTOR] refresh failed: {_e}")
            return {}

    def get_sector_bias(self, symbol):
        """
        Returns ("TOP"|"MIDDLE"|"BOTTOM", sector_name, pct).
        TOP    = sector ranked 1-3 (strong, boost score)
        MIDDLE = sector ranked 4-7 (neutral)
        BOTTOM = sector ranked 8-9 (weak, block entries)
        Returns (None, None, 0) if sector unknown or ranking not yet loaded.
        """
        sector = get_sector(symbol)
        if not sector or sector not in self._ranking:
            return (None, None, 0)
        rank = self._ranking[sector]["rank"]
        pct = self._ranking[sector]["pct"]
        total = len(self._ranking)
        if rank <= 3:
            return ("TOP", sector, pct)
        if rank >= total - 1:
            return ("BOTTOM", sector, pct)
        return ("MIDDLE", sector, pct)


class GapDetector:
    """
    PATCH_V6_GAP
    Detects pre-market gaps for all F&O stocks + indices at 08:50 IST.
    Tags each symbol as GAP_UP / GAP_DOWN / FLAT.
    Reviews pending AMOs and cancels any that contradict gap direction.
    """

    INDEX_SPOT_KEYS = {
        "NIFTY": "NSE:NIFTY 50",
        "BANKNIFTY": "NSE:NIFTY BANK",
        "FINNIFTY": "NSE:NIFTY FIN SERVICE",
        "MIDCPNIFTY": "NSE:NIFTY MID SELECT",
    }

    def __init__(self, fno_module):
        self.fno = fno_module
        self.gap_file = DATA_DIR / "gap_status.json"

    def detect_gaps(self):
        """Read previous close vs current LTP for all symbols. Save gap_status.json."""
        kite = KiteSession.kite()
        if not kite:
            log.warning("[PATCH_V6_GAP] no Kite session, skip")
            return {}
        gap_status = {"timestamp": now_ist().isoformat(), "stocks": {}, "indices": {}}
        # Build the full instrument list to query
        index_keys = list(self.INDEX_SPOT_KEYS.values())
        stock_keys = [f"NSE:{s}" for s in self.fno.fno_stocks]
        all_keys = index_keys + stock_keys
        # Batch ohlc calls
        ohlc_data = {}
        try:
            for i in range(0, len(all_keys), 250):
                batch = all_keys[i:i+250]
                resp = kite.ohlc(batch)
                ohlc_data.update(resp)
        except Exception as _e:
            log.error(f"[PATCH_V6_GAP] ohlc batch failed: {_e}")
            return {}
        # Process indices
        gap_up_idx = []
        gap_down_idx = []
        for idx_name, key in self.INDEX_SPOT_KEYS.items():
            try:
                d = ohlc_data.get(key, {})
                ltp = d.get("last_price", 0)
                ohlc = d.get("ohlc", {})
                prev_close = ohlc.get("close", 0)
                if ltp <= 0 or prev_close <= 0:
                    continue
                gap_pct = ((ltp - prev_close) / prev_close) * 100
                tag = "FLAT"
                if gap_pct >= GAP_INDEX_THRESHOLD:
                    tag = "GAP_UP"
                    gap_up_idx.append(f"{idx_name} {gap_pct:+.2f}%")
                elif gap_pct <= -GAP_INDEX_THRESHOLD:
                    tag = "GAP_DOWN"
                    gap_down_idx.append(f"{idx_name} {gap_pct:+.2f}%")
                gap_status["indices"][idx_name] = {"pct": round(gap_pct, 2), "tag": tag, "ltp": ltp, "prev_close": prev_close}
            except Exception:
                continue
        # Process stocks
        gap_up_stocks = []
        gap_down_stocks = []
        flat_count = 0
        for sym in self.fno.fno_stocks:
            try:
                key = f"NSE:{sym}"
                d = ohlc_data.get(key, {})
                ltp = d.get("last_price", 0)
                ohlc = d.get("ohlc", {})
                prev_close = ohlc.get("close", 0)
                if ltp <= 0 or prev_close <= 0:
                    continue
                gap_pct = ((ltp - prev_close) / prev_close) * 100
                tag = "FLAT"
                if gap_pct >= GAP_STOCK_THRESHOLD:
                    tag = "GAP_UP"
                    gap_up_stocks.append(f"{sym} {gap_pct:+.1f}%")
                elif gap_pct <= -GAP_STOCK_THRESHOLD:
                    tag = "GAP_DOWN"
                    gap_down_stocks.append(f"{sym} {gap_pct:+.1f}%")
                else:
                    flat_count += 1
                gap_status["stocks"][sym] = {"pct": round(gap_pct, 2), "tag": tag, "ltp": ltp, "prev_close": prev_close}
            except Exception:
                continue
        try:
            save_json(self.gap_file, gap_status)
        except Exception as _e:
            log.error(f"[PATCH_V6_GAP] save failed: {_e}")
        log.info(f"[PATCH_V6_GAP] gap detection complete: {len(gap_up_stocks)} GAP_UP, {len(gap_down_stocks)} GAP_DOWN, {flat_count} FLAT")
        if gap_up_idx:
            log.info(f"[PATCH_V6_GAP] indices GAP_UP: {', '.join(gap_up_idx)}")
        if gap_down_idx:
            log.info(f"[PATCH_V6_GAP] indices GAP_DOWN: {', '.join(gap_down_idx)}")
        if gap_up_stocks:
            log.info(f"[PATCH_V6_GAP] stocks GAP_UP ({len(gap_up_stocks)}): {', '.join(gap_up_stocks[:10])}{' ...' if len(gap_up_stocks)>10 else ''}")
        if gap_down_stocks:
            log.info(f"[PATCH_V6_GAP] stocks GAP_DOWN ({len(gap_down_stocks)}): {', '.join(gap_down_stocks[:10])}{' ...' if len(gap_down_stocks)>10 else ''}")
        return gap_status

    def review_pending_amos(self, gap_status):
        """
        Review all pending AMOs (BUY orders, equity + F&O).
        Cancel any that contradict gap direction.
        For equity: check stock gap.
        For F&O: extract underlying from tradingsymbol, check stock gap.
        """
        kite = KiteSession.kite()
        if not kite:
            return
        cancelled = []
        try:
            orders = kite.orders()
        except Exception as _e:
            log.error(f"[PATCH_V6_GAP] orders fetch failed: {_e}")
            return
        for o in orders:
            try:
                if o.get("variety") != "amo":
                    continue
                if o.get("status") not in ("OPEN", "TRIGGER PENDING", "AMO REQ RECEIVED"):
                    continue
                if o.get("transaction_type") != "BUY":
                    continue
                tsym = o.get("tradingsymbol", "")
                exch = o.get("exchange", "")
                order_id = o.get("order_id", "")
                # Determine underlying symbol
                if exch == "NSE":
                    underlying = tsym
                    is_call = True  # equity is always bullish
                elif exch == "NFO":
                    # Extract underlying from F&O tradingsymbol like NIFTY26APR25150CE
                    is_call = tsym.endswith("CE")
                    is_put = tsym.endswith("PE")
                    if not (is_call or is_put):
                        continue
                    # Strip CE/PE and try to match underlying from fno_stocks
                    stripped = tsym[:-2]
                    underlying = None
                    for s in sorted(self.fno.fno_stocks, key=len, reverse=True):
                        if stripped.startswith(s):
                            underlying = s
                            break
                    if not underlying:
                        # Index check
                        for idx in OI_INDICES:
                            if stripped.startswith(idx):
                                underlying = idx
                                break
                    if not underlying:
                        continue
                else:
                    continue
                # Look up gap status
                gap_info = None
                if underlying in gap_status.get("stocks", {}):
                    gap_info = gap_status["stocks"][underlying]
                elif underlying in gap_status.get("indices", {}):
                    gap_info = gap_status["indices"][underlying]
                if not gap_info:
                    continue
                tag = gap_info.get("tag", "FLAT")
                pct = gap_info.get("pct", 0)
                # Determine if AMO contradicts gap
                should_cancel = False
                reason = ""
                if exch == "NSE":
                    # Equity BUY contradicts only if GAP_DOWN
                    if tag == "GAP_DOWN":
                        should_cancel = True
                        reason = f"equity BUY but {underlying} GAP_DOWN {pct:+.2f}%"
                    # PATCH_V28_STRANDED_GAPUP: cancel equity BUY AMO when stock gapped UP
                    # above limit price by more than 2% (SURAJEST-class stranding)
                    elif tag == "GAP_UP":
                        try:
                            _limit_price = float(o.get("price", 0))
                            _ltp = float(gap_info.get("open", 0))
                            if _limit_price > 0 and _ltp > 0:
                                _stranded_pct = ((_ltp - _limit_price) / _limit_price) * 100
                                if _stranded_pct >= 2.0:
                                    should_cancel = True
                                    reason = f"equity BUY stranded: limit Rs.{_limit_price:.2f} vs open Rs.{_ltp:.2f} ({_stranded_pct:+.2f}%)"
                        except Exception as _stranded_err:
                            log.debug(f"[PATCH_V28_STRANDED] check failed for {underlying}: {_stranded_err}")
                    # END_PATCH_V28_STRANDED_GAPUP
                else:
                    # F&O: CE BUY contradicts GAP_DOWN, PE BUY contradicts GAP_UP
                    if is_call and tag == "GAP_DOWN":
                        should_cancel = True
                        reason = f"CE BUY but {underlying} GAP_DOWN {pct:+.2f}%"
                    elif is_put and tag == "GAP_UP":
                        should_cancel = True
                        reason = f"PE BUY but {underlying} GAP_UP {pct:+.2f}%"
                if should_cancel:
                    try:
                        kite.cancel_order(variety="amo", order_id=order_id)
                        log.warning(f"[PATCH_V6_GAP] CANCELLED AMO {tsym} id={order_id}: {reason}")
                        cancelled.append({"tsym": tsym, "reason": reason})
                    except Exception as _ce:
                        log.error(f"[PATCH_V6_GAP] cancel failed {tsym}: {_ce}")
            except Exception as _e:
                log.debug(f"[PATCH_V6_GAP] AMO review per-order error: {_e}")
                continue
        if cancelled:
            try:
                msg_lines = [f"*GAP REVIEW: {len(cancelled)} AMOs CANCELLED*"]
                for c in cancelled[:10]:
                    msg_lines.append(f"`{c['tsym']}` — {c['reason']}")
                send_telegram("\n".join(msg_lines), silent=False)
            except Exception as _e72:
                log.warning(f"review_pending_amos: telegram failed: {_e72} [PATCH_V72_COMPREHENSIVE]")
        else:
            log.info("[PATCH_V6_GAP] AMO review: no contradictions found")


class OISnapshotEngine:
    """
    PATCH_V5_OI_LOGGER
    Captures option chain Open Interest snapshots for all F&O stocks + indices.
    Stores one JSON file per day, auto-deletes files older than OI_RETENTION_DAYS.
    Provides signal detection: long buildup, short buildup, unwinding.
    """

    def __init__(self, fno_module):
        self._v5_oi_log_cache = {}  # PATCH_V74A_OI_DEDUP: (sym,direction) -> (score, last_log_ts)
        self.fno = fno_module
        self.snapshot_dir = DATA_DIR / "oi_snapshots"
        self.snapshot_dir.mkdir(parents=True, exist_ok=True)
        # PATCH_V70_AUDIT_CLEANUP: removed unused self._last_snapshot_minute (never read)

    def _get_snapshot_file(self, d=None):
        d = d or today_ist()
        return self.snapshot_dir / f"oi_{d.isoformat()}.json"

    def _cleanup_old_snapshots(self):
        """Delete snapshot files older than OI_RETENTION_DAYS."""
        try:
            cutoff = today_ist() - timedelta(days=OI_RETENTION_DAYS)
            for f in self.snapshot_dir.glob("oi_*.json"):
                try:
                    fname = f.stem.replace("oi_", "")
                    fdate = datetime.strptime(fname, "%Y-%m-%d").date()
                    if fdate < cutoff:
                        f.unlink()
                        log.info(f"[PATCH_V5_OI] deleted old snapshot {f.name}")
                except Exception:
                    pass
        except Exception as _e:
            log.warning(f"[PATCH_V5_OI] cleanup failed: {_e}")

    def _round_to_strike(self, price, symbol):
        """Round spot price to the nearest valid strike for this symbol."""
        if symbol == "NIFTY" or symbol == "FINNIFTY":
            return round(price / 50) * 50
        if symbol == "BANKNIFTY":
            return round(price / 100) * 100
        if symbol == "MIDCPNIFTY":
            return round(price / 25) * 25
        if price < 100:
            return round(price / 2.5) * 2.5
        if price < 500:
            return round(price / 5) * 5
        if price < 1000:
            return round(price / 10) * 10
        if price < 2500:
            return round(price / 20) * 20
        return round(price / 50) * 50

    def _get_atm_strikes(self, symbol, spot):
        """Return list of ATM +/- OI_STRIKES_RANGE strikes that exist in NFO instruments."""
        atm = self._round_to_strike(spot, symbol)
        all_strikes = sorted({
            inst.get("strike", 0)
            for inst in self.fno._nfo_instruments
            if inst.get("name") == symbol and inst.get("instrument_type") in ("CE", "PE")
            and inst.get("strike", 0) > 0
        })
        if not all_strikes:
            return []
        try:
            atm_idx = min(range(len(all_strikes)), key=lambda i: abs(all_strikes[i] - atm))
        except Exception:
            return []
        lo = max(0, atm_idx - OI_STRIKES_RANGE)
        hi = min(len(all_strikes), atm_idx + OI_STRIKES_RANGE + 1)
        return all_strikes[lo:hi]

    def _get_current_expiry_instruments(self, symbol, strikes):
        """Get nearest-expiry CE+PE tradingsymbols for given strikes."""
        today = today_ist()
        candidates = []
        for inst in self.fno._nfo_instruments:
            if inst.get("name") != symbol:
                continue
            if inst.get("instrument_type") not in ("CE", "PE"):
                continue
            if inst.get("strike", 0) not in strikes:
                continue
            expiry = inst.get("expiry")
            if isinstance(expiry, str):
                try:
                    expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
                except Exception:
                    continue
            elif hasattr(expiry, "date"):
                expiry = expiry.date()
            dte = (expiry - today).days
            if dte < 0 or dte > 60:
                continue
            candidates.append({
                "tradingsymbol": inst["tradingsymbol"],
                "strike": inst["strike"],
                "instrument_type": inst["instrument_type"],
                "expiry": expiry.isoformat(),
                "dte": dte,
            })
        if not candidates:
            return []
        nearest_expiry = min(c["expiry"] for c in candidates)
        return [c for c in candidates if c["expiry"] == nearest_expiry]

    def take_snapshot(self, is_eod=False):
        """Hourly snapshot of OI for all symbols. Returns count of instruments captured.

        PATCH_V9_AMO_ARCH: is_eod=True flags the snapshot as end-of-day for
        day-over-day comparison in detect_buildup_signals(mode="daily").
        """
        kite = KiteSession.kite()
        if not kite:
            return 0
        all_symbols = list(OI_INDICES) + list(self.fno.fno_stocks)
        all_instruments = []
        symbol_meta = {}
        for sym in all_symbols:
            try:
                if sym in OI_INDICES:
                    spot_key = f"NSE:{sym} 50" if sym == "NIFTY" else f"NSE:NIFTY {sym.replace('NIFTY','').strip() or 'BANK'}"
                    if sym == "NIFTY":
                        spot_key = "NSE:NIFTY 50"
                    elif sym == "BANKNIFTY":
                        spot_key = "NSE:NIFTY BANK"
                    elif sym == "FINNIFTY":
                        spot_key = "NSE:NIFTY FIN SERVICE"
                    elif sym == "MIDCPNIFTY":
                        spot_key = "NSE:NIFTY MID SELECT"
                else:
                    spot_key = f"NSE:{sym}"
                spot_data = kite.ltp([spot_key])
                spot = spot_data.get(spot_key, {}).get("last_price", 0)
                if spot <= 0:
                    continue
                strikes = self._get_atm_strikes(sym, spot)
                if not strikes:
                    continue
                instruments = self._get_current_expiry_instruments(sym, strikes)
                if not instruments:
                    continue
                symbol_meta[sym] = {"spot": spot, "atm": self._round_to_strike(spot, sym), "instruments": instruments}
                all_instruments.extend(f"NFO:{i['tradingsymbol']}" for i in instruments)
            except Exception as _e:
                log.debug(f"[PATCH_V5_OI] {sym} prep failed: {_e}")
                continue
        if not all_instruments:
            log.warning("[PATCH_V5_OI] no instruments to snapshot")
            return 0
        # Batch quote calls (Kite limit: 250 per call)
        oi_data = {}
        try:
            for i in range(0, len(all_instruments), 250):
                batch = all_instruments[i:i+250]
                quotes = kite.quote(batch)
                if i > 0: time.sleep(0.3)  # FIX_V51: Kite rate limit between batches
                for k, v in quotes.items():
                    oi_data[k] = {
                        "ltp": v.get("last_price", 0),
                        "oi": v.get("oi", 0),
                        "volume": v.get("volume", 0),
                    }
        except Exception as _e:
            log.error(f"[PATCH_V5_OI] batch quote failed: {_e}")
            return 0
        # Build snapshot
        snapshot = {
            "timestamp": now_ist().isoformat(),
            "symbols": {},
        }
        # PATCH_V9_AMO_ARCH: mark EOD snapshots for day-over-day comparison
        if is_eod:
            snapshot["eod"] = True
        for sym, meta in symbol_meta.items():
            sym_entry = {"spot": meta["spot"], "atm": meta["atm"], "strikes": []}
            for inst in meta["instruments"]:
                key = f"NFO:{inst['tradingsymbol']}"
                q = oi_data.get(key, {})
                sym_entry["strikes"].append({
                    "strike": inst["strike"],
                    "type": inst["type"],
                    "ltp": q.get("ltp", 0),
                    "oi": q.get("oi", 0),
                    "volume": q.get("volume", 0),
                    "dte": inst["dte"],
                })
            snapshot["symbols"][sym] = sym_entry
        # Append to today's file
        snapshot_file = self._get_snapshot_file()
        try:
            existing = load_json(snapshot_file, {"date": str(today_ist()), "snapshots": []})
            existing["snapshots"].append(snapshot)
            save_json(snapshot_file, existing)
        except Exception as _e:
            log.error(f"[PATCH_V5_OI] save failed: {_e}")
            return 0
        # Periodic cleanup
        if now_ist().hour == 16:
            self._cleanup_old_snapshots()
        _eod_tag = " [EOD]" if is_eod else ""
        log.info(f"[PATCH_V5_OI] snapshot saved{_eod_tag}: {len(symbol_meta)} symbols, {len(all_instruments)} instruments, file={snapshot_file.name}")
        return len(all_instruments)

    def _find_recent_eod_snapshots(self, need=2):
        """
        PATCH_V9_AMO_ARCH: walk snapshot_dir, find the most recent `need` snapshots
        marked eod=True. Returns list ordered newest-first. May return empty list.
        """
        results = []
        try:
            files = sorted(self.snapshot_dir.glob("oi_*.json"), reverse=True)
            for f in files:
                try:
                    data = load_json(f, {"snapshots": []})
                    for snap in reversed(data.get("snapshots", [])):
                        if snap.get("eod") is True:
                            results.append(snap)
                            if len(results) >= need:
                                return results
                except Exception as _e:
                    log.debug(f"[PATCH_V9_AMO_ARCH] eod scan skipped {f.name}: {_e}")
                    continue
        except Exception as _e:
            log.warning(f"[PATCH_V9_AMO_ARCH] _find_recent_eod_snapshots failed: {_e}")
        return results

    def detect_buildup_signals(self, mode="intraday"):
        """
        Detect OI buildup signals.

        mode="intraday" (default, unchanged behavior):
          Compare latest 2 snapshots in today's file (~1h apart).
        mode="daily" (PATCH_V9_AMO_ARCH):
          Compare two most recent EOD snapshots across date files.
          Useful in overnight AMO scan window when intraday deltas are stale.

        Patterns detected in both modes:
          - LONG BUILDUP: spot up + total CE OI up significantly  -> BULLISH
          - SHORT BUILDUP: spot down + total PE OI up significantly -> BEARISH
          - LONG UNWINDING: spot down + total CE OI down -> BEARISH (weaker)
          - SHORT COVERING: spot up + total PE OI down -> BULLISH (weaker)

        Returns list of candidates compatible with F&O scanner format.
        """
        candidates = []
        try:
            if mode == "daily":
                # PATCH_V9_AMO_ARCH: day-over-day using EOD snapshots
                eod_snaps = self._find_recent_eod_snapshots(need=2)
                if len(eod_snaps) < 2:
                    log.debug(f"[PATCH_V9_AMO_ARCH] daily mode: only {len(eod_snaps)} EOD snapshots available, need 2")
                    return candidates
                latest = eod_snaps[0]
                prev = eod_snaps[1]
            else:
                # FIX_V76_OI_INTRADAY: bridge gap when today has only 1 snapshot.
                # OLD: required >=2 today snaps -> dead first 1-2 hrs of market.
                # NEW: if 1 today snap, fall back to yesterday's EOD as `prev`.
                snapshot_file = self._get_snapshot_file()
                data = load_json(snapshot_file, {"snapshots": []})
                snaps = data.get("snapshots", [])
                if len(snaps) >= 2:
                    latest = snaps[-1]
                    prev = snaps[-2]
                elif len(snaps) == 1:
                    try:
                        eod_snaps = self._find_recent_eod_snapshots(need=1)
                    except Exception as _e_v76b:
                        log.warning(f"[FIX_V76_OI_INTRADAY] EOD lookup failed: {_e_v76b}")
                        return candidates
                    if not eod_snaps:
                        log.info("[FIX_V76_OI_INTRADAY] only 1 today snap, no EOD fallback")
                        return candidates
                    latest = snaps[-1]
                    prev = eod_snaps[0]
                    log.info("[FIX_V76_OI_INTRADAY] bridging: today snap vs yesterday EOD")
                else:
                    log.debug("[PATCH_V5_OI] no snapshots in today's file")
                    return candidates
            for sym, latest_data in latest.get("symbols", {}).items():
                prev_data = prev.get("symbols", {}).get(sym)
                if not prev_data:
                    continue
                spot_now = latest_data.get("spot", 0)
                spot_prev = prev_data.get("spot", 0)
                if spot_now <= 0 or spot_prev <= 0:
                    continue
                spot_change_pct = ((spot_now - spot_prev) / spot_prev) * 100
                ce_oi_now = sum(s["oi"] for s in latest_data.get("strikes", []) if s["type"] == "CE")
                pe_oi_now = sum(s["oi"] for s in latest_data.get("strikes", []) if s["type"] == "PE")
                ce_oi_prev = sum(s["oi"] for s in prev_data.get("strikes", []) if s["type"] == "CE")
                pe_oi_prev = sum(s["oi"] for s in prev_data.get("strikes", []) if s["type"] == "PE")
                if ce_oi_prev <= 0 or pe_oi_prev <= 0:
                    continue
                ce_oi_change_pct = ((ce_oi_now - ce_oi_prev) / ce_oi_prev) * 100
                pe_oi_change_pct = ((pe_oi_now - pe_oi_prev) / pe_oi_prev) * 100
                pattern = None
                direction = None
                score = 0
                if spot_change_pct > 0.3 and pe_oi_change_pct > OI_BUILDUP_MIN_PCT:
                    pattern = "PUT_WRITING"
                    direction = "BULLISH"
                    score = min(95, 60 + int(pe_oi_change_pct / 2) + int(spot_change_pct * 5))
                elif spot_change_pct < -0.3 and ce_oi_change_pct > OI_BUILDUP_MIN_PCT:
                    pattern = "CALL_WRITING"
                    direction = "BEARISH"
                    score = min(95, 60 + int(ce_oi_change_pct / 2) + int(abs(spot_change_pct) * 5))
                elif spot_change_pct > 0.3 and ce_oi_change_pct < -OI_BUILDUP_MIN_PCT:
                    pattern = "SHORT_COVERING"
                    direction = "BULLISH"
                    score = min(85, 55 + int(abs(ce_oi_change_pct) / 2))
                elif spot_change_pct < -0.3 and pe_oi_change_pct < -OI_BUILDUP_MIN_PCT:
                    pattern = "LONG_UNWINDING"
                    direction = "BEARISH"
                    score = min(85, 55 + int(abs(pe_oi_change_pct) / 2))
                if pattern and direction and score >= OI_SIGNAL_MIN_SCORE:
                    _src = "OI_BUILDUP_DAILY" if mode == "daily" else "OI_BUILDUP"
                    _tag = "DAILY" if mode == "daily" else ""
                    candidates.append({
                        "symbol": sym,
                        "direction": direction,
                        "score": score,
                        "catalyst": f"OI {_tag} {pattern}: spot {spot_change_pct:+.2f}% CE_OI {ce_oi_change_pct:+.0f}% PE_OI {pe_oi_change_pct:+.0f}%".strip(),
                        "source": _src,
                    })
                    # PATCH_V74A_OI_DEDUP: skip if same sym/direction/score logged in last 30 min
                    _v74_key = (sym, direction)
                    _v74_now = time.time()
                    _v74_last = self._v5_oi_log_cache.get(_v74_key, (None, 0))
                    if _v74_last[0] != score or (_v74_now - _v74_last[1]) >= 1800:
                        log.info(f"[PATCH_V5_OI] signal ({mode}): {sym} {direction} {pattern} score={score}")
                        self._v5_oi_log_cache[_v74_key] = (score, _v74_now)
        except Exception as _e:
            log.error(f"[PATCH_V5_OI] detect_buildup failed: {_e}")
        return candidates


class FnoModule:
    """
    NFO NRML options trading.

    Entry: News catalyst → determine CE/PE → check F&O eligibility → smart strike
    Exit:  Trailing SL (activate +20%, trail 15%) + DTE cascade + GTT SL at -30%
    Protection: SL orders re-placed every morning at 9:16

    NO equity-style regime filter. Catalyst strength is the only entry gate.
    Own news scanner, own GTT manager, own risk manager, own capital.
    Zero awareness of equity module.
    """

    def __init__(self):
        self.positions = {}     # key -> position dict
        self.gtt = GTTManager("FNO")
        self.risk = None
        self.telegram = TelegramThrottle("FNO", max_per_hour=15)
        self.news = None
        self.filings = None     # FilingMonitor (interprets for F&O direction)
        self.brokerage = None   # BrokerageMonitor
        self.macro = None       # MacroDetector (RBI, war, oil → CE/PE on stocks)
        # PATCH_V63_INDEX_ROUTING: dedicated index news detector (NIFTY/BANKNIFTY/FINNIFTY)
        self.index_macro = None
        self.earnings = EarningsCalendar()
        # PATCH_V62_ECON_CALENDAR: hard-block F&O entries near high-impact macro events
        self.econ = EconCalendar(block_hours_before=2.0, block_hours_after=1.0)
        # PATCH_V62_PREMARKET: overnight global bias (GIFT Nifty + US futures + crude + DXY)
        self.premarket = GlobalPreMarket()
        self.capital = 0
        self.available = 0
        self.fno_stocks = set()
        self._nfo_instruments = []
        self._pos_file = DATA_DIR / "fno_positions.json"
        self._trades_file = DATA_DIR / "fno_trades.json"
        self._last_scan = None
        self._last_trail_check = None
        self._oi_engine = None  # PATCH_V5_OI_LOGGER: initialized in init()
        self._last_oi_snapshot_hour = -1
        self._gap_detector = None  # PATCH_V6_GAP: initialized in init()
        self._last_gap_check_day = None
        self._sector_rotation = None  # PATCH_V8_SECTOR_ROTATION: initialized in init()
        self.trendlyne = TrendlyneScorer()  # FIX_V51: F&O panel needs trendlyne (was crashing)

        # ================================================================
        # PATCH_V61_FNO_AUDIT_THREAD: F&O GTT audit on background thread
        # Root cause fix for Apr 23-24 frozen trail bug:
        #   NewsBrain blocks main tick loop 3-10 min on slow Gemini days.
        #   While blocked, hourly _fno_gtt_audit at :30 silently skips.
        #   HCLTECH/TRENT trail froze because audit never ran.
        # Fix: worker thread runs _fno_gtt_audit() every 10 sec forever.
        # Event call sites still fire (fast reaction). Lock prevents overlap.
        # ================================================================
        self._fno_audit_lock = threading.Lock()
        self._fno_audit_stop = threading.Event()
        self._fno_audit_thread = None  # spawned in init(), not __init__

    # ============================================================
    # PATCH_V1_FNO_AMO: Continuous overnight scanner + morning placer
    # Scans 16:01 -> 08:45 IST hourly, places AMO at 08:45
    # ============================================================
    def _fno_amo_candidates_file(self):
        return DATA_DIR / "fno_amo_candidates.json"

    def _fno_amo_load_candidates(self):
        try:
            data = load_json(self._fno_amo_candidates_file(), {})
            if data.get("date") == str(today_ist()):
                return data.get("candidates", [])
        except Exception:
            pass
        return []

    def _fno_amo_save_candidates(self, candidates):
        try:
            save_json(self._fno_amo_candidates_file(), {
                "date": str(today_ist()),
                "candidates": candidates,
                "updated": str(now_ist()) if "now_ist" in globals() else "",
            })
        except Exception as _e:
            log.error(f"[PATCH_V1_FNO_AMO] save candidates failed: {_e}")

    def _fno_amo_in_scan_window(self):
        """True between 16:01 and 08:45 IST (next day)."""
        try:
            from datetime import datetime
            try:
                from zoneinfo import ZoneInfo
                now = datetime.now(ZoneInfo("Asia/Kolkata"))
            except Exception:
                # V77_FIX9: refuse to act on UTC fallback
                log.error("[V77_FIX9] zoneinfo unavailable - refusing AMO scan window check")
                return False
            hm = now.hour * 60 + now.minute
            # 16:01 = 961, 08:45 = 525, 24:00 = 1440
            return hm >= 961 or hm <= 525
        except Exception:
            return False

    def _fno_amo_is_premarket_place_time(self):
        """True only between 08:45 and 08:55 IST."""
        try:
            from datetime import datetime
            try:
                from zoneinfo import ZoneInfo
                now = datetime.now(ZoneInfo("Asia/Kolkata"))
            except Exception:
                # V77_FIX9: refuse to act on UTC fallback
                log.error("[V77_FIX9] zoneinfo unavailable - refusing AMO place window check")
                return False
            hm = now.hour * 60 + now.minute
            return 525 <= hm <= 535
        except Exception:
            return False

    # ============================================================
    # PATCH_V55_FNO_AMO_STALE_CANCEL: cancel unfilled F&O AMOs at 09:20-09:30 IST
    # Mirrors equity AMOEngine.cancel_stale() pattern.
    # Fixes: Apr 30 2026 UNIONBANK 26MAY 145 PE AMO sat unfilled 90+ min
    # holding an FNO_MAX_POS slot hostage.
    # ============================================================
    def _fno_amo_stale_cancel_window(self):
        """True between 09:20 and 09:30 IST (mirrors equity stale window)."""
        try:
            from datetime import datetime
            try:
                from zoneinfo import ZoneInfo
                now = datetime.now(ZoneInfo("Asia/Kolkata"))
            except Exception:
                # V77_FIX9: refuse to act on UTC fallback
                log.error("[V77_FIX9] zoneinfo unavailable - refusing AMO stale window check")
                return False
            hm = now.hour * 60 + now.minute
            # 09:20 = 560, 09:30 = 570
            return 560 <= hm <= 570
        except Exception:
            return False

    def _fno_amo_cancel_stale(self):
        """Cancel pending F&O AMOs that did not fill at market open.

        PATCH_V55: Fires once per day in 09:20-09:30 IST window.
        Filters: variety=amo, exchange in NFO/BFO,
                 status in OPEN / TRIGGER PENDING / AMO REQ RECEIVED.
        """
        # Window guard
        if not self._fno_amo_stale_cancel_window():
            return 0

        # V78_FIX_AMO_TRADING_DAY: trading-day guard (PATCH_V55 forgot this — fired on Sat May 2 2026)
        if not is_trading_day():
            log.info("[V78_FIX_AMO_TRADING_DAY] FNO AMO stale cleanup skipped — non-trading day")
            return 0

        # Idempotency marker (once per day)
        marker_file = DATA_DIR / "fno_amo_stale_cancel_today.json"
        try:
            existing_marker = load_json(marker_file, {})
            if existing_marker.get("date") == str(today_ist()):
                return 0
        except Exception:
            pass

        kite = KiteSession.kite()
        if not kite:
            log.warning("[PATCH_V55_FNO_STALE] no kite session, skipping")
            return 0

        cancelled = []
        examined = 0
        try:
            orders = kite.orders() or []
            for o in orders:
                # Triple-filter: must be AMO + F&O exchange + pending status
                if (o.get("variety") == "amo"
                    and o.get("exchange") in ("NFO", "BFO")
                    and o.get("status") in ("OPEN", "TRIGGER PENDING",
                                            "AMO REQ RECEIVED",
                                            "AMO REQUEST RECEIVED",
                                            "PUT ORDER REQ RECEIVED")):
                    examined += 1
                    tsym = o.get("tradingsymbol", "?")
                    oid = o.get("order_id", "?")
                    qty = o.get("quantity", 0)
                    px = o.get("price", 0)
                    try:
                        kite.cancel_order(variety="amo", order_id=oid)
                        cancelled.append({"tsym": tsym, "qty": qty, "price": px, "id": oid})
                        log.warning(f"[PATCH_V55_FNO_STALE] CANCELLED stale F&O AMO {tsym} "
                                    f"qty={qty} px={px} id={oid}")
                    except Exception as _ce:
                        log.error(f"[PATCH_V55_FNO_STALE] cancel failed for {tsym} id={oid}: {_ce}")
        except Exception as _e:
            log.error(f"[PATCH_V55_FNO_STALE] orders fetch failed: {_e}")
            return 0

        # Mark done for the day even if zero cancelled (don't re-run all minute window)
        try:
            save_json(marker_file, {
                "date": str(today_ist()),
                "examined": examined,
                "cancelled": len(cancelled),
                "details": cancelled,
            })
        except Exception:
            pass

        if cancelled:
            log.info(f"[PATCH_V55_FNO_STALE] cancelled {len(cancelled)} stale F&O AMO(s) "
                     f"out of {examined} examined")
            # Telegram notification
            try:
                if hasattr(self, "telegram") and self.telegram:
                    msg_lines = []
                    msg_lines.append("*F&O AMO STALE CLEANUP*")
                    msg_lines.append(f"{len(cancelled)} unfilled F&O AMO(s) cancelled at market open:")
                    for c in cancelled:
                        msg_lines.append(f"  - {c['tsym']} x{c['qty']} @ Rs.{c['price']}")
                    msg_lines.append(f"FNO slots freed: {len(cancelled)}/{FNO_MAX_POS}")
                    self.telegram.send("\n".join(msg_lines), silent=False)
            except Exception as _te:
                log.debug(f"[PATCH_V55_FNO_STALE] telegram notify failed: {_te}")
        else:
            log.info(f"[PATCH_V55_FNO_STALE] no stale F&O AMOs found (examined {examined})")

        return len(cancelled)

    def _fno_amo_continuous_scan(self, threshold=65):
        """Hourly scan during 16:01-08:45 window. Accumulates catalysts.

        PATCH_V9_AMO_ARCH:
          - threshold param (default 65, was hardcoded 75) matches OI_SIGNAL_MIN_SCORE
          - adds daily-mode OI detection alongside intraday
        
        FIX_V33_EARNINGS_TRADER: earnings-driven candidates.
        """
        if not self._fno_amo_in_scan_window():
            return
        try:
            existing = self._fno_amo_load_candidates()
            existing_keys = {(c.get("symbol"), c.get("direction")) for c in existing}

            fresh = []
            if getattr(self, "filings", None):
                try: fresh.extend(self.filings.scan() or [])
                except Exception as _e: log.warning(f"[PATCH_V1_FNO_AMO] filings: {_e}")
            if getattr(self, "news", None):
                try: fresh.extend(self.news.scan() or [])
                except Exception as _e: log.warning(f"[PATCH_V1_FNO_AMO] news: {_e}")
            if getattr(self, "brokerage", None):
                try: fresh.extend(self.brokerage.scan(news_candidates=fresh) or [])
                except Exception as _e: log.warning(f"[PATCH_V1_FNO_AMO] brokerage: {_e}")
                try:
                    if self._oi_engine:
                        fresh.extend(self._oi_engine.detect_buildup_signals() or [])
                        try:
                            fresh.extend(self._oi_engine.detect_buildup_signals(mode="daily") or [])
                        except Exception as _e2:
                            log.warning(f"[PATCH_V9_AMO_ARCH] daily OI detect: {_e2}")
                except Exception as _e: log.warning(f"[PATCH_V1_FNO_AMO] brokerage: {_e}")
            # PATCH_V63_INDEX_ROUTING: add macro + index_macro to AMO continuous scan
            # Previously AMO missed macro news entirely. Big gap since Fed/RBI fire overnight.
            if getattr(self, "macro", None):
                try: fresh.extend(self.macro.scan() or [])
                except Exception as _e: log.warning(f"[PATCH_V63] AMO macro: {_e}")
            if getattr(self, "index_macro", None):
                try: fresh.extend(self.index_macro.scan() or [])
                except Exception as _e: log.warning(f"[PATCH_V63] AMO index_macro: {_e}")

            # PATCH_V72_COMPREHENSIVE: AMO now uses V68 LLM-decide for macro article-items
            # Previously these would all be threshold-filtered out (score=0). Now they're
            # decided by LLM with same logic as intraday F&O.
            _v68_amo_articles = [c for c in fresh if c.get("v68_article_item")]
            _legacy_amo = [c for c in fresh if not c.get("v68_article_item")]
            fresh = _legacy_amo
            if _v68_amo_articles and getattr(self, "news_brain", None) and hasattr(self.news_brain, "read_and_decide_parallel"):
                _amo_pairs = [
                    ({"title": a.get("title",""), "summary": a.get("summary",""),
                      "link": a.get("link",""), "source_feed": a.get("source_feed", a.get("source",""))},
                     a.get("hints", []),
                     a.get("source", "MACRO"))
                    for a in _v68_amo_articles
                ]
                try:
                    _amo_decided = self.news_brain.read_and_decide_parallel(_amo_pairs, max_workers=5)
                    log.info(f"[PATCH_V72_COMPREHENSIVE] AMO LLM-decide: {len(_v68_amo_articles)} articles -> {len(_amo_decided)} candidates")
                    fresh.extend(_amo_decided)
                except Exception as _ve:
                    log.error(f"[PATCH_V72_COMPREHENSIVE] AMO read_and_decide_parallel failed: {_ve}")

            # FIX_V33_EARNINGS_TRADER: earnings-driven candidates
            try:
                ef = self._earnings_amo_scan(news_candidates=fresh)
                if ef:
                    log.info(f"[FIX_V33_EARNINGS] {len(ef)} earnings candidates added")
                    fresh.extend(ef)
            except Exception as _ee:
                log.warning(f"[FIX_V33_EARNINGS] scan failed: {_ee}")

            added = 0
            for c in fresh:
                sym = c.get("symbol")
                direction = c.get("direction")
                score = c.get("score", 0)
                if not sym or direction not in ("BULLISH", "BEARISH"):
                    continue
                if sym not in self.fno_stocks:
                    continue
                if score < threshold:
                    continue
                key = (sym, direction)
                if key in existing_keys:
                    # Update score if higher
                    for e in existing:
                        if e.get("symbol") == sym and e.get("direction") == direction:
                            if score > e.get("score", 0):
                                e["score"] = score
                                e["catalyst"] = c.get("catalyst", "")[:120]
                                e["updated"] = str(now_ist()) if "now_ist" in globals() else ""
                            break
                else:
                    existing.append({
                        "symbol": sym,
                        "direction": direction,
                        "score": score,
                        "catalyst": c.get("catalyst", "")[:120],
                        "added": str(now_ist()) if "now_ist" in globals() else "",
                    })
                    existing_keys.add(key)
                    added += 1

            self._fno_amo_save_candidates(existing)
            log.info(f"[PATCH_V1_FNO_AMO] scan complete: {added} new, {len(existing)} total in pool (threshold={threshold})")
        except Exception as _e:
            log.error(f"[PATCH_V1_FNO_AMO] continuous scan failed: {_e}")

    def _earnings_amo_scan(self, news_candidates=None):
        """FIX_V33_EARNINGS_TRADER: earnings-driven candidates into AMO pool."""
        out = []
        if not self.earnings:
            return out
        # FIX_V82_THROTTLE: earnings calendar doesn't change minute-to-minute.
        # Per-minute AMO scanner was re-running full calendar scan every cycle,
        # wasting Trendlyne mcap lookups, IVR fetches, scoring CPU, log spam.
        # Throttle to once per hour. Still runs at startup (last_ts=0).
        _now_ts = time.time()
        _last_ts = getattr(self, "_last_earnings_scan_ts", 0)
        if _now_ts - _last_ts < 3600:
            return out
        self._last_earnings_scan_ts = _now_ts
        try:
            news_dirs = {}
            if news_candidates:
                for c in news_candidates:
                    s = c.get("symbol"); d = c.get("direction")
                    if s and d in ("BULLISH", "BEARISH"):
                        news_dirs.setdefault(s, []).append(d)
            today = today_ist()
            kite = KiteSession.kite()
            _nse_cache = None
            for sym, info in self.earnings._calendar.items():
                if sym not in self.fno_stocks:
                    continue
                if sym in ("NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"):
                    continue
                date_str = info.get("date", "")
                earn_date = None
                for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        earn_date = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if not earn_date:
                    continue
                days_to = (earn_date - today).days
                if not (3 <= days_to <= 7):
                    continue
                direction = None
                src_lbl = ""
                nd = news_dirs.get(sym, [])
                if nd and not ("BULLISH" in nd and "BEARISH" in nd):
                    direction = nd[0]
                    src_lbl = "news"
                # V81: PRIMARY direction source — OI buildup (smart-money signal)
                if not direction:
                    try:
                        if hasattr(self, "_oi_engine") and self._oi_engine:
                            _v81_oi = self._oi_engine.detect_buildup_signals(mode="daily") or []
                            for _o in _v81_oi:
                                if _o.get("symbol") == sym and _o.get("direction") in ("BULLISH", "BEARISH"):
                                    direction = _o["direction"]
                                    src_lbl = f"oi_buildup score={_o.get('score',0)}"
                                    break
                    except Exception as _v81oe:
                        log.debug(f"[V81_EARNINGS_OI] {sym} OI lookup failed: {_v81oe}")
                # V81: FALLBACK direction source — 20-SMA crossover (kept as safety net
                # for early-morning case when OI snapshots haven't accumulated yet)
                if not direction and kite:
                    try:
                        today_dt = datetime.combine(today, datetime.min.time())
                        from_dt = today_dt - timedelta(days=40)
                        tok = None
                        try:
                            if _nse_cache is None:
                                _nse_cache = kite.instruments("NSE")
                            for inst in _nse_cache:
                                if inst.get("tradingsymbol") == sym and inst.get("segment") == "NSE":
                                    tok = inst.get("instrument_token")
                                    break
                        except Exception:
                            pass
                        if tok:
                            h = kite.historical_data(tok, from_dt, today_dt, "day")
                            if h and len(h) >= 20:
                                rc = h[-1].get("close", 0)
                                sma = sum(x.get("close", 0) for x in h[-20:]) / 20
                                if rc > sma * 1.01:
                                    direction = "BULLISH"
                                    src_lbl = f"sma20_fallback {rc:.1f}>{sma:.1f}"
                                elif rc < sma * 0.99:
                                    direction = "BEARISH"
                                    src_lbl = f"sma20_fallback {rc:.1f}<{sma:.1f}"
                    except Exception as _te:
                        log.debug(f"[FIX_V33_EARNINGS] {sym} trend: {_te}")
                if not direction:
                    log.info(f"[FIX_V33_EARNINGS] {sym} skipped: no direction")
                    continue
                ivr = None
                iv_boost = 0
                try:
                    ivr = iv_tracker.iv_rank(sym)
                except Exception:
                    ivr = None
                if ivr is not None:
                    if ivr > 70:
                        log.info(f"[FIX_V33_EARNINGS] {sym} skipped: IVR={ivr:.0f}>70")
                        continue
                    if ivr < 30:
                        iv_boost = 5
                    elif ivr < 50:
                        iv_boost = 2
                score = 70 + (7 - days_to) * 3 + iv_boost
                # V81: tier-weighted boost (research: pre-earnings drift strongest in smaller caps)
                _v81_mcap = 0
                try:
                    _tl = getattr(self, "trendlyne", None) or globals().get("trendlyne")
                    if _tl and hasattr(_tl, "_data"):
                        _v81_mcap = _tl._data.get(sym, {}).get("market_cap", 0)
                except Exception:
                    pass
                if _v81_mcap and _v81_mcap < 5000:
                    score += 5      # small cap: biggest drift edge
                    src_lbl += " smallcap+5"
                elif _v81_mcap and _v81_mcap < 50000:
                    score += 3      # mid cap
                    src_lbl += " midcap+3"
                # large cap (>=50k Cr) and unknown (mcap=0): +0
                score = min(95, max(70, score))  # raised cap from 90 to 95 to allow boost
                out.append({
                    "symbol": sym,
                    "direction": direction,
                    "score": score,
                    "catalyst": f"EARNINGS {days_to}d dir={direction} ivr={ivr} {src_lbl}",
                    "source": "EARNINGS_TRADER_V33",
                })
                log.info(f"[FIX_V33_EARNINGS] queued {sym} {direction} score={score} days={days_to} ivr={ivr} mcap={_v81_mcap}")
        except Exception as e:
            log.error(f"[FIX_V33_EARNINGS] scan failed: {e}")
        return out

    def _fno_amo_premarket_place(self):
        """At 08:45 IST: read pool, apply GIFT Nifty filter, place top 3 AMOs.

        PATCH_V9_AMO_ARCH: if pool is empty, run one final aggressive scan at
        threshold 60 before giving up.
        PATCH_V28_FALLBACK_ONCE: fallback scan must fire only ONCE per day, not every minute.
        """
        try:
            if self._fno_amo_is_premarket_place_time():
                # PATCH_V28_FALLBACK_ONCE: idempotency marker for fallback scan
                _fb_marker = DATA_DIR / "fno_amo_fallback_today.json"
                _fb_existing = load_json(_fb_marker, {})
                if _fb_existing.get("date") != str(today_ist()):
                    _pool_check = self._fno_amo_load_candidates()
                    if not _pool_check:
                        log.info("[PATCH_V9_AMO_ARCH] pool empty at 08:45, running fallback scan at threshold=60 (ONCE)")
                        self._fno_amo_continuous_scan(threshold=60)
                    save_json(_fb_marker, {"date": str(today_ist()), "ran": True})
        except Exception as _e:
            log.error(f"[PATCH_V9_AMO_ARCH] fallback scan failed: {_e}")

        if not self._fno_amo_is_premarket_place_time():
            return
        kite = KiteSession.kite()
        if not kite:
            return

        # PATCH_V62_PREMARKET: log full overnight global bias (additive to existing GIFT filter)
        try:
            if getattr(self, "premarket", None):
                _bias = self.premarket.read()
                if _bias:
                    log.info(f"[PATCH_V62_PREMARKET] overnight bias={_bias['bias']} score={_bias['score']:+.2f} "
                             f"items={_bias['items']}")
        except Exception as _pe:
            log.debug(f"[PATCH_V62_PREMARKET] skipped: {_pe}")

        try:
            # Idempotency: only run once per day
            marker_file = DATA_DIR / "fno_amo_placed_today.json"
            existing_marker = load_json(marker_file, {})
            if existing_marker.get("date") == str(today_ist()):
                return

            candidates = self._fno_amo_load_candidates()
            log.info(f"[PATCH_V73_DIAG] pool loaded: {len(candidates)} candidates")
            for _ci, _c in enumerate(candidates[:10]):
                log.info(f"[PATCH_V73_DIAG] cand[{_ci}]: {_c.get('symbol')} {_c.get('direction')} conv={_c.get('conv', _c.get('conviction', '?'))} src={_c.get('src', _c.get('sources', '?'))} tag={_c.get('tag', _c.get('agreement', '?'))}")
            if not candidates:
                log.info("[PATCH_V1_FNO_AMO] no candidates to place")
                save_json(marker_file, {"date": str(today_ist()), "placed": []})
                return

            # GIFT Nifty direction filter
            gift_pct = 0.0
            try:
                gift_data = kite.ltp(["NSEIX:GIFT NIFTY"])
                gift_ltp = gift_data.get("NSEIX:GIFT NIFTY", {}).get("last_price", 0)
                nifty_ohlc = kite.ohlc(["NSE:NIFTY 50"]).get("NSE:NIFTY 50", {}).get("ohlc", {})
                yest_close = nifty_ohlc.get("close", 0)
                if gift_ltp > 0 and yest_close > 0:
                    gift_pct = ((gift_ltp - yest_close) / yest_close) * 100
                    log.info(f"[PATCH_V1_FNO_AMO] GIFT Nifty implied gap: {gift_pct:+.2f}%")
            except Exception as _e:
                log.warning(f"[PATCH_V1_FNO_AMO] GIFT Nifty fetch failed: {_e}, no direction filter")

            # PATCH_V54: confluence-based threshold
            v62_score = 0.0
            try:
                if getattr(self, "premarket", None):
                    _b = self.premarket.read()
                    if _b: v62_score = _b.get("score", 0.0)
            except Exception as _v54e:
                log.debug(f"[PATCH_V54] V62 read failed: {_v54e}")

            if gift_pct < 0 and v62_score <= -0.4:
                bear_threshold = -0.3
                bull_threshold = 0.5
                log.info(f"[PATCH_V54] CONFLUENCE bearish: GIFT {gift_pct:+.2f}%, V62 {v62_score:+.2f} -> bear_thr=-0.3%")
            elif gift_pct > 0 and v62_score >= 0.4:
                bear_threshold = -0.5
                bull_threshold = 0.3
                log.info(f"[PATCH_V54] CONFLUENCE bullish: GIFT {gift_pct:+.2f}%, V62 {v62_score:+.2f} -> bull_thr=+0.3%")
            else:
                bear_threshold = -0.5
                bull_threshold = 0.5
                if abs(v62_score) >= 0.4 or abs(gift_pct) >= 0.3:
                    log.info(f"[PATCH_V54] NO CONFLUENCE: GIFT {gift_pct:+.2f}%, V62 {v62_score:+.2f} -> legacy +/-0.5%")

            filtered = []
            _V91_INDICES = ("NIFTY", "BANKNIFTY", "FINNIFTY")
            for c in candidates:
                # V79C_COOLDOWN_HARD: skip if recent SL hit on this underlying
                # PATCH_V91_FIX_A: indices are exempt
                _sym = c.get("symbol", "")
                try:
                    _lt = get_loss_tracker()
                    _recent = _lt.count_recent_losses(_sym, days=21)
                    if _recent >= 1:
                        if _sym in _V91_INDICES:
                            log.info(f"[PATCH_V91_INDEX_COOLDOWN_EXEMPT] {_sym}: {_recent} loss(es) in 21d but index - allowing entry")
                        else:
                            log.info(f"[V79C_COOLDOWN_HARD] SKIP {_sym}: {_recent} loss(es) in 21d")
                            continue
                except Exception as _v79ce:
                    log.debug(f"[V79C_COOLDOWN_HARD] check failed for {_sym}: {_v79ce}")

                d = c.get("direction")
                _conv = c.get("conv", c.get("conviction", 0))
                _src = c.get("src", c.get("sources", c.get("cross_sources", 0)))
                _tag = c.get("tag", c.get("agreement", ""))
                if _conv >= 80 and _src >= 4 and _tag == "AGREED":
                    log.info(f"[PATCH_V73_GIFT_BYPASS] {c.get('symbol')} {d} bypasses GIFT filter (conv={_conv} src={_src} AGREED)")
                    filtered.append(c)
                    continue
                # PATCH_V91_FIX_B: high-conviction index verdicts bypass GIFT
                if _conv >= 75 and _sym in _V91_INDICES:
                    log.info(f"[PATCH_V91_INDEX_BYPASS] {_sym} {d} bypasses GIFT (conv={_conv} index)")
                    filtered.append(c)
                    continue
                if gift_pct <= bear_threshold and d == "BULLISH":
                    log.info(f"[PATCH_V1_FNO_AMO] SKIP {c.get('symbol')} bullish (GIFT {gift_pct:+.2f}% <= {bear_threshold}%)")
                    continue
                if gift_pct >= bull_threshold and d == "BEARISH":
                    log.info(f"[PATCH_V1_FNO_AMO] SKIP {c.get('symbol')} bearish (GIFT {gift_pct:+.2f}% >= {bull_threshold}%)")
                    continue
                filtered.append(c)
            log.info(f"[PATCH_V73_DIAG] after GIFT filter: {len(filtered)}/{len(candidates)} survived (gift_pct={gift_pct:+.2f}%, bear_thr={bear_threshold:+.2f}%, bull_thr={bull_threshold:+.2f}%)")

            filtered.sort(key=lambda c: c.get("score", 0), reverse=True)

            # FIX_V42_FNO_AMO_CAP: mirror V28_A2 equity fix — count live FNO positions + pending AMOs
            try:
                current_fno = len(self.positions)
                pending_fno_amos = 0
                for o in (kite.orders() or []):
                    if (o.get("variety") == "amo"
                        and o.get("transaction_type") == "BUY"
                        and o.get("status") in ("OPEN", "TRIGGER PENDING", "AMO REQ RECEIVED")
                        and o.get("exchange") in ("NFO", "BFO")):
                        pending_fno_amos += 1
                free_slots = max(0, FNO_MAX_POS - current_fno - pending_fno_amos)
                log.info(f"[FIX_V42_FNO_AMO_CAP] FNO slots: live={current_fno} pending_amo={pending_fno_amos} free={free_slots}/{FNO_MAX_POS}")
                if free_slots <= 0:
                    log.info("[FIX_V42_FNO_AMO_CAP] no free FNO slots, skipping AMO placement entirely")
                    save_json(marker_file, {"date": str(today_ist()), "placed": []})
                    return
                top = filtered[:min(3, free_slots)]
            except Exception as _v42e:
                log.error(f"[FIX_V42_FNO_AMO_CAP] cap check failed: {_v42e} — falling back to top 3")
                top = filtered[:3]
            log.info(f"[PATCH_V1_FNO_AMO] {len(top)} candidates selected for AMO from pool of {len(candidates)}")

            placed_list = []
            for cand in top:
                sym = cand["symbol"]
                direction = cand["direction"]
                score = cand.get("score", 0)
                catalyst = cand.get("catalyst", "")[:60]
                try:
                    # PATCH_V73: use index-aware spot key
                    _sk = _index_spot_key(sym)
                    spot_data = kite.ltp([_sk])
                    spot = spot_data.get(_sk, {}).get("last_price", 0)
                    if spot <= 0:
                        log.warning(f"[PATCH_V1_FNO_AMO] {sym}: no spot price (key={_sk}), skip")
                        continue
                    option = self._find_best_option(kite, sym, direction, spot, catalyst_score=score)
                    if not option:
                        log.info(f"[PATCH_V1_FNO_AMO] {sym}: no suitable option for {direction}")
                        continue
                    tsym = option["tradingsymbol"]
                    lot = option["lot_size"]
                    premium = option["premium"]
                    # V83_EARNINGS_SIZE: 1.5-2x lots for earnings AMO entries.
                    # Empirical: earnings IV-crush plays were the only profitable F&O
                    # setup (EXIDEIND +Rs.3,240 alone). Sizing up captures more of
                    # that edge while the 50% available cap below still bounds risk.
                    _v83_src = (cand.get("source") or "").upper()
                    _v83_is_earn = ("EARNINGS" in _v83_src) or ("IV_CRUSH" in _v83_src)
                    if _v83_is_earn:
                        qty = max(lot, (int(lot * 1.5) // lot) * lot)
                        if qty == lot:
                            _try_qty = lot * 2
                            _try_cost = _try_qty * premium * 1.01
                            if _try_cost <= self.available * 0.5:
                                qty = _try_qty
                        log.info(f"[V83_EARNINGS_SIZE] {sym}: earnings source detected, qty={qty} (vs base lot={lot})")
                    else:
                        qty = lot
                    cost = qty * premium * 1.01
                    if cost > self.available * 0.5:
                        log.info(f"[PATCH_V1_FNO_AMO] {sym}: cost {cost:.0f} > 50% of available {self.available:.0f}, skip")
                        continue
                    limit_price = round(premium * 1.02, 1)
                    order_id = kite.place_order(
                        variety="amo",
                        exchange="NFO",
                        tradingsymbol=tsym,
                        transaction_type="BUY",
                        quantity=qty,
                        order_type="LIMIT",
                        price=limit_price,
                        product="NRML",
                        validity="DAY",
                    )
                    log.info(f"[PATCH_V1_FNO_AMO] PLACED {tsym} x{qty} @ {limit_price} id={order_id} score={score}")
                    placed_list.append({"tsym": tsym, "qty": qty, "price": limit_price, "score": score, "catalyst": catalyst})
                    try:
                        self.telegram.send(
                            f"*FNO AMO PLACED*\n"
                            f"{tsym} x{qty} @ Rs.{limit_price}\n"
                            f"Direction: {direction} | Score: {score}\n"
                            f"Catalyst: {catalyst}\n"
                            f"GIFT Nifty: {gift_pct:+.2f}%",
                            silent=False,
                        )
                        # FIX_V29_GTT_ON_ENTRY: immediately place GTT on fresh F&O fill
                        try:
                            import time
                            time.sleep(2)  # let order register with Kite
                            self._fno_gtt_audit()
                            log.info(f"[FIX_V29_GTT_ON_ENTRY] GTT audit run immediately after {tsym} fill")
                        except Exception as _ge:
                            log.error(f"[FIX_V29_GTT_ON_ENTRY] failed: {_ge}")
                    except Exception:
                        pass
                except Exception as _e:
                    log.error(f"[PATCH_V1_FNO_AMO] place failed {sym}: {_e}")

            save_json(marker_file, {"date": str(today_ist()), "placed": placed_list})
            # Clear candidate pool for next night
            try:
                save_json(self._fno_amo_candidates_file(), {"date": str(today_ist()), "candidates": []})
            except Exception as _e71:
                log.warning(f"[PATCH_V1_FNO_AMO] candidate clear save_json failed: {_e71} [PATCH_V71_RESILIENCE]")
        except Exception as _e:
            log.error(f"[PATCH_V1_FNO_AMO] premarket place failed: {_e}")

    def init(self):
        """Initialize after Kite login."""
        kite = KiteSession.kite()
        if not kite:
            log.error("FNO: cannot init without Kite")
            return False

        # Load F&O eligible stocks
        try:
            self._nfo_instruments = kite.instruments("NFO")
            self.fno_stocks = set()
            for inst in self._nfo_instruments:
                name = inst.get("name", "")
                if name and inst.get("instrument_type") in ("CE", "PE"):
                    self.fno_stocks.add(name)
            # V78_FIXA1: ensure tradeable indices are always in fno_stocks
            # Safety net in case Kite returns index name differently
            for _idx in ("NIFTY", "BANKNIFTY", "FINNIFTY"):
                self.fno_stocks.add(_idx)
            log.info(f"FNO: {len(self.fno_stocks)} F&O eligible stocks loaded (incl. NIFTY/BANKNIFTY/FINNIFTY)")

            # FIX_V39_SCOUT_VALIDATION: init NewsBrain v3 with RSS feeds
            self.news_brain = None
            if USE_LLM_NEWS and _NEWS_BRAIN_AVAILABLE and NewsBrain:
                try:
                    try:
                        # FIX_NB_G2: upgrade F&O to v4 BOTH directions, relaxed filters
                        try:
                            self.news_brain = NewsBrain(
                                universe_stocks=self.fno_stocks,
                                news_feeds=NEWS_FEEDS,
                                module_name="FNO",
                                direction_mode="BOTH",
                                min_conviction=55,
                                require_cross_sources=1,
                                require_sentence_citation=False,
                                max_article_age_hours=48,
                            )
                            log.info(f"FNO: NewsBrain v4 [FIX_NB_G2] - BOTH, min_conv=55, cross>=1, no-cite, max_age=48h (DRY_RUN={LLM_DRY_RUN})")
                        except TypeError:
                            self.news_brain = NewsBrain(fno_stocks=self.fno_stocks, news_feeds=NEWS_FEEDS)
                            log.warning(f"FNO: NewsBrain v3 fallback (DRY_RUN={LLM_DRY_RUN})")
                    except TypeError:
                        self.news_brain = NewsBrain(fno_stocks=self.fno_stocks)
                        log.warning(f"FNO: NewsBrain v2 fallback (DRY_RUN={LLM_DRY_RUN})")
                except Exception as _nbe2:
                    log.error(f"FNO: NewsBrain init failed: {_nbe2}")
            else:
                log.info(f"FNO: NewsBrain disabled (USE_LLM_NEWS={USE_LLM_NEWS} available={_NEWS_BRAIN_AVAILABLE})")
        except Exception as e:
            log.error(f"FNO: instrument load failed: {e}")
            return False

        # Get capital
        self._refresh_capital()

        # Init risk manager
        self.risk = RiskManager("FNO", self.capital)

        # PATCH_V28_SYMBOL_REGEX: build name_map for F&O stocks from NFO instruments
        self._name_map = {}
        try:
            kite_inst = KiteSession.kite()
            if kite_inst:
                _nse_insts = kite_inst.instruments("NSE")
                for inst in _nse_insts:
                    _n = inst.get("name", "").upper().strip()
                    _s = inst.get("tradingsymbol", "")
                    if _n and _s and _s in self.fno_stocks:
                        self._name_map[_n] = _s
                        _words = _n.split()
                        if len(_words) >= 2:
                            self._name_map[" ".join(_words[:2])] = _s
                log.info(f"[PATCH_V28_SYMBOL_REGEX] FNO name_map with {len(self._name_map)} entries")
        except Exception as _nme:
            log.warning(f"[PATCH_V28_SYMBOL_REGEX] FNO name_map failed: {_nme}")

        # Init news scanner with F&O stocks + name mapping
        self.news = NewsScanner("FNO", self.fno_stocks, name_map=self._name_map)

        # Init filing monitor (for F&O-eligible stocks only)
        self.filings = FilingMonitor("FNO", self.fno_stocks)

        # Init brokerage monitor
        self.brokerage = BrokerageMonitor("FNO", self.fno_stocks)

        # Init macro detector (RBI, war, oil → specific stocks)
        self.macro = MacroDetector("FNO", self.fno_stocks, trendlyne_scorer=self.trendlyne)  # FIX_V4_DYNAMIC_MACRO

        # PATCH_V63_INDEX_ROUTING: init index macro detector (NIFTY/BANKNIFTY/FINNIFTY)
        # Emits index candidates from broad market news. Feeds same pipeline as other scouts.
        self.index_macro = IndexMacroDetector("FNO", self.fno_stocks)
        log.info(f"[PATCH_V63] IndexMacroDetector initialized (tradeable={IndexMacroDetector.TRADEABLE_INDICES & self.fno_stocks})")

        # Update earnings calendar
        try:
            self.earnings.update_from_filings()
        except Exception as _e72:
            log.warning(f"init: earnings.update_from_filings failed: {_e72} [PATCH_V72_COMPREHENSIVE]")

        # PATCH_V5_OI_LOGGER: initialize OI snapshot engine
        try:
            self._oi_engine = OISnapshotEngine(self)
            log.info("[PATCH_V5_OI] OISnapshotEngine initialized")
        except Exception as _e:
            log.error(f"[PATCH_V5_OI] init failed: {_e}")
            self._oi_engine = None

        # PATCH_V6_GAP: initialize gap detector
        try:
            self._gap_detector = GapDetector(self)
            log.info("[PATCH_V6_GAP] GapDetector initialized")
        except Exception as _e:
            log.error(f"[PATCH_V6_GAP] init failed: {_e}")
            self._gap_detector = None

        # PATCH_V8_SECTOR_ROTATION: initialize sector rotation tracker
        try:
            self._sector_rotation = SectorRotation(self)
            log.info("[PATCH_V8_SECTOR] SectorRotation initialized")
        except Exception as _e:
            log.error(f"[PATCH_V8_SECTOR] init failed: {_e}")
            self._sector_rotation = None

        # PATCH_V28_KITE_ONLY: DO NOT load from JSON. Kite is the only source of truth.
        # Previous behavior loaded phantom positions from live_fno.json that persisted
        # after manual Kite sells (Apr 16 2026 SURAJEST/BANKNIFTY phantom incident).
        self.positions = {}  # start empty, populate from Kite only
        self._sync_positions()

        # Run F&O GTT audit on startup
        self._fno_gtt_audit()
        # PATCH_V61: start background audit worker AFTER startup audit completes
        self._fno_audit_thread_start()
        
        log.info(f"FNO: init complete. Capital=Rs.{self.capital:,.0f} Positions={len(self.positions)}")
        self.telegram.send(
            f"F&O Module started\n"
            f"Capital: Rs.{self.capital:,.0f}\n"
            f"Positions: {len(self.positions)}\n"
            f"F&O stocks: {len(self.fno_stocks)}",
            force=True,
        )
        return True

    def _refresh_capital(self):
        """PATCH_V10_CAPITAL_RETRY: retry up to 3 times with 1s delay if Kite returns 0."""
        kite = KiteSession.kite()
        if not kite:
            return
        import time as _t
        total = 0
        for _attempt in range(1, 4):
            try:
                margins = kite.margins("equity")
                _t1 = margins.get("net", 0) or margins.get("available", {}).get("live_balance", 0)
                if not _t1:
                    _t1 = margins.get("available", {}).get("cash", 0)
                if _t1 and _t1 > 0:
                    total = _t1
                    break
                log.warning(f"[PATCH_V10_CAPITAL_RETRY] FNO attempt {_attempt}/3: Kite returned 0, retrying")
            except Exception as _e:
                log.warning(f"[PATCH_V10_CAPITAL_RETRY] FNO attempt {_attempt}/3: {type(_e).__name__}: {_e}")
            if _attempt < 3:
                _t.sleep(1)
        if total > 0:
            self.capital = round(total * FNO_PCT, 2)
            invested = sum(
                p.get("cost", 0) for p in self.positions.values()
            )
            self.available = max(0, self.capital - invested)
            log.info(f"[PATCH_V10_CAPITAL_RETRY] FNO refreshed: total={total} capital={self.capital} available={self.available}")
        else:
            log.warning(f"[PATCH_V10_CAPITAL_RETRY] FNO: all 3 attempts returned 0, preserving previous capital={self.capital}")


    def _save_positions(self):
        save_json(self._pos_file, {"positions": self.positions})

    def _sync_positions(self):
        """Sync with Kite F&O positions.

        PATCH_V28_KITE_ONLY: Now also PURGES phantom positions — any entry in
        self.positions that is NOT in Kite's NFO net book gets removed.
        This makes Kite the authoritative source.
        """
        kite = KiteSession.kite()
        if not kite:
            return
        try:
            kite_positions = kite.positions().get("net", [])

            # PATCH_V28_KITE_ONLY: build set of live NFO symbols from Kite
            _kite_nfo_syms = set()
            for _p in kite_positions:
                if _p.get("exchange") == "NFO" and _p.get("quantity", 0) > 0:
                    _kite_nfo_syms.add(_p.get("tradingsymbol", ""))

            # PATCH_V28_KITE_ONLY: purge any self.positions entries not in Kite
            _phantom_keys = [k for k in list(self.positions.keys()) if k not in _kite_nfo_syms]
            for _phantom in _phantom_keys:
                log.warning(f"[PATCH_V28_KITE_ONLY] FNO purge phantom: {_phantom} not in Kite NFO book")
                del self.positions[_phantom]
            if _phantom_keys:
                log.info(f"[PATCH_V28_KITE_ONLY] FNO purged {len(_phantom_keys)} phantom positions")

            for p in kite_positions:
                if p.get("exchange") == "NFO" and p.get("quantity", 0) > 0:
                    sym = p.get("tradingsymbol", "")
                    key = sym
                    if key not in self.positions:
                        # v26.3: Preserve existing trailed GTT SL instead of resetting to 30% default
                        _live_sl = _snap_tick(p.get("average_price", 0) * (1 - FNO_SL_PCT))
                        try:
                            _existing_gtts = kite.get_gtts() or []
                            for _g in _existing_gtts:
                                if _g.get("status") != "active":
                                    continue
                                _cond = _g.get("condition", {})
                                if _cond.get("tradingsymbol") == sym and _cond.get("exchange") == "NFO":
                                    _triggers = _cond.get("trigger_values", [])
                                    if _triggers and _triggers[0] > 0:
                                        _live_sl = float(_triggers[0])
                                        log.info(f"FNO sync: {sym} using live GTT SL={_live_sl} (not default)")
                                        break
                        except Exception as _e:
                            log.warning(f"FNO sync: GTT lookup failed for {sym}: {_e}")
                        # V27 PATCH (Bug #5): Enrich with expiry/strike from instrument master
                        # so DTE protection works for synced positions.
                        _expiry_str = ""
                        _strike = 0
                        _option_type = ""
                        try:
                            _instruments = kite.instruments("NFO")
                            for _inst in _instruments:
                                if _inst.get("tradingsymbol") == sym:
                                    _exp = _inst.get("expiry")
                                    if _exp:
                                        _expiry_str = _exp.strftime("%Y-%m-%d") if hasattr(_exp, "strftime") else str(_exp)
                                    _strike = _inst.get("strike", 0)
                                    _option_type = _inst.get("instrument_type", "")
                                    break
                            if not _expiry_str:
                                log.warning(f"FNO sync: {sym} expiry not found in instrument master - DTE protection DISABLED for this position")
                        except Exception as _ie:
                            log.error(f"FNO sync: instrument lookup for {sym} failed: {_ie}")

                        _dte_at_sync = 0
                        if _expiry_str:
                            try:
                                _exp_d = datetime.strptime(_expiry_str, "%Y-%m-%d").date()
                                _dte_at_sync = (_exp_d - today_ist()).days
                            except Exception:
                                pass

                        self.positions[key] = {
                            "tradingsymbol": sym,
                            "qty": p["quantity"],
                            "entry_price": p.get("average_price", 0),
                            "entry_time": now_ist().isoformat(),
                            "cost": p.get("quantity", 0) * p.get("average_price", 0),
                            "peak_premium": p.get("average_price", 0),
                            "sl_price": _live_sl,
                            "expiry": _expiry_str,
                            "dte_at_entry": _dte_at_sync,
                            "strike": _strike,
                            "option_type": _option_type,
                            "source": "KITE_SYNC",
                        }
                        log.info(f"FNO sync: {sym} expiry={_expiry_str} strike={_strike} dte={_dte_at_sync}")
            self._save_positions()
        except Exception as e:
            log.error(f"FNO: sync positions failed: {e}")

    def _find_best_option(self, kite, symbol, direction, spot_price, catalyst_score=60):
        """
        SMART F&O option selection using professional options analysis:
        1. IV Percentile — are options cheap or expensive?
        2. OI Analysis — where is smart money?
        3. Delta — probability of profit
        4. Theta — time decay vs expected move
        5. PCR — market direction confirmation
        6. Physical delivery check — exit stock options 2 days before expiry
        
        Returns dict with option details or None.
        PATCH_V63_STRICT_DIR: strict equality on direction; reject anything that is
        not exactly BULLISH or BEARISH (was: 'BULL' substring would let 'NEUTRAL'
        fall to PE silently).
        """
        # PATCH_V63_STRICT_DIR: strict direction validation
        _dir_norm = (direction or "").strip().upper()
        if _dir_norm == "BULLISH":
            opt_type = "CE"
        elif _dir_norm == "BEARISH":
            opt_type = "PE"
        else:
            log.warning(f"[PATCH_V63_STRICT_DIR] {symbol}: invalid direction '{direction}' "
                        f"(expected BULLISH/BEARISH). Skipping option selection.")
            return None
        today = today_ist()
        is_index = symbol in ("NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY")

        # ── Step 0: Gather all valid option contracts ──
        candidates = []
        for inst in self._nfo_instruments:
            if inst.get("name") != symbol or inst.get("instrument_type") != opt_type:
                continue

            expiry = inst.get("expiry")
            if not expiry:
                continue
            if isinstance(expiry, str):
                try:
                    expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
                except Exception:
                    continue
            elif hasattr(expiry, "date"):
                expiry = expiry.date()

            dte = (expiry - today).days

            # PATCH_V28_DTE_PREFER_FAR_MONTH: smart DTE selection
            # Stock options: if near-month DTE < EMERGENCY_DTE + 3 (=15), skip it.
            # Prevents buy-then-emergency-exit-same-day (killed HAL +27% on Apr 17).
            # Index options (cash-settled, liquid): can still use near-month.
            if is_index:
                min_dte = max(5, FNO_MIN_DTE - 5)
            else:
                # For stocks: never enter if DTE would trigger emergency exit soon
                _safe_min = FNO_EMERGENCY_DTE + 3  # 12 + 3 = 15 days minimum for stocks
                min_dte = max(FNO_MIN_DTE, _safe_min)
            if not (min_dte <= dte <= FNO_MAX_DTE):
                continue

            strike = inst.get("strike", 0)
            if strike <= 0:
                continue

            # Calculate OTM percentage
            if opt_type == "CE":
                otm_pct = (strike - spot_price) / spot_price if spot_price > 0 else 0
            else:
                otm_pct = (spot_price - strike) / spot_price if spot_price > 0 else 0

            # Wide range: 1-20% OTM
            if 0.01 <= otm_pct <= 0.20:
                candidates.append({
                    "tradingsymbol": inst["tradingsymbol"],
                    "strike": strike,
                    "expiry": expiry,
                    "dte": dte,
                    "lot_size": inst.get("lot_size", 1),
                    "otm_pct": otm_pct,
                    "instrument_token": inst.get("instrument_token"),
                    "is_index": is_index,
                })

        if not candidates:
            log.debug(f"FNO smart: {symbol} {opt_type} — no candidates in DTE/OTM range")
            return None

        # ── Step 1: Get live quotes for all candidates (batch) ──
        # Process in batches of 20 (Kite API limit)
        scored_candidates = []
        for i in range(0, min(len(candidates), 40), 20):
            if i > 0: time.sleep(0.3)  # FIX_V51: Kite rate limit between batches
            batch = candidates[i:i+20]
            sym_list = [f"NFO:{c['tradingsymbol']}" for c in batch]
            try:
                quotes = kite.quote(sym_list)
            except Exception:
                continue

            for cand in batch:
                tsym = cand["tradingsymbol"]
                q = quotes.get(f"NFO:{tsym}", {})
                if not q:
                    continue

                ltp = q.get("last_price", 0)
                oi = q.get("oi", 0)
                oi_change = q.get("oi_day_change", 0)  # Today's OI change
                volume = q.get("volume", 0)
                depth = q.get("depth", {})
                # PATCH_V70_AUDIT_CLEANUP: removed unused ohlc/prev_close pair (neither was read)

                # Best bid/ask from market depth
                bids = depth.get("buy", [])
                asks = depth.get("sell", [])
                best_bid = bids[0].get("price", 0) if bids else 0
                best_ask = asks[0].get("price", 0) if asks else 0

                # ── Filter: minimum OI ──
                if oi < FNO_MIN_OI:
                    continue

                # ── Filter: bid-ask spread ──
                if best_bid > 0 and best_ask > 0:
                    spread = (best_ask - best_bid) / best_ask
                    if spread > FNO_MAX_SPREAD:
                        continue
                elif ltp <= 0:
                    continue

                premium = best_ask if best_ask > 0 else ltp
                if premium <= 0:
                    continue

                # ── Filter: cost check ──
                cost = premium * cand["lot_size"]
                max_cost = self.available * FNO_CAP_PER_TRADE
                if cost > max_cost or cost > self.available:
                    continue

                # ── Score: IV analysis ──
                # Approximate IV percentile from price movement
                # If premium is low relative to spot and DTE, IV is low (cheap)
                iv_score = 0
                time_value_ratio = premium / (spot_price * 0.01) if spot_price > 0 else 0
                dte_factor = cand["dte"] / 30  # Normalize to 1 month
                
                if time_value_ratio < 1.5 * dte_factor:
                    iv_score = 30  # Options are CHEAP — good
                elif time_value_ratio < 3.0 * dte_factor:
                    iv_score = 15  # Normal pricing
                else:
                    iv_score = -10  # EXPENSIVE — avoid unless strong catalyst
                    if catalyst_score < 80:
                        continue  # Skip expensive options for weak catalysts

                # ── Score: OI analysis (follow smart money) ──
                oi_score = 0
                if oi_change > 0:
                    # OI increasing = new positions being built
                    oi_ratio = oi_change / oi if oi > 0 else 0
                    if oi_ratio > 0.10:
                        oi_score = 25  # >10% OI increase today — strong institutional activity
                    elif oi_ratio > 0.05:
                        oi_score = 15  # 5-10% increase — moderate
                    else:
                        oi_score = 5   # Some activity
                elif oi_change < 0:
                    oi_score = -5  # OI decreasing = unwinding positions

                # ── Score: Delta approximation ──
                # Rough delta from OTM percentage (Black-Scholes approximation)
                delta_score = 0
                otm = cand["otm_pct"]
                if otm <= 0.03:
                    approx_delta = 0.60  # Near ATM
                elif otm <= 0.05:
                    approx_delta = 0.45
                elif otm <= 0.08:
                    approx_delta = 0.30
                elif otm <= 0.12:
                    approx_delta = 0.20
                else:
                    approx_delta = 0.10  # Deep OTM

                # Strong catalyst → prefer higher delta (more probability)
                # Weak catalyst → prefer lower delta (cheaper, less risk)
                if catalyst_score >= 80:
                    # High conviction — want Delta 0.30-0.60
                    if 0.30 <= approx_delta <= 0.60:
                        delta_score = 25
                    elif 0.20 <= approx_delta < 0.30:
                        delta_score = 15
                    else:
                        delta_score = 5
                else:
                    # Lower conviction — want Delta 0.15-0.35 (cheaper)
                    if 0.15 <= approx_delta <= 0.35:
                        delta_score = 25
                    elif 0.10 <= approx_delta < 0.15:
                        delta_score = 15
                    else:
                        delta_score = 5

                # ── Score: Theta check ──
                # Approximate theta: premium / DTE
                theta_score = 0
                daily_decay = premium / cand["dte"] if cand["dte"] > 0 else premium
                # Expected move from catalyst (rough: score/100 * 3% of spot)
                expected_move_pct = (catalyst_score / 100) * 0.03
                expected_premium_gain = expected_move_pct * spot_price * approx_delta
                
                # Is expected gain > 3 days of theta decay?
                if expected_premium_gain > daily_decay * 3:
                    theta_score = 20  # Good — move should overcome decay
                elif expected_premium_gain > daily_decay * 1.5:
                    theta_score = 10  # Marginal
                else:
                    theta_score = -15  # Bad — theta will eat your profit
                    if catalyst_score < 75:
                        continue  # Skip if weak catalyst + bad theta

                # ── Score: Volume confirmation ──
                vol_score = 0
                if volume > 10000:
                    vol_score = 10
                elif volume > 1000:
                    vol_score = 5

                # ── Total score ──
                total_score = iv_score + oi_score + delta_score + theta_score + vol_score

                scored_candidates.append({
                    "tradingsymbol": tsym,
                    "lot_size": cand["lot_size"],
                    "premium": premium,
                    "expiry": cand["expiry"],
                    "dte": cand["dte"],
                    "strike": cand["strike"],
                    "oi": oi,
                    "oi_change": oi_change,
                    "spread": spread if best_bid > 0 and best_ask > 0 else 0,
                    "cost": cost,
                    "approx_delta": approx_delta,
                    "daily_theta": round(daily_decay, 2),
                    "smart_score": total_score,
                    "is_index": cand["is_index"],
                    "iv_score": iv_score,
                    "oi_score": oi_score,
                    "delta_score": delta_score,
                    "theta_score": theta_score,
                })

        if not scored_candidates:
            if candidates:
                log.info(f"FNO smart: {symbol} {opt_type} — {len(candidates)} in DTE/OTM range but ALL rejected by OI/spread/cost/IV/theta")
            else:
                log.info(f"FNO smart: {symbol} {opt_type} — 0 candidates in DTE range ({FNO_MIN_DTE}-{FNO_MAX_DTE}d) or OTM range — check expiry calendar")
            return None

        # ── Pick the BEST option (highest smart_score) ──
        scored_candidates.sort(key=lambda c: c["smart_score"], reverse=True)
        best = scored_candidates[0]

        log.info(
            f"FNO smart: {symbol} {opt_type} SELECTED {best['tradingsymbol']} "
            f"strike={best['strike']} premium={best['premium']} DTE={best['dte']} "
            f"delta~{best['approx_delta']:.2f} OI_chg={best['oi_change']} "
            f"score={best['smart_score']} (IV={best['iv_score']} OI={best['oi_score']} "
            f"D={best['delta_score']} T={best['theta_score']})"
        )

        return best

    def enter(self, symbol, catalyst, direction, score):
        """Place a real F&O BUY order."""
        if len(self.positions) >= FNO_MAX_POS:
            return False
        # Check per-stock limit
        stock_count = sum(1 for p in self.positions.values()
                          if p.get("tradingsymbol", "").startswith(symbol))
        if stock_count >= FNO_MAX_PER_STOCK:
            return False
        # PATCH_V3_SECTOR_CAP: enforce max positions per sector
        try:
            _new_sector = get_sector(symbol)
            _sector_count = sum(1 for p in self.positions.values()
                                if get_sector(p.get("symbol", "")) == _new_sector)
            if _sector_count >= FNO_MAX_PER_SECTOR:
                log.info(f"[PATCH_V3_SECTOR_CAP] FNO blocked {symbol}: sector {_new_sector} already has {_sector_count} positions")
                return False
        except Exception as _e:
            log.warning(f"[PATCH_V3_SECTOR_CAP] sector check failed for {symbol}: {_e}, allowing entry")
        # END_PATCH_V3_SECTOR_CAP

        # V28 M2: PATCH_V4_VIX_FILTER v4 — SMART VIX LOGIC (previously log-only)
        # HIGH VIX (>22): BLOCK CE entries (IV crush risk kills long calls)
        #                 ALLOW PE entries (puts profit in fear; direction dominates IV crush)
        # LOW VIX  (<12): allow both — option premiums are cheap
        # NORMAL   (12-22): allow both — standard regime
        # Cached 60s to avoid hammering kite.ltp().
        try:
            import time as _t
            _now = _t.time()
            _cache = getattr(self, "_vix_cache", None)
            if _cache and (_now - _cache[1]) < 60:
                _vix = _cache[0]
            else:
                # PATCH_V70_AUDIT_CLEANUP: was bare 'kite' (undefined) — VIX filter was non-functional
                _kite = KiteSession.kite()
                if _kite is None:
                    raise RuntimeError("Kite session unavailable for VIX check")
                _vix_data = _kite.ltp(["NSE:INDIA VIX"])
                _vix = _vix_data.get("NSE:INDIA VIX", {}).get("last_price", 0)
                self._vix_cache = (_vix, _now)
            if _vix > 0:
                _regime = "HIGH" if _vix > FNO_VIX_MAX else ("LOW" if _vix < 12 else "NORMAL")
                # V28 M2: block bullish CE in HIGH VIX only
                if _vix > FNO_VIX_MAX and direction == "BULLISH":
                    log.info(f"[V4_VIX_v4] BLOCKED bullish {symbol}: VIX={_vix:.2f} > {FNO_VIX_MAX} (IV crush risk on long CE)")
                    return False
                # Bearish PE at HIGH VIX = the correct direction; allow
                if _vix > FNO_VIX_MAX and direction == "BEARISH":
                    log.info(f"[V4_VIX_v4] ALLOW bearish {symbol}: VIX={_vix:.2f} HIGH — PE expected to profit in fear")
                else:
                    log.info(f"[V4_VIX_v4] {symbol} {direction} entering at VIX={_vix:.2f} ({_regime})")
        except Exception as _e:
            log.debug(f"[V4_VIX_v4] VIX check failed for {symbol}: {_e}, allowing entry")
        # END_PATCH_V4_VIX_FILTER

        # PATCH_V6_GAP: block contradictory entries based on per-stock gap status
        try:
            _gap_file = DATA_DIR / "gap_status.json"
            _gap_status = load_json(_gap_file, {})
            _gap_info = None
            if symbol in _gap_status.get("stocks", {}):
                _gap_info = _gap_status["stocks"][symbol]
            elif symbol in _gap_status.get("indices", {}):
                _gap_info = _gap_status["indices"][symbol]
            if _gap_info:
                _tag = _gap_info.get("tag", "FLAT")
                _pct = _gap_info.get("pct", 0)
                if direction == "BULLISH" and _tag == "GAP_DOWN":
                    log.info(f"[PATCH_V6_GAP] FNO blocked {symbol} BULLISH: stock GAP_DOWN {_pct:+.2f}%")
                    return False
                if direction == "BEARISH" and _tag == "GAP_UP":
                    log.info(f"[PATCH_V6_GAP] FNO blocked {symbol} BEARISH: stock GAP_UP {_pct:+.2f}%")
                    return False
                if (direction == "BULLISH" and _tag == "GAP_UP") or (direction == "BEARISH" and _tag == "GAP_DOWN"):
                    log.info(f"[PATCH_V6_GAP] OK {symbol} {direction}: aligned with {_tag} {_pct:+.2f}%")
        except Exception as _e:
            log.warning(f"[PATCH_V6_GAP] gap check failed for {symbol}: {_e}, allowing entry")
        # END_PATCH_V6_GAP

        # PATCH_V8_SECTOR_ROTATION: only trade stocks in strong sectors
        try:
            if self._sector_rotation and self._sector_rotation._ranking:
                _bias, _sec_name, _sec_pct = self._sector_rotation.get_sector_bias(symbol)
                if _bias == "BOTTOM" and SECTOR_ROTATION_BOTTOM_BLOCK:
                    log.info(f"[PATCH_V8_SECTOR] FNO blocked {symbol}: sector {_sec_name} in bottom-2 ({_sec_pct:+.2f}%)")
                    return False
                elif _bias == "TOP":
                    log.info(f"[PATCH_V8_SECTOR] OK {symbol}: sector {_sec_name} in top-3 ({_sec_pct:+.2f}%) — boosted")
        except Exception as _e:
            log.warning(f"[PATCH_V8_SECTOR] sector check failed for {symbol}: {_e}, allowing entry")
        # END_PATCH_V8_SECTOR_ROTATION
        if not self.risk or not self.risk.can_trade():
            return False

        kite = KiteSession.kite()
        if not kite:
            return False

        # Safety: crash protection only (no regime filter for F&O)
        if not SafetyFilters.check_nifty_crash(kite): return False

        # PATCH_V1_REGIME_GATE: refuse direction-contradicting entries during gap days
        try:
            _nifty_data = kite.ltp(["NSE:NIFTY 50"])
            _nifty_ltp = _nifty_data.get("NSE:NIFTY 50", {}).get("last_price", 0)
            _nifty_ohlc = kite.ohlc(["NSE:NIFTY 50"]).get("NSE:NIFTY 50", {}).get("ohlc", {})
            _yest_close = _nifty_ohlc.get("close", 0)
            if _nifty_ltp > 0 and _yest_close > 0:
                _intraday_pct = ((_nifty_ltp - _yest_close) / _yest_close) * 100
                if direction == "BULLISH" and _intraday_pct <= -1.5:
                    log.info(f"[PATCH_V73_REGIME] BLOCKED bullish {symbol}: NIFTY {_intraday_pct:.2f}% (strong gap down)")
                    return False
                if direction == "BEARISH" and _intraday_pct >= 1.5:
                    log.info(f"[PATCH_V73_REGIME] BLOCKED bearish {symbol}: NIFTY {_intraday_pct:.2f}% (strong gap up)")
                    return False
                log.debug(f"[PATCH_V1_REGIME] OK {symbol} {direction}: NIFTY {_intraday_pct:.2f}%")
        except Exception as _e:
            log.warning(f"[PATCH_V1_REGIME] check failed for {symbol}: {_e}, allowing entry")
        # END_PATCH_V1_REGIME_GATE

        # PATCH_V1_CONVICTION_LOCK: no contradictory CE+PE on same underlying (any expiry)
        # PATCH_V70_AUDIT_CLEANUP: removed dead _existing_expiry; comment matched intent but code never used it
        try:
            import re as _re
            _new_dir_is_call = (direction == "BULLISH")
            for _key, _pos in self.positions.items():
                _ptsym = _pos.get("tradingsymbol", "")
                if not _ptsym.startswith(symbol):
                    continue
                _m = _re.match(rf"^{_re.escape(symbol)}(\d{{2}}[A-Z]{{3}})", _ptsym)
                if not _m:
                    continue
                _existing_is_call = _ptsym.endswith("CE")
                _existing_is_put = _ptsym.endswith("PE")
                if _new_dir_is_call and _existing_is_put:
                    log.info(f"[PATCH_V1_CONVICTION] BLOCKED bullish {symbol}: already hold PE {_ptsym}")
                    return False
                if (not _new_dir_is_call) and _existing_is_call:
                    log.info(f"[PATCH_V1_CONVICTION] BLOCKED bearish {symbol}: already hold CE {_ptsym}")
                    return False
        except Exception as _e:
            log.warning(f"[PATCH_V1_CONVICTION] check failed for {symbol}: {_e}, allowing entry")
        # END_PATCH_V1_CONVICTION_LOCK

        # Order guard: duplicate + SEBI + pending check
        can_order, guard_reason = OrderGuard.can_place_order(symbol, is_fno=True)
        if not can_order:
            log.info(f"FNO: {symbol} blocked by OrderGuard: {guard_reason}")
            return False
        if OrderGuard.check_pending_orders(kite, symbol, is_fno=True):  # V28 A7
            return False

        try:
            # Get spot price
            # V83B_NIFTY_SPOT_KEY: use _index_spot_key() helper so NIFTY/BANKNIFTY/
            # FINNIFTY map to "NSE:NIFTY 50"/"NSE:NIFTY BANK"/"NSE:NIFTY FIN SERVICE".
            # Bug: f"NSE:{symbol}" was returning {} for indices, spot=0, silent return False.
            # Stocks are unaffected — _index_spot_key("RELIANCE") = "NSE:RELIANCE".
            _v83b_spot_key = _index_spot_key(symbol)
            ltp_data = kite.ltp([_v83b_spot_key])
            spot = ltp_data.get(_v83b_spot_key, {}).get("last_price", 0)
            if spot <= 0:
                log.info(f"[V83B_NIFTY_SPOT_KEY] {symbol}: no spot from key={_v83b_spot_key}, return False")
                return False

            # V28 C3: Freak spot protection — reject if spot >20% off prev close
            try:
                _q = kite.quote([_v83b_spot_key]).get(_v83b_spot_key, {})
                _prev = _q.get("ohlc", {}).get("close", 0)
                if _prev > 0 and abs(spot - _prev) / _prev > 0.20:
                    log.warning(f"FNO: {symbol} FREAK SPOT — spot={spot}, prev={_prev}, diff={(spot-_prev)/_prev*100:.1f}%, REJECTING")
                    return False
            except Exception:
                pass

            # Find best option
            option = self._find_best_option(kite, symbol, direction, spot, catalyst_score=score)
            if not option:
                log.info(f"FNO: no suitable option for {symbol} {direction}")
                return False

            tsym = option["tradingsymbol"]
            lot = option["lot_size"]
            premium = option["premium"]
            cost = lot * premium

            # V28 D4: Smart F&O sizing (respects your constraint that small F&O capital
            # can't always fit 30% — without allowing 95% catastrophic single-trade risk).
            # Rules:
            #   cost_pct = cost / available
            #   <= 0.30          -> OK (safe)
            #   0.30 < x <= 0.60 -> OK (allowed, but logged — lot-cost forces this)
            #   > 0.60           -> BLOCK (position too concentrated)
            # Keeps the original FNO_CAP_PER_TRADE as absolute ceiling too.
            if self.available <= 0:
                log.info(f"FNO: {tsym} no available capital")
                return False
            cost_pct = cost / self.available if self.available > 0 else 1.0
            if cost_pct > 0.60:
                log.info(f"FNO: {tsym} cost={cost:.0f} = {cost_pct*100:.0f}% of available Rs.{self.available:.0f} — BLOCKED (>60%)")
                return False
            if cost_pct > 0.30:
                log.info(f"FNO: {tsym} cost={cost:.0f} = {cost_pct*100:.0f}% of available (30-60% zone, allowed due to lot-cost)")
            # Absolute ceiling (legacy FNO_CAP_PER_TRADE)
            if cost > self.available * FNO_CAP_PER_TRADE or cost > self.available:
                log.info(f"FNO: {tsym} cost={cost:.0f} exceeds legacy ceiling")
                return False

            # Smart order pricing using market depth (v26.2: hardened against zero/stale quotes)
            # V77_FIX5: use TickSizeCache instead of hardcoded 0.05
            def _tick_round(p):
                return _snap_tick(p, symbol=tsym, exchange="NFO")
            
            smart_price = 0.0
            try:
                depth_data = kite.quote([f"NFO:{tsym}"])
                q = depth_data.get(f"NFO:{tsym}", {})
                ltp = q.get("last_price", 0) or 0
                depth = q.get("depth", {})
                asks = depth.get("sell", [])
                ask1_price = asks[0].get("price", 0) if asks else 0
                ask1_qty = asks[0].get("quantity", 0) if asks else 0
                ask2_price = asks[1].get("price", 0) if len(asks) > 1 else 0
                ask2_qty = asks[1].get("quantity", 0) if len(asks) > 1 else 0
                if ask1_price > 0:
                    if ask2_qty > ask1_qty * 5 and ask2_price > 0:
                        smart_price = (ask1_price + ask2_price) / 2
                        log.info(f"FNO: smart pricing {tsym} L1={ask1_price}x{ask1_qty} L2={ask2_price}x{ask2_qty}")
                    else:
                        smart_price = ask1_price * 1.005
                elif ltp > 0:
                    smart_price = ltp * 1.02
                    log.info(f"FNO: {tsym} no ask quote, using LTP={ltp} + 2%")
                elif premium > 0:
                    smart_price = premium * 1.02
                    log.info(f"FNO: {tsym} no market data, using theoretical premium={premium} + 2%")
            except Exception as _e:
                log.warning(f"FNO: quote fetch failed for {tsym}: {_e}")
                smart_price = premium * 1.02 if premium > 0 else 0.0
            
            smart_price = _tick_round(smart_price)
            if smart_price <= 0:
                log.warning(f"FNO: {tsym} smart_price zero after all fallbacks, ABORTING (premium was {premium})")
                return False
            
            # FIX_V53_RANGE: clamp to Kite's +/-40% execution range.
            # Reference price = ltp (live) or premium (theoretical).
            # If smart_price outside +/-40% of reference, Kite rejects LIMIT order
            # and position is never opened despite bot thinking it did.
            _ref_price = ltp if ltp > 0 else premium
            if _ref_price > 0:
                _clamped = _fno_safe_limit_price(smart_price, _ref_price, band=0.40)
                if abs(_clamped - smart_price) > 0.01:
                    log.warning(f"FNO: {tsym} smart_price {smart_price} outside +/-40% of ref={_ref_price}, clamped to {_clamped} [FIX_V53_RANGE]")
                    smart_price = _clamped
            
            # Place order with smart price
            order_id = kite.place_order(
                variety=kite.VARIETY_REGULAR,
                exchange="NFO",
                tradingsymbol=tsym,
                transaction_type="BUY",
                quantity=lot,
                order_type="LIMIT",
                price=smart_price,
                product="NRML",
                validity="DAY",
                market_protection=5,  # SEBI mandatory
            )
            log.info(f"FNO: ORDER PLACED {tsym} lot={lot} premium={premium} order_id={order_id}")
            OrderGuard.record_order(tsym)
            OrderGuard.record_order(symbol)  # Block both tradingsymbol and underlying

            # V27 PATCH (Bug #8): Poll for actual fill (not 3-sec assumption).
            # Saves position only on COMPLETE. Cancels and aborts on timeout.
            fill_price = 0.0
            order_status = None
            _poll_max_sec = 15
            _poll_interval = 1.5
            _elapsed = 0.0
            while _elapsed < _poll_max_sec:
                time.sleep(_poll_interval)
                _elapsed += _poll_interval
                try:
                    orders = kite.orders()
                    for o in orders:
                        if str(o.get("order_id")) == str(order_id):
                            order_status = o.get("status")
                            if o.get("average_price", 0) > 0:
                                fill_price = o["average_price"]
                            break
                except Exception as _oe:
                    log.warning(f"FNO: {tsym} order poll error: {_oe}")
                    continue

                if order_status == "COMPLETE":
                    log.info(f"FNO: {tsym} FILLED at {fill_price} after {_elapsed:.1f}s")
                    break
                if order_status == "REJECTED":
                    log.error(f"FNO: {tsym} order REJECTED")
                    try:
                        health.counters["orders_rejected"] += 1
                        health.fire(f"FNO REJECTED: {tsym}")
                    except Exception as _e71:
                        log.warning(f"FNO: telegram fire failed for REJECTED {tsym}: {_e71} [PATCH_V71_RESILIENCE]")
                    return False
                if order_status == "CANCELLED":
                    log.warning(f"FNO: {tsym} order CANCELLED externally")
                    try:
                        health.counters["orders_cancelled"] += 1
                        health.fire(f"FNO CANCELLED externally: {tsym}")
                    except Exception as _e71:
                        log.warning(f"FNO: telegram fire failed for CANCELLED {tsym}: {_e71} [PATCH_V71_RESILIENCE]")
                    return False

            if order_status != "COMPLETE":
                log.warning(f"FNO: {tsym} order status={order_status} after {_poll_max_sec}s, CANCELLING to avoid phantom position")
                try:
                    kite.cancel_order(variety=kite.VARIETY_REGULAR, order_id=order_id)
                except Exception as _ce:
                    log.error(f"FNO: {tsym} cancel failed: {_ce}")
                return False

            if fill_price <= 0:
                log.error(f"FNO: {tsym} COMPLETE but fill_price=0, ABORTING (will not save phantom position)")
                return False

            # Save position
            key = tsym
            self.positions[key] = {
                "tradingsymbol": tsym,
                "symbol": symbol,
                "qty": lot,
                "entry_price": fill_price,
                "entry_time": now_ist().isoformat(),
                "cost": lot * fill_price,
                "peak_premium": fill_price,
                "sl_price": _snap_tick(fill_price * (1 - FNO_SL_PCT)),
                "expiry": str(option["expiry"]),
                "dte_at_entry": option["dte"],
                "strike": option["strike"],
                "direction": direction,
                "catalyst": catalyst,
                "order_id": str(order_id),
                "source": getattr(self, "_pending_source", "UNKNOWN"),
            }
            self.available -= cost
            self._save_positions()

            # GTT SL (SL-only, no target cap - let winners run)
            # V27 PATCH (Bug #8b): order_status guaranteed COMPLETE here by poll above
            sl = _snap_tick(fill_price * (1 - FNO_SL_PCT))
            self.gtt.ensure_gtt(tsym, "NFO", lot, sl)

            self.telegram.send(
                f"F&O BUY {tsym}\n"
                f"Lot: {lot} | Premium: Rs.{fill_price:,.1f}\n"
                f"Cost: Rs.{lot * fill_price:,.0f}\n"
                f"SL: Rs.{sl:,.1f} | DTE: {option['dte']}\n"  # V77_FIX6
                f"Direction: {direction}\n"
                f"Catalyst: {catalyst[:60]}",
                force=True,
            )
            return True

        except Exception as e:
            log.error(f"FNO: entry {symbol} failed: {e}")
            return False



    def check_trailing_sl(self):
        """FIX_V32_FNO_UNIFIED: peaks-only maintenance. No Kite GTT writes.
        
        ALL F&O GTT writes go through _fno_gtt_audit with ONE formula.
        This function only updates fno_peaks.json (stage detection) and
        then delegates to audit. Eliminates dual-writer race.
        """
        try:
            kite = KiteSession.kite()
            if not kite:
                log.warning("FNO trail: no Kite session")
                return
            
            try:
                positions = kite.positions().get("net", [])
            except Exception as _pe:
                log.error(f"FNO trail: positions() failed: {_pe}")
                return
            
            fno_positions = [p for p in positions
                             if p.get("exchange") == "NFO" and p.get("quantity", 0) != 0]
            
            if not fno_positions:
                log.debug("FNO trail: no open positions")
                return
            
            peaks_file = Path("/home/globalbot/fno_peaks.json")
            try:
                peaks = load_json(str(peaks_file), {})
                if not isinstance(peaks, dict):
                    peaks = {}
            except Exception:
                peaks = {}
            
            log.info(f"FNO trail V32: updating peaks for {len(fno_positions)} positions")
            
            live_symbols = set()
            for p in fno_positions:
                try:
                    tsym = p.get("tradingsymbol")
                    qty = p.get("quantity", 0)
                    avg = p.get("average_price", 0)
                    live_symbols.add(tsym)
                    
                    ltp_data = kite.ltp([f"NFO:{tsym}"])
                    ltp = ltp_data.get(f"NFO:{tsym}", {}).get("last_price", 0)
                    if ltp <= 0:
                        log.warning(f"FNO trail: {tsym} no LTP, skip peak update")
                        continue
                    
                    prev_peak = peaks.get(tsym, {}).get("peak", avg if avg > 0 else ltp)
                    if qty > 0:
                        peak = max(prev_peak, ltp)
                    else:
                        peak = min(prev_peak, ltp) if prev_peak > 0 else ltp
                    
                    peaks[tsym] = {"peak": peak, "avg": avg, "last_update": now_ist().isoformat()}
                    
                    pnl_pct = ((ltp - avg) / avg * 100) if avg > 0 else 0
                    peak_gain = ((peak - avg) / avg * 100) if (qty > 0 and avg > 0) else ((avg - peak) / avg * 100) if avg > 0 else 0
                    log.info(f"FNO trail: {tsym} qty={qty} avg={avg} ltp={ltp} peak={peak} pnl={pnl_pct:.1f}% peak_gain={peak_gain:.1f}%")
                except Exception as e:
                    log.error(f"FNO trail per-position error {p.get('tradingsymbol')}: {e}")
            
            # Prune stale peaks for closed positions
            stale = [k for k in list(peaks.keys()) if k not in live_symbols]
            for k in stale:
                del peaks[k]
                log.info(f"FNO trail: pruned peak for closed position {k}")
            
            try:
                save_json(str(peaks_file), peaks)
            except Exception as _e:
                log.error(f"FNO trail: failed to save peaks: {_e}")
            
            # Delegate all GTT writes to audit (unified path)
            try:
                self._fno_gtt_audit()
            except Exception as _ae:
                log.error(f"FNO trail: audit delegation failed: {_ae}")
        
        except Exception as e:
            log.error(f"FNO trailing SL error: {e}")
            try:
                health.counters["trailing_sl_errors"] += 1
                health.fire(f"FNO trailing SL ERROR (safety-critical): {e}")
            except Exception as _e71:
                log.error(f"FNO trailing SL: telegram fire ALSO failed: {_e71} (original: {e}) [PATCH_V71_RESILIENCE]")

    def _try_roll_position(self, key, pos, kite):
        """
        Roll a profitable position to next month expiry.
        Sell current month, buy next month at same/similar strike.
        Only rolls if position is profitable (don't roll losers).
        Returns True if rolled, False if not.
        """
        tsym = pos.get("tradingsymbol", "")
        entry = pos.get("entry_price", 0)
        symbol = pos.get("symbol", "")
        direction = pos.get("direction", "BULLISH")
        
        if not tsym or not symbol or not kite:
            return False
        
        try:
            # Check if profitable
            ltp_data = kite.ltp([f"NFO:{tsym}"])
            ltp = ltp_data.get(f"NFO:{tsym}", {}).get("last_price", 0)
            if ltp <= 0 or ltp <= entry:
                return False  # Not profitable — don't roll, just exit
            
            profit_pct = (ltp - entry) / entry
            if profit_pct < 0.10:
                return False  # Less than 10% profit — not worth rolling
            
            # Find next month option at similar strike
            # PATCH_V73: use index-aware spot key
            _sk = _index_spot_key(symbol)
            spot_data = kite.ltp([_sk])
            spot = spot_data.get(_sk, {}).get("last_price", 0)
            if spot <= 0:
                return False
            
            next_option = self._find_best_option(kite, symbol, direction, spot)
            if not next_option:
                return False
            
            # Check cost fits capital
            roll_cost = next_option["premium"] * next_option["lot_size"]
            if roll_cost > self.available + (ltp * pos.get("qty", 0)):
                return False
            
            # Step 1: Sell current position
            qty = pos.get("qty", 0)
            sell_id = kite.place_order(
                variety=kite.VARIETY_REGULAR, exchange="NFO",
                tradingsymbol=tsym, transaction_type="SELL",
                quantity=qty, order_type="MARKET", product="NRML",
                validity="DAY", market_protection=5,
            )
            log.info(f"FNO ROLL: sold {tsym} x{qty} for roll — awaiting fill confirmation")
            OrderGuard.record_order(tsym)
            
            # FIX_V53_ROLL: Poll for SELL fill COMPLETE before buying next month.
            # Prior bug: 3-sec sleep then buy regardless. If sell failed (rate limit,
            # circuit, reject), bot placed BUY anyway → doubled exposure.
            _sell_status = None
            _sell_elapsed = 0.0
            _sell_poll_max = 12
            _sell_interval = 1.5
            while _sell_elapsed < _sell_poll_max:
                time.sleep(_sell_interval)
                _sell_elapsed += _sell_interval
                try:
                    _orders = kite.orders()
                    for _o in _orders:
                        if str(_o.get("order_id")) == str(sell_id):
                            _sell_status = _o.get("status")
                            break
                except Exception as _oe:
                    log.warning(f"FNO ROLL: sell poll error: {_oe}")
                    continue
                if _sell_status == "COMPLETE":
                    log.info(f"FNO ROLL: sell {tsym} CONFIRMED after {_sell_elapsed:.1f}s [FIX_V53_ROLL]")
                    break
                if _sell_status in ("REJECTED", "CANCELLED"):
                    log.error(f"FNO ROLL: sell {tsym} status={_sell_status} — ABORTING roll to prevent doubled exposure [FIX_V53_ROLL]")
                    try:
                        health.fire(f"FNO ROLL ABORTED: sell {tsym} {_sell_status}")
                    except Exception as _e72:
                        log.warning(f"FNO ROLL: telegram fire failed: {_e72} [PATCH_V72_COMPREHENSIVE]")
                    return False
            
            if _sell_status != "COMPLETE":
                log.error(f"FNO ROLL: sell {tsym} did not fill in {_sell_poll_max}s (status={_sell_status}) — ABORTING roll [FIX_V53_ROLL]")
                try:
                    health.fire(f"FNO ROLL ABORTED: sell {tsym} timeout")
                except Exception:
                    pass
                return False
            
            # Step 2: Buy next month (only runs if sell COMPLETE)
            new_tsym = next_option["tradingsymbol"]
            new_lot = next_option["lot_size"]
            new_premium = next_option["premium"]
            
            try:
                buy_id = kite.place_order(
                    variety=kite.VARIETY_REGULAR, exchange="NFO",
                    tradingsymbol=new_tsym, transaction_type="BUY",
                    quantity=new_lot, order_type="LIMIT",
                    price=round(new_premium * 1.005, 1),
                    product="NRML", validity="DAY", market_protection=5,
                )
                # PATCH_V70_AUDIT_CLEANUP: validate buy_id (was captured but discarded)
                if not buy_id:
                    raise RuntimeError(f"place_order returned empty order_id for {new_tsym}")
                log.info(f"FNO ROLL: bought {new_tsym} x{new_lot} @ {new_premium} order_id={buy_id}")
                OrderGuard.record_order(new_tsym)
                OrderGuard.record_order(symbol)
            except Exception as _be:
                log.error(f"FNO ROLL: buy {new_tsym} failed after successful sell: {_be} — position closed but not rolled [FIX_V53_ROLL]")
                try:
                    health.fire(f"FNO ROLL: BUY leg failed {new_tsym}: {_be}")
                except Exception:
                    pass
                # Sell succeeded — remove old position so we don't audit it anymore
                self.positions.pop(key, None)
                self._save_positions()
                return False
            
            # Update position
            del self.positions[key]
            self.positions[new_tsym] = {
                "tradingsymbol": new_tsym,
                "symbol": symbol,
                "qty": new_lot,
                "entry_price": new_premium,
                "entry_time": now_ist().isoformat(),
                "cost": new_lot * new_premium,
                "peak_premium": new_premium,
                "sl_price": _snap_tick(new_premium * (1 - FNO_SL_PCT)),
                "expiry": str(next_option["expiry"]),
                "direction": direction,
                "catalyst": f"ROLLED from {tsym}",
            }
            self._save_positions()
            
            self.telegram.send(
                f"F&O ROLLED\n"
                f"Sold: {tsym} (profit {profit_pct:.0%})\n"
                f"Bought: {new_tsym} DTE={next_option['dte']}\n"
                f"Reason: DTE low, position profitable — rolled to next month",
                force=True,
            )
            return True
            
        except Exception as e:
            log.error(f"FNO roll {tsym} failed: {e}")
            return False

    def check_dte_exits(self):
        """DTE cascade — exit or roll before expiry kills premium."""
        today = today_ist()
        kite = KiteSession.kite()
        
        for key in list(self.positions.keys()):
            pos = self.positions[key]
            expiry_str = pos.get("expiry", "")
            try:
                expiry = datetime.strptime(expiry_str, "%Y-%m-%d").date()
            except Exception:
                continue

            dte = (expiry - today).days

            # Physical delivery risk — stock options must exit 2 days before expiry
            is_stock_option = not any(idx in pos.get("tradingsymbol", "") for idx in ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"])
            if is_stock_option and dte <= 2:
                log.info(f"FNO: {pos['tradingsymbol']} DTE={dte} PHYSICAL DELIVERY RISK — exit now")
                self._exit_position(key, f"PHYSICAL_DELIVERY_DTE_{dte}")
                continue

            if dte <= FNO_EMERGENCY_DTE:
                # Try to roll if profitable, otherwise exit
                if kite and self._try_roll_position(key, pos, kite):
                    log.info(f"FNO: {pos.get('tradingsymbol','')} DTE={dte} — ROLLED to next month")
                    continue
                log.info(f"FNO: {pos['tradingsymbol']} DTE={dte} EMERGENCY EXIT (roll failed)")
                self._exit_position(key, f"DTE_EMERGENCY_{dte}")
            elif dte <= FNO_DTE_CUT_LOSERS:
                # Cut if losing
                entry = pos.get("entry_price", 0)
                if kite:
                    ltp_data = kite.ltp([f"NFO:{pos['tradingsymbol']}"])
                    ltp = ltp_data.get(f"NFO:{pos['tradingsymbol']}", {}).get("last_price", 0)
                    if ltp > 0 and ltp < entry:
                        log.info(f"FNO: {pos['tradingsymbol']} DTE={dte} losing, cutting")
                        self._exit_position(key, f"DTE_CUT_LOSER_{dte}")

    def _exit_position(self, key, reason):
        """Exit an F&O position."""
        pos = self.positions.get(key)
        if not pos:
            return

        kite = KiteSession.kite()
        if not kite:
            return

        try:
            tsym = pos.get("tradingsymbol", "")
            qty = pos.get("qty", 0)
            if qty <= 0 or not tsym:
                return

            order_id = kite.place_order(
                variety=kite.VARIETY_REGULAR,
                exchange="NFO",
                tradingsymbol=tsym,
                transaction_type="SELL",
                quantity=qty,
                order_type="MARKET",
                product="NRML",
                validity="DAY",
                market_protection=5,  # SEBI mandatory
            )
            log.info(f"FNO: SELL {tsym} x{qty} reason={reason}")

            # V85_EXIT_POLL: poll up to 30s for COMPLETE status
            exit_price = 0
            _v85_status = "UNKNOWN"
            for _v85_attempt in range(15):
                time.sleep(2)
                try:
                    _v85_orders = kite.orders()
                    for _v85_o in _v85_orders:
                        if str(_v85_o.get("order_id")) == str(order_id):
                            _v85_status = _v85_o.get("status", "UNKNOWN")
                            if _v85_status == "COMPLETE":
                                exit_price = _v85_o.get("average_price", 0)
                            break
                except Exception as _v85_e:
                    log.warning(f"[V85_EXIT_POLL] {tsym}: kite.orders() failed: {_v85_e}")
                if _v85_status == "COMPLETE" and exit_price > 0:
                    log.info(f"[V85_EXIT_POLL] {tsym}: filled @ Rs.{exit_price} after {(_v85_attempt+1)*2}s")
                    break
                if _v85_status in ("REJECTED", "CANCELLED"):
                    log.warning(f"[V85_EXIT_POLL] {tsym}: {_v85_status} after {(_v85_attempt+1)*2}s")
                    break
            else:
                log.warning(f"[V85_EXIT_POLL] {tsym}: no fill after 30s, last_status={_v85_status}")

            entry = pos.get("entry_price", 0)
            pnl = (exit_price - entry) * qty if exit_price > 0 else 0

            # Record trade
            trade = {
                "tradingsymbol": tsym,
                "entry_price": entry,
                "exit_price": exit_price,
                "qty": qty,
                "pnl": round(pnl, 2),
                "reason": reason,
                "entry_time": pos.get("entry_time", ""),
                "exit_time": now_ist().isoformat(),
                "source": pos.get("source", "UNKNOWN"),
                "symbol": pos.get("symbol", ""),
                "direction": pos.get("direction", ""),
            }
            trades = load_json(self._trades_file, {"trades": []})
            trades["trades"].append(trade)
            save_json(self._trades_file, trades)

            if self.risk:
                self.risk.record_trade(pnl)

            del self.positions[key]
            self._save_positions()

            self.telegram.send(
                f"F&O SELL {tsym}\n"
                f"P&L: Rs.{pnl:,.0f} | Reason: {reason}",
                force=True,
            )

        except Exception as e:
            log.error(f"FNO: exit {key} failed: {e}")

    def scan_and_trade(self):
        """
        Main F&O scan.
        ALL sources: news, filings, brokerage, macro, earnings calendar.
        Direction: BULLISH → CE, BEARISH → PE.
        NO regime filter. Catalyst score is the only gate.
        PATCH_V62_ECON_CALENDAR: hard-block fresh entries within 2h before / 1h after HIGH events.
        """
        if not self.risk or not self.risk.can_trade():
            return

        # PATCH_V62_ECON_CALENDAR: block fresh F&O entries near macro events
        if getattr(self, "econ", None) and self.econ.is_blocked():
            log.info(f"[PATCH_V62_ECON] F&O entries blocked: {self.econ.reason()}")
            # Still allow trailing/exit logic — this only blocks NEW entries
            return

        # FIX_V39_SCOUT_VALIDATION: scouts -> validator -> trade list
        scout_candidates = []

        if self.news:
            scout_candidates.extend(self.news.scan())
        if self.filings:
            scout_candidates.extend(self.filings.scan())
        if self.brokerage:
            scout_candidates.extend(self.brokerage.scan())
        if self.macro:
            scout_candidates.extend(self.macro.scan())
        # PATCH_V63_INDEX_ROUTING: add index macro candidates
        if getattr(self, "index_macro", None):
            try:
                scout_candidates.extend(self.index_macro.scan())
            except Exception as _ie:
                log.error(f"[PATCH_V63] index_macro.scan failed: {_ie}")

        # PATCH_V68_LLM_FIRST_DECISION: split V68 article-items from legacy scouts
        _v68_articles_fno = [c for c in scout_candidates if c.get("v68_article_item")]
        _legacy_fno = [c for c in scout_candidates if not c.get("v68_article_item")]
        scout_candidates = _legacy_fno
        _v68_decided_fno = []
        if _v68_articles_fno and getattr(self, "news_brain", None) and hasattr(self.news_brain, "read_and_decide_parallel"):
            _pairs_fno = [
                ({"title": a.get("title",""), "summary": a.get("summary",""),
                  "link": a.get("link",""), "source_feed": a.get("source_feed", a.get("source",""))},
                 a.get("hints", []),
                 a.get("source", "MACRO"))
                for a in _v68_articles_fno
            ]
            try:
                _v68_decided_fno = self.news_brain.read_and_decide_parallel(_pairs_fno, max_workers=5)
                log.info(f"[PATCH_V68] fno LLM-decide: {len(_v68_articles_fno)} articles -> {len(_v68_decided_fno)} candidates")
            except Exception as _ve:
                log.error(f"[PATCH_V68] fno read_and_decide_parallel failed: {_ve}")

        candidates = []
        # PATCH_V68: prepend V68 LLM-decided candidates
        for _vc in _v68_decided_fno:
            candidates.append(_vc)
        _use_v3 = os.getenv("USE_V3_VALIDATION", "True").lower() in ("true", "1", "yes")
        if _use_v3 and getattr(self, "news_brain", None) and hasattr(self.news_brain, "validate_scout_candidate"):
            _validated = 0
            _rejected = 0
            _corrected = 0
            for sc in scout_candidates:
                news_item = {
                    "title": sc.get("title", ""),
                    "summary": sc.get("summary", ""),
                    "link": sc.get("link", ""),
                    "source_feed": sc.get("source_feed", sc.get("source", "")),
                }
                try:
                    validated = self.news_brain.validate_scout_candidate(sc, news_item)
                except Exception as _ve:
                    log.error(f"FIX_V39: validator exception for {sc.get('symbol','?')}: {_ve}")
                    validated = []
                if not validated:
                    _rejected += 1
                    continue
                for v in validated:
                    if v.get("corrected"):
                        _corrected += 1
                    candidates.append(v)
                    _validated += 1
            log.info(f"FIX_V39: scouts={len(scout_candidates)} -> validated={_validated}, rejected={_rejected}, corrected={_corrected}")
        else:
            log.warning(f"FIX_V39: v3 validation UNAVAILABLE (use_v3={_use_v3}) - fallback to raw scouts")
            candidates = scout_candidates

        # ── V27 Stage 2: LLM news brain (dry-run by default) ──
        # Runs alongside keyword path. In dry-run, only logs decisions.
        # In live mode (LLM_DRY_RUN=False), candidates are added to the trade list.
        if getattr(self, "news_brain", None):
            try:
                # Re-scan news as raw items so NewsBrain can analyze full text
                _llm_news_items = []
                if self.news:
                    # Use existing scanner output — already deduplicated & filtered
                    _llm_news_items = self.news.scan() or []
                _llm_cands = self.news_brain.analyze_news_batch(_llm_news_items) if _llm_news_items else []
                if _llm_cands:
                    log.info(f"[LLM] NewsBrain returned {len(_llm_cands)} candidates (DRY_RUN={LLM_DRY_RUN})")
                    for _lc in _llm_cands[:5]:  # top 5 only to avoid spam
                        log.info(f"[LLM] {_lc['symbol']} {_lc['direction']} score={_lc['score']} reasoning={_lc.get('reasoning','')[:120]}")
                    # Telegram notify on every LLM decision in dry-run
                    if LLM_DRY_RUN:
                        try:
                            _msg = "[LLM DRY-RUN] NewsBrain candidates:\n"
                            for _lc in _llm_cands[:5]:
                                _msg += f"- {_lc['symbol']} {_lc['direction']} ({_lc['score']}): {_lc.get('reasoning','')[:80]}\n"
                            self.telegram.send(_msg, force=True)
                        except Exception as _te:
                            log.warning(f"[LLM] telegram notify failed: {_te}")
                    else:
                        # LIVE mode: merge into candidates list
                        candidates.extend(_llm_cands)
                        log.info(f"[LLM] LIVE mode: {len(_llm_cands)} candidates added to trade list")
                else:
                    log.debug("[LLM] NewsBrain returned 0 candidates this scan")
            except Exception as _le:
                log.error(f"[LLM] NewsBrain scan failed: {_le}")

        # PATCH_V5_OI_LOGGER: Source 6 — OI buildup signals
        if self._oi_engine:
            try:
                oi_candidates = self._oi_engine.detect_buildup_signals()
                if oi_candidates:
                    candidates.extend(oi_candidates)
                    log.info(f"[PATCH_V5_OI] {len(oi_candidates)} OI buildup candidates added to F&O scan")
            except Exception as _e:
                log.error(f"[PATCH_V5_OI] signal injection failed: {_e}")

        # PATCH_V28_MOMENTUM_SCANNER: Source 7 — Price momentum signals
        # Scans all 218 F&O stocks for intraday moves >= 3% with volume >= 2x average.
        # Score 58 (below catalyst 70+) so momentum plays only enter when gates are clear.
        # Catches Trent-type non-news moves that keyword scanner misses.
        try:
            kite = KiteSession.kite()
            if kite and is_market_hours():
                _mom_candidates = []
                _fno_list = list(self.fno_stocks)
                _mom_checked = 0
                for _bi in range(0, min(len(_fno_list), 220), 20):
                    if _bi > 0: time.sleep(0.3)  # FIX_V51: Kite rate limit between batches
                    _batch = _fno_list[_bi:_bi+20]
                    _sym_keys = [f"NSE:{s}" for s in _batch]
                    try:
                        _quotes = kite.quote(_sym_keys)
                    except Exception:
                        continue
                    for _sym in _batch:
                        _q = _quotes.get(f"NSE:{_sym}", {})
                        if not _q:
                            continue
                        _mom_checked += 1
                        _ltp = _q.get("last_price", 0)
                        _ohlc = _q.get("ohlc", {})
                        _prev_close = _ohlc.get("close", 0)
                        _volume = _q.get("volume", 0)
                        _avg_volume = _q.get("average_volume", 0)
                        if _ltp <= 0 or _prev_close <= 0:
                            continue
                        _change_pct = ((_ltp - _prev_close) / _prev_close) * 100
                        _vol_ratio = (_volume / _avg_volume) if _avg_volume > 0 else 0
                        # Momentum threshold: >= 3% move with >= 2x volume
                        if abs(_change_pct) >= 3.0 and _vol_ratio >= 2.0:
                            _direction = "BULLISH" if _change_pct > 0 else "BEARISH"
                            _score = min(58 + int(abs(_change_pct)), 68)  # 58-68 range
                            _mom_candidates.append({
                                "symbol": _sym,
                                "direction": _direction,
                                "score": _score,
                                "catalyst": f"MOMENTUM: {_change_pct:+.1f}% vol={_vol_ratio:.1f}x",
                                "source": "MOMENTUM",
                                "timestamp": now_ist().isoformat(),
                            })
                if _mom_candidates:
                    candidates.extend(_mom_candidates)
                    _mom_syms = ", ".join(c["symbol"] for c in _mom_candidates[:5])
                    log.info(f"[PATCH_V28_MOMENTUM] {len(_mom_candidates)} momentum signals from {_mom_checked} stocks: {_mom_syms}")
                else:
                    log.debug(f"[PATCH_V28_MOMENTUM] 0 momentum signals from {_mom_checked} stocks")
        except Exception as _me:
            log.error(f"[PATCH_V28_MOMENTUM] scanner failed: {_me}")

        # Source 5: Earnings plays — pre-earnings IV expansion strategy
        # PATCH_V7_SIGNAL_QUALITY: inherit direction from other signals on same stock, skip if none
        # V81: also inherit direction from OI buildup engine — this is what makes the
        # 14-21d early entry path actually fire. At T-21 there usually isn't news yet,
        # but institutional OI positioning is visible. OI buildup is the smart-money signal.
        if self.earnings:
            plays = self.earnings.get_earnings_plays(self.fno_stocks)
            # Build direction map from current candidates (news/macro/filing/brokerage already scanned)
            _existing_dirs = {}
            for _c in candidates:
                _sym = _c.get("symbol")
                _d = _c.get("direction")
                if _sym and _d in ("BULLISH", "BEARISH"):
                    if _sym not in _existing_dirs:
                        _existing_dirs[_sym] = []
                    _existing_dirs[_sym].append(_d)
            # V81: merge OI buildup directions into the same map
            try:
                if hasattr(self, "_oi_engine") and self._oi_engine:
                    _oi_sigs = self._oi_engine.detect_buildup_signals(mode="daily") or []
                    for _o in _oi_sigs:
                        _osym = _o.get("symbol"); _od = _o.get("direction")
                        if _osym and _od in ("BULLISH", "BEARISH"):
                            _existing_dirs.setdefault(_osym, []).append(_od)
                    if _oi_sigs:
                        log.info(f"[V81_EARNINGS_OI] merged {len(_oi_sigs)} OI buildup directions into earnings direction map")
            except Exception as _v81e:
                log.warning(f"[V81_EARNINGS_OI] OI direction merge failed: {_v81e}")
            for play in plays:
                sym = play["symbol"]
                if play["action"] == "BUY_IV_PLAY":
                    # Look up real direction from other signals
                    real_dirs = _existing_dirs.get(sym, [])
                    if not real_dirs:
                        log.info(f"[PATCH_V7_SQ] earnings {sym} skipped: no directional signal found")
                        continue
                    # If conflicting, skip (contradiction check handles it anyway)
                    if "BULLISH" in real_dirs and "BEARISH" in real_dirs:
                        log.info(f"[PATCH_V7_SQ] earnings {sym} skipped: conflicting signals")
                        continue
                    real_direction = real_dirs[0]
                    candidates.append({
                        "symbol": sym,
                        "direction": real_direction,
                        "score": play["score"],
                        "catalyst": f"EARNINGS+{real_direction}: {play['catalyst']}",
                        "source": "EARNINGS_PLAY",
                        "timestamp": now_ist().isoformat(),
                    })
                    log.info(f"[PATCH_V7_SQ] FNO earnings play: {sym} in {play['days_to']}d direction={real_direction} (inherited)")
                elif play["action"] == "BUY_IF_STRONG":
                    real_dirs = _existing_dirs.get(sym, [])
                    if not real_dirs or ("BULLISH" in real_dirs and "BEARISH" in real_dirs):
                        continue
                    candidates.append({
                        "symbol": sym,
                        "direction": real_dirs[0],
                        "score": play["score"],
                        "catalyst": f"EARNINGS+{real_dirs[0]}: {play['catalyst']}",
                        "source": "EARNINGS_MODERATE",
                        "timestamp": now_ist().isoformat(),
                    })
                elif play["action"] == "EXIT_IV_CRUSH":
                    # V84_IV_CRUSH_SMART_EXIT: conditional exit on result day.
                    _v84_kite = KiteSession.kite()
                    _v84_now = now_ist()
                    _v84_force_close_after = _v84_now.replace(hour=14, minute=0, second=0, microsecond=0)
                    for key, pos in list(self.positions.items()):
                        if play["symbol"] not in pos.get("tradingsymbol", ""):
                            continue
                        _v84_tsym = pos.get("tradingsymbol", "")
                        _v84_underlying = pos.get("symbol", play["symbol"])
                        _v84_entry = pos.get("entry_price", 0)
                        _v84_direction = pos.get("direction", "")
                        _v84_days_to = play["days_to"]
                        if _v84_now >= _v84_force_close_after:
                            log.warning(f"[V84_IV_CRUSH] {_v84_tsym}: past 14:00 IST, force-exit (theta protection)")
                            self._exit_position(key, f"IV_CRUSH_{_v84_days_to}d_before_results")
                            continue
                        if _v84_days_to != 0:
                            log.warning(f"FNO IV CRUSH: {_v84_tsym} — {play['symbol']} results in {_v84_days_to}d, EXITING")
                            self._exit_position(key, f"IV_CRUSH_{_v84_days_to}d_before_results")
                            continue
                        try:
                            if not _v84_kite or _v84_entry <= 0:
                                raise RuntimeError("no kite or no entry price")
                            _v84_oq = _v84_kite.ltp([f"NFO:{_v84_tsym}"])
                            _v84_opt_ltp = _v84_oq.get(f"NFO:{_v84_tsym}", {}).get("last_price", 0)
                            if _v84_opt_ltp <= 0:
                                raise RuntimeError("no option ltp")
                            _v84_premium_pct = (_v84_opt_ltp - _v84_entry) / _v84_entry * 100
                            _v84_uk = _index_spot_key(_v84_underlying)
                            _v84_uq = _v84_kite.quote([_v84_uk]).get(_v84_uk, {})
                            _v84_uspot = _v84_uq.get("last_price", 0)
                            _v84_uprev = _v84_uq.get("ohlc", {}).get("close", 0)
                            _v84_umove = ((_v84_uspot - _v84_uprev) / _v84_uprev * 100) if _v84_uprev > 0 else 0
                            _v84_dir_ok = (_v84_direction == "BULLISH" and _v84_umove >= 0.5) or (_v84_direction == "BEARISH" and _v84_umove <= -0.5)
                            if _v84_premium_pct >= 30 and _v84_dir_ok:
                                _v84_new_sl = round(_v84_opt_ltp * 0.85, 1)
                                _v84_old_sl = pos.get("sl_price", 0)
                                if _v84_new_sl > _v84_old_sl:
                                    pos["sl_price"] = _v84_new_sl
                                    self._save_positions()
                                    try:
                                        self.gtt.ensure_gtt(_v84_tsym, "NFO", pos.get("qty", 0), _v84_new_sl)
                                    except Exception as _v84ge:
                                        log.warning(f"[V84_IV_CRUSH] {_v84_tsym}: GTT update failed: {_v84ge}")
                                log.info(f"[V84_IV_CRUSH] HOLD {_v84_tsym}: prem +{_v84_premium_pct:.1f}%, {_v84_underlying} {_v84_umove:+.2f}% (dir {_v84_direction}) — trail SL to {_v84_new_sl}")
                            else:
                                log.warning(f"[V84_IV_CRUSH] EXIT {_v84_tsym}: prem {_v84_premium_pct:+.1f}%, {_v84_underlying} {_v84_umove:+.2f}% (need >=30% AND dir-confirm)")
                                self._exit_position(key, f"IV_CRUSH_{_v84_days_to}d_before_results")
                        except Exception as _v84e:
                            log.warning(f"[V84_IV_CRUSH] {_v84_tsym}: check failed ({_v84e}), defaulting to immediate exit")
                            self._exit_position(key, f"IV_CRUSH_{_v84_days_to}d_before_results")

        # PATCH_V63_INDEX_ROUTING: Index trading is handled by IndexMacroDetector
        # (initialized in init(), scanned via scout_candidates pipeline above).
        # The IndexMacroDetector emits NIFTY/BANKNIFTY/FINNIFTY candidates with proper
        # NewsBrain validation. No separate index logic needed here.
        # (Old broken V62 dual-direction blind scan removed; new directional logic
        # via IndexMacroDetector goes through the full scout → validation pipeline.)

        # V81: Source 6 (duplicate get_upcoming earnings boost) REMOVED.
        # It generated BULLISH-default candidates that conflicted with Source 5
        # (which uses real direction signals). Source 5 above is the correct path.

        if not candidates:
            return

        # PATCH_V7_SIGNAL_QUALITY: filter stale candidates (older than 4 hours)
        try:
            _cutoff = now_ist() - timedelta(hours=4)
            _fresh_candidates = []
            for c in candidates:
                ts = c.get("timestamp")
                if not ts:
                    _fresh_candidates.append(c)  # No timestamp = assume fresh (macro/filings)
                    continue
                try:
                    c_ts = datetime.fromisoformat(ts)
                    if c_ts.replace(tzinfo=None) >= _cutoff.replace(tzinfo=None):
                        _fresh_candidates.append(c)
                    else:
                        log.debug(f"[PATCH_V7_SQ] stale candidate skipped: {c.get('symbol')} ts={ts}")
                except Exception:
                    _fresh_candidates.append(c)
            candidates = _fresh_candidates
        except Exception as _e:
            log.warning(f"[PATCH_V7_SQ] time decay filter failed: {_e}")

        # PATCH_V28_CONTRADICTION_TIEBREAK: if same symbol has both bull and bear,
        # keep the stronger signal if score gap >= 10pts, veto both only if close.
        _scores_by_sym = {}
        for c in candidates:
            sym = c.get("symbol")
            d = c.get("direction")
            s = c.get("score", 0)
            if sym and d in ("BULLISH", "BEARISH"):
                if sym not in _scores_by_sym:
                    _scores_by_sym[sym] = {}
                if d not in _scores_by_sym[sym] or s > _scores_by_sym[sym][d]:
                    _scores_by_sym[sym][d] = s
        _conflicting = set()
        _tiebreak_winners = {}
        for sym, dir_scores in _scores_by_sym.items():
            if len(dir_scores) >= 2:
                bull_s = dir_scores.get("BULLISH", 0)
                bear_s = dir_scores.get("BEARISH", 0)
                gap = abs(bull_s - bear_s)
                if gap >= 10:
                    winner = "BULLISH" if bull_s > bear_s else "BEARISH"
                    _tiebreak_winners[sym] = winner
                    log.info(f"[PATCH_V28_TIEBREAK] {sym} contradiction resolved: {winner} wins (B={bull_s} vs R={bear_s}, gap={gap})")
                else:
                    _conflicting.add(sym)
                    log.info(f"[PATCH_V28_TIEBREAK] {sym} contradiction too close, vetoing both (B={bull_s} vs R={bear_s}, gap={gap})")
        # Remove vetoed symbols entirely, keep only winning direction for tiebreak symbols
        _filtered = []
        for c in candidates:
            sym = c.get("symbol")
            if sym in _conflicting:
                continue  # vetoed
            if sym in _tiebreak_winners:
                if c.get("direction") == _tiebreak_winners[sym]:
                    _filtered.append(c)
                # else: losing direction dropped
            else:
                _filtered.append(c)
        candidates = _filtered

        # PATCH_V94_FNO: bulk/block deal scout feeds F&O brain too.
        # Superstar/FII/DII buying a stock => BULLISH for that stock's CE.
        try:
            _v94_now = now_ist()
            if (_v94_now.hour, _v94_now.minute) >= (17, 30):
                _v94_bb = EquitySmartScanner.scan_bulk_block_deals(self.all_symbols if hasattr(self, "all_symbols") else set())
                for _vc in _v94_bb:
                    candidates.append(_vc)
                if _v94_bb:
                    log.info(f"[V94_FNO] +{len(_v94_bb)} bulk/block candidates fed to F&O brain")
        except Exception as _v94fe:
            log.debug(f"[V94_FNO] bulk/block feed error: {_v94fe}")

        # PATCH_V28_SOURCE_WEIGHT: boost scores based on source reliability
        # Filing/Brokerage = highest conviction, News/Momentum = lowest
        _SOURCE_WEIGHTS = {
            "FILING": 1.15,        # NSE filings — hard data, highest conviction
            "BROKERAGE": 1.12,     # Analyst upgrades/downgrades — professional research
            "PROMOTER_BUY": 1.10,  # Insider buying — strong signal
            "MACRO": 1.05,         # RBI/Fed/GDP — affects entire market
            "52WK_LOW": 1.0,       # Technical — no boost
            "BREAKOUT": 1.0,       # Technical — no boost
            "SECTOR_ROTATION": 1.0,# Sector play — no boost
            "NEWS": 0.95,          # RSS headlines — noisy, slight penalty
            "MOMENTUM": 1.00,      # Price momentum (V93: 0.90->1.00, V85+V83_HOURLY already filter noise)
            "SUPERSTAR_BUY": 1.20, # V94: superstar investor bulk/block buy — highest predictive signal
            "FII_BLOCK_BUY": 1.15, # V94: FII institutional bulk/block buy
            "DII_BLOCK_BUY": 1.10, # V94: DII (MF/insurance) bulk/block buy
        }
        for c in candidates:
            _src = c.get("source", "")
            _w = _SOURCE_WEIGHTS.get(_src, 1.0)
            if _w != 1.0:
                _old_score = c.get("score", 0)
                c["score"] = min(95, int(_old_score * _w))
                if c["score"] != _old_score:
                    log.debug(f"[PATCH_V28_SOURCE_WEIGHT] {c.get('symbol')} {_src}: {_old_score} -> {c['score']} (x{_w})")

        # Deduplicate by symbol (keep highest score)
        best_by_sym = {}
        for c in candidates:
            sym = c["symbol"]
            if sym not in best_by_sym or c.get("score", 0) > best_by_sym[sym].get("score", 0):
                best_by_sym[sym] = c
        candidates = list(best_by_sym.values())

        # Filter: score >= minimum catalyst score
        # V78_FIXB1: explicit index allow-list (belt-and-suspenders with V78_FIXA1)
        qualified = [c for c in candidates
                     if c.get("score", 0) >= FNO_MIN_CATALYST
                     and (c.get("symbol") in self.fno_stocks
                          or c.get("symbol") in {"NIFTY", "BANKNIFTY", "FINNIFTY"})
                     and c.get("direction") in ("BULLISH", "BEARISH")]

        # V70 IV CRUSH GUARD: block entries on stocks with results in 0-3 days
        if self.earnings:
            _blocked = []
            _safe = []
            from datetime import datetime as _dt
            for c in qualified:
                _sym = c.get("symbol", "")
                try:
                    _info = self.earnings._calendar.get(_sym, {})
                    _date_str = _info.get("date", "")
                    _earn_date = None
                    for _fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                        try:
                            _earn_date = _dt.strptime(_date_str, _fmt).date()
                            break
                        except ValueError:
                            continue
                    if _earn_date is not None:
                        _days_to = (_earn_date - today_ist()).days
                        if 0 <= _days_to <= 3:
                            _blocked.append(f"{_sym}({_days_to}d)")
                            continue
                except Exception:
                    pass
                _safe.append(c)
            if _blocked:
                log.warning(f"[V70_IV_CRUSH_GUARD] blocked {len(_blocked)} entries within 3d of results: {', '.join(_blocked)}")
            qualified = _safe

        qualified.sort(key=lambda c: c.get("score", 0), reverse=True)

        entered = 0
        for cand in qualified[:5]:
            self._pending_source = cand.get("source", "UNKNOWN")
            try:
                if not hasattr(self, "_expert_panel"):
                    self._expert_panel = ExpertPanel(getattr(self, "trendlyne", None), getattr(self, "_sector_rotation", None), getattr(self, "news_brain", None))
                kite_p = KiteSession.kite()
                opt_data = cand.get("option_data", {})
                pr = self._expert_panel.evaluate_fno(kite_p, cand["symbol"], cand["direction"], cand, opt_data)
                log.info(ExpertPanel.format_fno_log(pr))
                if not pr["allow"]:
                    continue
            except Exception as _pe:
                log.error(f"[FIX_V46_PANEL_FNO] {cand.get('symbol')} error: {_pe}")
                continue  # FIX_V51: panel crash = BLOCK, never fall through to enter
            success = self.enter(cand["symbol"], cand.get("catalyst", ""), cand["direction"], cand.get("score", 0))
            if success:
                entered += 1
                if entered >= 2:
                    break

    def tick(self):
        """Called every minute from main loop."""
        n = now_ist()

        # Morning sync at 08:01 (v26.4: skip audit to preserve trailed SLs)
        if n.hour == 8 and n.minute == 1:
            self._refresh_capital()
            self._sync_positions()
            # self._fno_gtt_audit()  # DISABLED — was wiping trailed SLs daily

        # PATCH_V6_GAP: gap detection + AMO review at 08:50 IST
        if n.hour == 8 and n.minute == 50 and is_trading_day():
            if self._gap_detector and self._last_gap_check_day != today_ist():
                self._last_gap_check_day = today_ist()
                try:
                    gap_status = self._gap_detector.detect_gaps()
                    if gap_status:
                        self._gap_detector.review_pending_amos(gap_status)
                except Exception as _e:
                    log.error(f"[PATCH_V6_GAP] morning routine failed: {_e}")

        # FIX_V44_GTT_AT_FILL_WINDOW: widened from minute==17 to 17-30 window
        # Original bug: scan_and_trade at 09:15 blocks main loop with Gemini calls,
        # causing tick() to skip minute 17 entirely. Now retries every minute in
        # the window until it runs successfully once per day.
        if n.hour == 9 and 17 <= n.minute <= 30:
            if not getattr(self, "_v28_amo_fill_synced_today", False):
                self._v28_amo_fill_synced_today = True
                try:
                    _before = len(self.positions)
                    self._sync_positions()
                    _after = len(self.positions)
                    if _after > _before:
                        log.info(f"[FIX_V44_GTT_AT_FILL] synced {_after - _before} new positions from AMO fills at {n.hour:02d}:{n.minute:02d}")
                    self._fno_gtt_audit()
                    log.info(f"[FIX_V44_GTT_AT_FILL] GTT audit complete after AMO fill sync at {n.hour:02d}:{n.minute:02d}")
                except Exception as _e:
                    log.error(f"[FIX_V44_GTT_AT_FILL] failed at {n.hour:02d}:{n.minute:02d}: {_e}")
                    self._v28_amo_fill_synced_today = False

        # FIX_V44_MIDNIGHT_RESET: reset daily flag at 00:01 IST so next morning triggers fresh
        if n.hour == 0 and n.minute == 1:
            self._v28_amo_fill_synced_today = False

        # SL re-place at 09:16 (v26.4: skip audit to preserve trailed SLs)
        if n.hour == 9 and n.minute == 16:
            self._replace_sl_orders()
            # self._fno_gtt_audit()  # DISABLED — was wiping trailed SLs daily

        # Scan every 15 minutes during market hours
        # PATCH_V69_RESTART_SCAN: also fire on first tick after a mid-market restart.
        _fno_first_tick = (self._last_scan is None)
        if is_market_hours() and (n.minute % FNO_SCAN_INTERVAL_MIN == 0 or _fno_first_tick):
            if self._last_scan != n.hour * 60 + n.minute:
                self._last_scan = n.hour * 60 + n.minute
                _trig = "first-tick-after-restart" if _fno_first_tick else f"interval={FNO_SCAN_INTERVAL_MIN}m"
                log.info(f"[PATCH_V69_RESTART_SCAN] fno scan fired at {n.hour:02d}:{n.minute:02d} ({_trig})")
                self.scan_and_trade()

        # PATCH_V72_COMPREHENSIVE: AMO hooks moved from main loop into tick().
        # Internal window guards (_fno_amo_in_scan_window, _fno_amo_is_premarket_place_time)
        # already gate when these actually run. Calling them here keeps tick() self-contained.
        try:
            # PATCH_V54: dynamic threshold based on V62 overnight bias
            _v54_thr = 65
            try:
                if getattr(self, "premarket", None):
                    _v54_b = self.premarket.read()
                    if _v54_b:
                        _v54_s = _v54_b.get("score", 0.0)
                        if _v54_s <= -0.4: _v54_thr = 72
                        elif _v54_s >= 0.4: _v54_thr = 60
            except Exception as _v54e:
                log.debug(f"[PATCH_V54] threshold scaling failed: {_v54e}")
            self._fno_amo_continuous_scan(threshold=_v54_thr)
        except Exception as _e72:
            log.error(f"[PATCH_V72_COMPREHENSIVE] AMO continuous scan hook failed: {_e72}")
        try:
            self._fno_amo_premarket_place()
        except Exception as _e72:
            log.error(f"[PATCH_V72_COMPREHENSIVE] AMO premarket place hook failed: {_e72}")
        # PATCH_V55_FNO_AMO_STALE_CANCEL: cancel unfilled F&O AMOs at 09:20-09:30 IST
        try:
            self._fno_amo_cancel_stale()
        except Exception as _e55:
            log.error(f"[PATCH_V55_FNO_STALE] cancel_stale hook failed: {_e55}")

        # Trailing SL every 5 minutes
        if is_market_hours() and n.minute % 5 == 0:
            if self._last_trail_check != n.hour * 60 + n.minute:
                self._last_trail_check = n.hour * 60 + n.minute
                self.check_trailing_sl()

        # FIX_V76_OI_INTRADAY: removed n.minute<5 gate.
        # OLD: required minute 0-4 AND new hour. Single-tick miss = lost hour.
        # NEW: hour-throttle alone (first scan in a new hour triggers save).
        if is_market_hours() and self._oi_engine:
            if self._last_oi_snapshot_hour != n.hour:
                self._last_oi_snapshot_hour = n.hour
                try:
                    self._oi_engine.take_snapshot()
                    log.info(f"[FIX_V76_OI_INTRADAY] hourly snapshot at {n.strftime('%H:%M')} IST")
                except Exception as _e:
                    log.error(f"[PATCH_V5_OI] hourly snapshot failed: {_e}")

        # PATCH_V9_AMO_ARCH: EOD OI snapshot at 15:25 IST (once per trading day)
        if self._oi_engine and n.hour == 15 and n.minute == 25 and is_trading_day():
            if getattr(self, "_last_eod_snapshot_day", None) != today_ist():
                self._last_eod_snapshot_day = today_ist()
                try:
                    self._oi_engine.take_snapshot(is_eod=True)
                except Exception as _e:
                    log.error(f"[PATCH_V9_AMO_ARCH] EOD snapshot failed: {_e}")

        # PATCH_V8_SECTOR_ROTATION: refresh sector ranking every 15 minutes
        if is_market_hours() and self._sector_rotation and n.minute % SECTOR_ROTATION_REFRESH_MIN == 0:
            _sr_key = n.hour * 60 + n.minute
            if self._sector_rotation._last_refresh_minute != _sr_key:
                self._sector_rotation._last_refresh_minute = _sr_key
                try:
                    self._sector_rotation.refresh()
                except Exception as _e:
                    log.error(f"[PATCH_V8_SECTOR] refresh failed: {_e}")

        # DTE check at 09:30 and 14:00
        if n.hour in (9, 14) and n.minute == 30:
            self.check_dte_exits()

        # FIX_V29_MIDDAY_AUDIT: midday GTT safety check at 12:30
        if n.hour == 12 and n.minute == 30:
            self._fno_gtt_audit()
        # FIX_V29_FNO_GTT_HOURLY: run F&O GTT audit every hour during market hours
        if n.hour in (9, 10, 11, 13, 14, 15) and n.minute == 30:
            self._fno_gtt_audit()

        # EOD report at 15:36
        if n.hour == 15 and n.minute == 36 and is_trading_day():
            report = EODReporter.generate_fno_report(self)
            self.telegram.send(report, force=True)

        # Update earnings calendar at 07:00 and 16:00
        if n.hour in (7, 16) and n.minute == 0:
            try:
                self.earnings.update_from_filings()
                plays = self.earnings.get_earnings_plays(self.fno_stocks)
                if plays:
                    buy_plays = [p for p in plays if p["action"] == "BUY_IV_PLAY"]
                    exit_plays = [p for p in plays if p["action"] == "EXIT_IV_CRUSH"]
                    if buy_plays or exit_plays:
                        msg = "EARNINGS CALENDAR UPDATE\n"
                        for p in buy_plays[:5]:
                            msg += f"BUY: {p['symbol']} results in {p['days_to']}d\n"
                        for p in exit_plays[:5]:
                            msg += f"EXIT: {p['symbol']} results in {p['days_to']}d (IV crush)\n"
                        self.telegram.send(msg, force=True)
            except Exception:
                pass

    # ================================================================
    # ================================================================
    # ================================================================
    # PATCH_V64_STUCK_RECOVERY: cascade stuck SELL limit orders to best bid
    # When GTT triggers but LIMIT doesnt fill (TATASTEEL Apr 24):
    #   - Detect SELL order, 0 filled, >30s old
    #   - Check market depth best bid
    #   - If best bid within 5% of LTP -> modify order to best bid (walk book)
    #   - If best bid outside 5% -> hold, dont chase freak crashes
    # ================================================================
    def _fno_stuck_order_recovery(self):
        """Walk the book on stuck SELL limit orders. Runs inside audit worker."""
        kite = KiteSession.kite()
        if not kite:
            return
        if not hasattr(self, "_stuck_mod_times"):
            self._stuck_mod_times = {}
        try:
            orders = kite.orders() or []
        except Exception as _oe:
            log.debug(f"[PATCH_V64] kite.orders() failed: {_oe}")
            return
        try:
            from datetime import datetime as _dt
            _now = time.time()
            _held = set()
            for _k, _p in (self.positions or {}).items():
                _ts = _p.get("tradingsymbol") or _p.get("symbol") or ""
                if _ts:
                    _held.add(_ts)
            for o in orders:
                try:
                    if o.get("exchange") != "NFO":
                        continue
                    if o.get("transaction_type") != "SELL":
                        continue
                    if o.get("order_type") != "LIMIT":
                        continue
                    status = (o.get("status") or "").upper()
                    if status not in ("OPEN", "TRIGGER PENDING"):
                        continue
                    if int(o.get("filled_quantity", 0)) > 0:
                        continue
                    tsym = o.get("tradingsymbol", "")
                    if tsym not in _held:
                        continue
                    oid = str(o.get("order_id", ""))
                    if not oid:
                        continue
                    _ots_str = o.get("order_timestamp", "")
                    if not _ots_str:
                        continue
                    try:
                        if isinstance(_ots_str, str):
                            _ots = _dt.strptime(_ots_str[:19], "%Y-%m-%d %H:%M:%S").timestamp()
                        else:
                            _ots = _ots_str.timestamp()
                    except Exception:
                        continue
                    _age = _now - _ots
                    if _age < 30:
                        continue
                    _last_mod = self._stuck_mod_times.get(oid, 0)
                    if _now - _last_mod < 30:
                        continue
                    try:
                        _qd = kite.quote([f"NFO:{tsym}"]).get(f"NFO:{tsym}", {})
                    except Exception as _qe:
                        log.warning(f"[PATCH_V64] quote failed for {tsym}: {_qe}")
                        continue
                    _ltp = float(_qd.get("last_price", 0) or 0)
                    _depth = _qd.get("depth", {}) or {}
                    _buys = _depth.get("buy", []) or []
                    if not _buys or _ltp <= 0:
                        continue
                    _best_bid = float(_buys[0].get("price", 0) or 0)
                    if _best_bid <= 0:
                        continue
                    _cur_price = float(o.get("price", 0) or 0)
                    _floor = _ltp * 0.95
                    if _best_bid < _floor:
                        log.warning(f"[PATCH_V64] {tsym} oid={oid} best_bid={_best_bid} < LTP*0.95={_floor:.4f} - NOT chasing (freak crash protection)")
                        self._stuck_mod_times[oid] = _now
                        continue
                    if abs(_cur_price - _best_bid) < 1e-6:
                        continue
                    try:
                        kite.modify_order(
                            variety=o.get("variety", "regular"),
                            order_id=oid,
                            price=_best_bid,
                        )
                        self._stuck_mod_times[oid] = _now
                        log.info(f"[PATCH_V64] CASCADED {tsym} oid={oid} age={_age:.0f}s: {_cur_price} -> {_best_bid} (LTP={_ltp}, floor={_floor:.4f})")
                    except Exception as _me:
                        log.error(f"[PATCH_V64] modify_order failed for {tsym} oid={oid}: {_me}")
                        self._stuck_mod_times[oid] = _now
                except Exception as _ie:
                    log.error(f"[PATCH_V64] inner loop error: {_ie}")
        except Exception as _oe2:
            log.error(f"[PATCH_V64] outer error: {_oe2}")

    # ================================================================
    # PATCH_V61_FNO_AUDIT_THREAD: background worker
    # ================================================================
    def _fno_audit_worker(self):
        """Runs _fno_gtt_audit() every 10s. Survives all exceptions.
        Only audits during/around market hours to skip useless overnight calls."""
        # PATCH_V70_AUDIT_CLEANUP: removed unused 'import time as _time'
        log.info("[PATCH_V61] F&O audit worker thread STARTED")
        _cycle = 0
        while not self._fno_audit_stop.is_set():
            try:
                _n = now_ist()
                _hm = _n.hour * 60 + _n.minute
                _active = (525 <= _hm <= 945) and is_trading_day()  # 08:45 to 15:45 IST
                if _active and self.positions:
                    if self._fno_audit_lock.acquire(blocking=False):
                        try:
                            self._fno_gtt_audit()
                            self._fno_stuck_order_recovery()  # PATCH_V64
                            _cycle += 1
                            if _cycle % 30 == 0:
                                log.info(f"[PATCH_V61] worker heartbeat cycle={_cycle} pos={len(self.positions)}")
                        finally:
                            self._fno_audit_lock.release()
                    else:
                        log.debug("[PATCH_V61] worker skip: lock held by event path")
            except Exception as _we:
                log.error(f"[PATCH_V61] worker exception (continuing): {_we}")
            self._fno_audit_stop.wait(10)
        log.info("[PATCH_V61] F&O audit worker thread STOPPED")

    def _fno_audit_thread_start(self):
        """Spawn worker thread if not already running. Idempotent."""
        if self._fno_audit_thread is not None and self._fno_audit_thread.is_alive():
            log.warning("[PATCH_V61] audit thread already running, skip spawn")
            return
        self._fno_audit_stop.clear()
        self._fno_audit_thread = threading.Thread(
            target=self._fno_audit_worker,
            name="FnoAuditWorker",
            daemon=True,
        )
        self._fno_audit_thread.start()
        log.info(f"[PATCH_V61] audit thread spawned: {self._fno_audit_thread.name}")

    def _fno_gtt_audit(self):
        """FIX_V32_FNO_UNIFIED: single source of truth for F&O GTT writes.
        
        ONE formula (shared with check_trailing_sl via this function):
            stage_pct = 0.20 if peak_gain >= 25% else 0.30
            formula_sl = LTP * (1 - stage_pct)
        
        FLOOR RULE (solves live + restart regression):
            final_sl = max(formula_sl, existing_gtt_trigger, position_sl_price)
        
        existing_gtt_trigger = live Kite GTT trigger (survives restart, honors
                               manual fixes).
        position_sl_price   = local backup in self.positions.
        
        SAFETY CAP: final_sl must be < LTP (never above market).
        
        OCO PRESERVATION: if existing GTT is 'two-leg', modify keeps the target
        leg unchanged and only raises stop. No forced conversion to SINGLE.
        
        API FAILURE SAFETY: if kite.positions() or kite.get_gtts() fails,
        abort immediately. Never mass-delete GTTs on empty/failed response.
        """
        kite = KiteSession.kite()
        if not kite:
            return
        try:
            # Step 1: Fetch positions (abort on failure — prevents orphan mass-delete)
            try:
                positions = kite.positions().get("net", [])
            except Exception as _pe:
                log.error(f"GTT FNO audit: positions() failed: {_pe} — ABORTING")
                return
            
            fno_positions = {}
            for p in positions:
                if p.get("exchange") == "NFO" and p.get("quantity", 0) > 0:
                    tsym = p["tradingsymbol"]
                    fno_positions[tsym] = {
                        "qty": p["quantity"],
                        "avg": p.get("average_price", 0),
                    }
            
            # Step 2: Fetch GTTs (abort on failure)
            try:
                all_gtts = kite.get_gtts() or []
            except Exception as _ge:
                log.error(f"GTT FNO audit: get_gtts() failed: {_ge} — ABORTING")
                return
            
            fno_gtts = defaultdict(list)
            for g in all_gtts:
                if g.get("status") == "active":
                    sym = g.get("condition", {}).get("tradingsymbol", "")
                    exch = g.get("condition", {}).get("exchange", "")
                    if exch == "NFO" and sym:
                        fno_gtts[sym].append(g)
            
            placed, modified, deleted = 0, 0, 0
            
            # Step 3: Delete duplicates — keep HIGHEST trigger
            def _trig_of(g):
                tv = g.get("condition", {}).get("trigger_values") or [0]
                try:
                    return float(tv[0])
                except Exception:
                    return 0.0
            
            for sym, gtts in list(fno_gtts.items()):
                if len(gtts) > 1:
                    gtts.sort(key=_trig_of, reverse=True)
                    keeper = gtts[0]
                    for dup in gtts[1:]:
                        try:
                            kite.delete_gtt(dup["id"])
                            deleted += 1
                            log.warning(f"GTT FNO: deleted dup {sym} id={dup['id']} trig={_trig_of(dup)} (kept id={keeper['id']} trig={_trig_of(keeper)})")
                        except Exception as _e71:
                            log.warning(f"GTT FNO: failed to delete dup {sym} id={dup['id']}: {_e71} [PATCH_V71_RESILIENCE]")
                    fno_gtts[sym] = [keeper]
            
            # Step 4: Delete orphans
            for sym in list(fno_gtts.keys()):
                if sym not in fno_positions:
                    for g in fno_gtts[sym]:
                        try:
                            kite.delete_gtt(g["id"])
                            deleted += 1
                            log.info(f"GTT FNO: deleted orphan {sym} id={g['id']}")
                        except Exception as _e71:
                            log.warning(f"GTT FNO: failed to delete orphan {sym} id={g['id']}: {_e71} [PATCH_V71_RESILIENCE]")
                    fno_gtts.pop(sym, None)
            
            if not fno_positions:
                if placed or modified or deleted:
                    log.info(f"GTT FNO AUDIT: placed={placed} modified={modified} deleted={deleted}")
                return
            
            # Load peaks (for stage detection only)
            try:
                peaks_map = load_json("/home/globalbot/fno_peaks.json", {}) or {}
                if not isinstance(peaks_map, dict):
                    peaks_map = {}
            except Exception:
                peaks_map = {}
            
            # Batch LTP fetch
            try:
                ltp_batch = kite.ltp([f"NFO:{s}" for s in fno_positions])
            except Exception:
                ltp_batch = {}
            
            # Step 5: For each position — compute SL via ONE formula, apply floor, act
            for tsym, info in fno_positions.items():
                qty = info["qty"]
                avg = info["avg"]
                if avg <= 0 or qty <= 0:
                    continue
                
                ltp = ltp_batch.get(f"NFO:{tsym}", {}).get("last_price", 0)
                if ltp <= 0:
                    ltp = avg
                
                # ── THE ONE FORMULA ──
                peak_info = peaks_map.get(tsym) if isinstance(peaks_map.get(tsym), dict) else {}
                peak = float(peak_info.get("peak", 0) or 0) if peak_info else 0.0
                peak_gain = ((peak - avg) / avg) if (peak > 0 and avg > 0) else 0.0
                # FIX_V45_BREAKEVEN: Stage 0 locks entry (breakeven) once peak reaches +10%
                # Prevents "almost winners" from becoming losers on reversal.
                # Peak >=25% -> 20% trail (Stage 1)
                # Peak 10-25% -> max(entry, LTP*0.7) (Stage 0 BREAKEVEN LOCK)
                # Peak <10% -> LTP * 0.7 (Initial 30% stop)
                # PATCH_V67_TIGHT_TRAIL: tighter SL trail at higher gains - locks more profit on big winners
                # Peak >=100% -> 5% trail | 50-100% -> 8% | 25-50% -> 12% | 10-25% breakeven | <10% INIT 30%
                if peak_gain >= 1.00:
                    stage_pct = 0.05
                    formula_sl = _snap_sl_tick_up(ltp * (1 - stage_pct), symbol=tsym, exchange="NFO")
                    stage_label = "S3_LOCK_95"
                elif peak_gain >= 0.50:
                    stage_pct = 0.08
                    formula_sl = _snap_sl_tick_up(ltp * (1 - stage_pct), symbol=tsym, exchange="NFO")
                    stage_label = "S2_LOCK_92"
                elif peak_gain >= 0.25:
                    stage_pct = 0.12
                    formula_sl = _snap_sl_tick_up(ltp * (1 - stage_pct), symbol=tsym, exchange="NFO")
                    stage_label = "S1_TRAIL"
                elif peak_gain >= 0.10:
                    # Stage 0: lock breakeven
                    stage_pct = 0.30
                    raw_sl = ltp * (1 - stage_pct)
                    breakeven_sl = avg  # entry price
                    formula_sl = _snap_sl_tick_up(max(raw_sl, breakeven_sl), symbol=tsym, exchange="NFO")
                    stage_label = "S0_BREAKEVEN"
                else:
                    stage_pct = 0.30
                    formula_sl = _snap_sl_tick_up(ltp * (1 - stage_pct), symbol=tsym, exchange="NFO")
                    stage_label = "INIT"
                # ───────────────────────────────────────
                
                # ── FLOOR RULE ──
                existing_gtts = fno_gtts.get(tsym, [])
                old_sl = 0.0
                existing_type = None
                existing_gtt_obj = None
                if existing_gtts:
                    existing_gtt_obj = existing_gtts[0]
                    old_sl = _trig_of(existing_gtt_obj)
                    existing_type = existing_gtt_obj.get("type")  # 'single' or 'two-leg'
                
                pos_sl = 0.0
                try:
                    pos_sl = float(self.positions.get(tsym, {}).get("sl_price", 0) or 0)
                except Exception:
                    pos_sl = 0.0
                
                final_sl = max(formula_sl, old_sl, _snap_sl_tick_up(pos_sl, symbol=tsym, exchange="NFO") if pos_sl > 0 else 0.0)
                
                # ── SAFETY CAP (SL must be below LTP) ──
                if ltp > 0 and final_sl >= ltp:
                    # PATCH_V69_PROFIT_LOCK: if pos_sl already locks profit, don't widen down. Let SL fire.
                    if pos_sl > 0 and pos_sl >= avg:
                        log.info(f"GTT FNO: [V69_LOCK] {tsym} pos_sl={pos_sl} locks profit (avg={avg}) — SL stays, fires on touch")
                        final_sl = pos_sl
                    else:
                        capped = _snap_sl_tick_up(ltp * 0.85, symbol=tsym, exchange="NFO")
                        log.warning(f"GTT FNO: [CAP] {tsym} final_sl={final_sl} >= ltp={ltp}, capping to {capped}")
                        final_sl = capped
                
                if final_sl <= 0:
                    log.warning(f"GTT FNO: {tsym} final_sl={final_sl} invalid, skip")
                    continue
                
                # FIX_V53_GAP: enforce Kite's 0.25% gap rule (floor rule preserved via max below)
                _safe_sl = _gtt_safe_trigger(final_sl, ltp, side="SELL", symbol=tsym, exchange="NFO")
                if _safe_sl <= 0:
                    log.warning(f"GTT FNO: {tsym} safe_sl invalid after gap check (final={final_sl} ltp={ltp}), skip [FIX_V53_GAP]")
                    continue
                if abs(_safe_sl - final_sl) > 0.01:
                    log.warning(f"GTT FNO: {tsym} SL adjusted {final_sl}→{_safe_sl} for gap rule [FIX_V53_GAP]")
                # Preserve no-regress: final is max(gap-adjusted, existing trigger) — never lower than old_sl
                final_sl = max(_safe_sl, old_sl) if old_sl > 0 else _safe_sl
                
                log.info(f"GTT FNO: {tsym} ltp={ltp} avg={avg} peak={peak} gain={peak_gain*100:.1f}% stage={stage_label} formula={formula_sl} old_sl={old_sl} pos_sl={pos_sl} → final_sl={final_sl} type={existing_type}")
                
                # ── ACTION ──
                if not existing_gtt_obj:
                    # PATCH_V66_RETRY_COOLDOWN: skip if 3+ failures in last 60s
                    if not hasattr(self, "_gtt_fail_log"):
                        self._gtt_fail_log = {}
                    _now = time.time()
                    _fails = [t for t in self._gtt_fail_log.get(tsym, []) if _now - t < 60]
                    self._gtt_fail_log[tsym] = _fails
                    if len(_fails) >= 3:
                        log.warning(f"[PATCH_V66_RETRY_COOLDOWN] {tsym} failed {len(_fails)}x in 60s, skipping place this cycle")
                        continue
                    try:
                        gtt_id = _safe_place_gtt(kite, 
                            trigger_type=kite.GTT_TYPE_SINGLE,
                            tradingsymbol=tsym, exchange="NFO",
                            trigger_values=[final_sl], last_price=ltp,
                            orders=[{"transaction_type": "SELL", "quantity": qty,
                                     "price": _snap_tick(final_sl * (1 - FNO_GTT_LIMIT_OFFSET), symbol=tsym, exchange="NFO"),
                                     "order_type": "LIMIT", "product": "NRML"}],
                        )
                        placed += 1
                        log.info(f"GTT FNO: placed {tsym} SL={final_sl} qty={qty} id={gtt_id}")
                        if tsym in self.positions:
                            self.positions[tsym]["sl_price"] = final_sl
                        self._gtt_fail_log.pop(tsym, None)
                    except Exception as e:
                        log.error(f"GTT FNO: place {tsym} failed: {e}")
                        self._gtt_fail_log.setdefault(tsym, []).append(_now)
                
                elif final_sl > old_sl:
                    # Modify up — preserve existing type (SINGLE or OCO)
                    if existing_type == "two-leg":
                        # OCO: keep target leg, only raise stop trigger
                        existing_triggers = existing_gtt_obj.get("condition", {}).get("trigger_values") or [old_sl, old_sl]
                        target_trig = existing_triggers[1] if len(existing_triggers) >= 2 else (ltp * 1.1)
                        existing_orders = existing_gtt_obj.get("orders") or []
                        
                        new_orders = []
                        if len(existing_orders) >= 2:
                            # Rebuild both legs, update stop qty/price, keep target untouched
                            stop_order = dict(existing_orders[0])
                            stop_order["quantity"] = qty
                            stop_order["price"] = _snap_tick(final_sl * (1 - FNO_GTT_LIMIT_OFFSET), symbol=tsym, exchange="NFO")
                            stop_order["order_type"] = "LIMIT"
                            stop_order["transaction_type"] = "SELL"
                            new_orders.append(stop_order)
                            target_order = dict(existing_orders[1])
                            target_order["quantity"] = qty
                            target_order["transaction_type"] = "SELL"
                            new_orders.append(target_order)
                        else:
                            new_orders = [
                                {"transaction_type": "SELL", "quantity": qty,
                                 "price": _snap_tick(final_sl * (1 - FNO_GTT_LIMIT_OFFSET), symbol=tsym, exchange="NFO"),
                                 "order_type": "LIMIT", "product": "NRML"},
                                {"transaction_type": "SELL", "quantity": qty,
                                 "price": _snap_tick(target_trig),
                                 "order_type": "LIMIT", "product": "NRML"},
                            ]
                        
                        try:
                            _safe_modify_gtt(kite, 
                                trigger_id=existing_gtt_obj["id"],
                                trigger_type=kite.GTT_TYPE_OCO,
                                tradingsymbol=tsym, exchange="NFO",
                                trigger_values=[final_sl, _snap_tick(target_trig)],
                                last_price=ltp,
                                orders=new_orders,
                            )
                            modified += 1
                            log.info(f"GTT FNO: modified OCO {tsym} stop {old_sl}→{final_sl} (target {target_trig} preserved)")
                            if tsym in self.positions:
                                self.positions[tsym]["sl_price"] = final_sl
                        except Exception as e1:
                            log.error(f"GTT FNO: OCO modify {tsym} failed: {e1} — leaving existing intact")
                    else:
                        # SINGLE modify
                        try:
                            _safe_modify_gtt(kite, 
                                trigger_id=existing_gtt_obj["id"],
                                trigger_type=kite.GTT_TYPE_SINGLE,
                                tradingsymbol=tsym, exchange="NFO",
                                trigger_values=[final_sl], last_price=ltp,
                                orders=[{"transaction_type": "SELL", "quantity": qty,
                                         "price": _snap_tick(final_sl * (1 - FNO_GTT_LIMIT_OFFSET), symbol=tsym, exchange="NFO"),
                                         "order_type": "LIMIT", "product": "NRML"}],
                            )
                            modified += 1
                            log.info(f"GTT FNO: modified SINGLE {tsym} SL {old_sl}→{final_sl}")
                            if tsym in self.positions:
                                self.positions[tsym]["sl_price"] = final_sl
                        except Exception as e1:
                            # Delete + re-place as SINGLE
                            try:
                                kite.delete_gtt(existing_gtt_obj["id"])
                                gtt_id = _safe_place_gtt(kite, 
                                    trigger_type=kite.GTT_TYPE_SINGLE,
                                    tradingsymbol=tsym, exchange="NFO",
                                    trigger_values=[final_sl], last_price=ltp,
                                    orders=[{"transaction_type": "SELL", "quantity": qty,
                                             "price": _snap_tick(final_sl * (1 - FNO_GTT_LIMIT_OFFSET), symbol=tsym, exchange="NFO"),
                                             "order_type": "LIMIT", "product": "NRML"}],
                                )
                                placed += 1
                                log.info(f"GTT FNO: re-placed {tsym} (modify failed: {e1}) SL={final_sl} id={gtt_id}")
                                if tsym in self.positions:
                                    self.positions[tsym]["sl_price"] = final_sl
                            except Exception as e2:
                                log.error(f"GTT FNO: re-place {tsym} failed: {e2}")
                # else: final_sl <= old_sl → leave alone (never regress)
            
            try:
                self._save_positions()
            except Exception:
                pass
            
            if placed or modified or deleted:
                log.info(f"GTT FNO AUDIT: placed={placed} modified={modified} deleted={deleted}")
        
        except Exception as e:
            log.error(f"GTT FNO audit error: {e}")

    def _replace_sl_orders(self):
        """PATCH_V12_SL_DEDUP: DISABLED. GTTManager (check_trailing_sl every 5 min) handles all SL.

        The legacy _replace_sl_orders ran daily at 09:16 IST and placed regular SL orders
        on top of existing GTTs without cancelling stale ones, causing duplicate stop-losses
        on the same position (e.g., BPCL had SL at 1.20 from old run AND GTT at 5.12 from
        trailing). GTTManager.ensure_gtt() already handles dedup, in-loss floor, and
        upward-only trailing. Running both systems in parallel created the duplicate-SL bug.

        Trusting GTT-only matches the standing rule: GTT is the most critical safety feature.
        """
        log.info("[PATCH_V12_SL_DEDUP] FNO _replace_sl_orders skipped — GTTManager handles SL via check_trailing_sl every 5 min")
        return


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN LOOP
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    log.info(f"{'='*60}")
    log.info(f"  {VERSION} — Starting")
    log.info(f"  Server: <server-ip> | Kite: {ZERODHA_USER_ID}")
    log.info("  ZERO paper trading | Real orders only")
    log.info(f"{'='*60}")

    # Load env
    env_path = Path("/home/globalbot/.env")
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())
        # Reload constants from env
        globals().update({
            "KITE_API_KEY": os.getenv("KITE_API_KEY", ""),   # V28 A4: reload from env
            "KITE_API_SECRET": os.getenv("KITE_API_SECRET", ""),
            "KITE_TOTP_SECRET": os.getenv("KITE_TOTP_SECRET", ""),
            "ZERODHA_PASSWORD": os.getenv("ZERODHA_PASSWORD", ""),
            "TELEGRAM_TOKEN": os.getenv("TELEGRAM_TOKEN", ""),
            "TELEGRAM_CHAT_ID": os.getenv("TELEGRAM_CHAT_ID", ""),
        })

    # V28 A4: validate KITE_API_KEY is present after env load
    if not globals().get("KITE_API_KEY"):
        log.error("FATAL: KITE_API_KEY missing from /home/globalbot/.env — add it before restart")
        sys.exit(1)

    # Login to Kite
    if not KiteSession.login():
        log.error("FATAL: Kite login failed. Retrying in 60s...")
        time.sleep(60)
        if not KiteSession.login():
            log.error("FATAL: Kite login failed twice. Exiting.")
            sys.exit(1)

    # Initialize modules — completely independent
    equity = EquityModule()
    fno = FnoModule()

    eq_ok = equity.init()
    fno_ok = fno.init()

    if not eq_ok and not fno_ok:
        log.error("FATAL: both modules failed to init")
        sys.exit(1)

    send_telegram(
        f"*{VERSION} STARTED*\n"
        f"Equity: {'OK' if eq_ok else 'FAILED'} | Rs.{equity.capital:,.0f}\n"
        f"F&O: {'OK' if fno_ok else 'FAILED'} | Rs.{fno.capital:,.0f}\n"
        f"Stocks: {len(equity.all_symbols)} | F&O: {len(fno.fno_stocks)}\n"
        f"Positions: EQ={len(equity.positions)} FNO={len(fno.positions)}\n"
        f"Mode: LIVE — real orders only",
    )

    # Main loop — runs every 60 seconds
    log.info("Main loop started")
    last_minute = -1
    reconnect_count = 0

    while True:
        try:
            n = now_ist()
            current_minute = n.hour * 60 + n.minute

            # Only process once per minute
            if current_minute == last_minute:
                time.sleep(5)
                continue
            last_minute = current_minute

            # Ensure Kite connected
            if not KiteSession.ensure_connected():
                reconnect_count += 1
                try:
                    health.on_reconnect_fail(reconnect_count)
                except Exception:
                    pass
                if reconnect_count > 5:
                    log.error("Kite: too many reconnect failures, sleeping 5 min")
                    # PATCH_V73: poll lifetime flag (one-shot manual trigger)
                    try:
                        EODReporter.check_lifetime_flag(equity, fno)
                    except Exception:
                        pass
                    time.sleep(300)
                    reconnect_count = 0
                continue
            try:
                health.on_reconnect_ok()
            except Exception:
                pass
            reconnect_count = 0

            # Tick both modules independently
            if eq_ok:
                try:
                    equity.tick()
                except Exception as e:
                    log.error(f"Equity tick error: {e}\n{traceback.format_exc()}")

            if fno_ok:
                try:
                    # PATCH_V72_COMPREHENSIVE: AMO hooks now inside fno.tick() — removed from main loop
                    fno.tick()
                except Exception as e:
                    log.error(f"FNO tick error: {e}\n{traceback.format_exc()}")

            # Daily log rotation at midnight
            if n.hour == 0 and n.minute == 0:
                log.info("Midnight: daily reset")

            # Kite re-login at 08:00 on trading days
            if n.hour == 8 and n.minute == 0 and is_trading_day():
                KiteSession.login()

            # M1: IV snapshot at 15:26 on trading days (after OI engine at 15:25)
            if n.hour == 15 and n.minute == 26 and is_trading_day():
                try:
                    if fno_ok:
                        _kite = KiteSession.kite()
                        if _kite:
                            _ok, _fail = daily_iv_snapshot(_kite, fno)
                            log.info(f"[M1] IV snapshot complete: ok={_ok} fail={_fail}")
                        else:
                            log.warning("[M1] IV snapshot skipped: no Kite session")
                    else:
                        log.warning("[M1] IV snapshot skipped: FNO not initialized")
                except Exception as _e:
                    log.error(f"[M1] IV snapshot hook failed: {_e}")

            # M11: Daily summary at 15:35 on trading days
            if n.hour == 15 and n.minute == 35 and is_trading_day():
                try:
                    c = health.counters
                    _eq_pos = len(equity.positions) if eq_ok else "N/A"
                    _fno_pos = len(fno.positions) if fno_ok else "N/A"
                    _daily = (
                        f"*DAILY SUMMARY {n.strftime('%Y-%m-%d')}*\n"
                        f"Orders placed: {c['orders_placed']}\n"
                        f"Rejected: {c['orders_rejected']}\n"
                        f"Cancelled: {c['orders_cancelled']}\n"
                        f"Kite login fails: {c['kite_login_fails']}\n"
                        f"Trailing SL errors: {c['trailing_sl_errors']}\n"
                        f"Equity positions: {_eq_pos}\n"
                        f"FNO positions: {_fno_pos}"
                    )
                    send_telegram(_daily, silent=False)
                    log.info("M11 daily summary sent")
                    health.reset_daily_counters()
                except Exception as _e:
                    log.error(f"M11 daily summary failed: {_e}")

            # PATCH_V2_WEEKLY_ATTRIBUTION: Sunday 09:00 IST report
            if n.weekday() == 6 and n.hour == 9 and n.minute == 0:
                try:
                    _wk_report = EODReporter.generate_weekly_attribution(equity, fno)
                    # PATCH_V73: lifetime report after weekly attribution
                    try:
                        _life_rpt = EODReporter.generate_lifetime_report(equity, fno)
                        send_telegram(_life_rpt, silent=False)
                    except Exception as _le:
                        log.error(f'[PATCH_V73] weekly lifetime send failed: {_le}')
                    send_telegram(_wk_report, silent=False)
                    log.info("Weekly attribution report sent")
                except Exception as _e:
                    log.error(f"Weekly attribution failed: {_e}")
            # END_PATCH_V2_WEEKLY_ATTRIBUTION

            time.sleep(10)

        except KeyboardInterrupt:
            log.info("Shutdown requested")
            send_telegram(f"*{VERSION} STOPPED* — manual shutdown")
            break
        except Exception as e:
            log.error(f"Main loop error: {e}\n{traceback.format_exc()}")
            time.sleep(30)


if __name__ == "__main__":
    main()
